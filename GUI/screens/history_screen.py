"""History screen component"""

import flet as ft
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename

class HistoryScreen:
    def __init__(self, app):
        self.app = app
        self.page = app.page
    
# GUI/screens/history_screen.py

    def build(self):
        """Build history screen UI"""
    
        # Filters
        self.brand_filter = ft.Dropdown(
            label="Filter by Brand",
            width=150,
            options=[
                ft.dropdown.Option("All"),
                ft.dropdown.Option("KROSS"),
                ft.dropdown.Option("Pinarello"),
                ft.dropdown.Option("Basso"),
                ft.dropdown.Option("Factor"),
                ft.dropdown.Option("LeeCougan"),
                ft.dropdown.Option("Rascal"),
                ft.dropdown.Option("Rondo"),
                ft.dropdown.Option("Octane"),
            ],
            value="All",
            on_change=lambda e: self.refresh_history()
        )
    
        self.status_filter = ft.Dropdown(
            label="Filter by Status",
            width=150,
            options=[
                ft.dropdown.Option("All"),
                ft.dropdown.Option("Success"),
                ft.dropdown.Option("Failed"),
            ],
            value="All",
            on_change=lambda e: self.refresh_history()
        )
    
        self.date_filter = ft.Dropdown(
            label="Date Range",
            width=150,
            options=[
                ft.dropdown.Option("All Time"),
                ft.dropdown.Option("Today"),
                ft.dropdown.Option("Last 7 Days"),
                ft.dropdown.Option("Last 30 Days"),
            ],
            value="All Time",
            on_change=lambda e: self.refresh_history()
        )
    
        # Export button
        self.export_button = ft.ElevatedButton(
            "Export to Excel",
            icon=ft.Icons.TABLE_CHART,
            on_click=self.export_to_excel
        )
    
        # Refresh button (manual)
        self.refresh_button = ft.IconButton(
            icon=ft.Icons.REFRESH,
            tooltip="Refresh",
            on_click=lambda e: self.refresh_history()
        )
    
        # Statistics
        self.stats_container = ft.Container(
            content=ft.Text("Loading statistics...", size=14),  # Placeholder
            padding=15,
            bgcolor=ft.Colors.BLUE_50,
            border_radius=10
        )
    
        # History list container
        self.history_list = ft.Column([], scroll=ft.ScrollMode.AUTO, expand=True)
    
    
        # Layout
        self.screen_content = ft.Column(
            [
                ft.Text("Apdorojimo Istorija", size=24, weight=ft.FontWeight.BOLD),
                ft.Container(height=10),
            
                # Statistics
                self.stats_container,
                ft.Container(height=20),
            
                # Filters
                ft.Row(
                    [
                        self.brand_filter,
                        self.status_filter,
                        self.date_filter,
                        self.refresh_button,
                        self.export_button
                    ],
                    spacing=10
                ),
                ft.Container(height=10),
            
                # History list
                self.history_list
            ],
            expand=True
        )
        
        return self.screen_content
    
    def build_stats(self):
        """Build statistics summary"""
        cursor = self.app.db.conn.cursor()
        
        # Total uploads
        total = cursor.execute("SELECT COUNT(*) FROM processing_history").fetchone()[0]
        
        # Success rate
        success = cursor.execute(
            "SELECT COUNT(*) FROM processing_history WHERE status = 'success'"
        ).fetchone()[0]
        
        success_rate = (success / total * 100) if total > 0 else 0
        
        # Average duration
        avg_duration = cursor.execute(
            "SELECT AVG(duration_seconds) FROM processing_history WHERE status = 'success'"
        ).fetchone()[0] or 0
        
        # Today's uploads
        today = datetime.now().strftime('%Y-%m-%d')
        today_count = cursor.execute(
            "SELECT COUNT(*) FROM processing_history WHERE DATE(processed_at) = ?",
            (today,)
        ).fetchone()[0]
        
        return ft.Row(
            [
                self.stat_box("Total Uploads", str(total), ft.Icons.UPLOAD),
                self.stat_box("Success Rate", f"{success_rate:.1f}%", ft.Icons.CHECK_CIRCLE),
                self.stat_box("Avg Duration", f"{avg_duration:.1f}s", ft.Icons.TIMER),
                self.stat_box("Today", str(today_count), ft.Icons.TODAY),
            ],
            spacing=20
        )
    
    def stat_box(self, label, value, icon):
        """Create a statistics box"""
        return ft.Container(
            content=ft.Column(
                [
                    ft.Icon(icon, size=30, color=ft.Colors.BLUE_700),
                    ft.Text(value, size=24, weight=ft.FontWeight.BOLD),
                    ft.Text(label, size=12, color=ft.Colors.GREY_700)
                ],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                spacing=5
            ),
            padding=15,
            bgcolor=ft.Colors.WHITE,
            border_radius=8,
            expand=True
        )
    
    def refresh_history(self):
        """Refresh history list based on filters"""
        cursor = self.app.db.conn.cursor()
    
        # Build query with filters
        query = "SELECT brand, product_code, url_or_code, status, duration_seconds, features_uploaded, images_uploaded, error_message, failed_stage, processed_at FROM processing_history WHERE 1=1"
        params = []
    
        # Brand filter - CASE INSENSITIVE
        if self.brand_filter.value != "All":
            query += " AND UPPER(brand) = UPPER(?)"  # ← Changed this line
            params.append(self.brand_filter.value)
    
        # Status filter
        if self.status_filter.value != "All":
            status_value = "success" if self.status_filter.value == "Success" else "failed"
            query += " AND status = ?"
            params.append(status_value)
        
        # Date filter
        if self.date_filter.value != "All Time":
            if self.date_filter.value == "Today":
                date_threshold = datetime.now().strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) = ?"
                params.append(date_threshold)
            elif self.date_filter.value == "Last 7 Days":
                date_threshold = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) >= ?"
                params.append(date_threshold)
            elif self.date_filter.value == "Last 30 Days":
                date_threshold = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) >= ?"
                params.append(date_threshold)
        
        query += " ORDER BY processed_at DESC LIMIT 100"
        
        # Execute query
        history = cursor.execute(query, params).fetchall()
        
        # Clear existing list
        self.history_list.controls.clear()
        
        # Add history items
        if history:
            for row in history:
                self.history_list.controls.append(self.build_history_item(row))
        else:
            self.history_list.controls.append(
                ft.Text("Nerasta įrašų pagal pasirinktus filtrus", color=ft.Colors.GREY_600)
            )
        
        # Update stats
        self.stats_container.content = self.build_stats()
        
        # Update UI
        self.page.update()
    
    def build_history_item(self, row):
        """Build a single history item"""
        status_icon = "✓" if row['status'] == 'success' else "✗"
        status_color = ft.Colors.GREEN if row['status'] == 'success' else ft.Colors.RED
    
        # Format duration
        duration = f"{row['duration_seconds']:.1f}s" if row['duration_seconds'] else "N/A"
    
        # Format features/images counts
        features = str(row['features_uploaded']) if row['features_uploaded'] else "0"
        images = str(row['images_uploaded']) if row['images_uploaded'] else "0"
    
        # Format failed stage
        stage_info = ""
        if row['status'] == 'failed' and row['failed_stage']:
            stage_info = f" (Failed at: {row['failed_stage']})"
    
        return ft.Container(
            content=ft.Column(
                [
                    ft.Row(
                        [
                            ft.Text(status_icon, size=20, color=status_color, width=30),
                            ft.Column(
                                [
                                    ft.Text(
                                        f"{row['brand']} - {row['product_code']}",
                                        size=16,
                                        weight=ft.FontWeight.BOLD
                                    ),
                                    ft.Text(
                                        f"Duration: {duration} | Features: {features} | Images: {images}{stage_info}",
                                        size=12,
                                        color=ft.Colors.GREY_700
                                    ),
                                    ft.Text(
                                        f"{row['processed_at']}",
                                        size=11,
                                        color=ft.Colors.GREY_600
                                    )
                                ],
                                spacing=2
                            )
                        ]
                    ),
                    # Show error if failed
                    ft.Text(
                        f"Error: {row['error_message']}",
                        size=11,
                        color=ft.Colors.RED_700,
                        italic=True
                    ) if row['status'] == 'failed' and row['error_message'] else ft.Container()
                ],
                spacing=5
            ),
            padding=15,
            border=ft.border.only(bottom=ft.BorderSide(1, ft.Colors.GREY_300))
        )
    
    def export_to_excel(self, e):
        """Export history to Excel file"""
        try:
            # Get filtered data
            cursor = self.app.db.conn.cursor()
        
            query = "SELECT brand, product_code, url_or_code, status, duration_seconds, features_uploaded, images_uploaded, error_message, failed_stage, processed_at FROM processing_history WHERE 1=1"
            params = []
        
            # Apply same filters as display - CASE INSENSITIVE
            if self.brand_filter.value != "All":
                query += " AND UPPER(brand) = UPPER(?)"  # ← Changed this line
                params.append(self.brand_filter.value)
            
            if self.status_filter.value != "All":
                status_value = "success" if self.status_filter.value == "Success" else "failed"
                query += " AND status = ?"
                params.append(status_value)
            
            if self.date_filter.value != "All Time":
                if self.date_filter.value == "Today":
                    date_threshold = datetime.now().strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) = ?"
                    params.append(date_threshold)
                elif self.date_filter.value == "Last 7 Days":
                    date_threshold = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) >= ?"
                    params.append(date_threshold)
                elif self.date_filter.value == "Last 30 Days":
                    date_threshold = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) >= ?"
                    params.append(date_threshold)
            
            query += " ORDER BY processed_at DESC"
            
            history = cursor.execute(query, params).fetchall()
            
            if not history:
                self.show_export_status("No data to export", ft.Colors.ORANGE)
                return
            
            # Open save dialog
            root = Tk()
            root.withdraw()
            root.attributes('-topmost', True)
            
            filename = asksaveasfilename(
                title="Save Export As",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"ultrabike_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            root.destroy()
            
            if not filename:
                return  # User cancelled
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Processing History"
            
            # Headers
            headers = ["Brand", "Product Code", "URL/Code", "Status", "Duration (s)", "Features", "Images", "Error Message", "Processed At"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Data rows
            for row in history:
                ws.append([
                    row['brand'],
                    row['product_code'],
                    row['url_or_code'],
                    row['status'].capitalize(),
                    row['duration_seconds'],
                    row['features_uploaded'],
                    row['images_uploaded'],
                    row['error_message'] or "",
                    row['processed_at']
                ])
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Color-code status cells
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                status_cell = row[3]  # Status column
                if status_cell.value == "Success":
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    status_cell.font = Font(color="006100")
                elif status_cell.value == "Failed":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    status_cell.font = Font(color="9C0006")
            
            # Save
            wb.save(filename)
            
            self.show_export_status(f"✓ Exported to {filename}", ft.Colors.GREEN)
            
        except Exception as ex:
            self.show_export_status(f"✗ Export failed: {str(ex)}", ft.Colors.RED)
    
    def show_export_status(self, message, color):
        """Show export status message"""
        # Create temporary status text
        status_text = ft.Text(message, color=color, size=14)
        
        # Add to history list temporarily
        self.history_list.controls.insert(0, status_text)
        self.page.update()
        
        # Remove after 3 seconds
        import threading
        def clear():
            import time
            time.sleep(3)
            try:
                self.history_list.controls.remove(status_text)
                self.page.update()
            except:
                pass
        
        threading.Thread(target=clear, daemon=True).start()