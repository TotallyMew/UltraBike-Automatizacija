"""Batch upload screen - modal for uploading multiple bikes"""

import flet as ft
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from Utilities.BatchProcessor import BatchProcessor
from uploaderFactory import getUploaderClass


class BatchUploadScreen:
    def __init__(self, app, on_close):
        self.app = app
        self.page = app.page
        self.on_close_callback = on_close
        self.batch_processor = None
        self.is_processing = False

        from Managers.DescriptionManager import DescriptionManager

        self.desc_manager = DescriptionManager(
            app.db,
            app.logger if hasattr(app, "logger") else None
        )

        self.available_descriptions = [
            ft.dropdown.Option(d["name"])
            for d in self.desc_manager.list_descriptions()
        ]

    def build(self):
        """Build batch upload modal"""

        self.tab_bar = ft.Tabs(
            selected_index=0,
            animation_duration=200,
            tabs=[
                ft.Tab(text="Manual Entry", icon=ft.Icons.EDIT),
                ft.Tab(text="Excel Upload", icon=ft.Icons.UPLOAD_FILE),
            ],
            on_change=self.handle_tab_change,
            expand=False,
        )

        self.manual_content = self.build_manual_entry()
        self.excel_content = self.build_excel_upload()

        self.content_area = ft.Container(
            content=self.manual_content,
            padding=8,
            expand=True,
        )

        self.start_button = ft.ElevatedButton(
            "Start Batch Upload",
            icon=ft.Icons.PLAY_ARROW,
            on_click=self.handle_start_batch,
            disabled=True,
        )

        self.cancel_button = ft.TextButton(
            "Cancel",
            on_click=lambda e: self.on_close_callback(),
        )

        return ft.AlertDialog(
            modal=True,
            content=ft.Container(
                width=1100,
                height=650,
                content=ft.Column(
                    [
                        ft.Text("Batch Upload", size=20, weight=ft.FontWeight.BOLD),
                        ft.Container(self.tab_bar, padding=ft.padding.only(bottom=4)),
                        ft.Divider(height=1),
                        self.content_area,
                    ],
                    expand=True,
                    spacing=6,
                ),
            ),
            actions=[self.cancel_button, self.start_button],
            actions_alignment=ft.MainAxisAlignment.END,
        )

    def handle_tab_change(self, e):
        if e.control.selected_index == 0:
            self.content_area.content = self.manual_content
        else:
            self.content_area.content = self.excel_content
        self.page.update()

    # -------------------- MANUAL ENTRY --------------------

    def build_manual_entry(self):
        # Header with select-all checkboxes
        frameset_select_all = ft.Checkbox(
            tooltip="Select/Deselect all frameset",
            value=False,
            on_change=lambda e: self.toggle_all_framesets(e.control.value)
        )

        disclaimer_select_all = ft.Checkbox(
            tooltip="Select/Deselect all disclaimers",
            value=False,
            on_change=lambda e: self.toggle_all_disclaimers(e.control.value)
        )

        header = ft.Row(
            [
                ft.Text("Brand", width=120, weight=ft.FontWeight.BOLD),
                ft.Text("Product Code", width=120, weight=ft.FontWeight.BOLD),
                ft.Text("URL or Code", width=280, weight=ft.FontWeight.BOLD),
                ft.Text("Description", width=200, weight=ft.FontWeight.BOLD),
                ft.Column([
                    ft.Text("Frameset", weight=ft.FontWeight.BOLD, size=12, tooltip="Pinarello: Only frameset"),
                    frameset_select_all
                ], width=70, spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Column([
                    ft.Text("Disc", weight=ft.FontWeight.BOLD, size=12, tooltip="Append disclaimer"),
                    disclaimer_select_all
                ], width=50, spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Text("", width=32),
            ],
            spacing=6,
        )

        self.manual_rows = ft.Column(
            [],
            spacing=4,
            expand=True,
            scroll=ft.ScrollMode.AUTO,
        )

        for _ in range(10):
            self.add_manual_row()

        add_button = ft.ElevatedButton(
            "Add Row",
            icon=ft.Icons.ADD,
            on_click=lambda e: self.add_manual_row(),
        )

        clear_button = ft.TextButton(
            "Clear All",
            icon=ft.Icons.CLEAR,
            on_click=lambda e: self.clear_manual_rows(),
        )

        return ft.Column(
            [
                header,
                ft.Divider(height=1),
                ft.Container(
                    content=self.manual_rows,
                    expand=True,
                    padding=ft.padding.only(top=4),
                ),
                ft.Container(
                    content=ft.Row([add_button, clear_button], spacing=10),
                    padding=ft.padding.only(top=4),
                ),
            ],
            expand=True,
            spacing=6,
        )

    # File: batch_upload_screen.py
    # Function: add_manual_row

    def add_manual_row(self):
        # Create the row first so we can reference it in callbacks
        row = ft.Row(spacing=6)

        brand_dropdown = ft.Dropdown(
            width=120,
            dense=True,
            options=[
                ft.dropdown.Option("KROSS"),
                ft.dropdown.Option("Pinarello"),
                ft.dropdown.Option("Basso"),
                ft.dropdown.Option("Factor"),
                ft.dropdown.Option("LeeCougan"),
                ft.dropdown.Option("Rascal"),
                ft.dropdown.Option("Rondo"),
                ft.dropdown.Option("Octane"),
            ],
            on_change=lambda e: self.on_brand_change(row),
        )

        code_field = ft.TextField(
            width=120,
            dense=True,
            hint_text="UB-XXXX",
            on_change=lambda e: self.validate_manual_entry(),
        )

        url_field = ft.TextField(
            width=280,
            dense=True,
            hint_text="https://... or config code",
            on_change=lambda e: self.validate_manual_entry(),
        )

        description_dropdown = ft.Dropdown(
            width=200,
            dense=True,
            hint_text="Description (optional)",
            options=self.available_descriptions,
        )

        frameset_checkbox = ft.Checkbox(
            tooltip="Pinarello: Tik frameset (Rėmas, Šakė, Balnakotis, Užspaustukas)",
            value=False,
            visible=False  # Hidden by default, shown only for Pinarello
        )

        disclaimer_checkbox = ft.Checkbox(
            tooltip="Append disclaimer"
        )

        delete_button = ft.IconButton(
            icon=ft.Icons.DELETE,
            icon_color=ft.Colors.RED,
            tooltip="Remove row",
            on_click=lambda e: self.remove_manual_row(row),
        )

        row.controls.extend(
            [
                brand_dropdown,
                code_field,
                url_field,
                description_dropdown,
                frameset_checkbox,
                disclaimer_checkbox,
                delete_button,
            ]
        )

        self.manual_rows.controls.append(row)
        self.page.update()

    def on_brand_change(self, row):
        """Handle brand selection change - show/hide frameset checkbox"""
        if len(row.controls) > 4:
            brand = row.controls[0].value
            frameset_checkbox = row.controls[4]
            # Show frameset checkbox only for Pinarello
            frameset_checkbox.visible = (brand == "Pinarello")
            if brand != "Pinarello":
                frameset_checkbox.value = False  # Reset value if not Pinarello
        self.validate_manual_entry()
        self.page.update()


    def remove_manual_row(self, row):
        self.manual_rows.controls.remove(row)
        self.validate_manual_entry()
        self.page.update()

    def clear_manual_rows(self):
        self.manual_rows.controls.clear()
        for _ in range(10):
            self.add_manual_row()
        self.validate_manual_entry()

    def toggle_all_framesets(self, value):
        """Toggle all frameset checkboxes in manual entry"""
        for row in self.manual_rows.controls:
            if len(row.controls) > 4:
                # Only toggle if brand is Pinarello
                brand = row.controls[0].value
                if brand == "Pinarello":
                    row.controls[4].value = value
        self.page.update()

    def toggle_all_disclaimers(self, value):
        """Toggle all disclaimer checkboxes in manual entry"""
        for row in self.manual_rows.controls:
            if len(row.controls) > 5:
                row.controls[5].value = value
        self.page.update()

    def validate_manual_entry(self):
        valid_rows = 0
        for row in self.manual_rows.controls:
            brand = row.controls[0].value
            code = row.controls[1].value
            url = row.controls[2].value
            if brand and code and url:
                valid_rows += 1

        self.start_button.disabled = valid_rows == 0
        self.page.update()

    # -------------------- EXCEL UPLOAD --------------------

    def build_excel_upload(self):
        self.excel_filename = ft.Text(
            "No file selected", size=14, color=ft.Colors.GREY_600
        )

        browse_button = ft.ElevatedButton(
            "Browse Excel File",
            icon=ft.Icons.FOLDER_OPEN,
            on_click=self.handle_browse_excel,
        )

        template_button = ft.TextButton(
            "Download Template",
            icon=ft.Icons.DOWNLOAD,
            on_click=self.handle_download_template,
        )

        self.excel_preview = ft.Column(
            [], spacing=4, expand=True, scroll=ft.ScrollMode.AUTO
        )

        self.excel_status = ft.Text("", size=14)

        # Select-all checkboxes for Excel preview
        self.excel_frameset_select_all = ft.Checkbox(
            tooltip="Select/Deselect all frameset",
            value=False,
            on_change=lambda e: self.toggle_all_excel_framesets(e.control.value)
        )

        self.excel_disclaimer_select_all = ft.Checkbox(
            tooltip="Select/Deselect all disclaimers",
            value=False,
            on_change=lambda e: self.toggle_all_excel_disclaimers(e.control.value)
        )

        # Store preview elements as instance variables for conditional display
        self.excel_header = ft.Row(
            [
                ft.Text("", width=24),
                ft.Text("Brand", width=120, weight=ft.FontWeight.BOLD),
                ft.Text("Code", width=120, weight=ft.FontWeight.BOLD),
                ft.Text("URL", width=280, weight=ft.FontWeight.BOLD),
                ft.Text("Description", width=200, weight=ft.FontWeight.BOLD),
                ft.Column([
                    ft.Text("Frameset", weight=ft.FontWeight.BOLD, size=12, tooltip="Pinarello only"),
                    self.excel_frameset_select_all
                ], width=70, spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Column([
                    ft.Text("Disc", weight=ft.FontWeight.BOLD, size=12, tooltip="Append disclaimer"),
                    self.excel_disclaimer_select_all
                ], width=50, spacing=0, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            ],
            spacing=6,
            visible=False,  # Hidden initially
        )

        self.excel_preview_label = ft.Text("Preview", weight=ft.FontWeight.BOLD, visible=False)

        self.excel_divider = ft.Divider(height=1, visible=False)

        self.excel_preview_container = ft.Container(
            content=self.excel_preview,
            expand=True,
            border=ft.border.all(1, ft.Colors.GREY_300),
            border_radius=5,
            padding=6,
            visible=False,  # Hidden initially
        )

        # Empty state - shown when no file loaded
        self.empty_state = ft.Column(
            [
                ft.Container(height=40),  # Spacer
                ft.Icon(ft.Icons.UPLOAD_FILE, size=64, color=ft.Colors.GREY_400),
                ft.Text(
                    "Upload an Excel file to begin",
                    size=16,
                    color=ft.Colors.GREY_600,
                    text_align=ft.TextAlign.CENTER,
                ),
                ft.Text(
                    "or download the template to get started",
                    size=14,
                    color=ft.Colors.GREY_500,
                    text_align=ft.TextAlign.CENTER,
                ),
            ],
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=10,
            visible=True,  # Visible initially
        )

        self.excel_status.visible = False  # Hidden initially

        return ft.Column(
            [
                ft.Row([browse_button, template_button], spacing=10),
                self.excel_filename,
                self.empty_state,
                self.excel_preview_label,
                self.excel_header,
                self.excel_divider,
                self.excel_preview_container,
                self.excel_status,
            ],
            expand=True,
            spacing=6,
        )

    def handle_browse_excel(self, e):
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        filename = askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )

        root.destroy()

        if filename:
            self.excel_filename.value = filename
            self.load_excel_preview(filename)

        self.page.update()

    # File: batch_upload_screen.py
    # Function: load_excel_preview

    def load_excel_preview(self, filename):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            self.excel_preview.controls.clear()

            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                values = [str(c.value).lower() if c.value else "" for c in row]
                if any("brand" in v for v in values):
                    header_row = idx
                    break

            if not header_row:
                self.excel_status.value = (
                    "Could not find header row (Brand, Product Code, URL)"
                )
                self.excel_status.color = ft.Colors.ORANGE
                self.start_button.disabled = True
                return

            valid_count = 0
            invalid_count = 0

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(row):
                    continue

                brand = str(row[0]) if len(row) > 0 and row[0] else ""
                code = str(row[1]) if len(row) > 1 and row[1] else ""
                url = str(row[2]) if len(row) > 2 and row[2] else ""

                is_valid = bool(brand and code and url)

                if is_valid:
                    valid_count += 1
                    bg = ft.Colors.GREEN_50
                    status = "✓"
                else:
                    invalid_count += 1
                    bg = ft.Colors.RED_50
                    status = "✗"

                # Read description and disclaimer from Excel
                excel_description = str(row[3]).strip() if len(row) > 3 and row[3] and str(row[3]).strip() else None
                excel_disclaimer_raw = row[4] if len(row) > 4 and row[4] else False

                # Normalize disclaimer (handle "Yes"/"No" strings)
                if isinstance(excel_disclaimer_raw, str):
                    excel_disclaimer = excel_disclaimer_raw.strip().lower() in ("yes", "true", "1")
                else:
                    excel_disclaimer = bool(excel_disclaimer_raw)

                # Create editable controls pre-populated from Excel
                brand_dropdown = ft.Dropdown(
                    width=120,
                    dense=True,
                    value=brand,
                    options=[
                        ft.dropdown.Option("KROSS"),
                        ft.dropdown.Option("Pinarello"),
                        ft.dropdown.Option("Basso"),
                        ft.dropdown.Option("Factor"),
                        ft.dropdown.Option("LeeCougan"),
                        ft.dropdown.Option("Rascal"),
                        ft.dropdown.Option("Rondo"),
                        ft.dropdown.Option("Octane"),
                    ],
                    on_change=lambda e: self.validate_excel_preview(),
                )

                code_field = ft.TextField(
                    width=120,
                    dense=True,
                    value=code,
                    hint_text="UB-XXXX",
                    on_change=lambda e: self.validate_excel_preview(),
                )

                url_field = ft.TextField(
                    width=280,
                    dense=True,
                    value=url,
                    hint_text="https://...",
                    on_change=lambda e: self.validate_excel_preview(),
                )

                desc_dropdown = ft.Dropdown(
                    width=200,
                    dense=True,
                    hint_text="Description",
                    value=excel_description,
                    options=self.available_descriptions,
                    on_change=lambda e: self.validate_excel_preview(),
                )

                # Read frameset from Excel column 5
                excel_frameset_raw = row[5] if len(row) > 5 and row[5] else False
                if isinstance(excel_frameset_raw, str):
                    excel_frameset = excel_frameset_raw.strip().lower() in ("yes", "true", "1", "frameset")
                else:
                    excel_frameset = bool(excel_frameset_raw)

                frameset_checkbox = ft.Checkbox(
                    value=excel_frameset,
                    tooltip="Pinarello: Frameset only",
                    visible=(brand == "Pinarello"),  # Only visible for Pinarello
                    on_change=lambda e: self.validate_excel_preview(),
                )

                disclaimer_checkbox = ft.Checkbox(
                    value=excel_disclaimer,
                    tooltip="Append disclaimer",
                    on_change=lambda e: self.validate_excel_preview(),
                )

                self.excel_preview.controls.append(
                    ft.Container(
                        bgcolor=bg,
                        padding=4,
                        border_radius=3,
                        data={
                            "brand": brand_dropdown,
                            "code": code_field,
                            "url": url_field,
                            "desc": desc_dropdown,
                            "frameset": frameset_checkbox,
                            "disc": disclaimer_checkbox,
                            "valid": is_valid,
                        },
                        content=ft.Row(
                            [
                                ft.Text(status, width=24),
                                brand_dropdown,
                                code_field,
                                url_field,
                                desc_dropdown,
                                frameset_checkbox,
                                disclaimer_checkbox,
                            ],
                            spacing=6,
                        ),
                    )
                )

            if invalid_count > 0:
                self.excel_status.value = (
                    f"{valid_count} valid, {invalid_count} invalid rows"
                )
                self.excel_status.color = ft.Colors.ORANGE
                self.start_button.disabled = True
            else:
                self.excel_status.value = f"{valid_count} valid rows ready"
                self.excel_status.color = ft.Colors.GREEN
                self.start_button.disabled = valid_count == 0

            # Show preview section, hide empty state
            self.empty_state.visible = False
            self.excel_preview_label.visible = True
            self.excel_header.visible = True
            self.excel_divider.visible = True
            self.excel_preview_container.visible = True
            self.excel_status.visible = True

            wb.close()
            self.page.update()

        except Exception as ex:
            self.excel_status.value = f"Error reading file: {ex}"
            self.excel_status.color = ft.Colors.RED
            self.start_button.disabled = True
            self.excel_status.visible = True  # Show error message

    def validate_excel_preview(self):
        """Validate current state of Excel preview controls"""
        valid_count = 0
        invalid_count = 0

        for container in self.excel_preview.controls:
            brand = container.data["brand"].value
            code = container.data["code"].value
            url = container.data["url"].value

            is_valid = bool(brand and code and url)
            container.data["valid"] = is_valid

            if is_valid:
                valid_count += 1
                container.bgcolor = ft.Colors.GREEN_50
            else:
                invalid_count += 1
                container.bgcolor = ft.Colors.RED_50

        if invalid_count > 0:
            self.excel_status.value = f"{valid_count} valid, {invalid_count} invalid rows"
            self.excel_status.color = ft.Colors.ORANGE
            self.start_button.disabled = True
        else:
            self.excel_status.value = f"{valid_count} valid rows ready"
            self.excel_status.color = ft.Colors.GREEN
            self.start_button.disabled = valid_count == 0

        self.page.update()

    def toggle_all_excel_framesets(self, value):
        """Toggle all frameset checkboxes in Excel preview (only for Pinarello)"""
        for container in self.excel_preview.controls:
            brand = container.data["brand"].value
            if brand == "Pinarello":
                container.data["frameset"].value = value
        self.page.update()

    def toggle_all_excel_disclaimers(self, value):
        """Toggle all disclaimer checkboxes in Excel preview"""
        for container in self.excel_preview.controls:
            container.data["disc"].value = value
        self.page.update()


    # File: batch_upload_screen.py
    # Function: handle_download_template

    def handle_download_template(self, e):
        root = Tk()
        root.withdraw()
        root.attributes("-topmost", True)

        filename = asksaveasfilename(
            title="Save Template As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="ultrabike_batch_template.xlsx",
        )

        root.destroy()

        if not filename:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Batch Upload"

        # Header row with all necessary columns
        ws.append(["Brand", "Product Code", "URL or Code", "Description", "Frameset", "Append Disclaimer"])

        # Example rows
        ws.append(["KROSS", "UB-1234", "https://kross.pl/product1", "Mountain Bike", "No", "Yes"])
        ws.append(["Pinarello", "UB-5678", "https://pinarello.com/product2", "Road Bike", "Yes", "No"])

        from openpyxl.styles import Font, PatternFill

        # Style header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        wb.save(filename)
        wb.close()

        self.excel_status.value = f"Template saved to {filename}"
        self.excel_status.color = ft.Colors.GREEN
        self.page.update()



        # -------------------- START BATCH --------------------

    def handle_start_batch(self, e):
        items = []

        if self.tab_bar.selected_index == 0:
            for row in self.manual_rows.controls:
                brand = row.controls[0].value
                code = row.controls[1].value
                url = row.controls[2].value
                description = row.controls[3].value if len(row.controls) > 3 else ""
                frameset_only = bool(row.controls[4].value) if len(row.controls) > 4 else False
                append_disclaimer = bool(row.controls[5].value) if len(row.controls) > 5 else False

                if brand and code and url:
                    items.append({
                        "brand": brand,
                        "code": code,
                        "url": url,
                        "description_name": description,
                        "frameset_only": frameset_only,
                        "append_disclaimer": append_disclaimer
                    })
        else:
            # Excel mode - read from preview controls, NOT from Excel file
            for container in self.excel_preview.controls:
                if not container.data.get("valid", False):
                    continue

                # Read from UI controls
                brand = container.data["brand"].value
                code = container.data["code"].value
                url = container.data["url"].value
                description = container.data["desc"].value
                frameset_only = container.data["frameset"].value or False
                append_disclaimer = container.data["disc"].value or False

                if not (brand and code and url):
                    continue

                items.append({
                    "brand": brand,
                    "code": code,
                    "url": url,
                    "description_name": description,
                    "frameset_only": frameset_only,
                    "append_disclaimer": append_disclaimer
                })

        if not items:
            return

        self.on_close_callback()
        self.app.start_batch_processing(items)


