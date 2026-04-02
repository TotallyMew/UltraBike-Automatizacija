"""GUI_Qt/screens/BatchDescriptionsScreen.py

Batch Descriptions uploader.

UX parity with BatchUploadScreen:
- Manual mode: table rows
- Excel mode: drag/drop + browse + template download
- Start enabled only when rows are valid
- Per-row Status/Error updates while running

Automation reuses the existing logged-in Selenium session from MainWindow.
"""

from __future__ import annotations

import os
import time
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QEvent, QThread, QTimer, Qt, Signal
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent, QStandardItemModel, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    CheckBox,
    ComboBox,
    FluentIcon,
    IconWidget,
    IndeterminateProgressRing,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    MessageBox,
    PillPushButton,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    TitleLabel,
    TransparentToolButton,
    isDarkTheme,
    qconfig,
)

from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import COLORS, COMPONENT_COLORS, FONTS, RADII, PADDINGS, SIZES
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS, PAGE_SPACING, ICON_TEXT_GAP, ROW_SPACING, TOOLBAR_MARGINS,
    CARD_SPACING, CONTENT_SPACING, TABLE_CELL_MARGINS, DETAILS_SPACING, MID_SPACING, SPACING,
    apply_screen_theme, enforce_transparent_labels,
    get_responsive_margins, get_responsive_spacing,
)
from Config.Selectors import ProductEditorSelectors, ProductListSelectors
from Managers.DescriptionManager import DescriptionManager
from Utilities.ProductNavigationHandler import ProductNavigationHandler
from Utilities.WebIntercationHandler import WebInteractionHandler


class DropZoneWidget(QWidget):
    """Drag & drop zone for a single .xlsx file."""

    file_dropped = Signal(str)

    def __init__(self, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(PAGE_SPACING)

        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES['icon_huge'], SIZES['icon_huge'])

        self.title_label = BodyLabel("")
        self.title_label.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']}; color: {COLORS['lavender_grey']};")

        self.subtitle_label = CaptionLabel("")
        self.subtitle_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        layout.addStretch()
        layout.addWidget(icon, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.title_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.subtitle_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addStretch()

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._apply_style(False)
        self.retranslate_ui()

    def retranslate_ui(self):
        self.title_label.setText(self.tr("batch.drop.title"))
        self.subtitle_label.setText(self.tr("batch.drop.subtitle"))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1 and urls[0].toLocalFile().endswith('.xlsx'):
                event.acceptProposedAction()
                self._apply_style(True)

    def dragLeaveEvent(self, event):
        self._apply_style(False)

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith('.xlsx'):
            self.file_dropped.emit(files[0])
        self._apply_style(False)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.file_dropped.emit("__browse__")

    def _apply_style(self, drag_active: bool):
        is_dark = isDarkTheme()
        border = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        dashed = COLORS['lavender_grey']
        from GUI_Qt.styles.theme_config import get_hover_bg
        hover_bg = get_hover_bg(is_dark)

        if drag_active:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: {hover_bg};
                    border: 2px solid {border};
                    border-radius: {RADII['lg']}px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: transparent;
                    border: 2px dashed {dashed};
                    border-radius: {RADII['lg']}px;
                }}
                DropZoneWidget:hover {{
                    border-color: {border};
                    background-color: {hover_bg};
                }}
            """)


class BatchDescriptionsWorker(QThread):
    row_update = Signal(int, str, str)  # rowIndex, status, error
    log = Signal(str)
    done = Signal(int, int)  # ok, total
    progress_update = Signal(int, int, float, float)  # current, total, speed (items/sec), eta_seconds

    def __init__(self, main_window, items: list[dict]):
        super().__init__()
        self.main = main_window
        self.items = items
        self._processed_times = []  # Track duration of each processed item

    def _ensure_products_page(self, driver) -> None:
        try:
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.ui import WebDriverWait

            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.NAV_PRODUCTS)
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
            )
        except Exception:
            pass

    def run(self):
        tr = self.main.i18n.tr
        driver = getattr(self.main, "driver", None)
        if driver is None:
            for it in self.items:
                self.row_update.emit(it["row"], tr("batchdesc.status.failed"), tr("batchdesc.no_session"))
            self.done.emit(0, len(self.items))
            return

        desc_manager = DescriptionManager(self.main.db, logger=self.main.logger)
        nav = None
        web = None
        if driver is not None:
            nav = ProductNavigationHandler(driver, logger=self.main.logger)
            web = WebInteractionHandler(driver)

        ok = 0
        total = len(self.items)

        batch_id = f"batchdesc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


        import time as _time

        for idx, it in enumerate(self.items, start=1):
            row = it["row"]
            brand = it["brand"]
            code = it["code"]
            desc = it["description"]
            disclaimer = bool(it.get("append_disclaimer"))

            self.row_update.emit(row, tr("batchdesc.status.running"), "")
            self.log.emit(tr("batchdesc.progress.open", current=idx, total=total, code=code))

            timings = {}
            t0 = _time.perf_counter()
            try:
                if driver is None or nav is None or web is None:
                    self.row_update.emit(row, tr("batchdesc.status.failed"), tr("batchdesc.no_session"))
                    continue

                t_start = _time.perf_counter()
                self._ensure_products_page(driver)
                timings["ensure_products_page"] = _time.perf_counter() - t_start

                t_start = _time.perf_counter()
                nav.navigate_to_product(brand, code)
                timings["navigate_to_product"] = _time.perf_counter() - t_start

                did_upload = False
                if desc:
                    self.log.emit(tr("batchdesc.progress.upload", current=idx, total=total, code=code))
                    t_start = _time.perf_counter()
                    uploaded = desc_manager.upload_description(
                        driver,
                        product_code=code,
                        name=desc,
                        append_disclaimer=disclaimer,
                    )
                    timings["upload_description"] = _time.perf_counter() - t_start
                    if not uploaded:
                        self.row_update.emit(row, tr("batchdesc.status.failed"), tr("batchdesc.item.failed_upload"))

                        # Record failure
                        try:
                            details = {
                                "workflow": "batch_descriptions",
                                "batch_id": batch_id,
                                "append_disclaimer": disclaimer,
                                "description_template": desc or None,
                            }
                            self.main.db.conn.execute(
                                """
                                INSERT INTO processing_history
                                (brand, product_code, url_or_code, status, error_message, duration_seconds,
                                 failed_stage, batch_id, details_json, processed_at)
                                VALUES (?, ?, ?, 'failed', ?, ?, ?, ?, ?, ?)
                                """,
                                (
                                    brand,
                                    code,
                                    None,
                                    tr("batchdesc.item.failed_upload"),
                                    float(_time.perf_counter() - t0),
                                    "upload_description",
                                    batch_id,
                                    json.dumps(details, ensure_ascii=False),
                                ),
                            )
                            self.main.db.conn.commit()
                        except Exception:
                            pass
                        continue
                    did_upload = True

                # Scroll to and click the product title field for the last language edited before saving
                t_start = _time.perf_counter()
                try:
                    from selenium.webdriver.support.ui import WebDriverWait
                    from selenium.webdriver.support import expected_conditions as EC
                    from Config.LanguageConfig import LANG_CODE_TO_PLATFORM_ID
                    last_lang_id = LANG_CODE_TO_PLATFORM_ID["lv"]
                    title_field = WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located(ProductEditorSelectors.name_field(last_lang_id))
                    )
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", title_field)
                    try:
                        title_field.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", title_field)
                except Exception:
                    pass
                timings["focus_title_field"] = _time.perf_counter() - t_start

                t_start = _time.perf_counter()
                self.log.emit(tr("batchdesc.progress.save", current=idx, total=total, code=code))
                web.save_information()
                timings["save_information"] = _time.perf_counter() - t_start

                t_start = _time.perf_counter()
                _time.sleep(0.5)
                timings["sleep"] = _time.perf_counter() - t_start

                timings["total"] = _time.perf_counter() - t0

                # Log timings for this row
                timing_report = ", ".join(f"{k}: {v:.3f}s" for k, v in timings.items())
                self.log.emit(f"Timing for {code}: {timing_report}")

                ok += 1
                self.row_update.emit(row, tr("batchdesc.status.ok"), "")

                # Record success
                try:
                    details = {
                        "workflow": "batch_descriptions",
                        "batch_id": batch_id,
                        "append_disclaimer": disclaimer,
                        "description_template": desc or None,
                        "timings": timings,
                    }
                    self.main.db.conn.execute(
                        """
                        INSERT INTO processing_history
                        (brand, product_code, url_or_code, status, duration_seconds,
                         failed_stage, batch_id, details_json, processed_at)
                        VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                        """,
                        (
                            brand,
                            code,
                            None,
                            float(_time.perf_counter() - t0),
                            None,
                            batch_id,
                            json.dumps(details, ensure_ascii=False),
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        ),
                    )
                    self.main.db.conn.commit()
                except Exception:
                    pass
            except Exception as e:
                self.row_update.emit(row, tr("batchdesc.status.failed"), str(e))

                # Record failure
                try:
                    details = {
                        "workflow": "batch_descriptions",
                        "batch_id": batch_id,
                        "append_disclaimer": disclaimer,
                        "description_template": desc or None,
                    }
                    self.main.db.conn.execute(
                        """
                        INSERT INTO processing_history
                        (brand, product_code, url_or_code, status, error_message, duration_seconds,
                         failed_stage, batch_id, details_json, processed_at)
                        VALUES (?, ?, ?, 'failed', ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            brand,
                            code,
                            None,
                            str(e),
                            float(_time.perf_counter() - t0),
                            "batch_descriptions",
                            batch_id,
                            json.dumps(details, ensure_ascii=False),
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        ),
                    )
                    self.main.db.conn.commit()
                except Exception:
                    pass

            # Track timing and emit progress update
            item_duration = _time.perf_counter() - t0
            self._processed_times.append(item_duration)

            # Calculate ETA (skip first 2 items for more accurate average)
            if len(self._processed_times) >= 2:
                recent_times = self._processed_times[1:]  # Skip first item (often slower)
                avg_time_per_item = sum(recent_times) / len(recent_times)
                remaining_items = total - idx
                eta_seconds = avg_time_per_item * remaining_items
                speed = 1.0 / avg_time_per_item if avg_time_per_item > 0 else 0

                self.progress_update.emit(idx, total, speed, eta_seconds)

        self.done.emit(ok, total)


class ParallelBatchDescriptionsWorker(QThread):
    """Parallel batch descriptions update using browser session pool"""
    row_update = Signal(int, str, str)  # rowIndex, status, error
    log = Signal(str)
    done = Signal(int, int)  # ok, total
    session_status = Signal(dict)  # session pool statistics

    def __init__(self, main_window, session_manager, items: list[dict]):
        super().__init__()
        self.main = main_window
        self.session_manager = session_manager
        self.items = items
        self._stop = False
        self._lock = __import__('threading').Lock()
        self._ok_count = 0
        self._work_index = 0

        self._batch_id = f"batchdesc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    def stop(self):
        self._stop = True

    def _ensure_products_page(self, driver) -> None:
        try:
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.ui import WebDriverWait

            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.NAV_PRODUCTS)
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
            )
        except Exception:
            pass

    def run(self):
        """Execute parallel batch descriptions update using session pool"""
        import threading

        if not self.session_manager or not self.session_manager.is_ready():
            tr = self.main.i18n.tr
            for it in self.items:
                self.row_update.emit(it["row"], tr("batchdesc.status.failed"), "Session pool not available")
            self.done.emit(0, len(self.items))
            return

        total = len(self.items)
        pool_size = self.session_manager.pool_size

        self.log.emit(f"Starting parallel processing with {pool_size} browsers")

        # Create worker threads
        threads = []
        for worker_id in range(pool_size):
            thread = threading.Thread(target=self._worker_thread, args=(worker_id,))
            thread.daemon = True
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

        self.log.emit(f"Completed. Success: {self._ok_count}/{total}")
        self.done.emit(self._ok_count, total)

    def _worker_thread(self, worker_id: int):
        """Worker thread that processes items from the queue"""
        tr = self.main.i18n.tr
        import time as _time

        while not self._stop:
            # Get next work item
            with self._lock:
                if self._work_index >= len(self.items):
                    break  # No more work
                idx = self._work_index
                self._work_index += 1

            it = self.items[idx]
            row = it["row"]
            brand = it["brand"]
            code = it["code"]
            desc = it["description"]
            disclaimer = bool(it.get("append_disclaimer"))
            t0 = _time.perf_counter()

            # Acquire a browser session
            session = self.session_manager.acquire_session(timeout=60.0)
            if session is None:
                self.row_update.emit(row, tr("batchdesc.status.failed"), "Could not acquire browser session")
                continue

            try:
                # Update status
                self.row_update.emit(row, tr("batchdesc.status.running"), "")
                self.log.emit(tr("batchdesc.progress.open", current=idx+1, total=len(self.items), code=code))

                # Emit session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)

                # Process description via Selenium
                desc_manager = DescriptionManager(self.main.db, logger=self.main.logger)
                nav = ProductNavigationHandler(session.driver, logger=self.main.logger)
                web = WebInteractionHandler(session.driver)

                self._ensure_products_page(session.driver)
                nav.navigate_to_product(brand, code)

                did_upload = False
                if desc:
                    self.log.emit(tr("batchdesc.progress.upload", current=idx+1, total=len(self.items), code=code))
                    uploaded = desc_manager.upload_description(
                        session.driver,
                        product_code=code,
                        name=desc,
                        append_disclaimer=disclaimer,
                    )
                    if not uploaded:
                        self.row_update.emit(row, tr("batchdesc.status.failed"), tr("batchdesc.item.failed_upload"))
                        continue
                    did_upload = True

                # Focus title field before saving
                try:
                    from selenium.webdriver.support.ui import WebDriverWait
                    from selenium.webdriver.support import expected_conditions as EC
                    from Config.LanguageConfig import LANG_CODE_TO_PLATFORM_ID
                    last_lang_id = LANG_CODE_TO_PLATFORM_ID["lv"]
                    title_field = WebDriverWait(session.driver, 3).until(
                        EC.presence_of_element_located(ProductEditorSelectors.name_field(last_lang_id))
                    )
                    session.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", title_field)
                    try:
                        title_field.click()
                    except Exception:
                        session.driver.execute_script("arguments[0].click();", title_field)
                except Exception:
                    pass

                self.log.emit(tr("batchdesc.progress.save", current=idx+1, total=len(self.items), code=code))
                web.save_information()
                _time.sleep(0.5)

                self.row_update.emit(row, tr("batchdesc.status.ok"), "")
                with self._lock:
                    self._ok_count += 1

                # Record success
                try:
                    details = {
                        "workflow": "batch_descriptions",
                        "batch_id": self._batch_id,
                        "append_disclaimer": disclaimer,
                        "description_template": desc or None,
                        "mode": "parallel",
                    }
                    with self._lock:
                        self.main.db.conn.execute(
                            """
                            INSERT INTO processing_history
                            (brand, product_code, url_or_code, status, duration_seconds,
                             failed_stage, batch_id, details_json, processed_at)
                            VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                            """,
                            (
                                brand,
                                code,
                                None,
                                float(_time.perf_counter() - t0),
                                None,
                                self._batch_id,
                                json.dumps(details, ensure_ascii=False),
                            ),
                        )
                        self.main.db.conn.commit()
                except Exception:
                    pass

            except Exception as e:
                error_msg = str(e)
                self.row_update.emit(row, tr("batchdesc.status.failed"), error_msg)
                session.error_count += 1

                # Record failure
                try:
                    details = {
                        "workflow": "batch_descriptions",
                        "batch_id": self._batch_id,
                        "append_disclaimer": disclaimer,
                        "description_template": desc or None,
                        "mode": "parallel",
                    }
                    with self._lock:
                        self.main.db.conn.execute(
                            """
                            INSERT INTO processing_history
                            (brand, product_code, url_or_code, status, error_message, duration_seconds,
                             failed_stage, batch_id, details_json, processed_at)
                            VALUES (?, ?, ?, 'failed', ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                brand,
                                code,
                                None,
                                error_msg,
                                float(_time.perf_counter() - t0),
                                "batch_descriptions",
                                self._batch_id,
                                json.dumps(details, ensure_ascii=False),
                            ),
                        )
                        self.main.db.conn.commit()
                except Exception:
                    pass

                # Reset session if too many errors
                if session.error_count >= 3:
                    self.log.emit(f"Resetting session {session.session_id} due to errors")
                    self.session_manager.reset_session(session)

            finally:
                # Always release the session
                self.session_manager.release_session(session)

                # Update session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)


class BatchDescriptionsScreen(ResponsiveWidget):

        # ...existing code...
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.worker: BatchDescriptionsWorker | None = None

        # Multi-session browser pool
        self.session_manager = None

        self.current_mode = "manual"
        self._base_table_rows = 5

        self.brands = [
            "KROSS", "Pinarello", "Basso", "Factor",
            "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan",
        ]

        desc_manager = DescriptionManager(main_window.db)
        self.descriptions = [d['name'] for d in desc_manager.list_descriptions()]

        # Store references for responsive UI
        self.scroll = None
        self.content_widget = None

        self._init_ui()
        self._setup_shortcuts()
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(0, self._ensure_table_fills_viewport)

    def hideEvent(self, event):
        """Cleanup session manager when screen is hidden"""
        super().hideEvent(event)

        # Shutdown session manager if it exists
        if self.session_manager is not None:
            try:
                self.session_manager.shutdown_all()
                self.session_manager = None
            except Exception as e:
                if hasattr(self.main, 'logger'):
                    self.main.logger.log("BatchDescriptionsScreen", f"Error shutting down session manager: {str(e)}")

    def eventFilter(self, obj, event):
        if obj is self.table.viewport() and event.type() == QEvent.Type.Resize:
            QTimer.singleShot(0, self._ensure_table_fills_viewport)
        return super().eventFilter(obj, event)

    def _apply_theme(self):
        is_dark = isDarkTheme()
        bg_color = COLORS['space_indigo'] if is_dark else COLORS['platinum']
        self.setStyleSheet(f"""
            BatchDescriptionsScreen {{
                background-color: {bg_color};
                font-family: {FONTS['family']};
            }}
        """)

    def _init_ui(self):
        # Root layout (for ScrollArea container)
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # Create ScrollArea
        self.scroll = ScrollArea(self)
        self.scroll.setWidgetResizable(True)
        root_layout.addWidget(self.scroll)

        # Content widget (inside ScrollArea)
        self.content_widget = QWidget()
        self.scroll.setWidget(self.content_widget)

        # Main layout (on content widget)
        root = QVBoxLayout(self.content_widget)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        # Apply theme
        apply_screen_theme(
            self,
            "BatchDescriptionsScreen",
            scroll=self.scroll,
            content=self.content_widget
        )

        header = QHBoxLayout()
        title_container = QHBoxLayout()
        title_container.setSpacing(ICON_TEXT_GAP)

        title_icon = TransparentToolButton(FluentIcon.TAG, self)
        title_icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        title_icon.setEnabled(False)

        self.title_label = TitleLabel("")
        title_container.addWidget(title_icon)
        title_container.addWidget(self.title_label)
        header.addLayout(title_container)
        header.addStretch(1)

        mode_container = QWidget()
        mode_container.setStyleSheet("background: transparent;")
        mode_layout = QHBoxLayout(mode_container)
        mode_layout.setContentsMargins(0, 0, 0, 0)
        mode_layout.setSpacing(ROW_SPACING)

        self.manual_pill = PillPushButton("")
        self.manual_pill.setCheckable(True)
        self.manual_pill.setChecked(True)
        self.manual_pill.clicked.connect(lambda: self._switch_mode("manual"))

        self.excel_pill = PillPushButton("")
        self.excel_pill.setCheckable(True)
        self.excel_pill.clicked.connect(lambda: self._switch_mode("excel"))

        mode_layout.addWidget(self.manual_pill)
        mode_layout.addWidget(self.excel_pill)
        header.addWidget(mode_container)
        root.addLayout(header)

        toolbar_card = CardWidget()
        toolbar_card.setBorderRadius(RADII['md'])
        # Keep this toolbar in a single row (like Batch Titles) so the card doesn't
        # stretch in the Y axis at the default window size.
        toolbar_layout = QHBoxLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self.add_btn = PushButton("")
        self.add_btn.setIcon(FluentIcon.ADD.icon())
        self.add_btn.clicked.connect(self._add_row)

        self.clear_btn = PushButton("")
        self.clear_btn.setIcon(FluentIcon.DELETE.icon())
        self.clear_btn.clicked.connect(self._clear_all)

        self.bulk_label = CaptionLabel("")
        self.bulk_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.disclaimer_bulk = CheckBox("")
        self.disclaimer_bulk.stateChanged.connect(self._toggle_all_disclaimers)

        self.template_btn = PushButton("")
        self.template_btn.setIcon(FluentIcon.DOWNLOAD.icon())
        self.template_btn.clicked.connect(self._download_template)

        self.browse_btn = PushButton("")
        self.browse_btn.setIcon(FluentIcon.FOLDER_ADD.icon())
        self.browse_btn.clicked.connect(self._browse_excel)

        self.refresh_btn = PushButton("")
        self.refresh_btn.setIcon(FluentIcon.SYNC.icon())
        self.refresh_btn.clicked.connect(self._refresh_descriptions)
        self.refresh_btn.setToolTip(self.main.i18n.tr("batch.refresh_descriptions"))

        self.excel_file_label = CaptionLabel("")
        self.excel_file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        # Expandable spacer that pushes status/start to the right.
        # This keeps the Excel controls on the left in Excel mode.
        self._toolbar_spacer = QWidget()
        self._toolbar_spacer.setStyleSheet("background: transparent;")
        self._toolbar_spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        self.status_label = BodyLabel("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self.progress_ring = IndeterminateProgressRing()
        self.progress_ring.setFixedSize(SIZES['progress_ring'], SIZES['progress_ring'])
        self.progress_ring.setVisible(False)

        self.start_btn = PrimaryPushButton("")
        self.start_btn.setIcon(FluentIcon.PLAY.icon())
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self._start_batch)

        # Spacer between Clear and Bulk Select (hide/show with manual controls)
        self._manual_sep = QWidget()
        self._manual_sep.setStyleSheet("background: transparent;")
        self._manual_sep.setFixedWidth(ROW_SPACING)

        # Spacer between excel file label and status (hide/show with excel controls)
        self._excel_status_sep = QWidget()
        self._excel_status_sep.setStyleSheet("background: transparent;")
        self._excel_status_sep.setFixedWidth(CONTENT_SPACING)

        # Single row: manual controls + bulk selectors + excel controls + status/start
        # (show/hide is handled by _switch_mode).
        toolbar_layout.addWidget(self.add_btn)
        toolbar_layout.addWidget(self.clear_btn)
        toolbar_layout.addWidget(self._manual_sep)
        toolbar_layout.addWidget(self.bulk_label)
        toolbar_layout.addWidget(self.disclaimer_bulk)
        toolbar_layout.addWidget(self.template_btn)
        toolbar_layout.addWidget(self.browse_btn)
        toolbar_layout.addWidget(self.refresh_btn)
        toolbar_layout.addWidget(self.excel_file_label)
        toolbar_layout.addWidget(self._excel_status_sep)
        toolbar_layout.addWidget(self._toolbar_spacer)
        toolbar_layout.addWidget(self.status_label)
        toolbar_layout.addWidget(self.progress_ring)
        toolbar_layout.addWidget(self.start_btn)

        root.addWidget(toolbar_card)

        table_card = CardWidget()
        table_card.setBorderRadius(RADII['md'])
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.excel_drop = DropZoneWidget(self.main.i18n.tr)
        self.excel_drop.file_dropped.connect(self._handle_file_drop)

        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Smooth scrolling (avoid row/column snapping).
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

        # Enable horizontal scrolling for responsive design
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        try:
            self.table.verticalScrollBar().setSingleStep(12)
            self.table.horizontalScrollBar().setSingleStep(12)
        except Exception:
            pass
        # Fixed columns (horizontal scrolling as needed).
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.viewport().installEventFilter(self)

        # Wider brand column so placeholder doesn't get truncated
        self.table.setColumnWidth(0, SIZES['col_w_240'])  # Brand
        self.table.setColumnWidth(1, SIZES['col_w_180'])  # Code
        self.table.setColumnWidth(2, SIZES['col_w_420'])  # Description
        self.table.setColumnWidth(3, SIZES['col_w_120'])  # Disclaimer
        self.table.setColumnWidth(4, SIZES['col_w_160'])  # Status
        self.table.setColumnWidth(5, SIZES['col_w_420'])  # Error
        self.table.setColumnWidth(6, SIZES['col_w_56'])   # Delete

        self.table.verticalHeader().setDefaultSectionSize(SIZES['table_row_height_lg'])

        table_layout.addWidget(self.excel_drop)
        table_layout.addWidget(self.table)
        root.addWidget(table_card, 1)

        self.table.setRowCount(self._base_table_rows)
        for r in range(self.table.rowCount()):
            self._setup_table_row(r)

        self._update_table_theme()
        self._switch_mode("manual")
        self.retranslate_ui()
        enforce_transparent_labels(self)

    def retranslate_ui(self):
        tr = self.main.i18n.tr
        self.title_label.setText(tr("batchdesc.title"))
        self.manual_pill.setText(tr("batch.mode.manual"))
        self.excel_pill.setText(tr("batch.mode.excel"))
        self.add_btn.setText(tr("batch.add_row"))
        self.add_btn.setToolTip(tr("batch.add_row.tip"))
        self.clear_btn.setText(tr("batch.clear_all"))
        self.clear_btn.setToolTip(tr("batch.clear_all.tip"))
        self.bulk_label.setText(tr("batch.bulk_select"))
        self.disclaimer_bulk.setText(tr("batch.bulk.disclaimer"))
        self.disclaimer_bulk.setToolTip(tr("batch.bulk.disclaimer.tip"))
        self.template_btn.setText(tr("batch.download_template"))
        self.template_btn.setToolTip(tr("batch.download_template.tip"))
        self.browse_btn.setText(tr("batch.browse_excel"))
        self.browse_btn.setToolTip(tr("batch.browse_excel.tip"))
        self.excel_file_label.setText(tr("batch.no_file"))
        # Keep the Start button area clean; don't show the "Ready" hint.
        self.status_label.setText("")
        self.status_label.setVisible(False)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.start_btn.setText(f"{tr('batchdesc.start')} (Ctrl+Enter)")
        self.start_btn.setToolTip(tr("batchdesc.start.tip"))

        self.table.setHorizontalHeaderLabels([
            tr("batch.table.brand"),
            tr("batch.table.code"),
            tr("batch.table.desc"),
            tr("batch.table.disclaimer"),
            tr("batchdesc.table.status"),
            tr("batchdesc.table.error"),
            "",
        ])

        self.excel_drop.retranslate_ui()

        # Update placeholder labels for existing row widgets
        for row in range(self.table.rowCount()):
            brand_cell = self.table.cellWidget(row, 0)
            if not brand_cell:
                continue
            brand_combo = brand_cell.findChild(ComboBox)
            if brand_combo:
                self._ensure_combo_placeholder_item(brand_combo, tr("batch.select_brand"), self.brands)

    def _on_breakpoint_changed(self, breakpoint: str):
        """Respond to breakpoint changes - adjust margins and spacing."""
        margins = get_responsive_margins(breakpoint)
        spacing = get_responsive_spacing(breakpoint)
        if hasattr(self, 'content_widget') and self.content_widget and self.content_widget.layout():
            self.content_widget.layout().setContentsMargins(*margins)
            self.content_widget.layout().setSpacing(spacing)

    def _on_theme_changed(self):
        apply_screen_theme(
            self,
            "BatchDescriptionsScreen",
            scroll=self.scroll,
            content=self.content_widget
        )
        enforce_transparent_labels(self)
        self._update_table_theme()

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        table_colors = COMPONENT_COLORS['table']
        bg_color = table_colors['row_alt_bg_dark'] if is_dark else table_colors['row_alt_bg_light']
        alt_bg = table_colors['row_bg_dark'] if is_dark else table_colors['row_bg_light']
        border_color = table_colors['border_dark'] if is_dark else table_colors['border_light']

        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']
        text_color = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']
        from GUI_Qt.styles.theme_config import (
            get_embedded_input_border,
            get_scrollbar_handle_bg,
            get_scrollbar_handle_hover_bg,
            get_selection_bg,
        )
        embedded_input_border = get_embedded_input_border(is_dark)
        embedded_input_bg = COMPONENT_COLORS['input']['bg_dark'] if is_dark else COMPONENT_COLORS['input']['bg_light']

        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg_color};
                alternate-background-color: {alt_bg};
                border: 1px solid {border_color};
                border-radius: {RADII['md']}px;
                gridline-color: {border_color};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::viewport {{
                background-color: {bg_color};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableView::item:focus, QAbstractItemView::item:focus {{
                outline: none;
            }}
            QTableWidget QWidget:focus {{
                outline: none;
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
                letter-spacing: 0.5px;
                text-transform: uppercase;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}
            QHeaderView::section {{
                border-right: 1px solid {border_color};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:last {{
                border-top-right-radius: {RADII['md']}px;
            }}

            /* Keep scrollbars visually inside rounded corners */
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}

            QScrollBar:horizontal {{
                background: transparent;
                height: {SIZES['scrollbar_thickness']}px;
                margin: 0px {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:horizontal {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-width: {SIZES['scrollbar_handle_min']}px;
                margin: {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                border: none;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
            QTableWidget QWidget[ubTableCell="true"] LineEdit,
            QTableWidget QWidget[ubTableCell="true"] QLineEdit,
            QTableWidget QWidget[ubTableCell="true"] ComboBox,
            QTableWidget QWidget[ubTableCell="true"] QComboBox {{
                background-color: {embedded_input_bg};
                border: 1px solid {embedded_input_border};
                border-bottom: 2px solid {embedded_input_border};
                border-radius: {RADII['sm']}px;
                padding: {PADDINGS['combo']};
                outline: none;
            }}
        """)

    def _switch_mode(self, mode: str):
        self.current_mode = mode
        self.manual_pill.setChecked(mode == "manual")
        self.excel_pill.setChecked(mode == "excel")

        is_manual = mode == "manual"
        self.add_btn.setVisible(is_manual)
        self.clear_btn.setVisible(is_manual)
        self.bulk_label.setVisible(is_manual)
        self.disclaimer_bulk.setVisible(is_manual)
        if hasattr(self, "_manual_sep"):
            self._manual_sep.setVisible(is_manual)

        is_excel = mode == "excel"
        self.template_btn.setVisible(is_excel)
        self.browse_btn.setVisible(is_excel)
        self.excel_file_label.setVisible(is_excel)
        if hasattr(self, "_excel_status_sep"):
            self._excel_status_sep.setVisible(is_excel)

        no_file = self.excel_file_label.text() == self.main.i18n.tr("batch.no_file")
        self.excel_drop.setVisible(is_excel and no_file)
        self.table.setVisible((not is_excel) or (is_excel and not no_file))
        self._validate()

    def _handle_file_drop(self, filepath: str):
        if filepath == "__browse__":
            self._browse_excel()
        else:
            self._load_excel(filepath)

    def _browse_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self,
            self.main.i18n.tr("batch.file.select_excel.title"),
            "",
            self.main.i18n.tr("batch.file.select_excel.filter"),
        )
        if filename:
            self._load_excel(filename)

    def _load_excel(self, filename: str):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                values = [str(c.value).lower() if c.value else "" for c in row]
                if any("brand" in v for v in values) and any("code" in v or "product" in v for v in values):
                    header_row = idx
                    break
            if not header_row:
                raise Exception(self.main.i18n.tr("batch.excel.header_missing_in_file"))

            data_rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if any(row):
                    data_rows.append(row)

            self.table.setRowCount(len(data_rows) if data_rows else self._base_table_rows)
            for r in range(self.table.rowCount()):
                self._setup_table_row(r)

            valid_count = 0
            invalid_count = 0

            for table_row, excel_row in enumerate(data_rows):
                brand = str(excel_row[0]).strip() if len(excel_row) > 0 and excel_row[0] else ""
                code_raw = str(excel_row[1]).strip() if len(excel_row) > 1 and excel_row[1] else ""
                code = self._normalize_code(code_raw)
                desc = str(excel_row[2]).strip() if len(excel_row) > 2 and excel_row[2] else ""

                disclaimer_raw = excel_row[3] if len(excel_row) > 3 else False
                if isinstance(disclaimer_raw, str):
                    disclaimer = disclaimer_raw.strip().lower() in ("yes", "true", "1")
                else:
                    disclaimer = bool(disclaimer_raw)

                brand_widget = self._get_widget_from_cell(table_row, 0, ComboBox)
                if brand_widget and brand:
                    brand_widget.setCurrentText(brand)

                code_widget = self._get_widget_from_cell(table_row, 1, LineEdit)
                if code_widget:
                    code_widget.setText(code)

                desc_widget = self._get_widget_from_cell(table_row, 2, ComboBox)
                if desc_widget and desc:
                    desc_widget.setCurrentText(desc)

                disclaimer_widget = self._get_widget_from_cell(table_row, 3, CheckBox)
                if disclaimer_widget:
                    disclaimer_widget.setChecked(disclaimer)

                # Row is valid if it has brand + code + (description OR disclaimer)
                desc_selected = bool(desc) and desc != self.main.i18n.tr("batch.optional")
                disclaimer_checked = disclaimer_widget and disclaimer_widget.isChecked()

                if brand and code and (desc_selected or disclaimer_checked):
                    valid_count += 1
                else:
                    invalid_count += 1

            wb.close()

            self.excel_file_label.setText(os.path.basename(filename))
            self.excel_drop.setVisible(False)
            self.table.setVisible(True)

            if invalid_count > 0:
                self.status_label.setText(
                    self.main.i18n.tr(
                        "batch.excel.rows_fix_errors",
                        valid=valid_count,
                        invalid=invalid_count,
                    )
                )
                self.status_label.setStyleSheet(f"color: {COLORS['warning']}; font-weight: 500;")
                self.status_label.setVisible(True)
                self.start_btn.setEnabled(False)
            else:
                # Don't show a "Ready" hint; keep Start enabled silently.
                self.status_label.setText("")
                self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
                self.status_label.setVisible(False)
                self.start_btn.setEnabled(valid_count > 0)
        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.excel.load_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self,
            )

    def _download_template(self):
        filename, _ = QFileDialog.getSaveFileName(
            self,
            self.main.i18n.tr("batch.template.save.title"),
            "batch_descriptions_template.xlsx",
            self.main.i18n.tr("batch.template.filter"),
        )
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.main.i18n.tr("batchdesc.title")
            ws.append(["Brand", "Product Code", "Description", "Append Disclaimer", "Append Order Note"])
            ws.append(["KROSS", "UB-1234", "MyTemplateName", "Yes", "Yes"])

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="8D99AE", end_color="8D99AE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 22

            wb.save(filename)
            wb.close()

            InfoBar.success(
                title=self.main.i18n.tr("common.success"),
                content=self.main.i18n.tr("batch.template.downloaded"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self,
            )
        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.template.save_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self,
            )

    def _refresh_descriptions(self):
        """Reload description list from database and update all ComboBoxes in the table."""
        try:
            # Reload descriptions from database
            desc_manager = DescriptionManager(self.main.db)
            self.descriptions = [d['name'] for d in desc_manager.list_descriptions()]

            # Update all description ComboBoxes in the table
            for row in range(self.table.rowCount()):
                desc_widget = self._get_widget_from_cell(row, 2, ComboBox)
                if desc_widget:
                    # Save current selection
                    current_selection = desc_widget.currentText()

                    # Rebuild combo items
                    desc_widget.blockSignals(True)
                    desc_widget.clear()
                    desc_widget.addItem(self.main.i18n.tr("batch.optional"))
                    for d in self.descriptions:
                        desc_widget.addItem(d)

                    # Restore selection if still valid
                    if current_selection in self.descriptions:
                        desc_widget.setCurrentText(current_selection)
                    else:
                        desc_widget.setCurrentIndex(0)  # Reset to placeholder

                    desc_widget.blockSignals(False)

            InfoBar.success(
                title=self.main.i18n.tr("common.success"),
                content=self.main.i18n.tr("batch.descriptions_refreshed"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=2000,
                parent=self,
            )
        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=f"Failed to refresh descriptions: {str(ex)}",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self,
            )

    def _add_row(self):
        r = self.table.rowCount()
        self.table.setRowCount(r + 1)
        self._setup_table_row(r)
        self._validate()

    def _clear_all(self):
        """Clear all rows with confirmation dialog"""
        # Show confirmation dialog
        w = MessageBox(
            title=self.main.i18n.tr("common.confirm"),
            content=self.main.i18n.tr("batch.clear.confirm"),
            parent=self
        )
        if w.exec():
            self.table.setRowCount(self._base_table_rows)
            for r in range(self.table.rowCount()):
                self._setup_table_row(r)
            self.excel_file_label.setText(self.main.i18n.tr("batch.no_file"))
            if self.current_mode == "excel":
                self.excel_drop.setVisible(True)
                self.table.setVisible(False)
            self._validate()

    def _toggle_all_disclaimers(self, state):
        checked = state == Qt.CheckState.Checked.value
        for row in range(self.table.rowCount()):
            cb = self._get_widget_from_cell(row, 3, CheckBox)
            if cb:
                cb.setChecked(checked)

    def _delete_row(self, row: int):
        if self.table.rowCount() <= 1:
            return
        self.table.removeRow(row)
        self._validate()

    def _get_widget_from_cell(self, row: int, col: int, widget_type):
        cell = self.table.cellWidget(row, col)
        return cell.findChild(widget_type) if cell else None

    def _setup_shortcuts(self):
        """Setup keyboard shortcuts for batch descriptions screen"""
        # Ctrl+Enter to start upload
        start_shortcut = QShortcut(QKeySequence("Ctrl+Return"), self)
        start_shortcut.activated.connect(self._handle_start_shortcut)

        # Escape to cancel/stop upload
        cancel_shortcut = QShortcut(QKeySequence("Escape"), self)
        cancel_shortcut.activated.connect(self._handle_cancel_shortcut)

        # Ctrl+R to retry failed items
        retry_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        retry_shortcut.activated.connect(self._handle_retry_shortcut)

    def _handle_start_shortcut(self):
        """Handle Ctrl+Enter shortcut to start upload"""
        if self.start_btn.isEnabled():
            self._start_upload()

    def _handle_cancel_shortcut(self):
        """Handle Escape shortcut to cancel/stop upload"""
        if self.worker is not None and self.worker.isRunning():
            self.worker.stop()
            InfoBar.warning(
                title=self.main.i18n.tr("batch.cancel.title"),
                content=self.main.i18n.tr("batch.cancel.content"),
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _handle_retry_shortcut(self):
        """Handle Ctrl+R shortcut to retry failed items"""
        if self.retry_btn.isVisible() and self.retry_btn.isEnabled():
            self._retry_failed()

    def _setup_table_row(self, row: int):
        def wrap(widget: QWidget):
            c = QWidget()
            c.setStyleSheet("background: transparent;")
            c.setProperty("ubTableCell", True)
            c.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

            # Allow embedded widgets to shrink to the available cell width.
            try:
                widget.setMinimumWidth(0)
            except Exception:
                pass
            try:
                widget.setSizePolicy(QSizePolicy.Policy.Expanding, widget.sizePolicy().verticalPolicy())
            except Exception:
                pass

            l = QHBoxLayout(c)
            l.setContentsMargins(*TABLE_CELL_MARGINS)
            l.setSpacing(0)
            l.addWidget(widget, 1, Qt.AlignmentFlag.AlignVCenter)
            return c

        brand = ComboBox()
        self._ensure_combo_placeholder_item(
            brand,
            self.main.i18n.tr("batch.select_brand"),
            self.brands,
            force_placeholder=True,
        )
        brand.setMinimumHeight(SIZES['input_height'])
        brand.currentTextChanged.connect(self._validate)
        self.table.setCellWidget(row, 0, wrap(brand))

        code = LineEdit()
        code.setPlaceholderText(self.main.i18n.tr("upload.code.placeholder"))
        code.setMinimumHeight(SIZES['input_height'])
        code.textChanged.connect(lambda _t: self._validate())
        self.table.setCellWidget(row, 1, wrap(code))

        desc = ComboBox()
        desc.setPlaceholderText(self.main.i18n.tr("batch.optional"))
        desc.setMinimumHeight(SIZES['input_height'])
        desc.currentTextChanged.connect(self._validate)
        desc.addItem(self.main.i18n.tr("batch.optional"))
        for d in self.descriptions:
            desc.addItem(d)
        desc.setCurrentIndex(0)
        self.table.setCellWidget(row, 2, wrap(desc))

        disc = CheckBox("")
        disc.setProperty("ubTableCheck", True)
        disc.stateChanged.connect(self._validate)
        disc_container = QWidget()
        disc_container.setStyleSheet("background: transparent;")
        disc_container.setProperty("ubTableCell", True)
        dl = QHBoxLayout(disc_container)
        dl.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        disc_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        dl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        dl.addWidget(disc)
        self.table.setCellWidget(row, 3, disc_container)

        status = BodyLabel(self.main.i18n.tr("batchdesc.status.pending"))
        status.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.table.setCellWidget(row, 4, wrap(status))

        error = LineEdit()
        error.setReadOnly(True)
        error.setMinimumHeight(SIZES['input_height'])
        self.table.setCellWidget(row, 5, wrap(error))

        delete_btn = TransparentToolButton(FluentIcon.DELETE)
        delete_btn.clicked.connect(lambda _=False, r=row: self._delete_row(r))
        delete_btn.setFixedSize(SIZES['button_height_sm'], SIZES['button_height_sm'])
        dc = QWidget()
        dc.setStyleSheet("background: transparent;")
        dc.setProperty("ubTableCell", True)
        dcl = QHBoxLayout(dc)
        dcl.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        dcl.addStretch(1)
        dcl.addWidget(delete_btn, 0, Qt.AlignmentFlag.AlignVCenter)
        dcl.addStretch(1)
        self.table.setCellWidget(row, 6, dc)

        self.table.setRowHeight(row, SIZES['table_row_height_lg'])

    def _collect_items(self) -> list[dict]:
        items = []
        for row in range(self.table.rowCount()):
            b = self._get_widget_from_cell(row, 0, ComboBox)
            c = self._get_widget_from_cell(row, 1, LineEdit)
            d = self._get_widget_from_cell(row, 2, ComboBox)
            disc = self._get_widget_from_cell(row, 3, CheckBox)

            brand = b.currentText().strip() if b else ""
            code_raw = c.text().strip() if c else ""
            code = self._normalize_code(code_raw)
            desc = d.currentText().strip() if d else ""
            disclaimer = disc.isChecked() if disc else False

            desc_selected = bool(desc) and desc != self.main.i18n.tr("batch.optional")
            if brand in self.brands and code and (desc_selected or disclaimer):
                items.append({
                    "row": row,
                    "brand": brand,
                    "code": code,
                    "description": desc if desc_selected else "",
                    "append_disclaimer": disclaimer,
                })
        return items

    def _ensure_combo_placeholder_item(
        self,
        combo: ComboBox,
        placeholder_text: str,
        valid_values: list[str],
        force_placeholder: bool = False,
    ) -> None:
        combo.blockSignals(True)
        try:
            current = combo.currentText().strip()
            keep_current = (not force_placeholder) and (current in valid_values)

            combo.clear()
            combo.addItem(placeholder_text)
            combo.addItems(valid_values)

            try:
                model = combo.model()
                if isinstance(model, QStandardItemModel):
                    item = model.item(0)
                    if item is not None:
                        item.setEnabled(False)
            except Exception:
                pass

            if keep_current:
                combo.setCurrentText(current)
            else:
                combo.setCurrentIndex(0)
        finally:
            combo.blockSignals(False)

    def _validate(self):
        valid = len(self._collect_items())
        self.start_btn.setEnabled(valid > 0 and self.worker is None)

        if self.worker is not None:
            return

        # Don't show a "Ready" hint; keep Start enabled/disabled silently.
        self.status_label.setText("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(False)

    def _set_busy(self, busy: bool):
        self.progress_ring.setVisible(busy)
        self.add_btn.setEnabled(not busy)
        self.clear_btn.setEnabled(not busy)
        self.template_btn.setEnabled(not busy)
        self.browse_btn.setEnabled(not busy)
        self.refresh_btn.setEnabled(not busy)
        self.manual_pill.setEnabled(not busy)
        self.excel_pill.setEnabled(not busy)
        # Keep table enabled so user can see status updates and scroll
        # self.table.setEnabled(not busy)  # Commented out - table stays interactive
        self.disclaimer_bulk.setEnabled(not busy)
        self.start_btn.setEnabled(not busy and len(self._collect_items()) > 0)

    def _start_batch(self):
        tr = self.main.i18n.tr
        if getattr(self.main, "driver", None) is None:
            InfoBar.error(
                title=tr("common.error"),
                content=tr("batchdesc.no_session"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=5000,
                parent=self,
            )
            return

        items = self._collect_items()
        if not items:
            return

        # Initialize tracking variable for row highlighting
        self._current_processing_row = None

        for row in range(self.table.rowCount()):
            # Clear any previous row highlighting
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(Qt.GlobalColor.transparent)

            # Reset status (column 4)
            s = self._get_widget_from_cell(row, 4, BodyLabel)
            if s:
                s.setText(tr("batchdesc.status.pending"))
                s.setStyleSheet(f"color: {COLORS['text_secondary']};")
            # Reset error (column 5)
            e = self._get_widget_from_cell(row, 5, LineEdit)
            if e:
                e.setText("")
                e.setStyleSheet("")

        # Check if multi-session is enabled
        multi_session_enabled = self.main.settings.get('multi_session_enabled', False)
        browser_count = self.main.settings.get('browser_count', 2)

        # Initialize session manager if multi-session enabled
        if multi_session_enabled:
            if self.session_manager is None or not self.session_manager.is_ready():
                from Managers.BrowserSessionManager import BrowserSessionManager

                # Get browser type from settings
                browser_type = self.main.settings.get('browser_choice', 'Chrome')

                InfoBar.info(
                    title=tr("common.info"),
                    content=f"Initializing {browser_count} browser instances...",
                    orient=Qt.Orientation.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                    parent=self
                )

                # Create and initialize session pool
                self.session_manager = BrowserSessionManager(
                    browser_type=browser_type,
                    pool_size=browser_count,
                    logger=self.main.logger
                )

                if not self.session_manager.initialize_pool():
                    InfoBar.error(
                        title=tr("common.error"),
                        content="Failed to initialize browser pool. Falling back to single-session mode.",
                        orient=Qt.Orientation.Horizontal,
                        isClosable=True,
                        position=InfoBarPosition.TOP,
                        duration=5000,
                        parent=self
                    )
                    self.session_manager = None
                    multi_session_enabled = False

        InfoBar.info(
            title=tr("batchdesc.run.started.title"),
            content=tr("batchdesc.run.started.content", count=len(items)),
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self,
        )

        self._set_busy(True)

        # Choose worker based on multi-session setting
        if multi_session_enabled and self.session_manager and self.session_manager.is_ready():
            # Use parallel worker with session pool
            self.worker = ParallelBatchDescriptionsWorker(self.main, self.session_manager, items)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self._on_log)
            self.worker.done.connect(self._on_done)
            self.worker.session_status.connect(self._on_session_status_update)
            self.worker.start()
        else:
            # Use single-session worker
            self.worker = BatchDescriptionsWorker(self.main, items)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self._on_log)
            self.worker.progress_update.connect(self._on_progress_update)
            self.worker.done.connect(self._on_done)
            self.worker.start()

    def _on_progress_update(self, current: int, total: int, speed: float, eta_seconds: float):
        """Handle progress updates with ETA from worker thread."""
        percentage = (current / total * 100) if total > 0 else 0

        # Format ETA
        if eta_seconds < 60:
            eta_text = f"{int(eta_seconds)}s"
        elif eta_seconds < 3600:
            minutes = int(eta_seconds / 60)
            seconds = int(eta_seconds % 60)
            eta_text = f"{minutes}m {seconds}s"
        else:
            hours = int(eta_seconds / 3600)
            minutes = int((eta_seconds % 3600) / 60)
            eta_text = f"{hours}h {minutes}m"

        # Format speed (items per minute)
        items_per_min = speed * 60

        # Update status label with progress information
        message = f"Processing ({current}/{total}) {percentage:.0f}% - ETA: {eta_text} | {items_per_min:.1f} items/min"
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(True)

    def _on_log(self, message: str):
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(bool((message or "").strip()))

    def _on_session_status_update(self, stats: dict):
        """Handle session status updates from parallel worker (optional)"""
        # BatchDescriptionsScreen doesn't have a session status label yet
        # This is just for future extensibility
        pass

    def _on_row_update(self, row: int, status_text: str, error_text: str):
        """Update row status and highlight currently processing row."""
        # Clear previous row highlighting
        if hasattr(self, '_current_processing_row') and self._current_processing_row is not None:
            for col in range(self.table.columnCount()):
                item = self.table.item(self._current_processing_row, col)
                if item:
                    item.setBackground(Qt.GlobalColor.transparent)

        # Update status
        status = self._get_widget_from_cell(row, 4, BodyLabel)
        if status:
            status.setText(status_text)
            tr = self.main.i18n.tr

            # Highlight current row if running
            if status_text == tr("batchdesc.status.running"):
                status.setStyleSheet(f"color: {COLORS['blush_rose']}; font-weight: 600;")
                # Highlight the entire row
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table.setItem(row, col, item)
                    item.setBackground(QColor(COLORS['blush_rose']).lighter(180) if isDarkTheme() else QColor(COLORS['blush_rose']).lighter(195))
                self._current_processing_row = row
                # Scroll to current row
                self.table.scrollToItem(self.table.item(row, 0), QTableWidget.ScrollHint.PositionAtCenter)
            elif status_text == tr("batchdesc.status.ok"):
                status.setStyleSheet(f"color: {COLORS['success']}; font-weight: 600;")
                # Light green background for success
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table.setItem(row, col, item)
                    item.setBackground(QColor(COLORS['success']).lighter(190) if isDarkTheme() else QColor(COLORS['success']).lighter(195))
            elif status_text == tr("batchdesc.status.failed"):
                status.setStyleSheet(f"color: {COLORS['flag_red']}; font-weight: 600;")
                # Light red background for failure
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    if not item:
                        item = QTableWidgetItem()
                        self.table.setItem(row, col, item)
                    item.setBackground(QColor(COLORS['flag_red']).lighter(190) if isDarkTheme() else QColor(COLORS['flag_red']).lighter(195))
            else:
                status.setStyleSheet(f"color: {COLORS['text_secondary']};")

        # Update error message
        err = self._get_widget_from_cell(row, 5, LineEdit)
        if err:
            err.setText(error_text or "")
            if error_text:
                err.setStyleSheet(f"color: {COLORS['flag_red']}; font-weight: 500;")

    def _on_done(self, ok: int, total: int):
        tr = self.main.i18n.tr
        self.worker = None
        self._set_busy(False)
        self._validate()

        # Clear the last processing row highlight
        if hasattr(self, '_current_processing_row') and self._current_processing_row is not None:
            # Don't clear it - keep the final status color
            self._current_processing_row = None

        failed = total - ok
        if failed > 0:
            # Show warning if some failed
            InfoBar.warning(
                title=tr("batchdesc.run.complete.title"),
                content=tr("batchdesc.run.complete.partial", ok=ok, failed=failed, total=total),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=8000,
                parent=self,
            )
        else:
            # Show success if all succeeded
            InfoBar.success(
                title=tr("batchdesc.run.complete.title"),
                content=tr("batchdesc.run.complete.content", ok=ok, total=total),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=6000,
                parent=self,
            )

        # Windows notification
        self._show_windows_notification(ok, total)

    def _show_windows_notification(self, ok: int, total: int):
        """Show Windows desktop notification for batch completion."""
        # NOTE: win10toast relies on pywin32 WNDPROC callbacks that are known to
        # break on newer Python versions (e.g. 3.12+), causing hard crashes like:
        # "WNDPROC return value cannot be converted to LRESULT".
        try:
            import sys
            if sys.version_info >= (3, 12):
                return
        except Exception:
            return

        try:
            from win10toast import ToastNotifier
            toaster = ToastNotifier()

            failed = total - ok
            if failed > 0:
                title = "Batch Descriptions Update Completed with Errors"
                message = f"Success: {ok}/{total}\nFailed: {failed}"
            else:
                title = "Batch Descriptions Update Completed Successfully"
                message = f"All {total} descriptions updated successfully"

            # Show notification (non-blocking, 10 second duration)
            toaster.show_toast(
                title=title,
                msg=message,
                duration=10,
                threaded=True,
                icon_path=None  # Uses default Windows icon
            )
        except ImportError:
            # win10toast not installed, silently skip
            pass
        except Exception:
            # Any other error, silently skip
            pass

    def _ensure_table_fills_viewport(self):
        if not self.table.isVisible() or self.excel_drop.isVisible():
            return
        viewport_height = self.table.viewport().height()
        if viewport_height <= 0:
            return
        row_height = self.table.verticalHeader().defaultSectionSize() or 56
        target = (viewport_height + row_height - 1) // row_height + 1
        # Remove minimum row limit
        # Remove both minimum and maximum row limits
        target = int(target)
        current = self.table.rowCount()
        if current >= target:
            return
        self.table.setRowCount(target)
        for r in range(current, target):
            self._setup_table_row(r)


# Place static method at the end of the class
    @staticmethod
    def _normalize_code(raw: str) -> str:
        """Normalize product code to always include UB- prefix.
        Mirrors BatchTitles/BatchUploadDialog behavior.
        """
        code = (raw or "").strip()
        if not code:
            return ""
        if code.upper().startswith("UB-"):
            return code
        return f"UB-{code}"
