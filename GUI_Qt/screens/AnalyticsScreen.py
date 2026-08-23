"""
Analytics Screen - Comprehensive dashboard with charts and insights
Modern analytics interface with visual data representation
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QBoxLayout, QGridLayout, QLabel, QFileDialog
)
from PySide6.QtCore import Qt
from PySide6.QtGui import QPainter, QColor, QPen, QFont
from qfluentwidgets import (
    CardWidget, TitleLabel, StrongBodyLabel, BodyLabel, CaptionLabel,
    ComboBox, PushButton, TransparentToolButton, FluentIcon,
    InfoBar, InfoBarPosition, ScrollArea, isDarkTheme, IconWidget,
    ProgressRing, IndeterminateProgressRing, qconfig
)
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import (
    COLORS, FONTS, RADII, SPACING, SIZES, rgba_from_hex,
    get_status_text_color, get_surface_color, get_text_color,
)
from GUI_Qt.styles.screen_theme import (
    apply_screen_theme, PAGE_MARGINS, PAGE_SPACING, CARD_MARGINS, CARD_SPACING,
    ICON_TEXT_GAP, TIGHT_SPACING, get_responsive_margins, get_responsive_spacing
)
from GUI_Qt.styles.screen_theme import enforce_transparent_labels
from GUI_Qt.earnings.presentation import money
from Managers.AnalyticsManager import AnalyticsManager


class MetricCard(CardWidget):
    """Large metric card with trend indicator"""

    def __init__(self, title, value, subtitle, icon, trend=None, trend_positive=True, parent=None):
        super().__init__(parent)
        self.setBorderRadius(RADII['md'])

        self._trend_positive = True

        # Apply proper card styling
        self._apply_theme()

        layout = QVBoxLayout(self)
        layout.setContentsMargins(*CARD_MARGINS)
        layout.setSpacing(SPACING['md'])

        # Header with icon
        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)

        icon_widget = IconWidget(icon)
        icon_widget.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        icon_widget.setStyleSheet(f"color: {COLORS['space_indigo']}; background: transparent; background-color: transparent;")

        # Keep title explicit and readable (icons alone are ambiguous)
        self.title_label = BodyLabel(str(title))
        self.title_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')}; font-weight: 600; background: transparent; background-color: transparent;")

        header.addWidget(icon_widget)
        header.addWidget(self.title_label)
        header.addStretch()

        layout.addLayout(header)

        # Value
        self.value_label = QLabel(str(value))
        self.value_label.setStyleSheet(f"""
            font-size: 36px;
            font-weight: 700;
            color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']};
        """)
        layout.addWidget(self.value_label)

        # Subtitle with optional trend
        bottom_layout = QHBoxLayout()
        self.subtitle_label = CaptionLabel(str(subtitle))
        self.subtitle_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')}; background: transparent; background-color: transparent;")
        bottom_layout.addWidget(self.subtitle_label)

        self.trend_label = CaptionLabel("")
        bottom_layout.addWidget(self.trend_label)

        bottom_layout.addStretch()
        layout.addLayout(bottom_layout)

        self.update_values(value=value, subtitle=subtitle, trend=trend, trend_positive=trend_positive)

        # Connect to theme changes
        qconfig.themeChangedFinished.connect(self._apply_theme)

    def update_values(self, *, value, subtitle, trend=None, trend_positive=True):
        self.value_label.setText(str(value))
        self.subtitle_label.setText(str(subtitle))

        if trend is None:
            self.trend_label.setVisible(False)
            self.trend_label.setText("")
            return

        self.trend_label.setVisible(True)
        trend_color = get_status_text_color(
            "success" if trend_positive else "error", isDarkTheme()
        )
        trend_icon = "↑" if trend_positive else "↓"
        self.trend_label.setText(f"{trend_icon} {abs(trend):.1f}%")
        self.trend_label.setStyleSheet(f"color: {trend_color}; font-weight: 600; background: transparent; background-color: transparent;")

    def _apply_theme(self):
        """Apply proper theming to card"""
        is_dark = isDarkTheme()
        card_bg = rgba_from_hex(COLORS['text_white'], 0.03) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.03)
        card_border = COLORS['border_dark'] if is_dark else COLORS['border_light']

        # Apply directly to this widget (avoid fragile class-name selectors in QSS)
        self.setStyleSheet(f"""
            background-color: {card_bg};
            border: 1px solid {card_border};
            border-radius: {RADII['md']}px;
        """)


class SimpleBarChart(QWidget):
    """Simple bar chart widget"""

    def __init__(self, data, labels, max_value=None, parent=None):
        super().__init__(parent)
        self.data = data  # List of values
        self.labels = labels  # List of labels
        self.max_value = max_value or (max(data) if data else 1)
        self.setMinimumHeight(200)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        width = self.width()
        height = self.height()

        if not self.data or not self.labels:
            return

        bar_count = len(self.data)
        bar_spacing = 10
        total_spacing = bar_spacing * (bar_count + 1)
        bar_width = (width - total_spacing) / bar_count

        # Draw bars
        for i, (value, label) in enumerate(zip(self.data, self.labels)):
            x = bar_spacing + i * (bar_width + bar_spacing)
            bar_height = (value / self.max_value) * (height - 60) if self.max_value > 0 else 0
            y = height - bar_height - 40

            # Bar
            is_dark = isDarkTheme()
            bar_color = QColor(COLORS['space_indigo'])
            painter.setBrush(bar_color)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(int(x), int(y), int(bar_width), int(bar_height), 4, 4)

            # Value label
            font = QFont("Segoe UI", 10, QFont.Weight.Bold)
            painter.setFont(font)
            value_text = str(int(value))
            label_h = 20
            if y >= label_h + 2:
                # Enough room above the bar — draw above
                painter.setPen(QColor(COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']))
                painter.drawText(int(x), int(y - label_h), int(bar_width), label_h, Qt.AlignmentFlag.AlignCenter, value_text)
            else:
                # Bar is too tall — draw inside the bar at the top
                painter.setPen(QColor(COLORS['text_white']))
                painter.drawText(int(x), int(y + 4), int(bar_width), label_h, Qt.AlignmentFlag.AlignCenter, value_text)

            # Label at bottom
            painter.setPen(QColor(get_text_color(isDarkTheme(), 'secondary')))
            font = QFont("Segoe UI", 9)
            painter.setFont(font)
            painter.drawText(int(x), height - 30, int(bar_width), 20, Qt.AlignmentFlag.AlignCenter, label)


class DonutChart(QWidget):
    """Simple donut chart for success rate"""

    def __init__(self, percentage, parent=None):
        super().__init__(parent)
        self.percentage = max(0, min(100, percentage))
        self.setFixedSize(180, 180)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        def _blend(hex_a: str, hex_b: str, t: float) -> QColor:
            """Blend color A towards color B by t (0..1)."""
            t = max(0.0, min(1.0, float(t)))
            a = QColor(hex_a)
            b = QColor(hex_b)
            return QColor(
                int(a.red() * (1 - t) + b.red() * t),
                int(a.green() * (1 - t) + b.green() * t),
                int(a.blue() * (1 - t) + b.blue() * t),
                255,
            )

        # Center and radius
        cx, cy = self.width() // 2, self.height() // 2
        outer_radius = 70
        inner_radius = 50

        bg_hex = COLORS['space_indigo'] if isDarkTheme() else COLORS['platinum']

        # Background ring
        painter.setBrush(_blend(get_text_color(isDarkTheme(), 'tertiary'), bg_hex, 0.85))
        painter.setPen(Qt.PenStyle.NoPen)
        painter.drawEllipse(cx - outer_radius, cy - outer_radius, outer_radius * 2, outer_radius * 2)

        # Segments (softened towards background, but opaque to avoid muddy overlap)
        success_fill = _blend(COLORS['success'], bg_hex, 0.25)
        error_fill = _blend(COLORS['error'], bg_hex, 0.35)
        span_angle = int(360 * 16 * (self.percentage / 100))
        # Draw success first (from top), then remaining (fail) after it.
        if span_angle > 0:
            painter.setBrush(success_fill)
            painter.drawPie(
                cx - outer_radius,
                cy - outer_radius,
                outer_radius * 2,
                outer_radius * 2,
                90 * 16,
                -span_angle,
            )
        if span_angle < 360 * 16:
            painter.setBrush(error_fill)
            painter.drawPie(
                cx - outer_radius,
                cy - outer_radius,
                outer_radius * 2,
                outer_radius * 2,
                90 * 16 - span_angle,
                -(360 * 16 - span_angle),
            )

        # Inner circle (creates donut)
        painter.setBrush(QColor(bg_hex))
        painter.drawEllipse(cx - inner_radius, cy - inner_radius, inner_radius * 2, inner_radius * 2)

        # Percentage text
        painter.setPen(QColor(COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']))
        font = QFont("Segoe UI", 24, QFont.Weight.Bold)
        painter.setFont(font)
        painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, f"{int(self.percentage)}%")


class AnalyticsScreen(ResponsiveWidget):
    """Analytics dashboard with comprehensive metrics and visualizations"""

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.db = main_window.db
        self.analytics = AnalyticsManager(self.db)

        self._init_ui()
        self._apply_theme()
        self._apply_card_themes()  # Apply proper theming to all cards
        enforce_transparent_labels(self)
        self.load_analytics()
        self.update_translations()

        # Connect to theme change signal
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def _init_ui(self):
        """Initialize UI"""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(*PAGE_MARGINS)
        main_layout.setSpacing(PAGE_SPACING)

        # === HEADER ===
        header_layout = QHBoxLayout()
        header_layout.setSpacing(ICON_TEXT_GAP)

        # Header icon
        header_icon = IconWidget(FluentIcon.PIE_SINGLE)
        header_icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])

        # Title
        self.title_label = TitleLabel("")

        header_layout.addWidget(header_icon)
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()

        main_layout.addLayout(header_layout)

        # Scroll area
        scroll = ScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(ScrollArea.Shape.NoFrame)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        # Reserve space so the (overlay) scrollbar doesn't cover card content
        content_layout.setContentsMargins(0, 0, SIZES['scrollbar_thickness'] + SPACING['xs'], 0)
        content_layout.setSpacing(CARD_SPACING)

        # Time range selector
        toolbar = self._create_toolbar()
        content_layout.addWidget(toolbar)

        # Business totals. Earnings is the product ledger, so these include
        # manually recorded work completed outside the application.
        self.metrics_grid = QGridLayout()
        self.metrics_grid.setSpacing(CARD_SPACING)

        # Make cards share available width evenly
        for col in range(4):
            self.metrics_grid.setColumnStretch(col, 1)

        self.products_metric = MetricCard("Recorded products", "0", "", FluentIcon.SHOPPING_CART)
        self.earnings_metric = MetricCard("Earnings", "€0.00", "", FluentIcon.TAG)
        self.manual_products_metric = MetricCard("Outside the app", "0", "", FluentIcon.GLOBE)
        self.app_products_metric = MetricCard("App-linked", "0", "", FluentIcon.CONNECT)

        self.metric_cards = [
            self.products_metric,
            self.earnings_metric,
            self.manual_products_metric,
            self.app_products_metric,
        ]
        for column, card in enumerate(self.metric_cards):
            self.metrics_grid.addWidget(card, 0, column)

        content_layout.addLayout(self.metrics_grid)

        # Charts row
        self.charts_layout = QHBoxLayout()
        self.charts_layout.setSpacing(CARD_SPACING)

        # Brand performance chart
        self.brand_chart_card = CardWidget()
        self.brand_chart_card.setBorderRadius(RADII['md'])
        brand_chart_layout = QVBoxLayout(self.brand_chart_card)
        brand_chart_layout.setContentsMargins(*CARD_MARGINS)
        brand_chart_layout.setSpacing(SPACING['md'])

        self.brand_title = StrongBodyLabel("Brand Performance")
        self.brand_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        brand_chart_layout.addWidget(self.brand_title)

        self.brand_chart = SimpleBarChart([], [])
        brand_chart_layout.addWidget(self.brand_chart)

        self.charts_layout.addWidget(self.brand_chart_card, 2)

        # Success donut chart
        self.success_chart_card = CardWidget()
        self.success_chart_card.setBorderRadius(RADII['md'])
        success_chart_layout = QVBoxLayout(self.success_chart_card)
        success_chart_layout.setContentsMargins(*CARD_MARGINS)
        success_chart_layout.setSpacing(SPACING['md'])
        success_chart_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        self.success_title = StrongBodyLabel("Success Rate")
        self.success_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        success_chart_layout.addWidget(self.success_title, 0, Qt.AlignmentFlag.AlignCenter)

        self.success_donut = DonutChart(0)
        success_chart_layout.addWidget(self.success_donut, 0, Qt.AlignmentFlag.AlignCenter)

        self.success_caption = CaptionLabel("Last 30 days")
        self.success_caption.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'tertiary')}; background: transparent; background-color: transparent;")
        success_chart_layout.addWidget(self.success_caption, 0, Qt.AlignmentFlag.AlignCenter)

        self.charts_layout.addWidget(self.success_chart_card, 1)

        content_layout.addLayout(self.charts_layout)

        # Product mix from the Earnings ledger.
        self.breakdowns_layout = QHBoxLayout()
        self.breakdowns_layout.setSpacing(CARD_SPACING)

        self.source_chart_card = CardWidget()
        self.source_chart_card.setBorderRadius(RADII['md'])
        source_layout = QVBoxLayout(self.source_chart_card)
        source_layout.setContentsMargins(*CARD_MARGINS)
        source_layout.setSpacing(SPACING['md'])
        self.source_title = StrongBodyLabel("Products by source")
        self.source_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        source_layout.addWidget(self.source_title)
        self.source_chart = SimpleBarChart([], [])
        source_layout.addWidget(self.source_chart)
        self.breakdowns_layout.addWidget(self.source_chart_card, 1)

        self.type_chart_card = CardWidget()
        self.type_chart_card.setBorderRadius(RADII['md'])
        type_layout = QVBoxLayout(self.type_chart_card)
        type_layout.setContentsMargins(*CARD_MARGINS)
        type_layout.setSpacing(SPACING['md'])
        self.type_title = StrongBodyLabel("Products by type")
        self.type_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        type_layout.addWidget(self.type_title)
        self.type_chart = SimpleBarChart([], [])
        type_layout.addWidget(self.type_chart)
        self.breakdowns_layout.addWidget(self.type_chart_card, 1)

        content_layout.addLayout(self.breakdowns_layout)

        # Recent activity card
        self.activity_card = CardWidget()
        self.activity_card.setBorderRadius(RADII['md'])
        activity_layout = QVBoxLayout(self.activity_card)
        activity_layout.setContentsMargins(*CARD_MARGINS)
        activity_layout.setSpacing(SPACING['md'])

        activity_header = QHBoxLayout()
        self.activity_title = StrongBodyLabel("Recent Activity")
        self.activity_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        activity_header.addWidget(self.activity_title)
        activity_header.addStretch()

        self.view_all_btn = TransparentToolButton(FluentIcon.RIGHT_ARROW)
        self.view_all_btn.setToolTip("View full history")
        self.view_all_btn.clicked.connect(self._view_earnings)
        activity_header.addWidget(self.view_all_btn)

        activity_layout.addLayout(activity_header)

        self.activity_container = QVBoxLayout()
        self.activity_container.setSpacing(TIGHT_SPACING)
        activity_layout.addLayout(self.activity_container)

        content_layout.addWidget(self.activity_card)

        # Top errors card
        self.errors_card = CardWidget()
        self.errors_card.setBorderRadius(RADII['md'])
        errors_layout = QVBoxLayout(self.errors_card)
        errors_layout.setContentsMargins(*CARD_MARGINS)
        errors_layout.setSpacing(SPACING['md'])

        self.errors_title = StrongBodyLabel("Common Issues")
        self.errors_title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']};")
        errors_layout.addWidget(self.errors_title)

        self.errors_container = QVBoxLayout()
        self.errors_container.setSpacing(TIGHT_SPACING)
        errors_layout.addLayout(self.errors_container)

        content_layout.addWidget(self.errors_card)

        content_layout.addStretch()
        scroll.setWidget(content)
        main_layout.addWidget(scroll)

        self.scroll = scroll
        self.content_widget = content

    def _apply_theme(self):
        """Apply consistent screen theming (background + scroll viewport + label transparency)."""
        apply_screen_theme(
            self,
            "AnalyticsScreen",
            scroll=getattr(self, "scroll", None),
            content=getattr(self, "content_widget", None),
            transparent_labels=True,
        )

        # Rounded scrollbars for the main scroll area
        try:
            scroll = getattr(self, "scroll", None)
            if scroll is not None:
                is_dark = isDarkTheme()
                bg_color = get_surface_color(is_dark, 'canvas')
                handle = rgba_from_hex(COLORS['lavender_grey'], 0.35) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.22)
                handle_hover = rgba_from_hex(COLORS['lavender_grey'], 0.55) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.32)
                scroll.setStyleSheet(
                    (scroll.styleSheet() or "")
                    + f"""
                    QScrollBar:vertical {{
                        background: transparent;
                        width: {SIZES['scrollbar_thickness']}px;
                        margin: 0px;
                    }}
                    QScrollBar::handle:vertical {{
                        background: {handle};
                        border-radius: {RADII['sm']}px;
                        min-height: {SIZES['scrollbar_handle_min']}px;
                    }}
                    QScrollBar::handle:vertical:hover {{
                        background: {handle_hover};
                    }}
                    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                        height: 0px;
                    }}
                    QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                        background: none;
                    }}

                    QScrollBar:horizontal {{
                        background: transparent;
                        height: {SIZES['scrollbar_thickness']}px;
                        margin: 0px;
                    }}
                    QScrollBar::handle:horizontal {{
                        background: {handle};
                        border-radius: {RADII['sm']}px;
                        min-width: {SIZES['scrollbar_handle_min']}px;
                    }}
                    QScrollBar::handle:horizontal:hover {{
                        background: {handle_hover};
                    }}
                    QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                        width: 0px;
                    }}
                    QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                        background: none;
                    }}
                    """
                )
                try:
                    scroll.setViewportMargins(0, 0, SIZES['scrollbar_thickness'] + SPACING['xs'], 0)
                except Exception:
                    pass
        except Exception:
            pass
        enforce_transparent_labels(self)

    def _create_toolbar(self):
        """Create toolbar with filters"""
        self.toolbar_card = CardWidget()
        self.toolbar_card.setBorderRadius(RADII['md'])
        toolbar_layout = QHBoxLayout(self.toolbar_card)
        toolbar_layout.setContentsMargins(*CARD_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        # Time range filter
        self.period_label = BodyLabel("Period:")
        toolbar_layout.addWidget(self.period_label)

        self.period_combo = ComboBox()
        self.period_combo.addItem("Last 7 days", userData=7)
        self.period_combo.addItem("Last 30 days", userData=30)
        self.period_combo.addItem("Last 90 days", userData=90)
        self.period_combo.addItem("All time", userData=None)
        self.period_combo.setCurrentIndex(1)  # Default: Last 30 days
        self.period_combo.currentIndexChanged.connect(lambda _index: self.load_analytics())
        toolbar_layout.addWidget(self.period_combo)

        toolbar_layout.addStretch()

        # Refresh button
        self.refresh_btn = TransparentToolButton(FluentIcon.SYNC)
        self.refresh_btn.setToolTip("Refresh analytics")
        self.refresh_btn.clicked.connect(self.load_analytics)
        toolbar_layout.addWidget(self.refresh_btn)

        # Export button
        self.export_btn = PushButton("Export")
        self.export_btn.setIcon(FluentIcon.SAVE.icon())
        self.export_btn.clicked.connect(self._export_analytics)
        toolbar_layout.addWidget(self.export_btn)

        return self.toolbar_card

    def _apply_card_themes(self):
        """Apply proper theming to all cards"""
        is_dark = isDarkTheme()

        # Chart/content cards - subtle tinted backgrounds
        card_bg = rgba_from_hex(COLORS['text_white'], 0.03) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.03)
        card_border = COLORS['border_dark'] if is_dark else COLORS['border_light']

        # Apply directly on each card instance (avoid selector scoping issues)
        card_style = f"""
            background-color: {card_bg};
            border: 1px solid {card_border};
            border-radius: {RADII['md']}px;
        """

        # Apply to all content cards
        for card_attr in ['toolbar_card', 'brand_chart_card', 'success_chart_card',
                         'source_chart_card', 'type_chart_card', 'activity_card', 'errors_card']:
            if hasattr(self, card_attr):
                card_widget = getattr(self, card_attr)
                if card_widget is not None:
                    card_widget.setStyleSheet(card_style)

    def load_analytics(self):
        """Load business totals from Earnings and reliability from app history."""
        period = self.period_combo.currentText()
        days = self.period_combo.currentData()
        values = self.analytics.snapshot(days=days)
        earnings = values["earnings"]
        processing = values["processing"]
        tr = self.main.i18n.tr

        self._update_metric(
            self.products_metric, f"{earnings['product_count']:,}", period
        )
        self._update_metric(
            self.earnings_metric, money(earnings["earned_cents"]), period
        )
        self._update_metric(
            self.manual_products_metric,
            f"{earnings['manual_count']:,}",
            tr("analytics.metric.manual.subtitle"),
        )
        self._update_metric(
            self.app_products_metric,
            f"{earnings['app_linked_count']:,}",
            tr("analytics.metric.app_linked.subtitle"),
        )

        self.success_donut.percentage = processing["success_rate"]
        self.success_donut.update()
        self.success_caption.setText(
            tr(
                "analytics.success.caption",
                period=period,
                count=processing["total"],
            )
        )

        self._update_bar_chart(
            self.brand_chart,
            values["brands"],
            lambda label: (
                tr("analytics.brand.none") if label == "No brand" else str(label)
            ),
        )
        self._update_bar_chart(
            self.source_chart,
            values["sources"],
            lambda label: tr(
                {
                    "manual": "analytics.source.manual",
                    "regular_upload": "analytics.source.regular",
                    "batch_upload": "analytics.source.batch",
                }.get(str(label), "analytics.source.other")
            ),
        )
        self._update_bar_chart(
            self.type_chart,
            values["product_types"],
            lambda label: tr(
                {
                    "bicycle": "analytics.type.bicycle",
                    "frameset": "analytics.type.frameset",
                    "other": "analytics.type.other",
                }.get(str(label), "analytics.type.other")
            ),
        )

        self._clear_layout(self.activity_container)
        for row in values["recent"]:
            self.activity_container.addWidget(self._create_activity_item(row))
        if not values["recent"]:
            empty = CaptionLabel(tr("analytics.activity.empty"))
            self.activity_container.addWidget(empty)

        self._clear_layout(self.errors_container)
        if values["errors"]:
            for row in values["errors"]:
                self.errors_container.addWidget(
                    self._create_error_item(row["message"], row["count"])
                )
        else:
            no_errors = CaptionLabel(tr("analytics.errors.empty"))
            no_errors.setStyleSheet(
                f"color: {get_status_text_color('success', isDarkTheme())}; "
                "font-style: italic; background: transparent; background-color: transparent;"
            )
            self.errors_container.addWidget(no_errors)

    def showEvent(self, event):
        """Refresh when returning from Earnings or an upload workflow."""
        super().showEvent(event)
        self.load_analytics()

    @staticmethod
    def _update_bar_chart(chart, rows, label_formatter=lambda value: str(value)):
        chart.data = [int(row["count"] or 0) for row in rows]
        chart.labels = [label_formatter(row["label"]) for row in rows]
        chart.max_value = max(chart.data) if chart.data else 1
        chart.update()

    def _update_metric(self, card, value, subtitle, trend=None, trend_positive=True):
        """Update a metric card's values"""
        if hasattr(card, "update_values"):
            card.update_values(value=value, subtitle=subtitle, trend=trend, trend_positive=trend_positive)
            return

        # Fallback for older card variants
        for child in card.findChildren(QLabel):
            if "font-size: 36px" in child.styleSheet():
                child.setText(str(value))
                break

    def _create_activity_item(self, row):
        """Create a recent paid-product item from the Earnings ledger."""
        item = QWidget()
        item.setStyleSheet("background: transparent; background-color: transparent; border: none;")
        layout = QHBoxLayout(item)
        layout.setContentsMargins(SPACING['sm'], SPACING['sm'], SPACING['sm'], SPACING['sm'])
        layout.setSpacing(CARD_SPACING)

        tr = self.main.i18n.tr
        source = str(row.get("source") or "manual")
        source_key = {
            "manual": "analytics.source.manual",
            "regular_upload": "analytics.source.regular",
            "batch_upload": "analytics.source.batch",
        }.get(source, "analytics.source.other")
        source_icon = {
            "manual": FluentIcon.GLOBE,
            "regular_upload": FluentIcon.UP,
            "batch_upload": FluentIcon.FOLDER,
        }.get(source, FluentIcon.TAG)
        status_color = get_status_text_color('success', isDarkTheme())
        icon = IconWidget(source_icon)
        icon.setFixedSize(SIZES['icon_sm'], SIZES['icon_sm'])
        icon.setStyleSheet(f"color: {status_color}; background: transparent; background-color: transparent;")
        layout.addWidget(icon)

        brand = row.get("brand") or tr("analytics.brand.none")
        if brand == "No brand":
            brand = tr("analytics.brand.none")
        text = BodyLabel(
            f"{tr(source_key)} • {brand} • {row.get('sku', '')} • {money(row.get('payout_cents'))}"
        )
        if row.get("product_name"):
            text.setToolTip(str(row["product_name"]))
        text.setStyleSheet(
            f"color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']};"
            "background: transparent; background-color: transparent; border: none;"
        )
        layout.addWidget(text)

        layout.addStretch()

        # Time
        try:
            timestamp = row.get("earned_at")
            if isinstance(timestamp, str):
                dt = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
            else:
                dt = timestamp
            time_ago = self._time_ago(dt)
            time_label = CaptionLabel(time_ago)
            time_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'tertiary')}; background: transparent; background-color: transparent; border: none;")
            layout.addWidget(time_label)
        except:
            pass

        return item

    def _create_error_item(self, error_msg, count):
        """Create an error item"""
        item = QWidget()
        item.setStyleSheet("background: transparent; background-color: transparent; border: none;")
        layout = QHBoxLayout(item)
        layout.setContentsMargins(SPACING['sm'], SPACING['sm'], SPACING['sm'], SPACING['sm'])
        layout.setSpacing(CARD_SPACING)

        # Count badge
        count_label = StrongBodyLabel(f"{count}×")
        count_label.setFixedWidth(50)
        count_label.setStyleSheet(f"color: {get_status_text_color('error', isDarkTheme())}; font-weight: 700; background: transparent; background-color: transparent; border: none;")
        layout.addWidget(count_label)

        # Error message (truncated)
        display_error = error_msg if len(error_msg) <= 80 else error_msg[:77] + "..."
        error_label = BodyLabel(display_error)
        error_label.setWordWrap(False)
        error_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')}; background: transparent; background-color: transparent; border: none;")
        layout.addWidget(error_label, 1)

        return item

    def _time_ago(self, dt):
        """Convert datetime to relative time"""
        now = datetime.now(dt.tzinfo) if getattr(dt, "tzinfo", None) else datetime.now()
        diff = now - dt
        if diff.total_seconds() < 0:
            diff = timedelta(0)

        if diff.days > 0:
            return self.main.i18n.tr("analytics.time.days", count=diff.days)
        elif diff.seconds >= 3600:
            return self.main.i18n.tr("analytics.time.hours", count=diff.seconds // 3600)
        elif diff.seconds >= 60:
            return self.main.i18n.tr("analytics.time.minutes", count=diff.seconds // 60)
        else:
            return self.main.i18n.tr("analytics.time.now")

    def _clear_layout(self, layout):
        """Clear all widgets from layout"""
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

    def _view_earnings(self):
        """Open the ledger behind the business analytics."""
        self.main.open_route("earnings")

    def _export_analytics(self):
        """Export the same combined period currently shown on the dashboard."""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                self.main.i18n.tr("analytics.export.dialog"),
                f"analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Summary"

            headers = ["Metric", "Value"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4F46E5", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            values = self.analytics.snapshot(days=self.period_combo.currentData())
            earnings = values["earnings"]
            processing = values["processing"]
            data = [
                ["Period", self.period_combo.currentText()],
                ["Recorded products", earnings["product_count"]],
                ["Earnings", earnings["earned_cents"] / 100.0],
                ["Outside-app products", earnings["manual_count"]],
                ["App-linked products", earnings["app_linked_count"]],
                ["App activity records", processing["total"]],
                ["Successful app records", processing["succeeded"]],
                ["App success rate", processing["success_rate"] / 100.0],
                ["Failed app records", processing["failed"]],
            ]

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row_idx, col_idx, value)
            ws.cell(4, 2).number_format = '€#,##0.00'
            ws.cell(9, 2).number_format = '0.0%'
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 22

            def add_breakdown(title, rows):
                sheet = wb.create_sheet(title)
                sheet.append(["Name", "Products", "Earnings"])
                for cell in sheet[1]:
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.fill = PatternFill(start_color="4F46E5", fill_type="solid")
                for row in rows:
                    sheet.append([row["label"], row["count"], row["cents"] / 100.0])
                    sheet.cell(sheet.max_row, 3).number_format = '€#,##0.00'
                sheet.column_dimensions["A"].width = 28
                sheet.column_dimensions["B"].width = 14
                sheet.column_dimensions["C"].width = 16

            add_breakdown("Brands", values["brands"])
            add_breakdown("Sources", values["sources"])
            add_breakdown("Product types", values["product_types"])

            issues = wb.create_sheet("App issues")
            issues.append(["Issue", "Count"])
            for cell in issues[1]:
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="4F46E5", fill_type="solid")
            for row in values["errors"]:
                issues.append([row["message"], row["count"]])
            issues.column_dimensions["A"].width = 70
            issues.column_dimensions["B"].width = 12

            wb.save(file_path)

            InfoBar.success(
                title=self.main.i18n.tr("analytics.export.success.title"),
                content=self.main.i18n.tr("analytics.export.success.content", path=file_path),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000
            )
        except Exception as e:
            InfoBar.error(
                title=self.main.i18n.tr("analytics.export.failed.title"),
                content=self.main.i18n.tr("analytics.export.failed.content", error=str(e)),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000
            )

    def update_translations(self):
        """Update UI text for current language"""
        tr = self.main.i18n.tr
        self.title_label.setText(tr("analytics.title"))
        self.products_metric.title_label.setText(tr("analytics.metric.products"))
        self.earnings_metric.title_label.setText(tr("analytics.metric.earnings"))
        self.manual_products_metric.title_label.setText(tr("analytics.metric.manual"))
        self.app_products_metric.title_label.setText(tr("analytics.metric.app_linked"))
        self.brand_title.setText(tr("analytics.brand.title"))
        self.success_title.setText(tr("analytics.success.title"))
        self.source_title.setText(tr("analytics.source.title"))
        self.type_title.setText(tr("analytics.type.title"))
        self.activity_title.setText(tr("analytics.activity.title"))
        self.errors_title.setText(tr("analytics.errors.title"))
        self.period_label.setText(tr("common.period"))
        self.refresh_btn.setToolTip(tr("analytics.refresh"))
        self.export_btn.setText(tr("analytics.export"))
        self.view_all_btn.setToolTip(tr("analytics.activity.view_all"))

        selected_days = self.period_combo.currentData()
        self.period_combo.blockSignals(True)
        self.period_combo.clear()
        choices = (
            (tr("common.last_7_days"), 7),
            (tr("common.last_30_days"), 30),
            (tr("common.last_90_days"), 90),
            (tr("common.all_time"), None),
        )
        selected_index = 1
        for index, (label, days) in enumerate(choices):
            self.period_combo.addItem(label, userData=days)
            if days == selected_days:
                selected_index = index
        self.period_combo.setCurrentIndex(selected_index)
        self.period_combo.blockSignals(False)
        self.success_caption.setText(self.period_combo.currentText())
        self.load_analytics()

    def retranslate_ui(self):
        self.update_translations()

    def _on_breakpoint_changed(self, breakpoint: str):
        """Respond to breakpoint changes - adjust margins and spacing."""
        margins = get_responsive_margins(breakpoint)
        spacing = get_responsive_spacing(breakpoint)

        # Update main layout margins
        if self.layout():
            self.layout().setContentsMargins(*margins)
            self.layout().setSpacing(spacing)

        # Update content widget if it exists
        if hasattr(self, 'content_widget') and self.content_widget.layout():
            # For AnalyticsScreen, the content widget already has custom margins
            # Only update spacing to maintain the scrollbar offset
            self.content_widget.layout().setSpacing(spacing)

        metric_compact = breakpoint in ("xs", "sm", "md", "lg")
        chart_compact = breakpoint in ("xs", "sm")
        columns = 2 if metric_compact else 4
        if hasattr(self, "metrics_grid"):
            for card in self.metric_cards:
                self.metrics_grid.removeWidget(card)
            for index, card in enumerate(self.metric_cards):
                self.metrics_grid.addWidget(card, index // columns, index % columns)
            for column in range(4):
                self.metrics_grid.setColumnStretch(column, 1 if column < columns else 0)
            self.metrics_grid.setSpacing(spacing)
        if hasattr(self, "charts_layout"):
            self.charts_layout.setDirection(
                QBoxLayout.Direction.TopToBottom if chart_compact
                else QBoxLayout.Direction.LeftToRight
            )
            self.charts_layout.setSpacing(spacing)
        if hasattr(self, "breakdowns_layout"):
            self.breakdowns_layout.setDirection(
                QBoxLayout.Direction.TopToBottom if chart_compact
                else QBoxLayout.Direction.LeftToRight
            )
            self.breakdowns_layout.setSpacing(spacing)

    def _on_theme_changed(self):
        """Handle theme change"""
        self._apply_theme()
        self._apply_card_themes()  # Re-apply card styling on theme change
        enforce_transparent_labels(self)
        self.load_analytics()  # Reload to update colors
