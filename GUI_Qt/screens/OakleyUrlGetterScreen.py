"""GUI_Qt/screens/OakleyUrlGetterScreen.py

Looks up Oakley product URLs from UltraBike product codes.
"""

from __future__ import annotations

import os
import re
import time
from datetime import datetime
from urllib.parse import quote_plus, urljoin

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QFileDialog, QHeaderView, QTableWidgetItem

from qfluentwidgets import FluentIcon, InfoBar, InfoBarPosition, isDarkTheme

from Config.Selectors import OakleySelectors
from GUI_Qt.screens.CastelliUrlGetterScreen import CastelliUrlGetterScreen
from GUI_Qt.widgets import show_file_saved_bar
from GUI_Qt.styles.theme_config import get_text_color


def extract_oakley_code(raw_code: str) -> str:
    """Extract OO9208-4738 from UB-OO9208-4738."""
    value = (raw_code or "").strip()
    ub_match = re.search(r"\bUB-(.+)$", value, re.IGNORECASE)
    if ub_match:
        return ub_match.group(1).strip()
    return value


class OakleyUrlGetterWorker(QThread):
    row_update = Signal(int, str, str)
    progress_update = Signal(int, int)
    done = Signal(int, int)
    log = Signal(str)

    def __init__(self, driver, rows: list[dict[str, str]]):
        super().__init__()
        self.driver = driver
        self.rows = rows
        self._stop = False

    def request_stop(self):
        self._stop = True

    def run(self):
        found = 0
        errors = 0

        for idx, row in enumerate(self.rows):
            if self._stop:
                self.row_update.emit(idx, "Skipped", "")
                continue

            search_code = row["search_code"]
            self.progress_update.emit(idx + 1, len(self.rows))
            self.row_update.emit(idx, "Searching", "")
            self.log.emit(f"Searching {search_code}")

            try:
                urls = self._lookup_urls(search_code)
                if urls:
                    found += 1
                    status = "Found" if len(urls) == 1 else "Multiple results"
                    self.row_update.emit(idx, status, "\n".join(urls))
                else:
                    errors += 1
                    self.row_update.emit(idx, "Not found", "")
            except Exception as e:
                errors += 1
                self.row_update.emit(idx, f"Error: {e}", "")

        self.done.emit(found, errors)

    def _lookup_urls(self, search_code: str) -> list[str]:
        target_url = OakleySelectors.SEARCH_URL.format(query=quote_plus(search_code))
        self.log.emit(f"Opening {target_url}")
        self._navigate_to(target_url)
        self._accept_cookies()

        urls = self._wait_for_search_results(search_code)
        if urls or self._has_no_results():
            return urls

        return self._search_with_page_ui(search_code)

    def _navigate_to(self, url: str):
        try:
            self.driver.switch_to.alert.accept()
            time.sleep(0.2)
        except Exception:
            pass

        try:
            self.driver.get(url)
        except Exception:
            self.driver.execute_script("window.location.href = arguments[0];", url)

    def _accept_cookies(self) -> bool:
        for locator in OakleySelectors.COOKIE_BUTTON_CANDIDATES:
            try:
                for btn in self.driver.find_elements(*locator):
                    if btn.is_displayed() and btn.is_enabled():
                        self.driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.2)
                        return True
            except Exception:
                continue
        return False

    def _wait_for_search_results(self, search_code: str) -> list[str]:
        from selenium.webdriver.support.ui import WebDriverWait

        try:
            WebDriverWait(self.driver, 8).until(
                lambda d: self._collect_matching_urls(search_code) or self._has_no_results()
            )
        except Exception:
            pass
        return self._collect_matching_urls(search_code)

    def _search_with_page_ui(self, search_code: str) -> list[str]:
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        try:
            opener = WebDriverWait(self.driver, 3).until(EC.element_to_be_clickable(OakleySelectors.SEARCH_OPEN))
            self.driver.execute_script("arguments[0].click();", opener)
        except Exception:
            pass

        input_el = WebDriverWait(self.driver, 5).until(EC.visibility_of_element_located(OakleySelectors.SEARCH_INPUT))
        try:
            input_el.clear()
        except Exception:
            pass
        input_el.send_keys(search_code)

        try:
            submit = self.driver.find_element(*OakleySelectors.SEARCH_SUBMIT)
            self.driver.execute_script("arguments[0].click();", submit)
        except Exception:
            input_el.send_keys(Keys.ENTER)

        return self._wait_for_search_results(search_code)

    def _collect_matching_urls(self, search_code: str) -> list[str]:
        seen = set()
        exact_matches = []
        fallback_matches = []

        lowered_code = search_code.lower().strip()
        compact_code = re.sub(r"[^a-z0-9]", "", lowered_code)

        try:
            tiles = self.driver.find_elements("css selector", ".prod-tile")
        except Exception:
            tiles = []

        for tile in tiles:
            try:
                if not tile.is_displayed():
                    continue
            except Exception:
                pass

            try:
                sku = (tile.get_attribute("data-ct-sku") or "").strip()
                upc = (tile.get_attribute("data-ct-upc") or "").strip()
                name = (tile.get_attribute("data-ct-name") or "").strip()
                html = (tile.get_attribute("outerHTML") or "").strip()
            except Exception:
                sku = ""
                upc = ""
                name = ""
                html = ""

            href = ""
            for selector in [
                "a.product-name[href*='/product/']",
                "a.product-url[href*='/product/']",
                "a.preview_image[href*='/product/']",
                "a[href*='/product/']",
            ]:
                try:
                    link = tile.find_element("css selector", selector)
                    href = (link.get_attribute("href") or "").strip()
                    if href:
                        break
                except Exception:
                    continue

            if not href:
                try:
                    footer_href = tile.find_element("css selector", ".prod-tile_footer").get_attribute("data-href") or ""
                    href = footer_href.strip()
                except Exception:
                    href = ""

            if not href:
                continue

            url = urljoin(OakleySelectors.BASE_URL, href)
            if url in seen:
                continue

            seen.add(url)
            fallback_matches.append(url)

            haystack = " ".join([sku, upc, name, url, html]).lower()
            compact_haystack = re.sub(r"[^a-z0-9]", "", haystack)

            if lowered_code in haystack or compact_code in compact_haystack:
                exact_matches.append(url)

        if exact_matches:
            return exact_matches
        if fallback_matches:
            return fallback_matches

        try:
            links = self.driver.find_elements(*OakleySelectors.PRODUCT_LINK)
        except Exception:
            links = []

        for link in links:
            try:
                href = (link.get_attribute("href") or "").strip()
                if not href:
                    continue

                url = urljoin(OakleySelectors.BASE_URL, href)
                if url in seen:
                    continue

                seen.add(url)
                fallback_matches.append(url)
            except Exception:
                continue

        return fallback_matches

    def _has_no_results(self) -> bool:
        try:
            visible_tiles = self.driver.find_elements("css selector", ".prod-tile")
            if any(tile.is_displayed() for tile in visible_tiles):
                return False
        except Exception:
            pass

        try:
            lists = self.driver.find_elements("css selector", ".prod-tiles-list[data-items-length]")
            for list_el in lists:
                raw_count = (list_el.get_attribute("data-items-length") or "").strip()
                if raw_count.isdigit() and int(raw_count) > 0:
                    return False
        except Exception:
            pass

        try:
            counter_text = self.driver.find_element(
                "css selector",
                ".search-header-title-container .counter"
            ).text.lower()
            match = re.search(r"(\d+)\s*results?", counter_text)
            if match:
                return int(match.group(1)) == 0
        except Exception:
            pass

        try:
            header_text = self.driver.find_element(
                "css selector",
                ".search-header-title-container"
            ).text.lower()
            match = re.search(r"\((\d+)\s*results?\)", header_text)
            if match:
                return int(match.group(1)) == 0
        except Exception:
            pass

        try:
            return any(el.is_displayed() for el in self.driver.find_elements(*OakleySelectors.NO_RESULTS))
        except Exception:
            return False



class OakleyUrlGetterScreen(CastelliUrlGetterScreen):
    def __init__(self, main_window):
        super().__init__(main_window)
        self._worker: OakleyUrlGetterWorker | None = None

    def _build_ui(self):
        super()._build_ui()
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setColumnWidth(1, 180)
        self._table.setColumnWidth(2, 150)
        self._table.setColumnWidth(3, 130)

    def retranslate_ui(self):
        self._title.setText(self.tr("oakleyurl.title"))
        self._subtitle.setText(self.tr("oakleyurl.subtitle"))
        self._browse_btn.setText(self.tr("oakleyurl.browse"))
        self._start_btn.setText(self.tr("oakleyurl.start"))
        self._export_btn.setText(self.tr("oakleyurl.export"))
        self._table.setHorizontalHeaderLabels([
            self.tr("oakleyurl.col.index"),
            self.tr("oakleyurl.col.original_code"),
            self.tr("oakleyurl.col.search_code"),
            self.tr("oakleyurl.col.status"),
            self.tr("oakleyurl.col.url"),
        ])

    def _load_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            rows: list[dict[str, str]] = []

            for excel_row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                value = excel_row[0]
                if value is None:
                    continue
                original_code = str(value).strip()
                if not original_code:
                    continue
                if original_code.strip().lower() in {"code", "product code", "original code"}:
                    continue
                rows.append({
                    "original_code": original_code,
                    "search_code": extract_oakley_code(original_code),
                })
            wb.close()

            if not rows:
                InfoBar.warning(
                    self.tr("common.error"),
                    self.tr("oakleyurl.no_codes"),
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                )
                return

            self._rows = rows
            self._file_path = path
            self._file_label.setText(os.path.basename(path))
            self._table.setRowCount(len(rows))

            for idx, row in enumerate(rows):
                self._table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
                self._table.setItem(idx, 1, QTableWidgetItem(row["original_code"]))
                self._table.setItem(idx, 2, QTableWidgetItem(row["search_code"]))
                status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
                status.setForeground(QColor(get_text_color(isDarkTheme(), "secondary")))
                self._table.setItem(idx, 3, status)
                self._table.setItem(idx, 4, QTableWidgetItem(""))

            self._start_btn.setEnabled(True)
            self._export_btn.setEnabled(False)
            self._drop_zone.setVisible(False)
            self._progress_label.setText(self.tr("oakleyurl.loaded", count=len(rows)))
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        if not self._rows:
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        for idx in range(self._table.rowCount()):
            status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
            status.setForeground(QColor(get_text_color(isDarkTheme(), "secondary")))
            self._table.setItem(idx, 3, status)
            self._table.setItem(idx, 4, QTableWidgetItem(""))

        self._start_btn.setText(self.tr("oakleyurl.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._browse_btn.setEnabled(False)
        self._export_btn.setEnabled(False)

        self._worker = OakleyUrlGetterWorker(driver, self._rows)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[OakleyUrlGetter] {msg}"))
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._worker, "url_scanner", "oakley_url_getter", total=len(self._rows)
            )
        self._worker.start()

    def _on_progress(self, current: int, total: int):
        self._progress_label.setText(self.tr("oakleyurl.progress", current=current, total=total))

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("oakleyurl.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._browse_btn.setEnabled(True)
        self._export_btn.setEnabled(True)

        total = found + errors
        text = self.tr("oakleyurl.done", found=found, errors=errors, total=total)
        self._progress_label.setText(text)
        InfoBar.success(
            self.tr("oakleyurl.done_title"),
            text,
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("oakleyurl.export"),
            f"oakley_urls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Oakley URLs"

            headers = ["#", "Original Code", "Search Code", "Status", "URL"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(5):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 100

            wb.save(path)
            show_file_saved_bar(self, self.tr("common.success"), self.tr("oakleyurl.exported"), path)
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )
