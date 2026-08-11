"""
Full History Screen - Comprehensive detailed view
Complements Analytics dashboard with searchable, sortable record table
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QFileDialog, QAbstractItemView
)
from PySide6.QtCore import Qt, QUrl, Signal, QEvent, QObject
from PySide6.QtGui import QColor, QCursor, QDesktopServices
from qfluentwidgets import (
    LineEdit, ComboBox, PushButton, TransparentToolButton, FluentIcon,
    BodyLabel, TitleLabel, CardWidget, InfoBar, InfoBarPosition,
    isDarkTheme, qconfig, ScrollArea
)
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.widgets import enable_table_copy
from GUI_Qt.styles.theme_config import COLORS, FONTS, RADII, SPACING, SIZES, rgba_from_hex, get_text_color
from GUI_Qt.styles.theme_config import COMPONENT_COLORS, PADDINGS, get_selection_bg
from GUI_Qt.styles.screen_theme import (
    apply_screen_theme,
    enforce_transparent_labels,
    PAGE_MARGINS,
    PAGE_SPACING,
    CARD_MARGINS,
    CARD_SPACING,
    get_responsive_margins,
    get_responsive_spacing,
)


class _ScrollBarStyleFilter(QObject):
    def __init__(self, qss: str, parent: QObject | None = None):
        super().__init__(parent)
        self._qss = qss
        self._applying = False

    def eventFilter(self, watched, event):  # noqa: N802 (Qt naming)
        try:
            if self._applying:
                return super().eventFilter(watched, event)

            # Avoid StyleChange: calling setStyleSheet() emits StyleChange again and
            # can recurse/crash when the widget is being polished.
            if event.type() in (
                QEvent.Type.Polish,
                QEvent.Type.PaletteChange,
                QEvent.Type.Show,
            ):
                current = watched.styleSheet() or ""
                if current != self._qss:
                    self._applying = True
                    try:
                        watched.setStyleSheet(self._qss)
                    finally:
                        self._applying = False
        except Exception:
            pass
        return super().eventFilter(watched, event)


class FullHistoryScreen(ResponsiveWidget):
    """Comprehensive full history view with search, sort, and advanced filtering"""

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self._FILTER_ALL = "__all__"
        self.all_history = []
        self.filtered_history = []
        self._sort_column = 10  # processed_at column (descending by default)
        self._sort_order = Qt.SortOrder.DescendingOrder

        self._init_ui()
        self.refresh_history()
        self.update_translations()

        # Connect to theme change
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def _init_ui(self):
        """Initialize comprehensive history UI"""
        # Root layout (for ScrollArea container)
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)

        # Create ScrollArea
        self.scroll = ScrollArea(self)
        self.scroll.setWidgetResizable(True)
        root_layout.addWidget(self.scroll)

        # Content widget (inside ScrollArea)
        self.content_widget = QWidget()
        self.scroll.setWidget(self.content_widget)

        # Main layout (on content widget)
        layout = QVBoxLayout(self.content_widget)
        layout.setContentsMargins(*PAGE_MARGINS)
        layout.setSpacing(PAGE_SPACING)

        # === HEADER ===
        header_layout = QHBoxLayout()

        # Back button
        self.back_btn = TransparentToolButton(FluentIcon.RETURN, self)
        self.back_btn.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        self.back_btn.clicked.connect(self._go_back_to_analytics)

        self.title_label = TitleLabel("")
        self.title_label.setStyleSheet(f"color: {get_text_color(isDarkTheme())}; font-weight: 600;")

        header_layout.addWidget(self.back_btn)
        header_layout.addSpacing(SPACING['md'])
        header_layout.addWidget(self.title_label)
        header_layout.addStretch()

        layout.addLayout(header_layout)

        # === SEARCH & FILTERS CARD ===
        self.filter_card = CardWidget()
        self.filter_card.setBorderRadius(RADII['md'])
        filter_layout = QVBoxLayout(self.filter_card)
        filter_layout.setContentsMargins(*CARD_MARGINS)
        filter_layout.setSpacing(CARD_SPACING)

        # Search bar
        search_layout = QHBoxLayout()
        search_layout.setSpacing(SPACING['md'])

        self.search_label = BodyLabel("")
        self.search_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")

        self.search_input = LineEdit()
        self.search_input.setClearButtonEnabled(True)
        self.search_input.textChanged.connect(self._on_search_changed)

        search_layout.addWidget(self.search_label)
        search_layout.addWidget(self.search_input, 1)

        filter_layout.addLayout(search_layout)

        # Filters row (keep controls readable; actions in separate row)
        filters_row = QHBoxLayout()
        filters_row.setSpacing(SPACING['md'])

        # Brand filter
        self.brand_label = BodyLabel("")
        self.brand_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self.brand_filter = ComboBox()
        self.brand_filter.setMinimumWidth(SIZES['filter_min_width'])
        self.brand_filter.currentTextChanged.connect(self._apply_filters)

        # Status filter
        self.status_label = BodyLabel("")
        self.status_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self.status_filter = ComboBox()
        self.status_filter.setMinimumWidth(SIZES['field_min_width_sm'])
        self.status_filter.currentTextChanged.connect(self._apply_filters)

        # Type filter
        self.type_label = BodyLabel("")
        self.type_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self.type_filter = ComboBox()
        self.type_filter.setMinimumWidth(SIZES['filter_min_width'])
        self.type_filter.currentTextChanged.connect(self._apply_filters)

        # Date range filter
        self.date_label = BodyLabel("")
        self.date_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self.date_filter = ComboBox()
        self.date_filter.setMinimumWidth(SIZES['filter_min_width'])
        self.date_filter.currentTextChanged.connect(self._apply_filters)

        filters_row.addWidget(self.brand_label)
        filters_row.addWidget(self.brand_filter)
        filters_row.addWidget(self.status_label)
        filters_row.addWidget(self.status_filter)
        filters_row.addWidget(self.type_label)
        filters_row.addWidget(self.type_filter)
        filters_row.addWidget(self.date_label)
        filters_row.addWidget(self.date_filter)
        filters_row.addStretch()

        filter_layout.addLayout(filters_row)

        # Action buttons row
        actions_row = QHBoxLayout()
        actions_row.setSpacing(SPACING['md'])
        actions_row.addStretch()

        self.refresh_btn = TransparentToolButton(FluentIcon.SYNC, self)
        self.refresh_btn.clicked.connect(self.refresh_history)

        self.clear_filters_btn = PushButton("")
        self.clear_filters_btn.setIcon(FluentIcon.CANCEL.icon())
        self.clear_filters_btn.clicked.connect(self._clear_filters)

        self.export_btn = PushButton("")
        self.export_btn.setIcon(FluentIcon.DOCUMENT.icon())
        self.export_btn.clicked.connect(self._export_to_excel)

        actions_row.addWidget(self.refresh_btn)
        actions_row.addWidget(self.clear_filters_btn)
        actions_row.addWidget(self.export_btn)
        filter_layout.addLayout(actions_row)

        # Results count
        self.results_label = BodyLabel("")
        self.results_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'tertiary')}; font-size: {FONTS['size_caption']};")
        filter_layout.addWidget(self.results_label)

        layout.addWidget(self.filter_card)

        # === TABLE ===
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "#", "Type", "Brand", "Product Code", "Status", "Duration",
            "Features", "Images", "Error", "Failed Stage", "URL/Code", "Processed At"
        ])

        # Table appearance
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        enable_table_copy(self.table)
        # Smooth scrolling (avoid snapping to grid/rows)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        try:
            self.table.verticalScrollBar().setSingleStep(12)
            self.table.horizontalScrollBar().setSingleStep(12)
        except Exception:
            pass
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        # Don't stretch the last column; we'll stretch the Error column instead
        header.setStretchLastSection(False)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)

        # Column widths/resizing
        self.table.setColumnWidth(0, 50)    # #
        self.table.setColumnWidth(1, 140)   # Type
        self.table.setColumnWidth(2, 120)   # Brand
        self.table.setColumnWidth(3, 170)   # Product Code
        self.table.setColumnWidth(4, 110)   # Status
        self.table.setColumnWidth(5, 90)    # Duration
        self.table.setColumnWidth(6, 80)    # Features
        self.table.setColumnWidth(7, 80)    # Images
        self.table.setColumnWidth(9, 140)   # Failed Stage
        self.table.setColumnWidth(10, 190)  # URL/Code
        self.table.setColumnWidth(11, 160)  # Processed At

        # Make Error take remaining space so it's visible without horizontal scrolling
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        header.setMinimumSectionSize(60)

        # Ensure column sizing is applied consistently (Qt/QFluent can repolish headers)
        self._apply_column_sizing()

        # Connect double-click to open URL
        self.table.cellDoubleClicked.connect(self._on_cell_double_clicked)

        # Apply theme-aware styling
        self._apply_table_theme()

        # Apply consistent screen theme (background + label transparency)
        self._apply_theme()

        layout.addWidget(self.table, 1)

        # Apply theme
        apply_screen_theme(
            self,
            "FullHistoryScreen",
            scroll=self.scroll,
            content=self.content_widget
        )

        # Populate filters
        self._populate_filters()

    def showEvent(self, event):  # noqa: N802 (Qt naming)
        # Re-apply after Qt/QFluent polishes widgets (can reset scrollbars)
        try:
            self._apply_table_theme()
            self._apply_column_sizing()
        except Exception:
            pass
        super().showEvent(event)

    def _apply_column_sizing(self) -> None:
        """Size columns so key fields aren't half-hidden until horizontal scrolling."""
        try:
            header = self.table.horizontalHeader()

            # Keep last column normal; Error takes remaining width
            header.setStretchLastSection(False)
            header.setMinimumSectionSize(60)
            header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            header.setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)

            # Auto-fit compact columns (do NOT auto-fit Error/URL/Product Code)
            auto_cols = (0, 2, 4, 5, 6, 7, 9, 11)
            for col in auto_cols:
                self.table.resizeColumnToContents(col)

            # Clamp to reasonable minimums so columns don't get too tight
            min_widths = {
                0: 50,   # #
                2: 110,  # Brand
                4: 100,  # Status
                5: 90,   # Duration
                6: 80,   # Features
                7: 80,   # Images
                9: 130,  # Failed Stage
                11: 160, # Processed At
            }
            for col, min_w in min_widths.items():
                current = self.table.columnWidth(col)
                if current < min_w:
                    self.table.setColumnWidth(col, min_w)

            # Keep these stable/predictable
            self.table.setColumnWidth(1, 140)   # Type
            self.table.setColumnWidth(3, 170)   # Product Code
            self.table.setColumnWidth(10, 220)  # URL/Code
        except Exception:
            return

    def _apply_table_theme(self):
        """Apply theme-aware styling to table"""
        is_dark = isDarkTheme()

        table_colors = COMPONENT_COLORS['table']
        # Standard: base row + alternate row tint
        bg = table_colors['row_bg_dark'] if is_dark else table_colors['row_bg_light']
        alt_bg = table_colors['row_alt_bg_dark'] if is_dark else table_colors['row_alt_bg_light']
        border = table_colors['border_dark'] if is_dark else table_colors['border_light']
        text = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']

        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']

        selection_bg = get_selection_bg()

        scrollbar_handle = rgba_from_hex(COLORS['lavender_grey'], 0.35) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.22)
        scrollbar_handle_hover = rgba_from_hex(COLORS['lavender_grey'], 0.55) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.32)

        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg};
                alternate-background-color: {alt_bg};
                gridline-color: {border};
                border: 1px solid {border};
                border-radius: {RADII['md']}px;
                color: {text};
                font-size: {FONTS['size_body']};
                selection-background-color: {selection_bg};
            }}
            QTableWidget::viewport {{
                background-color: {bg};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableView::item:focus, QAbstractItemView::item:focus {{
                outline: none;
            }}
            QTableWidget::item:selected {{
                background-color: {selection_bg};
                color: {text};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-right: 1px solid {border};
                border-bottom: 1px solid {border};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}

            /* Rounded scrollbars (scoped to this table) */
            QTableWidget QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: 0px;
            }}
            QTableWidget QScrollBar::handle:vertical {{
                background: {scrollbar_handle};
                border: none;
                border-radius: {RADII['md']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 2px;
            }}
            QTableWidget QScrollBar::handle:vertical:hover {{
                background: {scrollbar_handle_hover};
            }}
            QTableWidget QScrollBar::add-line:vertical, QTableWidget QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QTableWidget QScrollBar::add-page:vertical, QTableWidget QScrollBar::sub-page:vertical {{
                background: none;
            }}

            QTableWidget QScrollBar:horizontal {{
                background: transparent;
                height: {SIZES['scrollbar_thickness']}px;
                margin: 0px;
            }}
            QTableWidget QScrollBar::handle:horizontal {{
                background: {scrollbar_handle};
                border: none;
                border-radius: {RADII['md']}px;
                min-width: {SIZES['scrollbar_handle_min']}px;
                margin: 2px;
            }}
            QTableWidget QScrollBar::handle:horizontal:hover {{
                background: {scrollbar_handle_hover};
            }}
            QTableWidget QScrollBar::add-line:horizontal, QTableWidget QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
            QTableWidget QScrollBar::add-page:horizontal, QTableWidget QScrollBar::sub-page:horizontal {{
                background: none;
            }}
        """)

        # Apply directly to the scrollbar widgets as well (more reliable on Windows styles)
        sb_qss = f"""
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {scrollbar_handle};
                background-color: {scrollbar_handle};
                border: none;
                border-radius: {RADII['md']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 2px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {scrollbar_handle_hover};
                background-color: {scrollbar_handle_hover};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}

            QScrollBar:horizontal {{
                background: transparent;
                height: {SIZES['scrollbar_thickness']}px;
                margin: 0px;
            }}
            QScrollBar::handle:horizontal {{
                background: {scrollbar_handle};
                background-color: {scrollbar_handle};
                border: none;
                border-radius: {RADII['md']}px;
                min-width: {SIZES['scrollbar_handle_min']}px;
                margin: 2px;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {scrollbar_handle_hover};
                background-color: {scrollbar_handle_hover};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
        """
        try:
            self.table.verticalScrollBar().setStyleSheet(sb_qss)
            self.table.horizontalScrollBar().setStyleSheet(sb_qss)

            flt = getattr(self, "_ub_table_scrollbar_filter", None)
            if flt is None:
                flt = _ScrollBarStyleFilter(sb_qss, self)
                setattr(self, "_ub_table_scrollbar_filter", flt)
            try:
                self.table.verticalScrollBar().installEventFilter(flt)
                self.table.horizontalScrollBar().installEventFilter(flt)
            except Exception:
                pass
        except Exception:
            pass

    def _record_type(self, batch_id: str | None) -> str:
        """Return a stable type key based on batch_id prefix."""
        if not batch_id:
            return "regular"
        s = str(batch_id)
        if s.startswith("batchdesc_"):
            return "batch_desc"
        if s.startswith("batchtitle_"):
            return "batch_titles"
        if s.startswith("batch_"):
            return "batch_upload"
        return "batch"

    def _record_type_label(self, batch_id: str | None) -> str:
        t = self._record_type(batch_id)
        if t == "regular":
            return "Regular Upload"
        if t == "batch_upload":
            return "Batch Upload"
        if t == "batch_desc":
            return "Batch Descriptions"
        if t == "batch_titles":
            return "Batch Titles"
        return "Batch"

    def _apply_theme(self):
        apply_screen_theme(self, "FullHistoryScreen", transparent_labels=True)
        enforce_transparent_labels(self)

        # Card theme to match AnalyticsScreen-style subtle tint
        is_dark = isDarkTheme()
        card_bg = rgba_from_hex(COLORS['text_white'], 0.03) if is_dark else rgba_from_hex(COLORS['space_indigo'], 0.03)
        card_border = COLORS['border_dark'] if is_dark else COLORS['border_light']
        try:
            if getattr(self, 'filter_card', None) is not None:
                self.filter_card.setStyleSheet(f"""
                    background-color: {card_bg};
                    border: 1px solid {card_border};
                    border-radius: {RADII['md']}px;
                """)
        except Exception:
            pass

    def _populate_filters(self):
        """Populate filter dropdowns"""
        # Brand filter
        self.brand_filter.clear()
        self.brand_filter.addItem("All Brands", self._FILTER_ALL)
        for brand in ["KROSS", "Pinarello", "Basso", "Factor", "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan"]:
            self.brand_filter.addItem(brand, brand)

        # Status filter
        self.status_filter.clear()
        self.status_filter.addItem("All Statuses", self._FILTER_ALL)
        self.status_filter.addItem("✓ Success", "success")
        self.status_filter.addItem("✓ Saved manually", "saved_manually")
        self.status_filter.addItem("◷ Ready for review", "ready_for_review")
        self.status_filter.addItem("— Blocked non-Draft", "blocked_non_draft")
        self.status_filter.addItem("— Discarded", "discarded")
        self.status_filter.addItem("✗ Failed", "failed")

        # Type filter
        self.type_filter.clear()
        self.type_filter.addItem("All Types", self._FILTER_ALL)
        self.type_filter.addItem("Regular Uploads", "regular")
        self.type_filter.addItem("Batch Uploads", "batch_upload")
        self.type_filter.addItem("Batch Descriptions", "batch_desc")
        self.type_filter.addItem("Batch Titles", "batch_titles")

        # Date filter
        self.date_filter.clear()
        self.date_filter.addItem("All Time", self._FILTER_ALL)
        self.date_filter.addItem("Today", "today")
        self.date_filter.addItem("Last 7 Days", "last_7")
        self.date_filter.addItem("Last 30 Days", "last_30")
        self.date_filter.addItem("Last 90 Days", "last_90")

    def refresh_history(self):
        """Load all history from database"""
        cursor = self.main.db.conn.cursor()

        query = """
             SELECT brand, product_code, url_or_code, status, duration_seconds,
                 features_uploaded, images_uploaded, error_message,
                 failed_stage, batch_id, details_json, processed_at
            FROM processing_history
            ORDER BY processed_at DESC
        """

        self.all_history = cursor.execute(query).fetchall()
        self._apply_filters()

    def _apply_filters(self):
        """Apply all filters and search"""
        filtered = self.all_history

        # Brand filter
        brand = self.brand_filter.currentData()
        if brand and brand != self._FILTER_ALL:
            filtered = [r for r in filtered if r['brand'].upper() == brand.upper()]

        # Status filter
        status = self.status_filter.currentData()
        if status and status != self._FILTER_ALL:
            filtered = [r for r in filtered if r['status'] == status]

        # Type filter
        record_type = self.type_filter.currentData()
        if record_type and record_type != self._FILTER_ALL:
            filtered = [r for r in filtered if self._record_type(r['batch_id']) == record_type]

        # Date filter
        date_range = self.date_filter.currentData()
        if date_range and date_range != self._FILTER_ALL:
            now = datetime.now()
            if date_range == "today":
                threshold = now.strftime('%Y-%m-%d')
                filtered = [r for r in filtered if r['processed_at'] and r['processed_at'].startswith(threshold)]
            elif date_range.startswith("last_"):
                days = int(date_range.split("_")[1])
                threshold = (now - timedelta(days=days)).strftime('%Y-%m-%d')
                filtered = [r for r in filtered if r['processed_at'] and r['processed_at'] >= threshold]

        # Search filter
        search_text = self.search_input.text().strip().lower()
        if search_text:
            filtered = [
                r for r in filtered
                if (search_text in str(r['brand']).lower() or
                    search_text in str(r['product_code']).lower() or
                    (r['error_message'] and search_text in str(r['error_message']).lower()))
            ]

        self.filtered_history = filtered
        self._populate_table()

    def _on_search_changed(self, text):
        """Handle search text change"""
        self._apply_filters()

    def _populate_table(self):
        """Populate table with filtered results"""
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(self.filtered_history))

        for row_idx, record in enumerate(self.filtered_history):
            # Column 0: Row number
            item = QTableWidgetItem(str(row_idx + 1))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setForeground(QColor(COLORS['text_tertiary']))
            self.table.setItem(row_idx, 0, item)

            # Column 1: Type
            batch_id = record['batch_id']
            type_label = self._record_type_label(batch_id)
            item = QTableWidgetItem(type_label)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 1, item)

            # Column 2: Brand
            item = QTableWidgetItem(str(record['brand'] or ''))
            self.table.setItem(row_idx, 2, item)

            # Column 3: Product Code
            item = QTableWidgetItem(str(record['product_code'] or ''))
            self.table.setItem(row_idx, 3, item)

            # Column 4: Status
            status = record['status']
            status_labels = {
                'success': "✓ Success",
                'saved_manually': "✓ Saved manually",
                'ready_for_review': "◷ Ready for review",
                'blocked_non_draft': "— Blocked non-Draft",
                'discarded': "— Discarded",
                'failed': "✗ Failed",
            }
            item = QTableWidgetItem(status_labels.get(status, str(status or '')))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if status in ('success', 'saved_manually'):
                item.setForeground(QColor(COLORS['success']))
            elif status == 'ready_for_review':
                item.setForeground(QColor(COLORS['warning']))
            elif status in ('blocked_non_draft', 'discarded'):
                item.setForeground(QColor(COLORS['text_tertiary']))
            else:
                item.setForeground(QColor(COLORS['error']))
            self.table.setItem(row_idx, 4, item)

            # Column 5: Duration
            duration = record['duration_seconds']
            item = QTableWidgetItem(f"{duration:.1f}s" if duration else "N/A")
            item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.table.setItem(row_idx, 5, item)

            # Column 6: Features
            features = record['features_uploaded'] or 0
            item = QTableWidgetItem(str(features))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 6, item)

            # Column 7: Images
            images = record['images_uploaded'] or 0
            item = QTableWidgetItem(str(images))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 7, item)

            # Column 8: Error Message
            error = record['error_message'] or ''
            item = QTableWidgetItem(error[:200] + '...' if len(error) > 200 else error)
            item.setToolTip(error)
            if error:
                item.setForeground(QColor(COLORS['error']))
            self.table.setItem(row_idx, 8, item)

            # Column 9: Failed Stage
            failed_stage = record['failed_stage'] or ''
            item = QTableWidgetItem(failed_stage)
            if failed_stage:
                item.setForeground(QColor(COLORS['warning']))
            self.table.setItem(row_idx, 9, item)

            # Column 10: URL/Code
            url = record['url_or_code'] or ''
            item = QTableWidgetItem(url[:100] + '...' if len(url) > 100 else url)
            item.setToolTip(url)
            if url.startswith('http'):
                item.setForeground(QColor(COLORS['lavender_grey']))
                item.setData(Qt.ItemDataRole.UserRole, url)  # Store full URL
            self.table.setItem(row_idx, 10, item)

            # Column 11: Processed At
            processed_at = record['processed_at'] or ''
            # Format timestamp nicely
            try:
                dt = datetime.fromisoformat(str(processed_at).replace('Z', '').replace('T', ' '))
                formatted = dt.strftime('%Y-%m-%d %H:%M')
            except (ValueError, TypeError, AttributeError):
                formatted = str(processed_at)
            item = QTableWidgetItem(formatted)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row_idx, 11, item)

        self.table.setSortingEnabled(True)

        # After data is in place, auto-fit compact columns and give Error remaining space
        self._apply_column_sizing()

        # Update results count
        total = len(self.all_history)
        showing = len(self.filtered_history)
        self.results_label.setText(f"Showing {showing:,} of {total:,} records")

    def _on_cell_double_clicked(self, row, col):
        """Handle cell double-click - open URL if URL column"""
        if col == 10:  # URL column
            item = self.table.item(row, col)
            if item:
                url = item.data(Qt.ItemDataRole.UserRole)
                if url and url.startswith('http'):
                    QDesktopServices.openUrl(QUrl(url))

    def _clear_filters(self):
        """Reset all filters"""
        self.brand_filter.setCurrentIndex(0)
        self.status_filter.setCurrentIndex(0)
        self.date_filter.setCurrentIndex(0)
        self.search_input.clear()

    def _export_to_excel(self):
        """Export filtered results to Excel"""
        if not self.filtered_history:
            InfoBar.warning(
                title="No Data",
                content="No records to export with current filters.",
                parent=self,
                position=InfoBarPosition.TOP
            )
            return

        try:
            # Get save path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export History to Excel",
                f"history_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Upload History"

            # Headers
            headers = [
                "Type", "Brand", "Product Code", "Status", "Duration (s)",
                "Features", "Images", "Error Message", "Failed Stage",
                "URL/Code", "Processed At"
            ]
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
            header_font = Font(bold=True, color="8D99AE")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows
            for record in self.filtered_history:
                ws.append([
                    self._record_type_label(record['batch_id']),
                    record['brand'],
                    record['product_code'],
                    record['status'],
                    record['duration_seconds'],
                    record['features_uploaded'] or 0,
                    record['images_uploaded'] or 0,
                    record['error_message'] or '',
                    record['failed_stage'] or '',
                    record['url_or_code'] or '',
                    record['processed_at']
                ])

                # Color-code status
                status_cell = ws.cell(ws.max_row, 4)
                if record['status'] in ('success', 'saved_manually'):
                    status_cell.font = Font(color="10B981", bold=True)
                elif record['status'] == 'ready_for_review':
                    status_cell.font = Font(color="F59E0B", bold=True)
                elif record['status'] in ('blocked_non_draft', 'discarded'):
                    status_cell.font = Font(color="8D99AE", bold=True)
                else:
                    status_cell.font = Font(color="C81D25", bold=True)

            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except (AttributeError, TypeError):
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

            wb.save(file_path)

            InfoBar.success(
                title="Export Successful",
                content=f"Exported {len(self.filtered_history)} records to {file_path}",
                parent=self,
                position=InfoBarPosition.TOP
            )

        except Exception as e:
            InfoBar.error(
                title="Export Failed",
                content=str(e),
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _go_back_to_analytics(self):
        """Navigate back to Analytics screen"""
        # Navigate to Analytics screen
        from GUI_Qt.screens.AnalyticsScreen import AnalyticsScreen

        # Show analytics screen (which is stored as history_screen in MainWindow)
        if hasattr(self.main, 'history_screen') and isinstance(self.main.history_screen, AnalyticsScreen):
            self.main._show_screen(self.main.history_screen)

    def _on_theme_changed(self):
        """Handle theme change"""
        self._apply_theme()
        self._apply_table_theme()

    def update_translations(self):
        """Update UI text for current language"""
        tr = self.main.i18n.tr

        # Header
        self.title_label.setText(tr("history.full.title", default="Full Upload History"))
        self.back_btn.setToolTip(tr("history.full.back", default="Back to Analytics"))

        # Search and filters
        self.search_label.setText(tr("history.full.search", default="Search:"))
        self.search_input.setPlaceholderText(tr("history.full.search.placeholder", default="Search by brand, product code, or error..."))

        self.brand_label.setText(tr("history.full.brand", default="Brand:"))
        self.status_label.setText(tr("history.full.status", default="Status:"))
        self.type_label.setText(tr("history.full.type", default="Type:"))
        self.date_label.setText(tr("history.full.period", default="Period:"))

        # Buttons
        self.refresh_btn.setToolTip(tr("history.full.refresh", default="Refresh history"))
        self.export_btn.setText(tr("history.full.export", default="Export to Excel"))
        self.clear_filters_btn.setText(tr("history.full.clear", default="Clear Filters"))

        # Re-populate filters with translated text
        self._populate_filters()

        # Update results label
        total = len(self.all_history)
        showing = len(self.filtered_history)
        self.results_label.setText(tr("history.full.results", default=f"Showing {showing:,} of {total:,} records").format(showing=showing, total=total))

    def _on_breakpoint_changed(self, breakpoint: str):
        """Respond to breakpoint changes - adjust margins and spacing."""
        margins = get_responsive_margins(breakpoint)
        spacing = get_responsive_spacing(breakpoint)
        if hasattr(self, 'content_widget') and self.content_widget.layout():
            self.content_widget.layout().setContentsMargins(*margins)
            self.content_widget.layout().setSpacing(spacing)


# Alias for backward compatibility
HistoryScreen = FullHistoryScreen
