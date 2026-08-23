"""GUI_Qt/screens/CodeGetterScreen.py

Walks through all product list pages, opens each product, reads its ERP code,
and exports the collected codes to Excel.
"""

from __future__ import annotations

import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Qt, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    FluentIcon,
    IconWidget,
    InfoBar,
    InfoBarPosition,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    TitleLabel,
    isDarkTheme,
)

from Config.Selectors import ProductListSelectors
from Managers.PimboProductEditor import PimboProductEditor
from GUI_Qt.styles.screen_theme import (
    CARD_SPACING,
    CONTENT_SPACING,
    ICON_TEXT_GAP,
    PAGE_MARGINS,
    PAGE_SPACING,
    TOOLBAR_MARGINS,
    apply_screen_theme,
    enforce_transparent_labels,
)
from GUI_Qt.styles.theme_config import (
    COLORS,
    COMPONENT_COLORS,
    FONTS,
    PADDINGS,
    RADII,
    SIZES,
    SPACING as THEME_SPACING,
    get_status_text_color,
)
from GUI_Qt.widgets import enable_table_copy, show_file_saved_bar
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget


class CodeGetterWorker(QThread):
    row_found = Signal(int, str, str, str, str)  # index, code, title, page, status
    progress_update = Signal(int, str)
    log = Signal(str)
    done = Signal(int, int)  # found, errors

    def __init__(self, driver):
        super().__init__()
        self.driver = driver
        self._stop = False
        self._found = 0
        self._errors = 0

    def request_stop(self):
        self._stop = True

    def run(self):
        try:
            self._ensure_products_page()
            page_no = 1

            while not self._stop:
                rows_count = self._wait_for_rows()
                self.log.emit(f"Page {page_no}: {rows_count} products")

                for row_index in range(rows_count):
                    if self._stop:
                        break
                    self._process_row(page_no, row_index)

                if self._stop or not self._go_next_page():
                    break
                page_no += 1

            self.done.emit(self._found, self._errors)
        except Exception as e:
            self.log.emit(f"Error: {e}")
            self.done.emit(self._found, self._errors + 1)

    def _ensure_products_page(self):
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        if "/dashboard/products/" in (self.driver.current_url or ""):
            if PimboProductEditor(self.driver).is_dirty():
                raise RuntimeError(
                    "Atidaryta forma turi neišsaugotų pakeitimų; read-only įrankis jų neatmes."
                )

        try:
            products_link = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable(ProductListSelectors.NAV_PRODUCTS)
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
        except Exception:
            self.driver.get(ProductListSelectors.URL)

        WebDriverWait(self.driver, 15).until(
            EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
        )

    def _wait_for_rows(self) -> int:
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        WebDriverWait(self.driver, 15).until(
            EC.presence_of_element_located(ProductListSelectors.PRODUCT_ROW)
        )
        time.sleep(0.5)
        return len(self.driver.find_elements(*ProductListSelectors.PRODUCT_ROWS))

    def _get_row_snapshot(self, index: int):
        rows = self.driver.find_elements(*ProductListSelectors.PRODUCT_ROWS)
        if index >= len(rows):
            raise IndexError(f"Row {index + 1} is no longer available")

        row = rows[index]
        title = ""
        visible_code = ""

        try:
            title = row.find_element(*ProductListSelectors.product_link_by_brand("")).text.strip()
        except Exception:
            pass

        try:
            visible_code = row.find_element(
                "xpath",
                ".//span[contains(concat(' ', normalize-space(@class), ' '), ' font-mono ')]",
            ).text.strip()
        except Exception:
            pass

        return row, title, visible_code

    def _process_row(self, page_no: int, row_index: int):
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        display_index = self._found + self._errors + 1
        page_label = str(page_no)

        try:
            row, title, visible_code = self._get_row_snapshot(row_index)
            self.progress_update.emit(display_index, f"Page {page_no}, row {row_index + 1}")
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", row)
            self.driver.execute_script("arguments[0].click();", row)

            editor = PimboProductEditor(self.driver)
            editor.wait_ready()
            code = editor.external_id()
            product_url = self.driver.current_url

            if not code:
                code = visible_code

            status = "Found" if code else "Missing code"
            if code:
                self._found += 1
            else:
                self._errors += 1

            self.row_found.emit(display_index, code, title, page_label, status)
            self.log.emit(f"{display_index}: {code or '(missing)'}")

            self.driver.back()
            WebDriverWait(self.driver, 15).until(
                EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
            )
            self._wait_for_rows()

            # Keep the local variable useful for future debugging without
            # adding another column to the user-facing table.
            _ = product_url
        except Exception as e:
            self._errors += 1
            self.row_found.emit(display_index, "", "", page_label, f"Error: {e}")
            self.log.emit(f"{display_index}: error {e}")
            try:
                if not self.driver.find_elements(*ProductListSelectors.PRODUCT_ROWS):
                    self.driver.back()
                self._wait_for_rows()
            except Exception:
                pass

    def _go_next_page(self) -> bool:
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        try:
            next_btn = WebDriverWait(self.driver, 5).until(
                EC.element_to_be_clickable(ProductListSelectors.NEXT_PAGE)
            )
        except Exception:
            return False

        old_first = ""
        try:
            rows = self.driver.find_elements(*ProductListSelectors.PRODUCT_ROWS)
            if rows:
                old_first = rows[0].text
        except Exception:
            pass

        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_btn)
        self.driver.execute_script("arguments[0].click();", next_btn)

        try:
            WebDriverWait(self.driver, 15).until(
                lambda d: len(d.find_elements(*ProductListSelectors.PRODUCT_ROWS)) > 0
                and (not old_first or d.find_elements(*ProductListSelectors.PRODUCT_ROWS)[0].text != old_first)
            )
        except Exception:
            time.sleep(1.0)

        return True


class CodeGetterScreen(ResponsiveWidget):
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main = main_window
        self.tr = main_window.i18n.tr
        self._worker: CodeGetterWorker | None = None

        self._build_ui()
        self.retranslate_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        self._scroll = ScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        root.addWidget(self._scroll)

        self._container = QWidget()
        self._layout = QVBoxLayout(self._container)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(CONTENT_SPACING)
        self._scroll.setWidget(self._container)

        apply_screen_theme(self, "CodeGetterScreen", scroll=self._scroll, content=self._container)

        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)
        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES["icon_lg"], SIZES["icon_lg"])
        header.addWidget(icon)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self._title = TitleLabel("")
        self._subtitle = CaptionLabel("")
        self._subtitle.setStyleSheet(f"color: {COLORS['text_secondary']};")
        title_col.addWidget(self._title)
        title_col.addWidget(self._subtitle)
        header.addLayout(title_col)
        header.addStretch()
        self._layout.addLayout(header)

        toolbar_card = CardWidget()
        toolbar_layout = QGridLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self._status_label = BodyLabel("")
        self._status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self._export_btn = PushButton(FluentIcon.SAVE, "")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_results)

        self._start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._start_btn.clicked.connect(self._on_start_stop)

        toolbar_layout.addWidget(self._status_label, 0, 0, 1, 2)
        toolbar_layout.addWidget(self._export_btn, 1, 0)
        toolbar_layout.addWidget(self._start_btn, 1, 1)
        toolbar_layout.setColumnStretch(0, 1)
        self._layout.addWidget(toolbar_card)

        self._progress_label = CaptionLabel("")
        self._progress_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self._layout.addWidget(self._progress_label)

        self._table = QTableWidget(0, 5)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Interactive)
        self._table.setColumnWidth(1, 180)
        self._table.setColumnWidth(4, 150)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setMinimumHeight(320)
        enable_table_copy(self._table)
        self._layout.addWidget(self._table, 1)

        self._update_table_theme()
        enforce_transparent_labels(self)

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        tc = COMPONENT_COLORS["table"]
        bg = tc["row_bg_dark"] if is_dark else tc["row_bg_light"]
        alt_bg = tc["row_alt_bg_dark"] if is_dark else tc["row_alt_bg_light"]
        border = tc["border_dark"] if is_dark else tc["border_light"]
        header_bg = tc["header_bg_dark"] if is_dark else tc["header_bg_light"]
        header_text = tc["header_text_dark"] if is_dark else tc["header_text_light"]
        text_color = COLORS["text_primary_dark"] if is_dark else COLORS["text_primary_light"]

        from GUI_Qt.styles.theme_config import (
            get_scrollbar_handle_bg,
            get_scrollbar_handle_hover_bg,
            get_selection_bg,
        )

        self._table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg};
                alternate-background-color: {alt_bg};
                border: 1px solid {border};
                border-radius: {RADII['md']}px;
                gridline-color: {border};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-bottom: 1px solid {border};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {THEME_SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
        """)

    def retranslate_ui(self):
        self._title.setText(self.tr("codegetter.title"))
        self._subtitle.setText(self.tr("codegetter.subtitle"))
        self._start_btn.setText(self.tr("codegetter.start"))
        self._export_btn.setText(self.tr("codegetter.export"))
        self._status_label.setText(self.tr("codegetter.ready"))
        self._table.setHorizontalHeaderLabels([
            self.tr("codegetter.col.index"),
            self.tr("codegetter.col.code"),
            self.tr("codegetter.col.title"),
            self.tr("codegetter.col.page"),
            self.tr("codegetter.col.status"),
        ])

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        self._table.setRowCount(0)
        self._export_btn.setEnabled(False)
        self._start_btn.setText(self.tr("codegetter.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._status_label.setText(self.tr("codegetter.running"))

        self._worker = CodeGetterWorker(driver)
        self._worker.row_found.connect(self._on_row_found)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[CodeGetter] {msg}"))
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(self._worker, "code_scanner", "code_getter")
        self._worker.start()

    def _on_row_found(self, index: int, code: str, title: str, page: str, status: str):
        row = self._table.rowCount()
        self._table.insertRow(row)

        values = [str(index), code, title, page, status]
        for col, value in enumerate(values):
            item = QTableWidgetItem(value)
            if col == 4:
                if status == "Found":
                    item.setForeground(QColor(get_status_text_color("success", isDarkTheme())))
                elif status.startswith("Error"):
                    item.setForeground(QColor(get_status_text_color("error", isDarkTheme())))
                else:
                    item.setForeground(QColor(get_status_text_color("warning", isDarkTheme())))
            self._table.setItem(row, col, item)

        self._table.scrollToItem(self._table.item(row, 0))

    def _on_progress(self, current: int, label: str):
        self._progress_label.setText(self.tr("codegetter.progress", current=current, label=label))

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("codegetter.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._export_btn.setEnabled(self._table.rowCount() > 0)

        total = found + errors
        done_text = self.tr("codegetter.done", found=found, errors=errors, total=total)
        self._status_label.setText(done_text)
        self._progress_label.setText(done_text)

        InfoBar.success(
            self.tr("codegetter.done_title"),
            done_text,
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("codegetter.export"),
            f"product_codes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)",
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Product Codes"

            headers = ["#", "Code", "Title", "Page", "Status"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(5):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 60
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 22

            wb.save(path)
            show_file_saved_bar(self, self.tr("common.success"), self.tr("codegetter.exported"), path)
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )
