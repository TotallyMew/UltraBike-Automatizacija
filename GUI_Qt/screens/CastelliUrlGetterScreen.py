"""GUI_Qt/screens/CastelliUrlGetterScreen.py

Looks up Castelli product URLs from UltraBike product codes.
"""

from __future__ import annotations

import os
import re
import time
from datetime import datetime
from urllib.parse import quote_plus

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Qt, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    FluentIcon,
    IconWidget,
    InfoBar,
    InfoBarPosition,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    TitleLabel,
    isDarkTheme,
)

from Config.Selectors import CastelliSelectors
from GUI_Qt.screens.NameGetterScreen import DropZoneWidget
from GUI_Qt.styles.screen_theme import (
    CARD_SPACING,
    CONTENT_SPACING,
    ICON_TEXT_GAP,
    PAGE_MARGINS,
    PAGE_SPACING,
    TOOLBAR_MARGINS,
    apply_screen_theme,
    enforce_transparent_labels,
)
from GUI_Qt.styles.theme_config import (
    COLORS,
    COMPONENT_COLORS,
    FONTS,
    PADDINGS,
    RADII,
    SIZES,
    SPACING as THEME_SPACING,
    get_status_text_color,
)
from GUI_Qt.widgets import enable_table_copy, show_file_saved_bar
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget


def extract_castelli_code(raw_code: str) -> str:
    """Extract 4401237 from UB-4401237-002."""
    value = (raw_code or "").strip()
    match = re.search(r"UB-(\d+)-", value, re.IGNORECASE)
    if match:
        return match.group(1)
    digits = re.search(r"\d{6,}", value)
    return digits.group(0) if digits else value


def extract_castelli_color(raw_code: str) -> str:
    """Extract 002 from UB-4401237-002."""
    value = (raw_code or "").strip()
    match = re.search(r"UB-\d+-([A-Za-z0-9]+)", value, re.IGNORECASE)
    return match.group(1) if match else ""


class CastelliUrlGetterWorker(QThread):
    row_update = Signal(int, str, str)
    progress_update = Signal(int, int)
    done = Signal(int, int)
    log = Signal(str)

    def __init__(self, driver, rows: list[dict[str, str]]):
        super().__init__()
        self.driver = driver
        self.rows = rows
        self._stop = False

    def request_stop(self):
        self._stop = True

    def run(self):
        found = 0
        errors = 0

        for idx, row in enumerate(self.rows):
            if self._stop:
                self.row_update.emit(idx, "Skipped", "")
                continue

            search_code = row["search_code"]
            color_code = row.get("color_code", "")
            self.progress_update.emit(idx + 1, len(self.rows))
            self.row_update.emit(idx, "Searching", "")
            self.log.emit(f"Searching {search_code}")

            try:
                status, urls = self._lookup_urls(search_code, color_code)
                if urls:
                    found += 1
                    status = status or ("Found" if len(urls) == 1 else "Multiple results")
                    self.row_update.emit(idx, status, "\n".join(urls))
                else:
                    errors += 1
                    self.row_update.emit(idx, status or "Not found", "")
            except Exception as e:
                errors += 1
                self.row_update.emit(idx, f"Error: {e}", "")

        self.done.emit(found, errors)

    def _lookup_urls(self, search_code: str, color_code: str = "") -> tuple[str, list[str]]:
        target_url = CastelliSelectors.SEARCH_URL.format(query=quote_plus(search_code))
        self.log.emit(f"Opening {target_url}")
        self._navigate_to(target_url)
        self._dismiss_castelli_overlays()

        # The country chooser can redirect after "Choose"; return to the exact
        # search URL once overlays are gone.
        if search_code not in (self.driver.current_url or ""):
            self._navigate_to(target_url)
            self._dismiss_castelli_overlays()

        return self._wait_for_result(search_code, color_code)

    def _navigate_to(self, url: str):
        try:
            self.driver.switch_to.alert.accept()
            time.sleep(0.2)
        except Exception:
            pass

        try:
            self.driver.get(url)
        except Exception:
            try:
                self.driver.execute_script("window.location.href = arguments[0];", url)
            except Exception:
                raise

    def _dismiss_castelli_overlays(self):
        """Dismiss cookies/country popups in whichever order they appear."""
        for _ in range(3):
            changed = False
            if self._accept_cookies():
                changed = True
            if self._handle_country_modal():
                changed = True
            if self._accept_cookies():
                changed = True
            if not changed:
                break

    def _accept_cookies(self) -> bool:
        for locator in CastelliSelectors.COOKIE_BUTTON_CANDIDATES:
            try:
                for btn in self.driver.find_elements(*locator):
                    if btn.is_displayed() and btn.is_enabled():
                        self.driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.2)
                        return True
            except Exception:
                continue
        return False

    def _handle_country_modal(self) -> bool:
        from selenium.webdriver.support.ui import Select

        try:
            country_elements = self.driver.find_elements(*CastelliSelectors.COUNTRY_SELECT)
            if not country_elements:
                return False
            country_select = country_elements[0]
            if not country_select.is_displayed():
                return False
        except Exception:
            return False

        try:
            Select(country_select).select_by_value("LT")
        except Exception:
            pass

        try:
            language_select = self.driver.find_element(*CastelliSelectors.LANGUAGE_SELECT)
            Select(language_select).select_by_value("en")
        except Exception:
            pass

        try:
            for choose_btn in self.driver.find_elements(*CastelliSelectors.COUNTRY_CHOOSE_BUTTON):
                if choose_btn.is_displayed() and choose_btn.is_enabled():
                    self.driver.execute_script("arguments[0].click();", choose_btn)
                    time.sleep(0.5)
                    return True
        except Exception:
            pass
        return False

    def _open_search_box(self):
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        try:
            WebDriverWait(self.driver, 2).until(
                EC.visibility_of_element_located(CastelliSelectors.SEARCH_INPUT)
            )
            return
        except Exception:
            pass

        for locator in CastelliSelectors.SEARCH_OPEN_CANDIDATES:
            try:
                opener = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable(locator))
                self.driver.execute_script("arguments[0].click();", opener)
                WebDriverWait(self.driver, 4).until(
                    EC.visibility_of_element_located(CastelliSelectors.SEARCH_INPUT)
                )
                return
            except Exception:
                continue

        raise RuntimeError("Castelli search field was not found")

    def _wait_for_result(self, search_code: str, color_code: str = "") -> tuple[str, list[str]]:
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        time.sleep(0.5)

        try:
            result = WebDriverWait(self.driver, 6).until(
                lambda d: (
                    d.find_elements(*CastelliSelectors.PRODUCT_LINK)
                    or d.find_elements(*CastelliSelectors.NO_RESULTS)
                )
            )
        except Exception:
            return "Not found", []

        product_links = self.driver.find_elements(*CastelliSelectors.PRODUCT_LINK)
        if product_links:
            return self._select_matching_urls(product_links, search_code, color_code)

        _ = result
        return "Not found", []

    def _select_matching_urls(self, links, search_code: str, color_code: str = "") -> tuple[str, list[str]]:
        base_urls = []
        matched_urls = []
        seen = set()
        match_token = f"{search_code}_{color_code}_".lower() if color_code else ""

        for link in links:
            url = (link.get_attribute("href") or "").strip()
            if not url or url in seen:
                continue

            seen.add(url)

            # Castelli reuses onclick productClick data across color slides, so
            # match only against the actual href/title variant code.
            haystack = " ".join([
                url,
                link.get_attribute("title") or "",
            ]).lower()

            if search_code.lower() in haystack:
                base_urls.append(url)

            if match_token and match_token in haystack:
                matched_urls.append(url)

        if matched_urls:
            return "Found", [matched_urls[0]]

        if color_code and base_urls:
            return "Not found second part", []

        if base_urls:
            return ("Found" if len(base_urls) == 1 else "Multiple results"), base_urls

        return "Not found", []


class CastelliUrlGetterScreen(ResponsiveWidget):
    def __init__(self, main_window):
        super().__init__(main_window)
        self.main = main_window
        self.tr = main_window.i18n.tr
        self._worker: CastelliUrlGetterWorker | None = None
        self._rows: list[dict[str, str]] = []
        self._file_path: str | None = None

        self._build_ui()
        self.retranslate_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        self._scroll = ScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        root.addWidget(self._scroll)

        self._container = QWidget()
        self._layout = QVBoxLayout(self._container)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(CONTENT_SPACING)
        self._scroll.setWidget(self._container)

        apply_screen_theme(self, "CastelliUrlGetterScreen", scroll=self._scroll, content=self._container)

        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)
        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES["icon_lg"], SIZES["icon_lg"])
        header.addWidget(icon)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self._title = TitleLabel("")
        self._subtitle = CaptionLabel("")
        self._subtitle.setStyleSheet(f"color: {COLORS['text_secondary']};")
        title_col.addWidget(self._title)
        title_col.addWidget(self._subtitle)
        header.addLayout(title_col)
        header.addStretch()
        self._layout.addLayout(header)

        toolbar_card = CardWidget()
        toolbar_layout = QGridLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self._browse_btn = PushButton(FluentIcon.FOLDER, "")
        self._browse_btn.clicked.connect(self._browse_file)

        self._file_label = CaptionLabel("")
        self._file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self._export_btn = PushButton(FluentIcon.SAVE, "")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_results)

        self._start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._start_btn.setEnabled(False)
        self._start_btn.clicked.connect(self._on_start_stop)

        toolbar_layout.addWidget(self._browse_btn, 0, 0)
        toolbar_layout.addWidget(self._file_label, 0, 1)
        toolbar_layout.addWidget(self._export_btn, 1, 0)
        toolbar_layout.addWidget(self._start_btn, 1, 1)
        toolbar_layout.setColumnStretch(1, 1)
        self._layout.addWidget(toolbar_card)

        self._drop_zone = DropZoneWidget(self.tr)
        self._drop_zone.setMinimumHeight(120)
        self._drop_zone.file_dropped.connect(self._on_file_dropped)
        self._layout.addWidget(self._drop_zone)

        self._progress_label = CaptionLabel("")
        self._progress_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self._layout.addWidget(self._progress_label)

        self._table = QTableWidget(0, 5)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(1, 170)
        self._table.setColumnWidth(2, 140)
        self._table.setColumnWidth(3, 130)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setMinimumHeight(300)
        enable_table_copy(self._table)
        self._layout.addWidget(self._table, 1)

        self._update_table_theme()
        enforce_transparent_labels(self)

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        tc = COMPONENT_COLORS["table"]
        bg = tc["row_bg_dark"] if is_dark else tc["row_bg_light"]
        alt_bg = tc["row_alt_bg_dark"] if is_dark else tc["row_alt_bg_light"]
        border = tc["border_dark"] if is_dark else tc["border_light"]
        header_bg = tc["header_bg_dark"] if is_dark else tc["header_bg_light"]
        header_text = tc["header_text_dark"] if is_dark else tc["header_text_light"]
        text_color = COLORS["text_primary_dark"] if is_dark else COLORS["text_primary_light"]

        from GUI_Qt.styles.theme_config import get_scrollbar_handle_bg, get_scrollbar_handle_hover_bg, get_selection_bg

        self._table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg};
                alternate-background-color: {alt_bg};
                border: 1px solid {border};
                border-radius: {RADII['md']}px;
                gridline-color: {border};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-bottom: 1px solid {border};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {THEME_SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
        """)

    def retranslate_ui(self):
        self._title.setText(self.tr("castelliurl.title"))
        self._subtitle.setText(self.tr("castelliurl.subtitle"))
        self._browse_btn.setText(self.tr("castelliurl.browse"))
        self._start_btn.setText(self.tr("castelliurl.start"))
        self._export_btn.setText(self.tr("castelliurl.export"))
        self._table.setHorizontalHeaderLabels([
            self.tr("castelliurl.col.index"),
            self.tr("castelliurl.col.original_code"),
            self.tr("castelliurl.col.search_code"),
            self.tr("castelliurl.col.status"),
            self.tr("castelliurl.col.url"),
        ])

    def _on_file_dropped(self, path: str):
        if path == "__browse__":
            self._browse_file()
            return
        self._load_file(path)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("batch.file.select_excel.title"),
            "", "Excel files (*.xlsx);;All files (*.*)"
        )
        if path:
            self._load_file(path)

    def _load_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            rows: list[dict[str, str]] = []

            for excel_row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                value = excel_row[0]
                if value is None:
                    continue
                original_code = str(value).strip()
                if not original_code:
                    continue
                if original_code.strip().lower() in {"code", "product code", "original code"}:
                    continue
                rows.append({
                    "original_code": original_code,
                    "search_code": extract_castelli_code(original_code),
                    "color_code": extract_castelli_color(original_code),
                })
            wb.close()

            if not rows:
                InfoBar.warning(
                    self.tr("common.error"),
                    self.tr("castelliurl.no_codes"),
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                )
                return

            self._rows = rows
            self._file_path = path
            self._file_label.setText(os.path.basename(path))
            self._table.setRowCount(len(rows))

            for idx, row in enumerate(rows):
                self._table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
                self._table.setItem(idx, 1, QTableWidgetItem(row["original_code"]))
                self._table.setItem(idx, 2, QTableWidgetItem(row["search_code"]))
                status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
                status.setForeground(QColor(COLORS["text_secondary"]))
                self._table.setItem(idx, 3, status)
                self._table.setItem(idx, 4, QTableWidgetItem(""))

            self._start_btn.setEnabled(True)
            self._export_btn.setEnabled(False)
            self._drop_zone.setVisible(False)
            self._progress_label.setText(self.tr("castelliurl.loaded", count=len(rows)))
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        if not self._rows:
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        for idx in range(self._table.rowCount()):
            status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
            status.setForeground(QColor(COLORS["text_secondary"]))
            self._table.setItem(idx, 3, status)
            self._table.setItem(idx, 4, QTableWidgetItem(""))

        self._start_btn.setText(self.tr("castelliurl.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._browse_btn.setEnabled(False)
        self._export_btn.setEnabled(False)

        self._worker = CastelliUrlGetterWorker(driver, self._rows)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[CastelliUrlGetter] {msg}"))
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._worker, "url_scanner", "castelli_url_getter", total=len(self._rows)
            )
        self._worker.start()

    def _on_row_update(self, row: int, status: str, url: str):
        item = QTableWidgetItem(status)
        if status == "Found":
            item.setForeground(QColor(get_status_text_color("success", isDarkTheme())))
        elif status == "Multiple results":
            item.setForeground(QColor(COLORS['focus_ring_dark'] if isDarkTheme() else COLORS['focus_ring_light']))
        elif status in ("Searching", "Skipped", "Not found second part"):
            item.setForeground(QColor(get_status_text_color("warning", isDarkTheme())))
        else:
            item.setForeground(QColor(get_status_text_color("error", isDarkTheme())))

        self._table.setItem(row, 3, item)
        self._table.setItem(row, 4, QTableWidgetItem(url))
        self._table.scrollToItem(item)

    def _on_progress(self, current: int, total: int):
        self._progress_label.setText(self.tr("castelliurl.progress", current=current, total=total))

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("castelliurl.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._browse_btn.setEnabled(True)
        self._export_btn.setEnabled(True)

        total = found + errors
        text = self.tr("castelliurl.done", found=found, errors=errors, total=total)
        self._progress_label.setText(text)
        InfoBar.success(
            self.tr("castelliurl.done_title"),
            text,
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("castelliurl.export"),
            f"castelli_urls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Castelli URLs"

            headers = ["#", "Original Code", "Search Code", "Status", "URL"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(5):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 100

            wb.save(path)
            show_file_saved_bar(self, self.tr("common.success"), self.tr("castelliurl.exported"), path)
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )
