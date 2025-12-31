"""GUI_Qt/screens/BatchTitlesScreen.py

Batch Lithuanian titles updater.

UX parity with BatchUploadScreen / BatchDescriptionsScreen:
- Manual mode: table rows
- Excel mode: drag/drop + browse + template download
- Start enabled only when rows are valid
- Per-row Status/Error updates while running

Automation reuses the existing logged-in Selenium session from MainWindow.

Excel columns:
- Brand
- Code (without UB- prefix; we add it automatically)
- Title (Lithuanian title / full product name)
"""

from __future__ import annotations

import os
import time
import json
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QEvent, QThread, QTimer, Qt, Signal
from PySide6.QtGui import QDragEnterEvent, QDropEvent, QStandardItemModel, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QTableWidget,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    ComboBox,
    FluentIcon,
    IconWidget,
    IndeterminateProgressRing,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    MessageBox,
    PillPushButton,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    TitleLabel,
    TransparentToolButton,
    isDarkTheme,
    qconfig,
)

from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import COLORS, COMPONENT_COLORS, FONTS, RADII, PADDINGS, SIZES
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS, PAGE_SPACING, ICON_TEXT_GAP, ROW_SPACING, TOOLBAR_MARGINS,
    CARD_SPACING, CONTENT_SPACING, TABLE_CELL_MARGINS, SPACING,
    apply_screen_theme, get_responsive_margins, get_responsive_spacing
)
from Utilities.ProductNavigationHandler import ProductNavigationHandler
from Utilities.WebIntercationHandler import WebInteractionHandler
from Managers.FeatureUploader.LanguageSwitcher import LanguageSwitcher


class DropZoneWidget(QWidget):
    """Drag & drop zone for a single .xlsx file."""

    file_dropped = Signal(str)

    def __init__(self, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(PAGE_SPACING)

        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES['icon_huge'], SIZES['icon_huge'])

        self.title_label = BodyLabel("")
        self.title_label.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']}; color: {COLORS['lavender_grey']};")

        self.subtitle_label = CaptionLabel("")
        self.subtitle_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        layout.addStretch()
        layout.addWidget(icon, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.title_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.subtitle_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addStretch()

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._apply_style(False)
        self.retranslate_ui()

    def retranslate_ui(self):
        self.title_label.setText(self.tr("batch.drop.title"))
        self.subtitle_label.setText(self.tr("batch.drop.subtitle"))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1 and urls[0].toLocalFile().endswith('.xlsx'):
                event.acceptProposedAction()
                self._apply_style(True)

    def dragLeaveEvent(self, event):
        self._apply_style(False)

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith('.xlsx'):
            self.file_dropped.emit(files[0])
        self._apply_style(False)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.file_dropped.emit("__browse__")

    def _apply_style(self, drag_active: bool):
        is_dark = isDarkTheme()
        border = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        dashed = COLORS['lavender_grey']
        from GUI_Qt.styles.theme_config import get_hover_bg
        hover_bg = get_hover_bg(is_dark)

        if drag_active:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: {hover_bg};
                    border: 2px solid {border};
                    border-radius: {RADII['lg']}px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: transparent;
                    border: 2px dashed {dashed};
                    border-radius: {RADII['lg']}px;
                }}
                DropZoneWidget:hover {{
                    border-color: {border};
                    background-color: {hover_bg};
                }}
            """)


class BatchTitlesWorker(QThread):
    row_update = Signal(int, str, str)  # rowIndex, status, error
    log = Signal(str)
    done = Signal(int, int)  # ok, total
    progress_update = Signal(int, int, float, float)  # current, total, speed (items/sec), eta_seconds

    def __init__(self, main_window, items: list[dict]):
        super().__init__()
        self.main = main_window
        self.items = items
        self._processed_times = []  # Track duration of each processed item

        self._prestashop_api_base_url = (self.main.settings.get('prestashop_url', '') or '').strip()
        self._prestashop_api_key = (self.main.settings.get('prestashop_api_key', '') or '').strip()
        self._prestashop_api_enabled = bool(self.main.settings.get('prestashop_api_enabled', False))

    def _update_title_via_api(self, code: str, title_lt: str, title_en: str, title_lv: str) -> bool:
        try:
            if not self._prestashop_api_enabled:
                return False
            if not (self._prestashop_api_base_url and self._prestashop_api_key):
                return False

            from Managers.PrestaShopAPI import PrestaShopAPI

            api = PrestaShopAPI(self._prestashop_api_base_url, self._prestashop_api_key, logger=getattr(self.main, 'logger', None))
            lang_map = api.get_language_id_map() or {}
            en_id = str(lang_map.get('en') or '1')
            lt_id = str(lang_map.get('lt') or '2')
            lv_id = str(lang_map.get('lv') or '3')

            payload = {}
            if title_en and en_id:
                payload[en_id] = title_en
            if title_lt and lt_id:
                payload[lt_id] = title_lt
            if title_lv and lv_id:
                payload[lv_id] = title_lv
            if not payload:
                return False

            product_id = api.search_product_by_reference(code)
            if not product_id:
                extra = (api.last_error_summary() or '').strip()
                self.log.emit(f"API: product not found for {code}" + (f" ({extra})" if extra else ""))
                return False

            ok = api.update_product_name(int(product_id), payload)
            if not ok:
                extra = (api.last_error_summary() or '').strip()
                self.log.emit(f"API: title update failed for {code}" + (f" ({extra})" if extra else ""))
                return False

            self.log.emit(f"API: updated title for {code}")
            return True
        except Exception as e:
            self.log.emit(f"API: title update crashed for {code}: {e}")
            return False

    def _ensure_products_page(self, driver) -> None:
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.ui import WebDriverWait

            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "subtab-AdminProducts"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table"))
            )
        except Exception:
            pass

    def _set_title(self, driver, lang_code: str, title: str) -> None:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        from Config.LanguageConfig import LANG_CODE_TO_PRESTASHOP_ID, SUPPORTED_LANGUAGES

        lang_id_map = LANG_CODE_TO_PRESTASHOP_ID
        if lang_code not in SUPPORTED_LANGUAGES:
            raise Exception(f"Unsupported language: {lang_code}")

        # Make sure the correct language is active so we don't type into a hidden tab.
        try:
            switcher = LanguageSwitcher(driver, logger=getattr(self.main, "logger", None))
            switcher.switchTo(lang_code)
            time.sleep(0.4)
        except Exception:
            # Best-effort; we'll still try to set by ID below.
            pass

        el = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, f"form_step1_name_{lang_id_map[lang_code]}"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)

        # Try JS first (more reliable with some reactive inputs), then fallback to keystrokes.
        try:
            driver.execute_script(
                """
                const el = arguments[0];
                const value = arguments[1];
                el.focus();
                el.value = value;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                """,
                el,
                title,
            )
            return
        except Exception:
            pass

        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)

        try:
            el.send_keys(Keys.CONTROL, "a")
            el.send_keys(Keys.BACKSPACE)
            el.send_keys(title)
        except Exception:
            # Last resort: clear() + send_keys
            try:
                el.clear()
            except Exception:
                pass
            el.send_keys(title)

    def run(self):
        tr = self.main.i18n.tr
        driver = getattr(self.main, "driver", None)
        if driver is None and not (self._prestashop_api_enabled and self._prestashop_api_base_url and self._prestashop_api_key):
            for it in self.items:
                self.row_update.emit(it["row"], tr("batchtitle.status.failed"), tr("batchtitle.no_session"))
            self.done.emit(0, len(self.items))
            return

        nav = None
        web = None
        if driver is not None:
            nav = ProductNavigationHandler(driver, logger=self.main.logger)
            web = WebInteractionHandler(driver)

        ok = 0
        total = len(self.items)

        batch_id = f"batchtitle_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        for idx, it in enumerate(self.items, start=1):
            row = it["row"]
            brand = it["brand"]
            code = it["code"]
            title_lt = it.get("title_lt") or ""
            title_en = it.get("title_en") or ""
            title_lv = it.get("title_lv") or ""

            self.row_update.emit(row, tr("batchtitle.status.running"), "")
            self.log.emit(tr("batchtitle.progress.open", current=idx, total=total, code=code))

            t0 = time.perf_counter()

            try:
                # Try API first if enabled.
                if self._update_title_via_api(code, title_lt, title_en, title_lv):
                    ok += 1
                    self.row_update.emit(row, tr("batchtitle.status.ok"), "")
                    try:
                        details = {
                            "workflow": "batch_titles",
                            "batch_id": batch_id,
                            "set_lt": bool(title_lt),
                            "set_en": bool(title_en),
                            "set_lv": bool(title_lv),
                            "mode": "api",
                        }
                        self.main.db.conn.execute(
                            """
                            INSERT INTO processing_history
                            (brand, product_code, url_or_code, status, duration_seconds,
                             failed_stage, batch_id, details_json, processed_at)
                            VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                            """,
                            (
                                brand,
                                code,
                                None,
                                float(time.perf_counter() - t0),
                                None,
                                batch_id,
                                json.dumps(details, ensure_ascii=False),
                                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            ),
                        )
                        self.main.db.conn.commit()
                    except Exception:
                        pass
                    continue

                # Fall back to browser automation.
                if driver is None or nav is None or web is None:
                    self.row_update.emit(row, tr("batchtitle.status.failed"), tr("batchtitle.no_session"))
                    continue
                self._ensure_products_page(driver)
                nav.navigate_to_product(brand, code)

                # Update each provided language title.
                if title_lt:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx, total=total, code=code, lang="LT"))
                    self._set_title(driver, "lt", title_lt)
                if title_en:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx, total=total, code=code, lang="EN"))
                    self._set_title(driver, "en", title_en)
                if title_lv:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx, total=total, code=code, lang="LV"))
                    self._set_title(driver, "lv", title_lv)

                self.log.emit(tr("batchtitle.progress.save", current=idx, total=total, code=code))
                web.save_information()
                time.sleep(0.5)

                ok += 1
                self.row_update.emit(row, tr("batchtitle.status.ok"), "")

                try:
                    details = {
                        "workflow": "batch_titles",
                        "batch_id": batch_id,
                        "set_lt": bool(title_lt),
                        "set_en": bool(title_en),
                        "set_lv": bool(title_lv),
                        "mode": "sequential",
                    }
                    self.main.db.conn.execute(
                        """
                        INSERT INTO processing_history
                        (brand, product_code, url_or_code, status, duration_seconds,
                         failed_stage, batch_id, details_json, processed_at)
                        VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                        """,
                        (
                            brand,
                            code,
                            None,
                            float(time.perf_counter() - t0),
                            None,
                            batch_id,
                            json.dumps(details, ensure_ascii=False),
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        ),
                    )
                    self.main.db.conn.commit()
                except Exception:
                    pass
            except Exception as e:
                self.row_update.emit(row, tr("batchtitle.status.failed"), str(e))

                try:
                    details = {
                        "workflow": "batch_titles",
                        "batch_id": batch_id,
                        "set_lt": bool(title_lt),
                        "set_en": bool(title_en),
                        "set_lv": bool(title_lv),
                        "mode": "sequential",
                    }
                    self.main.db.conn.execute(
                        """
                        INSERT INTO processing_history
                        (brand, product_code, url_or_code, status, error_message, duration_seconds,
                         failed_stage, batch_id, details_json, processed_at)
                        VALUES (?, ?, ?, 'failed', ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            brand,
                            code,
                            None,
                            str(e),
                            float(time.perf_counter() - t0),
                            "batch_titles",
                            batch_id,
                            json.dumps(details, ensure_ascii=False),
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        ),
                    )
                    self.main.db.conn.commit()
                except Exception:
                    pass

            # Track timing and emit progress update
            item_duration = time.perf_counter() - t0
            self._processed_times.append(item_duration)

            # Calculate ETA (skip first 2 items for more accurate average)
            if len(self._processed_times) >= 2:
                recent_times = self._processed_times[1:]  # Skip first item (often slower)
                avg_time_per_item = sum(recent_times) / len(recent_times)
                remaining_items = total - idx
                eta_seconds = avg_time_per_item * remaining_items
                speed = 1.0 / avg_time_per_item if avg_time_per_item > 0 else 0

                self.progress_update.emit(idx, total, speed, eta_seconds)

        self.done.emit(ok, total)


class ParallelBatchTitlesWorker(QThread):
    """Parallel batch titles update using browser session pool"""
    row_update = Signal(int, str, str)  # rowIndex, status, error
    log = Signal(str)
    done = Signal(int, int)  # ok, total
    session_status = Signal(dict)  # session pool statistics

    def __init__(self, main_window, session_manager, items: list[dict]):
        super().__init__()
        self.main = main_window
        self.session_manager = session_manager
        self.items = items
        self._stop = False
        self._lock = __import__('threading').Lock()
        self._ok_count = 0
        self._work_index = 0

        self._batch_id = f"batchtitle_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        self._prestashop_api_base_url = (self.main.settings.get('prestashop_url', '') or '').strip()
        self._prestashop_api_key = (self.main.settings.get('prestashop_api_key', '') or '').strip()
        self._prestashop_api_enabled = bool(self.main.settings.get('prestashop_api_enabled', False))

    def stop(self):
        self._stop = True

    def _update_title_via_api(self, code: str, title_lt: str, title_en: str, title_lv: str) -> bool:
        try:
            if not self._prestashop_api_enabled:
                return False
            if not (self._prestashop_api_base_url and self._prestashop_api_key):
                return False

            from Managers.PrestaShopAPI import PrestaShopAPI

            api = PrestaShopAPI(self._prestashop_api_base_url, self._prestashop_api_key, logger=getattr(self.main, 'logger', None))
            lang_map = api.get_language_id_map() or {}
            en_id = str(lang_map.get('en') or '1')
            lt_id = str(lang_map.get('lt') or '2')
            lv_id = str(lang_map.get('lv') or '3')

            payload = {}
            if title_en and en_id:
                payload[en_id] = title_en
            if title_lt and lt_id:
                payload[lt_id] = title_lt
            if title_lv and lv_id:
                payload[lv_id] = title_lv
            if not payload:
                return False

            product_id = api.search_product_by_reference(code)
            if not product_id:
                return False

            return bool(api.update_product_name(int(product_id), payload))
        except Exception:
            return False

    def _ensure_products_page(self, driver) -> None:
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support import expected_conditions as EC
            from selenium.webdriver.support.ui import WebDriverWait

            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "subtab-AdminProducts"))
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table"))
            )
        except Exception:
            pass

    def _set_title(self, driver, lang_code: str, title: str) -> None:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        from Config.LanguageConfig import LANG_CODE_TO_PRESTASHOP_ID, SUPPORTED_LANGUAGES

        lang_id_map = LANG_CODE_TO_PRESTASHOP_ID
        if lang_code not in SUPPORTED_LANGUAGES:
            raise Exception(f"Unsupported language: {lang_code}")

        try:
            switcher = LanguageSwitcher(driver, logger=getattr(self.main, "logger", None))
            switcher.switchTo(lang_code)
            time.sleep(0.4)
        except Exception:
            pass

        el = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, f"form_step1_name_{lang_id_map[lang_code]}"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", el)

        try:
            driver.execute_script(
                """
                const el = arguments[0];
                const value = arguments[1];
                el.focus();
                el.value = value;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
                """,
                el,
                title,
            )
            return
        except Exception:
            pass

        try:
            el.click()
        except Exception:
            driver.execute_script("arguments[0].click();", el)

        try:
            el.send_keys(Keys.CONTROL, "a")
            el.send_keys(Keys.BACKSPACE)
            el.send_keys(title)
        except Exception:
            try:
                el.clear()
            except Exception:
                pass
            el.send_keys(title)

    def run(self):
        """Execute parallel batch titles update using session pool"""
        import threading

        if not self.session_manager or not self.session_manager.is_ready():
            tr = self.main.i18n.tr
            for it in self.items:
                self.row_update.emit(it["row"], tr("batchtitle.status.failed"), "Session pool not available")
            self.done.emit(0, len(self.items))
            return

        total = len(self.items)
        pool_size = self.session_manager.pool_size

        self.log.emit(f"Starting parallel processing with {pool_size} browsers")

        # Create worker threads
        threads = []
        for worker_id in range(pool_size):
            thread = threading.Thread(target=self._worker_thread, args=(worker_id,))
            thread.daemon = True
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

        self.log.emit(f"Completed. Success: {self._ok_count}/{total}")
        self.done.emit(self._ok_count, total)

    def _worker_thread(self, worker_id: int):
        """Worker thread that processes items from the queue"""
        tr = self.main.i18n.tr

        import time as _time

        while not self._stop:
            # Get next work item
            with self._lock:
                if self._work_index >= len(self.items):
                    break  # No more work
                idx = self._work_index
                self._work_index += 1

            it = self.items[idx]
            row = it["row"]
            brand = it["brand"]
            code = it["code"]
            title_lt = it.get("title_lt") or ""
            title_en = it.get("title_en") or ""
            title_lv = it.get("title_lv") or ""

            t0 = _time.perf_counter()

            # Try API first if enabled.
            if self._update_title_via_api(code, title_lt, title_en, title_lv):
                self.row_update.emit(row, tr("batchtitle.status.ok"), "")
                with self._lock:
                    self._ok_count += 1
                try:
                    details = {
                        "workflow": "batch_titles",
                        "batch_id": self._batch_id,
                        "set_lt": bool(title_lt),
                        "set_en": bool(title_en),
                        "set_lv": bool(title_lv),
                        "mode": "api",
                    }
                    self.main.db.conn.execute(
                        """
                        INSERT INTO processing_history
                        (brand, product_code, url_or_code, status, duration_seconds,
                         failed_stage, batch_id, details_json, processed_at)
                        VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                        """,
                        (
                            brand,
                            code,
                            None,
                            float(_time.perf_counter() - t0),
                            None,
                            self._batch_id,
                            json.dumps(details, ensure_ascii=False),
                        ),
                    )
                    self.main.db.conn.commit()
                except Exception:
                    pass
                continue

            # Acquire a browser session
            session = self.session_manager.acquire_session(timeout=60.0)
            if session is None:
                self.row_update.emit(row, tr("batchtitle.status.failed"), "Could not acquire browser session")
                continue

            try:
                # Update status
                self.row_update.emit(row, tr("batchtitle.status.running"), "")
                self.log.emit(tr("batchtitle.progress.open", current=idx+1, total=len(self.items), code=code))

                # Emit session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)

                # Navigate and update titles via Selenium
                nav = ProductNavigationHandler(session.driver, logger=self.main.logger)
                web = WebInteractionHandler(session.driver)

                self._ensure_products_page(session.driver)
                nav.navigate_to_product(brand, code)

                # Update each provided language title
                if title_lt:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx+1, total=len(self.items), code=code, lang="LT"))
                    self._set_title(session.driver, "lt", title_lt)
                if title_en:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx+1, total=len(self.items), code=code, lang="EN"))
                    self._set_title(session.driver, "en", title_en)
                if title_lv:
                    self.log.emit(tr("batchtitle.progress.fill", current=idx+1, total=len(self.items), code=code, lang="LV"))
                    self._set_title(session.driver, "lv", title_lv)

                self.log.emit(tr("batchtitle.progress.save", current=idx+1, total=len(self.items), code=code))
                web.save_information()
                time.sleep(0.5)

                self.row_update.emit(row, tr("batchtitle.status.ok"), "")
                with self._lock:
                    self._ok_count += 1

                try:
                    details = {
                        "workflow": "batch_titles",
                        "batch_id": self._batch_id,
                        "set_lt": bool(title_lt),
                        "set_en": bool(title_en),
                        "set_lv": bool(title_lv),
                        "mode": "parallel",
                    }
                    with self._lock:
                        self.main.db.conn.execute(
                            """
                            INSERT INTO processing_history
                            (brand, product_code, url_or_code, status, duration_seconds,
                             failed_stage, batch_id, details_json, processed_at)
                            VALUES (?, ?, ?, 'success', ?, ?, ?, ?, ?)
                            """,
                            (
                                brand,
                                code,
                                None,
                                float(_time.perf_counter() - t0),
                                None,
                                self._batch_id,
                                json.dumps(details, ensure_ascii=False),
                            ),
                        )
                        self.main.db.conn.commit()
                except Exception:
                    pass

            except Exception as e:
                error_msg = str(e)
                self.row_update.emit(row, tr("batchtitle.status.failed"), error_msg)
                session.error_count += 1

                try:
                    details = {
                        "workflow": "batch_titles",
                        "batch_id": self._batch_id,
                        "set_lt": bool(title_lt),
                        "set_en": bool(title_en),
                        "set_lv": bool(title_lv),
                        "mode": "parallel",
                    }
                    with self._lock:
                        self.main.db.conn.execute(
                            """
                            INSERT INTO processing_history
                            (brand, product_code, url_or_code, status, error_message, duration_seconds,
                             failed_stage, batch_id, details_json, processed_at)
                            VALUES (?, ?, ?, 'failed', ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                brand,
                                code,
                                None,
                                error_msg,
                                float(_time.perf_counter() - t0),
                                "batch_titles",
                                self._batch_id,
                                json.dumps(details, ensure_ascii=False),
                            ),
                        )
                        self.main.db.conn.commit()
                except Exception:
                    pass

                # Reset session if too many errors
                if session.error_count >= 3:
                    self.log.emit(f"Resetting session {session.session_id} due to errors")
                    self.session_manager.reset_session(session)

            finally:
                # Always release the session
                self.session_manager.release_session(session)

                # Update session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)


class BatchTitlesScreen(ResponsiveWidget):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.worker: BatchTitlesWorker | None = None

        # Multi-session browser pool
        self.session_manager = None

        self.current_mode = "manual"
        self._base_table_rows = 5

        self.brands = [
            "KROSS", "Pinarello", "Basso", "Factor",
            "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan",
        ]

        # Store references for responsive UI
        self.scroll = None
        self.content_widget = None

        self._init_ui()
        self._setup_shortcuts()
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(0, self._ensure_table_fills_viewport)

    def hideEvent(self, event):
        """Cleanup session manager when screen is hidden"""
        super().hideEvent(event)

        # Shutdown session manager if it exists
        if self.session_manager is not None:
            try:
                self.session_manager.shutdown_all()
                self.session_manager = None
            except Exception as e:
                if hasattr(self.main, 'logger'):
                    self.main.logger.log("BatchTitlesScreen", f"Error shutting down session manager: {str(e)}")

    def eventFilter(self, obj, event):
        if obj is self.table.viewport() and event.type() == QEvent.Type.Resize:
            QTimer.singleShot(0, self._ensure_table_fills_viewport)
        return super().eventFilter(obj, event)

    def _apply_theme(self):
        is_dark = isDarkTheme()
        bg_color = COLORS['space_indigo'] if is_dark else COLORS['platinum']
        self.setStyleSheet(f"""
            BatchTitlesScreen {{
                background-color: {bg_color};
                font-family: {FONTS['family']};
            }}
        """)

    def _init_ui(self):
        # Root layout (for ScrollArea container)
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # Create ScrollArea
        self.scroll = ScrollArea(self)
        self.scroll.setWidgetResizable(True)
        root_layout.addWidget(self.scroll)

        # Content widget (inside ScrollArea)
        self.content_widget = QWidget()
        self.scroll.setWidget(self.content_widget)

        # Main layout (on content widget)
        root = QVBoxLayout(self.content_widget)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        # Apply theme
        apply_screen_theme(
            self,
            "BatchTitlesScreen",
            scroll=self.scroll,
            content=self.content_widget
        )

        header = QHBoxLayout()
        title_container = QHBoxLayout()
        title_container.setSpacing(ICON_TEXT_GAP)

        title_icon = TransparentToolButton(FluentIcon.EDIT, self)
        title_icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        title_icon.setEnabled(False)

        self.title_label = TitleLabel("")
        title_container.addWidget(title_icon)
        title_container.addWidget(self.title_label)
        header.addLayout(title_container)
        header.addStretch(1)

        mode_container = QWidget()
        mode_container.setStyleSheet("background: transparent;")
        mode_layout = QHBoxLayout(mode_container)
        mode_layout.setContentsMargins(0, 0, 0, 0)
        mode_layout.setSpacing(ROW_SPACING)

        self.manual_pill = PillPushButton("")
        self.manual_pill.setCheckable(True)
        self.manual_pill.setChecked(True)
        self.manual_pill.clicked.connect(lambda: self._switch_mode("manual"))

        self.excel_pill = PillPushButton("")
        self.excel_pill.setCheckable(True)
        self.excel_pill.clicked.connect(lambda: self._switch_mode("excel"))

        mode_layout.addWidget(self.manual_pill)
        mode_layout.addWidget(self.excel_pill)
        header.addWidget(mode_container)
        root.addLayout(header)

        toolbar_card = CardWidget()
        toolbar_card.setBorderRadius(RADII['md'])
        toolbar_layout = QHBoxLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self.add_btn = PushButton("")
        self.add_btn.setIcon(FluentIcon.ADD.icon())
        self.add_btn.clicked.connect(self._add_row)

        self.clear_btn = PushButton("")
        self.clear_btn.setIcon(FluentIcon.DELETE.icon())
        self.clear_btn.clicked.connect(self._clear_all)

        self.template_btn = PushButton("")
        self.template_btn.setIcon(FluentIcon.DOWNLOAD.icon())
        self.template_btn.clicked.connect(self._download_template)

        self.browse_btn = PushButton("")
        self.browse_btn.setIcon(FluentIcon.FOLDER_ADD.icon())
        self.browse_btn.clicked.connect(self._browse_excel)

        self.excel_file_label = CaptionLabel("")
        self.excel_file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        # Expandable spacer that pushes status/start to the right.
        # This keeps the Excel controls on the left in Excel mode.
        self._toolbar_spacer = QWidget()
        self._toolbar_spacer.setStyleSheet("background: transparent;")
        self._toolbar_spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        self.status_label = BodyLabel("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self.progress_ring = IndeterminateProgressRing()
        self.progress_ring.setFixedSize(SIZES['progress_ring'], SIZES['progress_ring'])
        self.progress_ring.setVisible(False)

        self.start_btn = PrimaryPushButton("")
        self.start_btn.setIcon(FluentIcon.PLAY.icon())
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self._start_batch)

        toolbar_layout.addWidget(self.add_btn)
        toolbar_layout.addWidget(self.clear_btn)
        toolbar_layout.addWidget(self.template_btn)
        toolbar_layout.addWidget(self.browse_btn)
        toolbar_layout.addWidget(self.excel_file_label)
        # Spacer between excel file label and status (hide/show with excel controls)
        self._excel_status_sep = QWidget()
        self._excel_status_sep.setStyleSheet("background: transparent;")
        self._excel_status_sep.setFixedWidth(CONTENT_SPACING)
        toolbar_layout.addWidget(self._excel_status_sep)
        toolbar_layout.addWidget(self._toolbar_spacer)
        toolbar_layout.addWidget(self.status_label)
        toolbar_layout.addWidget(self.progress_ring)
        toolbar_layout.addWidget(self.start_btn)

        root.addWidget(toolbar_card)

        table_card = CardWidget()
        table_card.setBorderRadius(RADII['md'])
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.excel_drop = DropZoneWidget(self.main.i18n.tr)
        self.excel_drop.file_dropped.connect(self._handle_file_drop)

        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        # Smooth scrolling (avoid row/column snapping).
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

        # Enable horizontal scrolling for responsive design
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        try:
            self.table.verticalScrollBar().setSingleStep(12)
            self.table.horizontalScrollBar().setSingleStep(12)
        except Exception:
            pass
        # Fixed columns (horizontal scrolling as needed).
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.viewport().installEventFilter(self)

        # Wider brand column so placeholder doesn't get truncated
        self.table.setColumnWidth(0, SIZES['col_w_240'])  # Brand
        self.table.setColumnWidth(1, SIZES['col_w_180'])  # Code
        self.table.setColumnWidth(2, SIZES['col_w_320'])  # Title LT
        self.table.setColumnWidth(3, SIZES['col_w_320'])  # Title EN
        self.table.setColumnWidth(4, SIZES['col_w_320'])  # Title LV
        self.table.setColumnWidth(5, SIZES['col_w_160'])  # Status
        self.table.setColumnWidth(6, SIZES['col_w_320'])  # Error
        self.table.setColumnWidth(7, SIZES['col_w_56'])   # Delete

        self.table.verticalHeader().setDefaultSectionSize(SIZES['table_row_height_lg'])

        table_layout.addWidget(self.excel_drop)
        table_layout.addWidget(self.table)
        root.addWidget(table_card, 1)

        self.table.setRowCount(self._base_table_rows)
        for r in range(self.table.rowCount()):
            self._setup_table_row(r)

        self._update_table_theme()
        self._switch_mode("manual")
        self.retranslate_ui()

    def retranslate_ui(self):
        tr = self.main.i18n.tr
        self.title_label.setText(tr("batchtitle.title"))
        self.manual_pill.setText(tr("batch.mode.manual"))
        self.excel_pill.setText(tr("batch.mode.excel"))
        self.add_btn.setText(tr("batch.add_row"))
        self.add_btn.setToolTip(tr("batch.add_row.tip"))
        self.clear_btn.setText(tr("batch.clear_all"))
        self.clear_btn.setToolTip(tr("batch.clear_all.tip"))
        self.template_btn.setText(tr("batch.download_template"))
        self.template_btn.setToolTip(tr("batch.download_template.tip"))
        self.browse_btn.setText(tr("batch.browse_excel"))
        self.browse_btn.setToolTip(tr("batch.browse_excel.tip"))
        self.excel_file_label.setText(tr("batch.no_file"))
        # Keep the Start button area clean; don't show the "Ready" hint.
        self.status_label.setText("")
        self.status_label.setVisible(False)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.start_btn.setText(f"{tr('batchtitle.start')} (Ctrl+Enter)")
        self.start_btn.setToolTip(tr("batchtitle.start.tip"))

        self.table.setHorizontalHeaderLabels([
            tr("batch.table.brand"),
            tr("batch.table.code"),
            tr("batchtitle.table.title_lt"),
            tr("batchtitle.table.title_en"),
            tr("batchtitle.table.title_lv"),
            tr("batchtitle.table.status"),
            tr("batchtitle.table.error"),
            "",
        ])

        self.excel_drop.retranslate_ui()

        # Update placeholder labels for existing row widgets
        for row in range(self.table.rowCount()):
            brand_cell = self.table.cellWidget(row, 0)
            if not brand_cell:
                continue
            brand_combo = brand_cell.findChild(ComboBox)
            if brand_combo:
                self._ensure_combo_placeholder_item(brand_combo, tr("batch.select_brand"), self.brands)

    def _on_breakpoint_changed(self, breakpoint: str):
        """Respond to breakpoint changes - adjust margins and spacing."""
        margins = get_responsive_margins(breakpoint)
        spacing = get_responsive_spacing(breakpoint)
        if hasattr(self, 'content_widget') and self.content_widget and self.content_widget.layout():
            self.content_widget.layout().setContentsMargins(*margins)
            self.content_widget.layout().setSpacing(spacing)

    def _on_theme_changed(self):
        apply_screen_theme(
            self,
            "BatchTitlesScreen",
            scroll=self.scroll,
            content=self.content_widget
        )
        self._update_table_theme()

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        table_colors = COMPONENT_COLORS['table']
        bg_color = table_colors['row_alt_bg_dark'] if is_dark else table_colors['row_alt_bg_light']
        alt_bg = table_colors['row_bg_dark'] if is_dark else table_colors['row_bg_light']
        border_color = table_colors['border_dark'] if is_dark else table_colors['border_light']

        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']
        text_color = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']
        from GUI_Qt.styles.theme_config import (
            get_embedded_input_border,
            get_scrollbar_handle_bg,
            get_scrollbar_handle_hover_bg,
            get_selection_bg,
        )
        embedded_input_border = get_embedded_input_border(is_dark)
        embedded_input_bg = COMPONENT_COLORS['input']['bg_dark'] if is_dark else COMPONENT_COLORS['input']['bg_light']

        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg_color};
                alternate-background-color: {alt_bg};
                border: 1px solid {border_color};
                border-radius: {RADII['md']}px;
                gridline-color: {border_color};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::viewport {{
                background-color: {bg_color};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableView::item:focus, QAbstractItemView::item:focus {{
                outline: none;
            }}
            QTableWidget QWidget:focus {{
                outline: none;
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
                letter-spacing: 0.5px;
                text-transform: uppercase;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}
            QHeaderView::section {{
                border-right: 1px solid {border_color};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:last {{
                border-top-right-radius: {RADII['md']}px;
            }}

            /* Keep scrollbars visually inside rounded corners */
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}

            QScrollBar:horizontal {{
                background: transparent;
                height: {SIZES['scrollbar_thickness']}px;
                margin: 0px {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:horizontal {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-width: {SIZES['scrollbar_handle_min']}px;
                margin: {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                border: none;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
            QTableWidget QWidget[ubTableCell="true"] LineEdit,
            QTableWidget QWidget[ubTableCell="true"] QLineEdit,
            QTableWidget QWidget[ubTableCell="true"] ComboBox,
            QTableWidget QWidget[ubTableCell="true"] QComboBox {{
                background-color: {embedded_input_bg};
                border: 1px solid {embedded_input_border};
                border-bottom: 2px solid {embedded_input_border};
                border-radius: {RADII['sm']}px;
                padding: {PADDINGS['combo']};
                outline: none;
            }}
        """)

    def _switch_mode(self, mode: str):
        self.current_mode = mode
        self.manual_pill.setChecked(mode == "manual")
        self.excel_pill.setChecked(mode == "excel")

        is_manual = mode == "manual"
        self.add_btn.setVisible(is_manual)
        self.clear_btn.setVisible(is_manual)

        is_excel = mode == "excel"
        self.template_btn.setVisible(is_excel)
        self.browse_btn.setVisible(is_excel)
        self.excel_file_label.setVisible(is_excel)
        if hasattr(self, "_excel_status_sep"):
            self._excel_status_sep.setVisible(is_excel)

        no_file = self.excel_file_label.text() == self.main.i18n.tr("batch.no_file")
        self.excel_drop.setVisible(is_excel and no_file)
        self.table.setVisible((not is_excel) or (is_excel and not no_file))
        self._validate()

    def _handle_file_drop(self, filepath: str):
        if filepath == "__browse__":
            self._browse_excel()
        else:
            self._load_excel(filepath)

    def _browse_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self,
            self.main.i18n.tr("batch.file.select_excel.title"),
            "",
            self.main.i18n.tr("batch.file.select_excel.filter"),
        )
        if filename:
            self._load_excel(filename)

    @staticmethod
    def _normalize_code(raw: str) -> str:
        code = (raw or "").strip()
        if not code:
            return ""
        if code.upper().startswith("UB-"):
            return code
        return f"UB-{code}"

    def _load_excel(self, filename: str):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                values = [str(c.value).lower() if c.value else "" for c in row]
                if any("brand" in v for v in values) and any("code" in v or "product" in v for v in values):
                    header_row = idx
                    break
            if not header_row:
                raise Exception(self.main.i18n.tr("batch.excel.header_missing_in_file"))

            data_rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if any(row):
                    data_rows.append(row)

            self.table.setRowCount(len(data_rows) if data_rows else self._base_table_rows)
            for r in range(self.table.rowCount()):
                self._setup_table_row(r)

            valid_count = 0
            invalid_count = 0

            for table_row, excel_row in enumerate(data_rows):
                brand = str(excel_row[0]).strip() if len(excel_row) > 0 and excel_row[0] else ""
                code_raw = str(excel_row[1]).strip() if len(excel_row) > 1 and excel_row[1] else ""
                code = self._normalize_code(code_raw)
                title_lt = str(excel_row[2]).strip() if len(excel_row) > 2 and excel_row[2] else ""
                title_en = str(excel_row[3]).strip() if len(excel_row) > 3 and excel_row[3] else ""
                title_lv = str(excel_row[4]).strip() if len(excel_row) > 4 and excel_row[4] else ""

                brand_widget = self._get_widget_from_cell(table_row, 0, ComboBox)
                if brand_widget and brand:
                    brand_widget.setCurrentText(brand)

                code_widget = self._get_widget_from_cell(table_row, 1, LineEdit)
                if code_widget:
                    code_widget.setText(code_raw)

                lt_widget = self._get_widget_from_cell(table_row, 2, LineEdit)
                if lt_widget:
                    lt_widget.setText(title_lt)

                en_widget = self._get_widget_from_cell(table_row, 3, LineEdit)
                if en_widget:
                    en_widget.setText(title_en)

                lv_widget = self._get_widget_from_cell(table_row, 4, LineEdit)
                if lv_widget:
                    lv_widget.setText(title_lv)

                if brand and code and (title_lt or title_en or title_lv):
                    valid_count += 1
                else:
                    invalid_count += 1

            wb.close()

            self.excel_file_label.setText(os.path.basename(filename))
            self.excel_drop.setVisible(False)
            self.table.setVisible(True)

            if invalid_count > 0:
                self.status_label.setText(
                    self.main.i18n.tr(
                        "batch.excel.rows_fix_errors",
                        valid=valid_count,
                        invalid=invalid_count,
                    )
                )
                self.status_label.setStyleSheet(f"color: {COLORS['warning']}; font-weight: 500;")
                self.status_label.setVisible(True)
                self.start_btn.setEnabled(False)
            else:
                # Don't show a "Ready" hint; keep Start enabled silently.
                self.status_label.setText("")
                self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
                self.status_label.setVisible(False)
                self.start_btn.setEnabled(valid_count > 0)
        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.excel.load_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self,
            )

    def _download_template(self):
        filename, _ = QFileDialog.getSaveFileName(
            self,
            self.main.i18n.tr("batch.template.save.title"),
            "batch_titles_template.xlsx",
            self.main.i18n.tr("batch.template.filter"),
        )
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.main.i18n.tr("batchtitle.title")
            ws.append(["Brand", "Code", "Title (LT)", "Title (EN)", "Title (LV)"])
            ws.append(["KROSS", "1234", "KROSS Esker 6.0 2025 / ...", "KROSS Esker 6.0 2025 / ...", "KROSS Esker 6.0 2025 / ..."])

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="8D99AE", end_color="8D99AE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 64
            ws.column_dimensions['D'].width = 64
            ws.column_dimensions['E'].width = 64

            wb.save(filename)
            wb.close()

            InfoBar.success(
                title=self.main.i18n.tr("common.success"),
                content=self.main.i18n.tr("batch.template.downloaded"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self,
            )
        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.template.save_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self,
            )

    def _add_row(self):
        r = self.table.rowCount()
        self.table.setRowCount(r + 1)
        self._setup_table_row(r)
        self._validate()

    def _clear_all(self):
        """Clear all rows with confirmation dialog"""
        # Show confirmation dialog
        w = MessageBox(
            title=self.main.i18n.tr("common.confirm"),
            content=self.main.i18n.tr("batch.clear.confirm"),
            parent=self
        )
        if w.exec():
            self.table.setRowCount(self._base_table_rows)
            for r in range(self.table.rowCount()):
                self._setup_table_row(r)
            self.excel_file_label.setText(self.main.i18n.tr("batch.no_file"))
            if self.current_mode == "excel":
                self.excel_drop.setVisible(True)
                self.table.setVisible(False)
            self._validate()

    def _delete_row(self, row: int):
        if self.table.rowCount() <= 1:
            return
        self.table.removeRow(row)
        self._validate()

    def _get_widget_from_cell(self, row: int, col: int, widget_type):
        cell = self.table.cellWidget(row, col)
        return cell.findChild(widget_type) if cell else None

    def _setup_shortcuts(self):
        """Setup keyboard shortcuts for batch titles screen"""
        # Ctrl+Enter to start upload
        start_shortcut = QShortcut(QKeySequence("Ctrl+Return"), self)
        start_shortcut.activated.connect(self._handle_start_shortcut)

        # Escape to cancel/stop upload
        cancel_shortcut = QShortcut(QKeySequence("Escape"), self)
        cancel_shortcut.activated.connect(self._handle_cancel_shortcut)

        # Ctrl+R to retry failed items
        retry_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        retry_shortcut.activated.connect(self._handle_retry_shortcut)

    def _handle_start_shortcut(self):
        """Handle Ctrl+Enter shortcut to start upload"""
        if self.start_btn.isEnabled():
            self._start_upload()

    def _handle_cancel_shortcut(self):
        """Handle Escape shortcut to cancel/stop upload"""
        if self.worker is not None and self.worker.isRunning():
            self.worker.stop()
            InfoBar.warning(
                title=self.main.i18n.tr("batch.cancel.title"),
                content=self.main.i18n.tr("batch.cancel.content"),
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _handle_retry_shortcut(self):
        """Handle Ctrl+R shortcut to retry failed items"""
        if self.retry_btn.isVisible() and self.retry_btn.isEnabled():
            self._retry_failed()

    def _setup_table_row(self, row: int):
        def wrap(widget: QWidget):
            c = QWidget()
            c.setStyleSheet("background: transparent;")
            c.setProperty("ubTableCell", True)
            c.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

            # Allow embedded widgets to shrink to the available cell width.
            try:
                widget.setMinimumWidth(0)
            except Exception:
                pass
            try:
                widget.setSizePolicy(QSizePolicy.Policy.Expanding, widget.sizePolicy().verticalPolicy())
            except Exception:
                pass

            l = QHBoxLayout(c)
            l.setContentsMargins(*TABLE_CELL_MARGINS)
            l.setSpacing(0)
            l.addWidget(widget, 1, Qt.AlignmentFlag.AlignVCenter)
            return c

        brand = ComboBox()
        self._ensure_combo_placeholder_item(
            brand,
            self.main.i18n.tr("batch.select_brand"),
            self.brands,
            force_placeholder=True,
        )
        brand.setMinimumHeight(SIZES['input_height'])
        brand.currentTextChanged.connect(self._validate)
        self.table.setCellWidget(row, 0, wrap(brand))

        code = LineEdit()
        code.setPlaceholderText(self.main.i18n.tr("batchtitle.code.placeholder"))
        code.setMinimumHeight(SIZES['input_height'])
        code.textChanged.connect(lambda _t: self._validate())
        self.table.setCellWidget(row, 1, wrap(code))

        title_lt = LineEdit()
        title_lt.setPlaceholderText(self.main.i18n.tr("batchtitle.title_lt.placeholder"))
        title_lt.setMinimumHeight(SIZES['input_height'])
        title_lt.textChanged.connect(lambda _t: self._validate())
        self.table.setCellWidget(row, 2, wrap(title_lt))

        title_en = LineEdit()
        title_en.setPlaceholderText(self.main.i18n.tr("batchtitle.title_en.placeholder"))
        title_en.setMinimumHeight(SIZES['input_height'])
        title_en.textChanged.connect(lambda _t: self._validate())
        self.table.setCellWidget(row, 3, wrap(title_en))

        title_lv = LineEdit()
        title_lv.setPlaceholderText(self.main.i18n.tr("batchtitle.title_lv.placeholder"))
        title_lv.setMinimumHeight(SIZES['input_height'])
        title_lv.textChanged.connect(lambda _t: self._validate())
        self.table.setCellWidget(row, 4, wrap(title_lv))

        status = BodyLabel(self.main.i18n.tr("batchtitle.status.pending"))
        status.setStyleSheet(f"color: {COLORS['text_secondary']};")
        status.setProperty("ub_status", "pending")
        self.table.setCellWidget(row, 5, wrap(status))

        error = LineEdit()
        error.setReadOnly(True)
        error.setMinimumHeight(SIZES['input_height'])
        self.table.setCellWidget(row, 6, wrap(error))

        delete_btn = TransparentToolButton(FluentIcon.DELETE)
        delete_btn.clicked.connect(lambda _=False, r=row: self._delete_row(r))
        delete_btn.setFixedSize(SIZES['button_height_sm'], SIZES['button_height_sm'])
        dc = QWidget()
        dc.setStyleSheet("background: transparent;")
        dc.setProperty("ubTableCell", True)
        dcl = QHBoxLayout(dc)
        dcl.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        dcl.addStretch(1)
        dcl.addWidget(delete_btn, 0, Qt.AlignmentFlag.AlignVCenter)
        dcl.addStretch(1)
        self.table.setCellWidget(row, 7, dc)

        self.table.setRowHeight(row, SIZES['table_row_height_lg'])

    def _collect_items(self) -> list[dict]:
        items = []
        for row in range(self.table.rowCount()):
            b = self._get_widget_from_cell(row, 0, ComboBox)
            c = self._get_widget_from_cell(row, 1, LineEdit)
            t_lt = self._get_widget_from_cell(row, 2, LineEdit)
            t_en = self._get_widget_from_cell(row, 3, LineEdit)
            t_lv = self._get_widget_from_cell(row, 4, LineEdit)
            s = self._get_widget_from_cell(row, 5, BodyLabel)

            status_code = None
            if s is not None:
                try:
                    status_code = s.property("ub_status")
                except Exception:
                    status_code = None
            if status_code in ("running", "ok"):
                continue

            brand = b.currentText().strip() if b else ""
            code_raw = c.text().strip() if c else ""
            code = self._normalize_code(code_raw)
            title_lt = t_lt.text().strip() if t_lt else ""
            title_en = t_en.text().strip() if t_en else ""
            title_lv = t_lv.text().strip() if t_lv else ""

            if brand in self.brands and code and (title_lt or title_en or title_lv):
                items.append({
                    "row": row,
                    "brand": brand,
                    "code": code,
                    "title_lt": title_lt,
                    "title_en": title_en,
                    "title_lv": title_lv,
                })
        return items

    def _ensure_combo_placeholder_item(
        self,
        combo: ComboBox,
        placeholder_text: str,
        valid_values: list[str],
        force_placeholder: bool = False,
    ) -> None:
        combo.blockSignals(True)
        try:
            current = combo.currentText().strip()
            keep_current = (not force_placeholder) and (current in valid_values)

            combo.clear()
            combo.addItem(placeholder_text)
            combo.addItems(valid_values)

            try:
                model = combo.model()
                if isinstance(model, QStandardItemModel):
                    item = model.item(0)
                    if item is not None:
                        item.setEnabled(False)
            except Exception:
                pass

            if keep_current:
                combo.setCurrentText(current)
            else:
                combo.setCurrentIndex(0)
        finally:
            combo.blockSignals(False)

    def _set_row_locked(self, row: int, locked: bool) -> None:
        for col, widget_type in (
            (0, ComboBox),
            (1, LineEdit),
            (2, LineEdit),
            (3, LineEdit),
            (4, LineEdit),
        ):
            w = self._get_widget_from_cell(row, col, widget_type)
            if w is not None:
                w.setEnabled(not locked)

        delete_cell = self.table.cellWidget(row, 7)
        if delete_cell is not None:
            btn = delete_cell.findChild(TransparentToolButton)
            if btn is not None:
                btn.setEnabled(not locked)

    def _validate(self):
        valid = len(self._collect_items())
        self.start_btn.setEnabled(valid > 0 and self.worker is None)

        if self.worker is not None:
            return

        # Don't show a "Ready" hint; keep Start enabled/disabled silently.
        self.status_label.setText("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(False)

    def _set_busy(self, busy: bool):
        self.progress_ring.setVisible(busy)
        self.add_btn.setEnabled(not busy)
        self.clear_btn.setEnabled(not busy)
        self.template_btn.setEnabled(not busy)
        self.browse_btn.setEnabled(not busy)
        self.manual_pill.setEnabled(not busy)
        self.excel_pill.setEnabled(not busy)
        self.table.setEnabled(not busy)
        self.start_btn.setEnabled(not busy and len(self._collect_items()) > 0)

    def _start_batch(self):
        tr = self.main.i18n.tr
        if getattr(self.main, "driver", None) is None:
            InfoBar.error(
                title=tr("common.error"),
                content=tr("batchtitle.no_session"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=5000,
                parent=self,
            )
            return

        items = self._collect_items()
        if not items:
            return

        for row in range(self.table.rowCount()):
            s = self._get_widget_from_cell(row, 5, BodyLabel)
            if s:
                s.setText(tr("batchtitle.status.pending"))
                s.setStyleSheet(f"color: {COLORS['text_secondary']};")
                s.setProperty("ub_status", "pending")
            e = self._get_widget_from_cell(row, 6, LineEdit)
            if e:
                e.setText("")

        # Check if multi-session is enabled
        multi_session_enabled = self.main.settings.get('multi_session_enabled', False)
        browser_count = self.main.settings.get('browser_count', 2)

        # Initialize session manager if multi-session enabled
        if multi_session_enabled:
            if self.session_manager is None or not self.session_manager.is_ready():
                from Managers.BrowserSessionManager import BrowserSessionManager

                # Get browser type from settings
                browser_type = self.main.settings.get('browser_choice', 'Chrome')

                InfoBar.info(
                    title=tr("common.info"),
                    content=f"Initializing {browser_count} browser instances...",
                    orient=Qt.Orientation.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                    parent=self
                )

                # Create and initialize session pool
                self.session_manager = BrowserSessionManager(
                    browser_type=browser_type,
                    pool_size=browser_count,
                    logger=self.main.logger
                )

                if not self.session_manager.initialize_pool():
                    InfoBar.error(
                        title=tr("common.error"),
                        content="Failed to initialize browser pool. Falling back to single-session mode.",
                        orient=Qt.Orientation.Horizontal,
                        isClosable=True,
                        position=InfoBarPosition.TOP,
                        duration=5000,
                        parent=self
                    )
                    self.session_manager = None
                    multi_session_enabled = False

        InfoBar.info(
            title=tr("batchtitle.run.started.title"),
            content=tr("batchtitle.run.started.content", count=len(items)),
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self,
        )

        self._set_busy(True)

        # Choose worker based on multi-session setting
        if multi_session_enabled and self.session_manager and self.session_manager.is_ready():
            # Use parallel worker with session pool
            self.worker = ParallelBatchTitlesWorker(self.main, self.session_manager, items)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self._on_log)
            self.worker.done.connect(self._on_done)
            self.worker.session_status.connect(self._on_session_status_update)
            self.worker.start()
        else:
            # Use single-session worker
            self.worker = BatchTitlesWorker(self.main, items)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self._on_log)
            self.worker.progress_update.connect(self._on_progress_update)
            self.worker.done.connect(self._on_done)
            self.worker.start()

    def _on_progress_update(self, current: int, total: int, speed: float, eta_seconds: float):
        """Handle progress updates with ETA from worker thread."""
        percentage = (current / total * 100) if total > 0 else 0

        # Format ETA
        if eta_seconds < 60:
            eta_text = f"{int(eta_seconds)}s"
        elif eta_seconds < 3600:
            minutes = int(eta_seconds / 60)
            seconds = int(eta_seconds % 60)
            eta_text = f"{minutes}m {seconds}s"
        else:
            hours = int(eta_seconds / 3600)
            minutes = int((eta_seconds % 3600) / 60)
            eta_text = f"{hours}h {minutes}m"

        # Format speed (items per minute)
        items_per_min = speed * 60

        # Update status label with progress information
        message = f"Processing ({current}/{total}) {percentage:.0f}% - ETA: {eta_text} | {items_per_min:.1f} items/min"
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(True)

    def _on_log(self, message: str):
        self.status_label.setText(message)
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_label.setVisible(bool((message or "").strip()))

    def _on_session_status_update(self, stats: dict):
        """Handle session status updates from parallel worker (optional)"""
        # BatchTitlesScreen doesn't have a session status label yet
        # This is just for future extensibility
        pass

    def _on_row_update(self, row: int, status_text: str, error_text: str):
        status = self._get_widget_from_cell(row, 5, BodyLabel)
        if status:
            status.setText(status_text)
            if status_text == self.main.i18n.tr("batchtitle.status.ok"):
                status.setStyleSheet(f"color: {COLORS['success']}; font-weight: 500;")
                status.setProperty("ub_status", "ok")
                self._set_row_locked(row, True)
            elif status_text == self.main.i18n.tr("batchtitle.status.failed"):
                status.setStyleSheet(f"color: {COLORS['warning']}; font-weight: 500;")
                status.setProperty("ub_status", "failed")
            else:
                status.setStyleSheet(f"color: {COLORS['text_secondary']};")
                if status_text == self.main.i18n.tr("batchtitle.status.running"):
                    status.setProperty("ub_status", "running")
                else:
                    status.setProperty("ub_status", "pending")

        err = self._get_widget_from_cell(row, 6, LineEdit)
        if err:
            err.setText(error_text or "")

    def _on_done(self, ok: int, total: int):
        tr = self.main.i18n.tr
        self.worker = None
        self._set_busy(False)
        self._validate()

        InfoBar.success(
            title=tr("batchtitle.run.complete.title"),
            content=tr("batchtitle.run.complete.content", ok=ok, total=total),
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=6000,
            parent=self,
        )

        # Windows notification
        self._show_windows_notification(ok, total)

    def _show_windows_notification(self, ok: int, total: int):
        """Show Windows desktop notification for batch completion."""
        # NOTE: win10toast relies on pywin32 WNDPROC callbacks that are known to
        # break on newer Python versions (e.g. 3.12+), causing hard crashes like:
        # "WNDPROC return value cannot be converted to LRESULT".
        try:
            import sys
            if sys.version_info >= (3, 12):
                return
        except Exception:
            return

        try:
            from win10toast import ToastNotifier
            toaster = ToastNotifier()

            failed = total - ok
            if failed > 0:
                title = "Batch Titles Update Completed with Errors"
                message = f"Success: {ok}/{total}\nFailed: {failed}"
            else:
                title = "Batch Titles Update Completed Successfully"
                message = f"All {total} titles updated successfully"

            # Show notification (non-blocking, 10 second duration)
            toaster.show_toast(
                title=title,
                msg=message,
                duration=10,
                threaded=True,
                icon_path=None  # Uses default Windows icon
            )
        except ImportError:
            # win10toast not installed, silently skip
            pass
        except Exception:
            # Any other error, silently skip
            pass

    def _ensure_table_fills_viewport(self):
        if not self.table.isVisible() or self.excel_drop.isVisible():
            return
        viewport_height = self.table.viewport().height()
        if viewport_height <= 0:
            return
        row_height = self.table.verticalHeader().defaultSectionSize() or 56
        target = (viewport_height + row_height - 1) // row_height + 1
        target = max(self._base_table_rows, int(target))
        target = min(target, 50)
        current = self.table.rowCount()
        if current >= target:
            return
        self.table.setRowCount(target)
        for r in range(current, target):
            self._setup_table_row(r)
