"""
Batch Upload Screen - Complete Redesign
Fluent Design System with Space Indigo/Lavender color scheme
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFileDialog,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QFrame, QSizePolicy
)
from PySide6.QtCore import Qt, QPropertyAnimation, QEasingCurve, Signal, QTimer, QEvent, QThread
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent, QStandardItemModel, QKeySequence, QShortcut
from qfluentwidgets import (
    LineEdit, ComboBox, CheckBox, PrimaryPushButton, PushButton,
    BodyLabel, TitleLabel, StrongBodyLabel, CaptionLabel,
    TransparentToolButton, FluentIcon, InfoBar, InfoBarPosition,
    ScrollArea, CardWidget, PillPushButton, isDarkTheme, TransparentPushButton, IconWidget, qconfig,
    IndeterminateProgressRing, MessageBox
)
import openpyxl
import time
from openpyxl.styles import Font, PatternFill, Alignment
from Managers.DescriptionManager import DescriptionManager
from Utilities.ExcelHandler import ExcelHandler
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import COLORS, FONTS, COMPONENT_COLORS, RADII, PADDINGS, SIZES, SPACING
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS, PAGE_SPACING, CARD_MARGINS, CARD_SPACING, ICON_TEXT_GAP,
    ROW_SPACING, TOOLBAR_MARGINS, CONTENT_SPACING, TABLE_CELL_MARGINS,
    apply_screen_theme, get_responsive_margins, get_responsive_spacing
)


# Removed old ProductRow CardWidget class - replaced with modern table


class BatchUploadWorker(QThread):
    """Background worker for batch upload processing."""
    row_update = Signal(int, str, str)  # row, status_text, error_text
    log = Signal(str, str)  # source, message
    done = Signal(int, int)  # ok_count, total_count
    progress_update = Signal(int, int, float, float)  # current, total, speed (items/sec), eta_seconds

    def __init__(self, rows_data, main_window, master_password=None, table_row_indices=None):
        super().__init__()
        self.rows_data = rows_data  # List of dicts with row data
        self.main_window = main_window
        self.master_password = master_password
        self.table_row_indices = table_row_indices  # Optional: actual table row numbers for each item
        self._stop = False
        self._processed_times = []  # Track duration of each processed item

    def stop(self):
        self._stop = True

    def run(self):
        """Execute batch upload processing in background thread."""
        from Utilities.BatchProcessor import BatchProcessor

        driver = getattr(self.main_window, "driver", None)
        if driver is None:
            for idx in range(len(self.rows_data)):
                table_row = self.table_row_indices[idx] if self.table_row_indices else idx
                self.row_update.emit(table_row, "✗ Failed", "No browser session")
            self.done.emit(0, len(self.rows_data))
            return

        # Create batch processor
        batch_processor = BatchProcessor(
            driver=driver,
            db_manager=self.main_window.db,
            logger=self.main_window.logger,
        )

        ok = 0
        total = len(self.rows_data)

        # Process each row
        for idx, row_data in enumerate(self.rows_data):
            if self._stop:
                break

            # Get actual table row index (for proper UI updates)
            table_row = self.table_row_indices[idx] if self.table_row_indices else idx

            # Start timing this item
            item_start_time = time.perf_counter()

            try:
                self.row_update.emit(table_row, "Processing...", "")
                self.log.emit("BatchUpload", f"Processing {row_data['code']} ({idx + 1}/{total})")

                # Build items list for this single item
                items = [{
                    'brand': row_data['brand'],
                    'code': row_data['code'],
                    'url': row_data['url'],
                    'description_name': row_data.get('description_name'),
                    'frameset_only': row_data.get('frameset_only', False),
                    'append_disclaimer': row_data.get('append_disclaimer', False),
                }]

                # Process this single item
                result = batch_processor.process_batch(items, master_password=self.master_password)

                # Check if successful
                if result and result.get('success', 0) > 0:
                    self.row_update.emit(table_row, "✓ Success", "")
                    ok += 1
                else:
                    # Get error from results
                    failed_items = result.get('results', [])
                    error_msg = failed_items[0].get('error', 'Unknown error') if failed_items else 'Unknown error'
                    self.row_update.emit(table_row, "✗ Failed", error_msg)

            except Exception as e:
                error_msg = str(e)
                self.row_update.emit(table_row, "✗ Failed", error_msg)
                self.log.emit("BatchUpload", f"Error on row {idx}: {error_msg}")

            # Track timing and emit progress update
            item_duration = time.perf_counter() - item_start_time
            self._processed_times.append(item_duration)

            # Calculate ETA (skip first 2 items for more accurate average)
            if len(self._processed_times) >= 2:
                recent_times = self._processed_times[1:]  # Skip first item (often slower)
                avg_time_per_item = sum(recent_times) / len(recent_times)
                remaining_items = total - (idx + 1)
                eta_seconds = avg_time_per_item * remaining_items
                speed = 1.0 / avg_time_per_item if avg_time_per_item > 0 else 0

                self.progress_update.emit(idx + 1, total, speed, eta_seconds)

        self.done.emit(ok, total)


class ParallelBatchUploadWorker(QThread):
    """Parallel batch upload using browser session pool"""
    row_update = Signal(int, str, str)  # row, status_text, error_text
    log = Signal(str, str)  # source, message
    done = Signal(int, int)  # ok_count, total_count
    session_status = Signal(dict)  # session pool statistics

    def __init__(self, rows_data, session_manager, main_window, master_password=None, table_row_indices=None):
        super().__init__()
        self.rows_data = rows_data
        self.session_manager = session_manager
        self.main_window = main_window
        self.master_password = master_password
        self.table_row_indices = table_row_indices or list(range(len(rows_data)))
        self._stop = False
        self._lock = __import__('threading').Lock()
        self._ok_count = 0
        self._work_index = 0

    def stop(self):
        self._stop = True

    def run(self):
        """Execute parallel batch upload using session pool"""
        import threading

        if not self.session_manager or not self.session_manager.is_ready():
            # Fallback: no session manager, fail all
            for table_row in self.table_row_indices:
                self.row_update.emit(table_row, "✗ Failed", "Session pool not available")
            self.done.emit(0, len(self.rows_data))
            return

        total = len(self.rows_data)
        pool_size = self.session_manager.pool_size

        self.log.emit("ParallelBatchUpload", f"Starting parallel processing with {pool_size} browsers")

        # Create worker threads
        threads = []
        for worker_id in range(pool_size):
            thread = threading.Thread(target=self._worker_thread, args=(worker_id,))
            thread.daemon = True
            threads.append(thread)
            thread.start()

        # Wait for all threads to complete
        for thread in threads:
            thread.join()

        self.log.emit("ParallelBatchUpload", f"Completed. Success: {self._ok_count}/{total}")
        self.done.emit(self._ok_count, total)

    def _worker_thread(self, worker_id: int):
        """Worker thread that processes items from the queue"""
        from Utilities.BatchProcessor import BatchProcessor

        while not self._stop:
            # Get next work item
            with self._lock:
                if self._work_index >= len(self.rows_data):
                    break  # No more work
                idx = self._work_index
                self._work_index += 1

            row_data = self.rows_data[idx]
            table_row = self.table_row_indices[idx]

            # Acquire a browser session
            session = self.session_manager.acquire_session(timeout=60.0)
            if session is None:
                self.row_update.emit(table_row, "✗ Failed", "Could not acquire browser session")
                continue

            try:
                # Update status
                self.row_update.emit(table_row, "Processing...", "")
                self.log.emit("ParallelWorker", f"Worker {worker_id} processing {row_data['code']} (session {session.session_id})")

                # Emit session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)

                # Create batch processor for this session
                batch_processor = BatchProcessor(
                    driver=session.driver,
                    db_manager=self.main_window.db,
                    logger=self.main_window.logger,
                )

                # Build item
                items = [{
                    'brand': row_data['brand'],
                    'code': row_data['code'],
                    'url': row_data['url'],
                    'description_name': row_data.get('description_name'),
                    'frameset_only': row_data.get('frameset_only', False),
                    'append_disclaimer': row_data.get('append_disclaimer', False),
                }]

                # Process
                result = batch_processor.process_batch(items, master_password=self.master_password)

                # Check result
                if result and result.get('success', 0) > 0:
                    self.row_update.emit(table_row, "✓ Success", "")
                    with self._lock:
                        self._ok_count += 1
                else:
                    failed_items = result.get('results', [])
                    error_msg = failed_items[0].get('error', 'Unknown error') if failed_items else 'Unknown error'
                    self.row_update.emit(table_row, "✗ Failed", error_msg)
                    session.error_count += 1

            except Exception as e:
                error_msg = str(e)
                self.row_update.emit(table_row, "✗ Failed", error_msg)
                self.log.emit("ParallelWorker", f"Worker {worker_id} error: {error_msg}")
                session.error_count += 1

                # Reset session if too many errors
                if session.error_count >= 3:
                    self.log.emit("ParallelWorker", f"Resetting session {session.session_id} due to errors")
                    self.session_manager.reset_session(session)

            finally:
                # Always release the session
                self.session_manager.release_session(session)

                # Update session status
                stats = self.session_manager.get_session_stats()
                self.session_status.emit(stats)


class DropZoneWidget(QWidget):
    """Drag and drop zone for Excel files"""
    file_dropped = Signal(str)

    def __init__(self, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self.title_label = None
        self.subtitle_label = None
        self.setAcceptDrops(True)
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(PAGE_SPACING)

        # Icon
        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES['icon_huge'], SIZES['icon_huge'])

        # Title
        self.title_label = StrongBodyLabel("")
        title = self.title_label
        title.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']}; color: {COLORS['lavender_grey']};")

        # Subtitle
        self.subtitle_label = CaptionLabel("")
        subtitle = self.subtitle_label
        subtitle.setStyleSheet(f"color: {COLORS['text_secondary']};")

        layout.addStretch()
        layout.addWidget(icon, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(subtitle, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addStretch()

        # Style (theme-consistent)
        self._apply_style()

        # Make clickable
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        self.retranslate_ui()

    def retranslate_ui(self):
        if self.title_label is not None:
            self.title_label.setText(self.tr("batch.drop.title"))
        if self.subtitle_label is not None:
            self.subtitle_label.setText(self.tr("batch.drop.subtitle"))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1 and urls[0].toLocalFile().endswith('.xlsx'):
                event.acceptProposedAction()
                self._apply_style(is_drag_active=True)

    def dragLeaveEvent(self, event):
        self._apply_style()

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith('.xlsx'):
            self.file_dropped.emit(files[0])
        self._apply_style()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.file_dropped.emit("__browse__")  # Signal to open file dialog

    def _apply_style(self, is_drag_active: bool = False):
        is_dark = isDarkTheme()
        border = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        dashed = COLORS['lavender_grey']
        from GUI_Qt.styles.theme_config import get_hover_bg
        hover_bg = get_hover_bg(is_dark)

        if is_drag_active:
            # Stronger visual state when file is being dragged over
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: {hover_bg};
                    border: 2px solid {border};
                    border-radius: {RADII['lg']}px;
                }}
            """)
            return

        self.setStyleSheet(f"""
            DropZoneWidget {{
                background-color: transparent;
                border: 2px dashed {dashed};
                border-radius: {RADII['lg']}px;
            }}
            DropZoneWidget:hover {{
                border-color: {border};
                background-color: {hover_bg};
            }}
        """)


class BatchUploadScreen(ResponsiveWidget):
    """Complete redesign with Fluent Design System"""

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.current_mode = "manual"
        self.row_counter = 0
        self._base_table_rows = 5

        # Multi-session browser pool
        self.session_manager = None

        # Load descriptions
        desc_manager = DescriptionManager(main_window.db)
        self.descriptions = [d['name'] for d in desc_manager.list_descriptions()]

        # Brands
        self.brands = [
            "KROSS", "Pinarello", "Basso", "Factor",
            "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan"
        ]

        # Store references for responsive UI
        self.scroll = None
        self.content_widget = None
        self.toolbar_layout = None

        self._init_ui()
        self._setup_shortcuts()

        # Connect to theme change signal
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(0, self._ensure_table_fills_viewport)

    def hideEvent(self, event):
        """Cleanup session manager when screen is hidden"""
        super().hideEvent(event)

        # Shutdown session manager if it exists
        if self.session_manager is not None:
            try:
                self.session_manager.shutdown_all()
                self.session_manager = None
            except Exception as e:
                if hasattr(self.main, 'logger'):
                    self.main.logger.log("BatchUploadScreen", f"Error shutting down session manager: {str(e)}")

    def eventFilter(self, obj, event):
        if hasattr(self, "table") and obj is self.table.viewport() and event.type() == QEvent.Type.Resize:
            QTimer.singleShot(0, self._ensure_table_fills_viewport)
        return super().eventFilter(obj, event)

    def _ensure_table_fills_viewport(self):
        """Add empty rows so the table fills its visible height.

        This prevents a large blank area under the last row on big windows.
        """
        if not hasattr(self, "table"):
            return
        if not self.table.isVisible():
            return
        if hasattr(self, "excel_empty") and self.excel_empty.isVisible():
            return

        viewport = self.table.viewport()
        if viewport is None:
            return
        viewport_height = viewport.height()
        if viewport_height <= 0:
            return

        row_height = self.table.verticalHeader().defaultSectionSize() or 56
        target_rows = (viewport_height + row_height - 1) // row_height + 1
        target_rows = max(self._base_table_rows, int(target_rows))
        target_rows = min(target_rows, 50)

        current_rows = self.table.rowCount()
        if current_rows >= target_rows:
            return

        self.table.setRowCount(target_rows)
        for row in range(current_rows, target_rows):
            self._setup_table_row(row)

    def _find_row_for_sender(self, column: int, widget_type):
        sender = self.sender()
        if sender is None:
            return None
        for row in range(self.table.rowCount()):
            cell = self.table.cellWidget(row, column)
            if not cell:
                continue
            w = cell.findChild(widget_type)
            if w is sender:
                return row
        return None

    def _init_ui(self):
        """Initialize UI with proper Fluent Design"""
        # Root layout (for ScrollArea container)
        root_layout = QVBoxLayout(self)
        root_layout.setContentsMargins(0, 0, 0, 0)
        root_layout.setSpacing(0)

        # Create ScrollArea
        self.scroll = ScrollArea(self)
        self.scroll.setWidgetResizable(True)
        root_layout.addWidget(self.scroll)

        # Content widget (inside ScrollArea)
        self.content_widget = QWidget()
        self.scroll.setWidget(self.content_widget)

        # Main layout (on content widget)
        layout = QVBoxLayout(self.content_widget)
        layout.setContentsMargins(*PAGE_MARGINS)
        layout.setSpacing(PAGE_SPACING)

        # Apply theme
        apply_screen_theme(
            self,
            "BatchUploadScreen",
            scroll=self.scroll,
            content=self.content_widget
        )

        # === HEADER SECTION ===
        header = QHBoxLayout()

        # Title with icon
        title_container = QHBoxLayout()
        title_container.setSpacing(ICON_TEXT_GAP)

        title_icon = TransparentToolButton(FluentIcon.SYNC, self)
        title_icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        title_icon.setEnabled(False)

        self.title_label = TitleLabel("")
        title_container.addWidget(title_icon)
        title_container.addWidget(self.title_label)

        header.addLayout(title_container)
        header.addStretch()

        # Mode pills (Fluent Design style)
        mode_container = QWidget()
        mode_layout = QHBoxLayout(mode_container)
        mode_layout.setContentsMargins(0, 0, 0, 0)
        mode_layout.setSpacing(ROW_SPACING)
        mode_container.setStyleSheet("background: transparent;")

        self.manual_pill = PillPushButton("")
        self.manual_pill.setCheckable(True)
        self.manual_pill.setChecked(True)
        self.manual_pill.clicked.connect(lambda: self._switch_mode("manual"))

        self.excel_pill = PillPushButton("")
        self.excel_pill.setCheckable(True)
        self.excel_pill.clicked.connect(lambda: self._switch_mode("excel"))

        mode_layout.addWidget(self.manual_pill)
        mode_layout.addWidget(self.excel_pill)

        header.addWidget(mode_container)
        layout.addLayout(header)

        # === TOOLBAR SECTION ===
        # Keep this toolbar in a single row (like Batch Titles) so the card doesn't
        # stretch in the Y axis at the default window size.
        toolbar_card = CardWidget()
        toolbar_card.setBorderRadius(RADII['md'])
        self.toolbar_layout = QHBoxLayout(toolbar_card)
        toolbar_layout = self.toolbar_layout  # Local alias for existing code
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self.add_btn = PushButton("")
        add_btn = self.add_btn
        add_btn.setIcon(FluentIcon.ADD.icon())
        add_btn.clicked.connect(self._add_row)

        self.clear_btn = PushButton("")
        clear_btn = self.clear_btn
        clear_btn.setIcon(FluentIcon.DELETE.icon())
        clear_btn.clicked.connect(self._clear_all)

        # Bulk selectors
        self.bulk_label = CaptionLabel("")
        bulk_label = self.bulk_label
        bulk_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self.frameset_bulk = CheckBox("")
        frameset_bulk = self.frameset_bulk
        frameset_bulk.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        frameset_bulk.stateChanged.connect(self._toggle_all_framesets)

        self.disclaimer_bulk = CheckBox("")
        disclaimer_bulk = self.disclaimer_bulk
        disclaimer_bulk.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        disclaimer_bulk.stateChanged.connect(self._toggle_all_disclaimers)

        # Spacer between Clear and Bulk Select (hide/show with manual controls)
        self._manual_sep = QWidget()
        self._manual_sep.setStyleSheet("background: transparent;")
        self._manual_sep.setFixedWidth(ROW_SPACING)

        # Progress ring (hidden initially)
        self.progress_ring = IndeterminateProgressRing(self)
        self.progress_ring.setFixedSize(24, 24)
        self.progress_ring.hide()

        # Elapsed time label (hidden initially)
        self.elapsed_label = CaptionLabel("Elapsed: 00:00")
        self.elapsed_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.elapsed_label.hide()

        # Session status label (hidden initially, shown during multi-session processing)
        self.session_status_label = CaptionLabel("Browsers: 0/0")
        self.session_status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.session_status_label.hide()

        # Status + Start (kept in the top toolbar for parity with other batch screens)
        self.status_label = BodyLabel("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        self.start_btn = PrimaryPushButton("")
        self.start_btn.setIcon(FluentIcon.PLAY)
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self._start_upload)

        # Retry button (hidden initially, shown after batch with failures)
        self.retry_btn = PushButton("")
        self.retry_btn.setIcon(FluentIcon.SYNC.icon())
        self.retry_btn.setVisible(False)
        self.retry_btn.clicked.connect(self._retry_failed)

        self.browse_btn = PushButton("")
        browse_btn = self.browse_btn
        browse_btn.setIcon(FluentIcon.FOLDER_ADD.icon())
        browse_btn.clicked.connect(self._browse_excel)

        self.template_btn = PushButton("")
        template_btn = self.template_btn
        template_btn.setIcon(FluentIcon.DOWNLOAD.icon())
        template_btn.clicked.connect(self._download_template)

        # Export buttons
        self.export_table_btn = PushButton("")
        self.export_table_btn.setIcon(FluentIcon.SAVE.icon())
        self.export_table_btn.clicked.connect(self._export_table_data)
        self.export_table_btn.setVisible(False)  # Hidden by default, shown in manual mode

        self.export_failed_btn = PushButton("")
        self.export_failed_btn.setIcon(FluentIcon.SAVE.icon())
        self.export_failed_btn.clicked.connect(self._export_failed_items)
        self.export_failed_btn.setVisible(False)  # Hidden until there are failed items

        self.excel_file_label = CaptionLabel("")
        self.excel_file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.excel_file_label.setMinimumWidth(0)
        # Keep the excel controls packed to the right (like other batch screens).
        # If this label expands, it pushes the buttons left and looks "offset".
        self.excel_file_label.setMaximumWidth(SIZES['col_w_260'])
        self.excel_file_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)

        # Spacer between excel file label and status (hide/show with excel controls)
        self._excel_status_sep = QWidget()
        self._excel_status_sep.setStyleSheet("background: transparent;")
        self._excel_status_sep.setFixedWidth(CONTENT_SPACING)

        # Expandable spacer that pushes status/start to the right.
        # This keeps the Excel controls on the left in Excel mode.
        self._toolbar_spacer = QWidget()
        self._toolbar_spacer.setStyleSheet("background: transparent;")
        self._toolbar_spacer.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

        # Keep status messages from squeezing the bulk checkboxes.
        # (Long validation text should clip inside this area, not steal toolbar space.)
        self.status_label.setMinimumWidth(0)
        self.status_label.setMaximumWidth(SIZES['col_w_160'])
        self.status_label.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Preferred)

        # Single row: manual controls + bulk selectors + excel controls + status/start
        # (show/hide is handled by _switch_mode).
        toolbar_layout.addWidget(add_btn)
        toolbar_layout.addWidget(clear_btn)
        toolbar_layout.addWidget(self.export_table_btn)
        toolbar_layout.addWidget(self.export_failed_btn)
        toolbar_layout.addWidget(self._manual_sep)
        toolbar_layout.addWidget(bulk_label)
        toolbar_layout.addWidget(frameset_bulk)
        toolbar_layout.addWidget(disclaimer_bulk)
        toolbar_layout.addWidget(template_btn)
        toolbar_layout.addWidget(browse_btn)
        toolbar_layout.addWidget(self.excel_file_label)
        toolbar_layout.addWidget(self._excel_status_sep)
        toolbar_layout.addWidget(self._toolbar_spacer)
        toolbar_layout.addWidget(self.progress_ring)
        toolbar_layout.addWidget(self.session_status_label)
        toolbar_layout.addWidget(self.elapsed_label)
        toolbar_layout.addWidget(self.status_label)
        toolbar_layout.addWidget(self.retry_btn)
        toolbar_layout.addWidget(self.start_btn)
        layout.addWidget(toolbar_card)

        # === CONTENT AREA ===
        content_card = CardWidget()
        content_card.setBorderRadius(RADII['md'])
        content_layout = QVBoxLayout(content_card)
        # Keep content flush so the table doesn't sit inside a larger card background.
        # Other batch tables don't have the extra padded background.
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)

        # Modern Fluent table
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(["", "", "", "", "", "", "", "", "", ""])
        self.table.setRowCount(self._base_table_rows)

        # Apply table styling
        self._update_table_theme()

        # Table settings
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setMinimumHeight(SIZES['table_header_height'])
        self.table.verticalHeader().setVisible(False)
        # Slightly taller rows prevent border/focus clipping on high-DPI systems
        self.table.verticalHeader().setDefaultSectionSize(SIZES['table_row_height_lg'])
        # Show column separators to make the table easier to scan.
        self.table.setShowGrid(True)

        # Smooth scrolling (avoid row/column snapping).
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)

        # Enable horizontal scrolling for responsive design
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        try:
            self.table.verticalScrollBar().setSingleStep(12)
            self.table.horizontalScrollBar().setSingleStep(12)
        except Exception:
            pass

        # Auto-fill rows when the table viewport resizes
        self.table.viewport().installEventFilter(self)

        # Enable corner clipping for rounded borders
        self.table.setCornerButtonEnabled(False)

        # Column widths - fixed (horizontal scrolling as needed).
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Fixed)

        self.table.setColumnWidth(0, SIZES['col_w_240'])  # Brand
        self.table.setColumnWidth(1, SIZES['col_w_180'])  # Code
        self.table.setColumnWidth(2, SIZES['col_w_420'])  # URL / Code
        self.table.setColumnWidth(3, SIZES['col_w_420'])  # Description
        self.table.setColumnWidth(4, SIZES['col_w_120'])  # Frameset
        self.table.setColumnWidth(5, SIZES['col_w_120'])  # Disclaimer
        self.table.setColumnWidth(6, SIZES['col_w_120'])  # Status
        self.table.setColumnWidth(7, SIZES['col_w_260'])  # Error
        self.table.setColumnWidth(8, SIZES['col_w_56'])   # Delete

        # Populate initial rows
        for row in range(self._base_table_rows):
            self._setup_table_row(row)

        # Fill remaining visible area with extra empty rows (useful on 1080p+).
        QTimer.singleShot(0, self._ensure_table_fills_viewport)

        # Table container with max-width for better layout on large screens
        table_container = QWidget()
        table_container.setStyleSheet("background: transparent;")
        table_container_layout = QVBoxLayout(table_container)
        table_container_layout.setContentsMargins(0, 0, 0, 0)
        table_container_layout.addWidget(self.table, 1)

        content_layout.addWidget(table_container, 1)

        # Excel drop zone
        self.excel_empty = DropZoneWidget(self.main.i18n.tr)
        self.excel_empty.file_dropped.connect(self._handle_file_drop)
        self.excel_empty.setVisible(False)
        content_layout.addWidget(self.excel_empty, 1)

        layout.addWidget(content_card, 1)

        self.retranslate_ui()

        # Ensure initial toolbar visibility matches the active mode.
        # Without this, excel-mode controls can appear in manual mode until the user toggles pills.
        self._switch_mode(self.current_mode)

    def retranslate_ui(self):
        tr = self.main.i18n.tr

        self.title_label.setText(tr("batch.title"))
        self.manual_pill.setText(tr("batch.mode.manual"))
        self.excel_pill.setText(tr("batch.mode.excel"))

        self.add_btn.setText(tr("batch.add_row"))
        self.add_btn.setToolTip(tr("batch.add_row.tip"))
        self.clear_btn.setText(tr("batch.clear_all"))
        self.clear_btn.setToolTip(tr("batch.clear_all.tip"))
        self.bulk_label.setText(tr("batch.bulk_select"))
        self.frameset_bulk.setText(tr("batch.bulk.frameset"))
        self.frameset_bulk.setToolTip(tr("batch.bulk.frameset.tip"))
        self.disclaimer_bulk.setText(tr("batch.bulk.disclaimer"))
        self.disclaimer_bulk.setToolTip(tr("batch.bulk.disclaimer.tip"))
        self.browse_btn.setText(tr("batch.browse_excel"))
        self.browse_btn.setToolTip(tr("batch.browse_excel.tip"))
        self.template_btn.setText(tr("batch.download_template"))
        self.template_btn.setToolTip(tr("batch.download_template.tip"))
        self.export_table_btn.setText("Export Table")
        self.export_table_btn.setToolTip("Export current table data to Excel")
        self.export_failed_btn.setText("Export Failed")
        self.export_failed_btn.setToolTip("Export failed items with errors to Excel")
        self._set_excel_file_text(tr("batch.no_file"))

        # Keep the Start button area clean; don't show the "Ready" hint.
        self._set_status_text("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.start_btn.setText(f"{tr('batch.start')} (Ctrl+Enter)")
        self.start_btn.setToolTip(tr("batch.start.tip"))
        self.retry_btn.setText(f"{tr('batch.retry_failed')} (Ctrl+R)")
        self.retry_btn.setToolTip(tr("batch.retry_failed.tip"))

        # Table headers
        self.table.setHorizontalHeaderLabels([
            tr("batch.table.brand"),
            tr("batch.table.code"),
            tr("batch.table.url"),
            tr("batch.table.desc"),
            tr("batch.table.frameset"),
            tr("batch.table.disclaimer"),
            tr("batch.table.status"),
            tr("batch.table.error"),
            "",
        ])

        # Drop-zone text
        if hasattr(self, "excel_empty") and hasattr(self.excel_empty, "retranslate_ui"):
            self.excel_empty.retranslate_ui()

        # Update placeholders for existing table row widgets
        for row in range(self.table.rowCount()):
            brand_cell = self.table.cellWidget(row, 0)
            if brand_cell:
                brand_combo = brand_cell.findChild(ComboBox)
                if brand_combo:
                    self._ensure_combo_placeholder_item(brand_combo, tr("batch.select_brand"), self.brands)

            code_cell = self.table.cellWidget(row, 1)
            if code_cell:
                code_field = code_cell.findChild(LineEdit)
                if code_field:
                    code_field.setPlaceholderText(tr("upload.code.placeholder"))

            url_cell = self.table.cellWidget(row, 2)
            if url_cell:
                url_field = url_cell.findChild(LineEdit)
                if url_field:
                    url_field.setPlaceholderText(tr("upload.url.placeholder"))

            desc_cell = self.table.cellWidget(row, 3)
            if desc_cell:
                desc_combo = desc_cell.findChild(ComboBox)
                if desc_combo:
                    desc_combo.setPlaceholderText(tr("batch.optional"))

    def _setup_shortcuts(self):
        """Setup keyboard shortcuts for batch upload screen"""
        # Ctrl+Enter to start upload
        start_shortcut = QShortcut(QKeySequence("Ctrl+Return"), self)
        start_shortcut.activated.connect(self._handle_start_shortcut)

        # Escape to cancel/stop upload
        cancel_shortcut = QShortcut(QKeySequence("Escape"), self)
        cancel_shortcut.activated.connect(self._handle_cancel_shortcut)

        # Ctrl+R to retry failed items
        retry_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        retry_shortcut.activated.connect(self._handle_retry_shortcut)

    def _handle_start_shortcut(self):
        """Handle Ctrl+Enter shortcut to start upload"""
        if self.start_btn.isEnabled():
            self._start_upload()

    def _handle_cancel_shortcut(self):
        """Handle Escape shortcut to cancel/stop upload"""
        # Check if worker is running
        if hasattr(self, 'worker') and hasattr(self.worker, 'isRunning') and self.worker.isRunning():
            self.worker.stop()
            InfoBar.warning(
                title=self.main.i18n.tr("batch.cancel.title"),
                content=self.main.i18n.tr("batch.cancel.content"),
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _handle_retry_shortcut(self):
        """Handle Ctrl+R shortcut to retry failed items"""
        if self.retry_btn.isVisible() and self.retry_btn.isEnabled():
            self._retry_failed()

    def _setup_table_row(self, row):
        """Setup widgets for a table row"""
        # Helper to create centered container
        def create_centered_widget(widget):
            container = QWidget()
            container.setStyleSheet("background: transparent;")
            container.setProperty("ubTableCell", True)
            container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)

            # Allow embedded widgets to shrink to the available cell width.
            # Without this, QComboBox can keep a large implicit minimum width and get clipped.
            try:
                widget.setMinimumWidth(0)
            except Exception:
                pass
            try:
                widget.setSizePolicy(QSizePolicy.Policy.Expanding, widget.sizePolicy().verticalPolicy())
            except Exception:
                pass

            layout = QHBoxLayout(container)
            # Inset widgets from gridlines so their borders are never clipped/overdrawn
            # Extra vertical padding avoids clipping bottom borders/focus rings
            layout.setContentsMargins(*TABLE_CELL_MARGINS)
            layout.setSpacing(0)
            layout.addWidget(widget, 1, Qt.AlignmentFlag.AlignVCenter)
            return container

        # Brand combo
        brand_combo = ComboBox()
        self._ensure_combo_placeholder_item(
            brand_combo,
            self.main.i18n.tr("batch.select_brand"),
            self.brands,
            force_placeholder=True,
        )
        brand_combo.setMinimumHeight(SIZES['input_height'])
        brand_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        brand_combo.currentTextChanged.connect(self._on_brand_change)
        brand_combo.currentTextChanged.connect(self._validate)
        self.table.setCellWidget(row, 0, create_centered_widget(brand_combo))

        # Product code
        code_field = LineEdit()
        code_field.setPlaceholderText(self.main.i18n.tr("upload.code.placeholder"))
        code_field.setMinimumHeight(SIZES['input_height'])
        code_field.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        code_field.textChanged.connect(lambda text, field=code_field: self._on_code_changed(text, field))
        self.table.setCellWidget(row, 1, create_centered_widget(code_field))

        # URL
        url_field = LineEdit()
        url_field.setPlaceholderText(self.main.i18n.tr("upload.url.placeholder"))
        url_field.setMinimumHeight(SIZES['input_height'])
        url_field.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        url_field.textChanged.connect(lambda text, field=url_field: self._on_url_changed(text, field))
        self.table.setCellWidget(row, 2, create_centered_widget(url_field))

        # Description
        desc_combo = ComboBox()
        desc_combo.addItem("")
        desc_combo.addItems(self.descriptions)
        desc_combo.setPlaceholderText(self.main.i18n.tr("batch.optional"))
        desc_combo.setMinimumHeight(SIZES['input_height'])
        desc_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        desc_combo.currentTextChanged.connect(self._validate)
        self.table.setCellWidget(row, 3, create_centered_widget(desc_combo))

        # Frameset checkbox - centered with fixed approach
        frameset_check = CheckBox("")
        frameset_check.setProperty("ubTableCheck", True)
        frameset_check.setVisible(False)
        frameset_check.stateChanged.connect(self._validate)
        # Let the checkbox use its natural size; the container layout will center it.
        frameset_container = QWidget()
        frameset_container.setStyleSheet("background: transparent;")
        frameset_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        frameset_layout = QHBoxLayout(frameset_container)
        frameset_layout.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        frameset_layout.setSpacing(0)
        frameset_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        frameset_layout.addWidget(frameset_check)
        self.table.setCellWidget(row, 4, frameset_container)

        # Disclaimer checkbox - centered with fixed approach
        disclaimer_check = CheckBox("")
        disclaimer_check.setProperty("ubTableCheck", True)
        disclaimer_check.stateChanged.connect(self._validate)
        # Let the checkbox use its natural size; the container layout will center it.
        disclaimer_container = QWidget()
        disclaimer_container.setStyleSheet("background: transparent;")
        disclaimer_container.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        disclaimer_layout = QHBoxLayout(disclaimer_container)
        disclaimer_layout.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        disclaimer_layout.setSpacing(0)
        disclaimer_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        disclaimer_layout.addWidget(disclaimer_check)
        self.table.setCellWidget(row, 5, disclaimer_container)

        # Status column (initially empty)
        status_item = QTableWidgetItem("")
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(row, 6, status_item)

        # Error column (initially empty)
        error_item = QTableWidgetItem("")
        self.table.setItem(row, 7, error_item)

        # Delete button
        delete_btn = TransparentToolButton(FluentIcon.DELETE)
        delete_btn.setToolTip(self.main.i18n.tr("batch.row.remove.tip"))
        delete_btn.setFixedSize(SIZES['button_height_sm'], SIZES['button_height_sm'])
        delete_btn.clicked.connect(self._remove_row)
        delete_container = QWidget()
        delete_container.setStyleSheet("background: transparent;")
        delete_layout = QHBoxLayout(delete_container)
        # Tight margins so the icon isn't pushed into the table border/scrollbar.
        delete_layout.setContentsMargins(SPACING['xs'], 0, SPACING['xs'], 0)
        delete_layout.addStretch(1)
        delete_layout.addWidget(delete_btn, 0, Qt.AlignmentFlag.AlignVCenter)
        delete_layout.addStretch(1)
        self.table.setCellWidget(row, 8, delete_container)

    def _get_widget_from_cell(self, row, col, widget_type):
        """Helper to get widget from table cell (handles containers)"""
        cell_widget = self.table.cellWidget(row, col)
        if not cell_widget:
            return None
        # If it's a container, find the actual widget
        widget = cell_widget.findChild(widget_type)
        return widget if widget else (cell_widget if isinstance(cell_widget, widget_type) else None)

    def _on_brand_change(self, brand):
        """Handle brand change to show/hide frameset checkbox"""
        row = self._find_row_for_sender(0, ComboBox)
        if row is None:
            return
        checkbox = self._get_widget_from_cell(row, 4, CheckBox)
        if checkbox:
            is_pinarello = brand == "Pinarello"
            checkbox.setVisible(is_pinarello)
            if not is_pinarello:
                checkbox.setChecked(False)

    def _switch_mode(self, mode):
        """Switch between manual and Excel mode"""
        # Batch Upload has a weird sizing interaction on some systems where toggling
        # Excel → Manual causes the top-level window to grow horizontally.
        # To avoid that, temporarily lock the window width during the mode switch,
        # then restore the previous min/max constraints once the layout settles.
        win = self.window()
        lock_w = None
        prev_min_w = None
        prev_max_w = None
        try:
            if win is not None and win.isWindow() and not win.isMaximized():
                lock_w = int(win.width())
                prev_min_w = int(win.minimumWidth())
                prev_max_w = int(win.maximumWidth())
                win.setMinimumWidth(lock_w)
                win.setMaximumWidth(lock_w)
        except Exception:
            lock_w = None

        self.current_mode = mode

        # Update pills
        self.manual_pill.setChecked(mode == "manual")
        self.excel_pill.setChecked(mode == "excel")

        # Update toolbar groups
        is_manual = mode == "manual"
        self.add_btn.setVisible(is_manual)
        self.clear_btn.setVisible(is_manual)
        self.export_table_btn.setVisible(is_manual)  # Show export in manual mode
        self.bulk_label.setVisible(is_manual)
        self.frameset_bulk.setVisible(is_manual)
        self.disclaimer_bulk.setVisible(is_manual)
        self._manual_sep.setVisible(is_manual)

        is_excel = mode == "excel"
        self.browse_btn.setVisible(is_excel)
        self.template_btn.setVisible(is_excel)
        self.excel_file_label.setVisible(is_excel)
        self._excel_status_sep.setVisible(is_excel)

        # Update content visibility
        has_data = self._has_valid_rows()
        if mode == "excel" and not has_data:
            self.table.setVisible(False)
            self.excel_empty.setVisible(True)
        else:
            self.table.setVisible(True)
            self.excel_empty.setVisible(False)
            QTimer.singleShot(0, self._ensure_table_fills_viewport)

        if lock_w is not None:
            def _unlock_width(window=win, min_w=prev_min_w, max_w=prev_max_w):
                try:
                    if window is None or (not window.isWindow()) or window.isMaximized():
                        return
                    # Restore previous constraints.
                    if min_w is not None:
                        window.setMinimumWidth(int(min_w))
                    if max_w is not None:
                        window.setMaximumWidth(int(max_w))
                except Exception:
                    pass

            # Give Qt a moment to process pending layout/resizes.
            QTimer.singleShot(120, _unlock_width)

    def _has_valid_rows(self):
        """Check if table has any valid data"""
        for row in range(self.table.rowCount()):
            brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
            code_widget = self._get_widget_from_cell(row, 1, LineEdit)
            url_widget = self._get_widget_from_cell(row, 2, LineEdit)

            if (brand_widget and brand_widget.currentText() and
                code_widget and code_widget.text().strip() and
                url_widget and url_widget.text().strip()):
                return True
        return False

    def _add_row(self):
        """Add new row to table"""
        current_rows = self.table.rowCount()
        self.table.setRowCount(current_rows + 1)
        self._setup_table_row(current_rows)
        QTimer.singleShot(0, self._ensure_table_fills_viewport)
        self._validate()

    def _remove_row(self):
        """Remove a row from table"""
        row = self._find_row_for_sender(9, TransparentToolButton)
        if row is None:
            return
        if self.table.rowCount() > 1:
            self.table.removeRow(row)
            QTimer.singleShot(0, self._ensure_table_fills_viewport)
            self._validate()

    def _clear_all(self):
        """Clear all rows with confirmation dialog"""
        # Show confirmation dialog
        w = MessageBox(
            title=self.main.i18n.tr("common.confirm"),
            content=self.main.i18n.tr("batch.clear.confirm"),
            parent=self
        )
        if w.exec():
            self.table.setRowCount(self._base_table_rows)
            for row in range(self._base_table_rows):
                self._setup_table_row(row)
            QTimer.singleShot(0, self._ensure_table_fills_viewport)
            self._validate()

            # Hide retry button since table is cleared
            if hasattr(self, 'retry_btn'):
                self.retry_btn.setVisible(False)

    def _toggle_all_framesets(self, state):
        """Toggle all frameset checkboxes"""
        checked = state == Qt.CheckState.Checked.value
        for row in range(self.table.rowCount()):
            brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
            if brand_widget and brand_widget.currentText() == "Pinarello":
                checkbox = self._get_widget_from_cell(row, 4, CheckBox)
                if checkbox and checkbox.isVisible():
                    checkbox.setChecked(checked)

    def _toggle_all_disclaimers(self, state):
        """Toggle all disclaimer checkboxes"""
        checked = state == Qt.CheckState.Checked.value
        for row in range(self.table.rowCount()):
            checkbox = self._get_widget_from_cell(row, 5, CheckBox)
            if checkbox:
                checkbox.setChecked(checked)

    def _on_code_changed(self, text, field):
        """Handle product code field change with validation"""
        self._validate()

    def _on_url_changed(self, text, field):
        """Handle URL field change with validation"""
        self._validate()

    def _set_status_text(self, text: str) -> None:
        """Set status label text without allowing it to steal toolbar space."""
        if not hasattr(self, "status_label") or self.status_label is None:
            return

        # Don't reserve toolbar space when there's nothing to show.
        try:
            self.status_label.setVisible(bool((text or "").strip()))
        except Exception:
            pass

        # Keep full text accessible on hover.
        try:
            self.status_label.setToolTip(text)
        except Exception:
            pass

        # Elide to the label's max width so it can't expand the toolbar.
        try:
            max_w = int(self.status_label.maximumWidth() or 0)
        except Exception:
            max_w = 0
        if max_w <= 0:
            max_w = SIZES['col_w_160']

        try:
            fm = self.status_label.fontMetrics()
            self.status_label.setText(fm.elidedText(text, Qt.TextElideMode.ElideRight, max_w))
        except Exception:
            self.status_label.setText(text)

    def _set_excel_file_text(self, text: str) -> None:
        """Set excel file label without letting it stretch the toolbar."""
        if not hasattr(self, "excel_file_label") or self.excel_file_label is None:
            return

        try:
            self.excel_file_label.setToolTip(text)
        except Exception:
            pass

        try:
            max_w = int(self.excel_file_label.maximumWidth() or 0)
        except Exception:
            max_w = 0
        if max_w <= 0:
            max_w = SIZES['col_w_260']

        try:
            fm = self.excel_file_label.fontMetrics()
            self.excel_file_label.setText(fm.elidedText(text, Qt.TextElideMode.ElideRight, max_w))
        except Exception:
            self.excel_file_label.setText(text)

    def _validate(self):
        """Validate all rows"""
        valid_count = 0
        for row in range(self.table.rowCount()):
            brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
            code_widget = self._get_widget_from_cell(row, 1, LineEdit)
            url_widget = self._get_widget_from_cell(row, 2, LineEdit)

            brand = brand_widget.currentText().strip() if brand_widget else ""

            if (brand in self.brands and
                code_widget and code_widget.text().strip() and
                url_widget and url_widget.text().strip()):
                valid_count += 1

        self.start_btn.setEnabled(valid_count > 0)

        # Don't show a "Ready" hint; keep Start enabled/disabled quietly.
        self._set_status_text("")
        self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

    def _handle_file_drop(self, filepath):
        """Handle file drop or click on drop zone"""
        if filepath == "__browse__":
            self._browse_excel()
        else:
            self._load_excel(filepath)

    def _browse_excel(self):
        """Browse for Excel file"""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            self.main.i18n.tr("batch.file.select_excel.title"),
            "",
            self.main.i18n.tr("batch.file.select_excel.filter")
        )

        if filename:
            self._load_excel(filename)

    @staticmethod
    def _normalize_code(raw: str) -> str:
        """Normalize product code to always include UB- prefix.

        Mirrors BatchTitles behavior: users can type codes without the prefix,
        but uploads will always use UB-<code>.
        """
        code = (raw or "").strip()
        if not code:
            return ""
        if code.upper().startswith("UB-"):
            return code
        return f"UB-{code}"

    def _load_excel(self, filename):
        """Load Excel file"""
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # Find header
            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                values = [str(c.value).lower() if c.value else "" for c in row]
                if any("brand" in v for v in values):
                    header_row = idx
                    break

            if not header_row:
                raise Exception(self.main.i18n.tr("batch.excel.header_missing_in_file"))

            # Count data rows first
            data_rows = []
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if any(row):
                    data_rows.append(row)

            # Set table row count
            self.table.setRowCount(len(data_rows))

            # Load rows into table
            valid_count = 0
            invalid_count = 0

            for table_row, excel_row in enumerate(data_rows):
                brand = str(excel_row[0]) if len(excel_row) > 0 and excel_row[0] else ""
                code = str(excel_row[1]) if len(excel_row) > 1 and excel_row[1] else ""
                url = str(excel_row[2]) if len(excel_row) > 2 and excel_row[2] else ""
                desc = str(excel_row[3]).strip() if len(excel_row) > 3 and excel_row[3] else ""

                frameset_raw = excel_row[4] if len(excel_row) > 4 else False
                frameset = str(frameset_raw).lower() in ("yes", "true", "1") if isinstance(frameset_raw, str) else bool(frameset_raw)

                disclaimer_raw = excel_row[5] if len(excel_row) > 5 else False
                disclaimer = str(disclaimer_raw).lower() in ("yes", "true", "1") if isinstance(disclaimer_raw, str) else bool(disclaimer_raw)

                # Setup row
                self._setup_table_row(table_row)

                # Set data using helper
                brand_widget = self._get_widget_from_cell(table_row, 0, ComboBox)
                if brand_widget and brand:
                    brand_widget.setCurrentText(brand)

                code_widget = self._get_widget_from_cell(table_row, 1, LineEdit)
                if code_widget:
                    code_widget.setText(code)

                url_widget = self._get_widget_from_cell(table_row, 2, LineEdit)
                if url_widget:
                    url_widget.setText(url)

                desc_widget = self._get_widget_from_cell(table_row, 3, ComboBox)
                if desc_widget and desc:
                    desc_widget.setCurrentText(desc)

                frameset_checkbox = self._get_widget_from_cell(table_row, 4, CheckBox)
                if frameset_checkbox:
                    frameset_checkbox.setChecked(frameset)

                disclaimer_checkbox = self._get_widget_from_cell(table_row, 5, CheckBox)
                if disclaimer_checkbox:
                    disclaimer_checkbox.setChecked(disclaimer)

                if brand and code and url:
                    valid_count += 1
                else:
                    invalid_count += 1

            wb.close()

            # Update UI
            import os
            self._set_excel_file_text(os.path.basename(filename))
            self.table.setVisible(True)
            self.excel_empty.setVisible(False)

            # Update status
            if invalid_count > 0:
                self._set_status_text(
                    self.main.i18n.tr(
                        "batch.excel.rows_fix_errors",
                        valid=valid_count,
                        invalid=invalid_count,
                    )
                )
                self.status_label.setStyleSheet(f"color: {COLORS['warning']}; font-weight: 500;")
                self.start_btn.setEnabled(False)
            else:
                # Don't show a "Ready" hint; keep Start enabled silently.
                self._set_status_text("")
                self.status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
                self.start_btn.setEnabled(valid_count > 0)

        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.excel.load_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self
            )

    def _download_template(self):
        """Download Excel template"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            self.main.i18n.tr("batch.template.save.title"),
            self.main.i18n.tr("batch.template.default_name"),
            self.main.i18n.tr("batch.template.filter"),
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self.main.i18n.tr("batch.title")

            # Headers
            headers = ["Brand", "Product Code", "URL or Code", "Description", "Frameset", "Append Disclaimer", "Append Order Note"]
            ws.append(headers)

            # Examples
            ws.append(["KROSS", "UB-1234", "https://kross.pl/product1", "Mountain Bike", "No", "Yes", "Yes"])
            ws.append(["Pinarello", "UB-5678", "https://pinarello.com/product2", "Road Bike", "Yes", "No", "No"])

            # Style with our colors
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="8D99AE", end_color="8D99AE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-width
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 20

            wb.save(filename)
            wb.close()

            InfoBar.success(
                title=self.main.i18n.tr("common.success"),
                content=self.main.i18n.tr("batch.template.downloaded"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )

        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("common.error"),
                content=self.main.i18n.tr("batch.template.save_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=4000,
                parent=self
            )

    def _export_table_data(self):
        """Export current table data to Excel"""
        from datetime import datetime

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Table Data",
            f"batch_upload_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)",
        )

        if not filename:
            return

        try:
            # Collect all row data from table
            rows_data = []
            headers = ["Brand", "Code", "URL", "Description", "Frameset", "Disclaimer", "Order Note", "Status", "Error"]

            for row in range(self.table.rowCount()):
                row_data = {}

                # Brand (column 0)
                brand_cell = self.table.cellWidget(row, 0)
                brand_combo = brand_cell.findChild(ComboBox) if brand_cell else None
                row_data['brand'] = brand_combo.currentText() if brand_combo else ""

                # Code (column 1)
                code_cell = self.table.cellWidget(row, 1)
                code_field = code_cell.findChild(LineEdit) if code_cell else None
                row_data['code'] = code_field.text() if code_field else ""

                # URL (column 2)
                url_cell = self.table.cellWidget(row, 2)
                url_field = url_cell.findChild(LineEdit) if url_cell else None
                row_data['url'] = url_field.text() if url_field else ""

                # Description (column 3)
                desc_cell = self.table.cellWidget(row, 3)
                desc_combo = desc_cell.findChild(ComboBox) if desc_cell else None
                row_data['description'] = desc_combo.currentText() if desc_combo else ""

                # Frameset (column 4)
                frameset_cell = self.table.cellWidget(row, 4)
                frameset_check = frameset_cell.findChild(CheckBox) if frameset_cell else None
                row_data['frameset'] = "Yes" if (frameset_check and frameset_check.isChecked()) else "No"

                # Disclaimer (column 5)
                disclaimer_cell = self.table.cellWidget(row, 5)
                disclaimer_check = disclaimer_cell.findChild(CheckBox) if disclaimer_cell else None
                row_data['disclaimer'] = "Yes" if (disclaimer_check and disclaimer_check.isChecked()) else "No"

                # Status (column 6)
                status_item = self.table.item(row, 6)
                row_data['status'] = status_item.text() if status_item else ""

                # Error (column 7)
                error_item = self.table.item(row, 7)
                row_data['error'] = error_item.text() if error_item else ""

                rows_data.append(row_data)

            # Export using ExcelHandler
            success = ExcelHandler.export_table_data(headers, rows_data, filename, "Batch Upload Data")

            if success:
                InfoBar.success(
                    title="Export Successful",
                    content=f"Table data exported to {filename}",
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000
                )
            else:
                InfoBar.error(
                    title="Export Failed",
                    content="Failed to export table data",
                    parent=self,
                    position=InfoBarPosition.TOP
                )

        except Exception as ex:
            InfoBar.error(
                title="Export Error",
                content=f"Error exporting table: {str(ex)}",
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _export_failed_items(self):
        """Export failed items with error details to Excel"""
        from datetime import datetime

        # Collect failed items from table
        failed_items = []
        headers = ["Brand", "Code", "URL", "Description", "Frameset", "Disclaimer", "Status", "Error"]

        for row in range(self.table.rowCount()):
            status_item = self.table.item(row, 6)
            status_text = status_item.text() if status_item else ""

            # Only include failed items
            if "Failed" in status_text or "✗" in status_text:
                row_data = {}

                # Brand (column 0)
                brand_cell = self.table.cellWidget(row, 0)
                brand_combo = brand_cell.findChild(ComboBox) if brand_cell else None
                row_data['brand'] = brand_combo.currentText() if brand_combo else ""

                # Code (column 1)
                code_cell = self.table.cellWidget(row, 1)
                code_field = code_cell.findChild(LineEdit) if code_cell else None
                row_data['code'] = code_field.text() if code_field else ""

                # URL (column 2)
                url_cell = self.table.cellWidget(row, 2)
                url_field = url_cell.findChild(LineEdit) if url_cell else None
                row_data['url'] = url_field.text() if url_field else ""

                # Description (column 3)
                desc_cell = self.table.cellWidget(row, 3)
                desc_combo = desc_cell.findChild(ComboBox) if desc_cell else None
                row_data['description'] = desc_combo.currentText() if desc_combo else ""

                # Frameset (column 4)
                frameset_cell = self.table.cellWidget(row, 4)
                frameset_check = frameset_cell.findChild(CheckBox) if frameset_cell else None
                row_data['frameset'] = "Yes" if (frameset_check and frameset_check.isChecked()) else "No"

                # Disclaimer (column 5)
                disclaimer_cell = self.table.cellWidget(row, 5)
                disclaimer_check = disclaimer_cell.findChild(CheckBox) if disclaimer_cell else None
                row_data['disclaimer'] = "Yes" if (disclaimer_check and disclaimer_check.isChecked()) else "No"

                # Status (column 6)
                row_data['status'] = status_text

                # Error (column 7)
                error_item = self.table.item(row, 7)
                row_data['error'] = error_item.text() if error_item else ""

                failed_items.append(row_data)

        if not failed_items:
            InfoBar.warning(
                title="No Failed Items",
                content="There are no failed items to export",
                parent=self,
                position=InfoBarPosition.TOP
            )
            return

        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Export Failed Items",
            f"failed_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)",
        )

        if not filename:
            return

        try:
            success = ExcelHandler.export_failed_items(headers, failed_items, filename)

            if success:
                InfoBar.success(
                    title="Export Successful",
                    content=f"Failed items exported to {filename}",
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000
                )
            else:
                InfoBar.error(
                    title="Export Failed",
                    content="Failed to export failed items",
                    parent=self,
                    position=InfoBarPosition.TOP
                )

        except Exception as ex:
            InfoBar.error(
                title="Export Error",
                content=f"Error exporting failed items: {str(ex)}",
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _on_row_update(self, row: int, status_text: str, error_text: str):
        """Handle row status updates from worker thread."""

        # Clear previous row highlighting
        if hasattr(self, '_current_processing_row') and self._current_processing_row is not None:
            for col in range(self.table.columnCount()):
                item = self.table.item(self._current_processing_row, col)
                if item:
                    item.setBackground(Qt.GlobalColor.transparent)

        # Highlight current row based on status
        if "Processing" in status_text:
            bg_color = QColor(255, 182, 193, 100)  # Blush rose
            self._current_processing_row = row
        elif "Success" in status_text or "✓" in status_text:
            bg_color = QColor(144, 238, 144, 120)  # Light green
            self._current_processing_row = None
        elif "Failed" in status_text or "✗" in status_text:
            bg_color = QColor(255, 99, 71, 100)  # Tomato red
            self._current_processing_row = None
        else:
            bg_color = Qt.GlobalColor.transparent

        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item:
                item.setBackground(bg_color)

        # Update status column (column 6)
        status_item = self.table.item(row, 6)
        if not status_item:
            status_item = QTableWidgetItem()
            status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 6, status_item)
        status_item.setText(status_text)

        # Update error column (column 7)
        error_item = self.table.item(row, 7)
        if not error_item:
            error_item = QTableWidgetItem()
            self.table.setItem(row, 7, error_item)
        error_item.setText(error_text)

        # Auto-scroll to current row
        self.table.scrollToItem(self.table.item(row, 0), QTableWidget.ScrollHint.PositionAtCenter)

    def _on_progress_update(self, current: int, total: int, speed: float, eta_seconds: float):
        """Handle progress updates with ETA from worker thread."""
        percentage = (current / total * 100) if total > 0 else 0

        # Format ETA
        if eta_seconds < 60:
            eta_text = f"{int(eta_seconds)}s"
        elif eta_seconds < 3600:
            minutes = int(eta_seconds / 60)
            seconds = int(eta_seconds % 60)
            eta_text = f"{minutes}m {seconds}s"
        else:
            hours = int(eta_seconds / 3600)
            minutes = int((eta_seconds % 3600) / 60)
            eta_text = f"{hours}h {minutes}m"

        # Format speed (items per minute)
        items_per_min = speed * 60

        # Update status label with progress information
        message = f"Processing ({current}/{total}) {percentage:.0f}% - ETA: {eta_text} | {items_per_min:.1f} items/min"
        self._set_status_text(message)

    def _on_done(self, ok: int, total: int):
        """Handle batch completion."""
        # Stop elapsed timer
        if hasattr(self, '_elapsed_timer'):
            self._elapsed_timer.stop()
            if hasattr(self, 'elapsed_label'):
                self.elapsed_label.hide()

        # Hide session status label
        if hasattr(self, 'session_status_label'):
            self.session_status_label.hide()

        # Hide progress ring
        if hasattr(self, 'progress_ring'):
            self.progress_ring.stop()
            self.progress_ring.hide()

        # Re-enable controls
        self.start_btn.setEnabled(True)

        # Clear processing row tracking
        if hasattr(self, '_current_processing_row'):
            self._current_processing_row = None

        # Show completion message
        failed = total - ok
        if failed > 0:
            # Show retry button and export failed button
            self.retry_btn.setVisible(True)
            self.export_failed_btn.setVisible(True)

            InfoBar.warning(
                title=self.main.i18n.tr("batch.run.complete.title"),
                content=self.main.i18n.tr("batch.run.complete.partial", ok=ok, failed=failed, total=total),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=8000,
                parent=self
            )

            # Windows notification for partial completion
            self._show_windows_notification(ok, total)
        else:
            # Hide retry and export failed buttons if all succeeded
            self.retry_btn.setVisible(False)
            self.export_failed_btn.setVisible(False)

            InfoBar.success(
                title=self.main.i18n.tr("batch.run.complete.title"),
                content=self.main.i18n.tr("batch.run.complete.content", total=total),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=5000,
                parent=self
            )

        # Windows notification
        self._show_windows_notification(ok, total)

    def _show_windows_notification(self, ok: int, total: int):
        """Show Windows desktop notification for batch completion."""
        # NOTE: win10toast relies on pywin32 WNDPROC callbacks that are known to
        # break on newer Python versions (e.g. 3.12+), causing hard crashes like:
        # "WNDPROC return value cannot be converted to LRESULT".
        try:
            import sys
            if sys.version_info >= (3, 12):
                return
        except Exception:
            return

        try:
            from win10toast import ToastNotifier
            toaster = ToastNotifier()

            failed = total - ok
            if failed > 0:
                title = "Batch Upload Completed with Errors"
                message = f"Success: {ok}/{total}\nFailed: {failed}"
            else:
                title = "Batch Upload Completed Successfully"
                message = f"All {total} products uploaded successfully"

            # Show notification (non-blocking, 10 second duration)
            toaster.show_toast(
                title=title,
                msg=message,
                duration=10,
                threaded=True,
                icon_path=None  # Uses default Windows icon
            )
        except ImportError:
            # win10toast not installed, silently skip
            pass
        except Exception:
            # Any other error, silently skip
            pass

    def _update_elapsed(self):
        """Update elapsed time display."""
        if not hasattr(self, 'start_time') or not hasattr(self, 'elapsed_label'):
            return

        import time
        elapsed = int(time.time() - self.start_time)
        mins, secs = divmod(elapsed, 60)
        self.elapsed_label.setText(f"Elapsed: {mins:02d}:{secs:02d}")

    def _on_session_status_update(self, stats: dict):
        """Handle session status updates from parallel worker."""
        if hasattr(self, 'session_status_label'):
            busy = stats.get('busy', 0)
            total = stats.get('total', 0)
            self.session_status_label.setText(f"Browsers: {busy}/{total}")

    def _retry_failed(self):
        """Retry only the failed rows from the previous batch."""
        items = []
        failed_rows = []

        # Find all rows with "✗ Failed" status
        for row in range(self.table.rowCount()):
            status_item = self.table.item(row, 6)
            if status_item and ("✗" in status_item.text() or "Failed" in status_item.text()):
                # Collect data from this failed row
                brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
                code_widget = self._get_widget_from_cell(row, 1, LineEdit)
                url_widget = self._get_widget_from_cell(row, 2, LineEdit)

                if not (brand_widget and code_widget and url_widget):
                    continue

                brand = brand_widget.currentText()
                code_raw = code_widget.text().strip()
                code = self._normalize_code(code_raw)
                url = url_widget.text().strip()

                if not (brand in self.brands and code and url):
                    continue

                desc_widget = self._get_widget_from_cell(row, 3, ComboBox)
                desc = desc_widget.currentText() if desc_widget else ""

                frameset_checkbox = self._get_widget_from_cell(row, 4, CheckBox)
                frameset = frameset_checkbox.isChecked() if frameset_checkbox else False

                disclaimer_checkbox = self._get_widget_from_cell(row, 5, CheckBox)
                disclaimer = disclaimer_checkbox.isChecked() if disclaimer_checkbox else False

                items.append({
                    'brand': brand,
                    'code': code,
                    'url': url,
                    'description_name': desc,
                    'frameset_only': frameset,
                    'append_disclaimer': disclaimer,
                })
                failed_rows.append(row)

        if not items:
            InfoBar.info(
                title=self.main.i18n.tr("common.info"),
                content="No failed items to retry",
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )
            return

        # Hide retry button during processing
        self.retry_btn.setVisible(False)

        # Check for master password (same logic as _start_upload)
        master_password = getattr(self.main, "_unlocked_master_password", None)
        try:
            needs_master = False
            for it in items:
                if it.get('brand') in ['Basso', 'Lee Cougan']:
                    needs_master = True
                    break
            if needs_master:
                master_password = self.main.get_unlocked_master_password(parent=self)
                if not master_password:
                    self.retry_btn.setVisible(True)  # Show retry button again
                    InfoBar.error(
                        title=self.main.i18n.tr("common.error"),
                        content=self.main.i18n.tr("master.invalid.content"),
                        orient=Qt.Orientation.Horizontal,
                        isClosable=True,
                        position=InfoBarPosition.TOP,
                        duration=4000,
                        parent=self
                    )
                    return
        except Exception:
            pass

        # Initialize status/error columns for retry rows
        for row in failed_rows:
            status_item = self.table.item(row, 6)
            if status_item:
                status_item.setText("Pending")

            error_item = self.table.item(row, 7)
            if error_item:
                error_item.setText("")

        # Check if multi-session is enabled (same as _start_upload)
        multi_session_enabled = self.main.settings.get('multi_session_enabled', False)

        # Start worker thread (with table row indices for proper UI updates)
        if multi_session_enabled and self.session_manager and self.session_manager.is_ready():
            # Use parallel worker with session pool
            self.worker = ParallelBatchUploadWorker(
                items,
                self.session_manager,
                self.main,
                master_password=master_password,
                table_row_indices=failed_rows
            )
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self.main.logger.log)
            self.worker.done.connect(self._on_done)
            self.worker.session_status.connect(self._on_session_status_update)
            self.worker.start()

            # Show session status if we have the UI element
            if hasattr(self, 'session_status_label'):
                self.session_status_label.show()
        else:
            # Use single-session worker
            self.worker = BatchUploadWorker(items, self.main, master_password=master_password, table_row_indices=failed_rows)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self.main.logger.log)
            self.worker.progress_update.connect(self._on_progress_update)
            self.worker.done.connect(self._on_done)
            self.worker.start()

        # Update UI state
        self.start_btn.setEnabled(False)

        # Show progress ring
        if hasattr(self, 'progress_ring'):
            self.progress_ring.start()
            self.progress_ring.show()

        # Start elapsed time timer
        if hasattr(self, 'elapsed_label'):
            self.start_time = __import__('time').time()
            self._elapsed_timer = QTimer()
            self._elapsed_timer.timeout.connect(self._update_elapsed)
            self._elapsed_timer.start(1000)
            self.elapsed_label.show()

        InfoBar.info(
            title=self.main.i18n.tr("batch.run.started.title"),
            content=f"Retrying {len(items)} failed item(s)...",
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self
        )

    def _start_upload(self):
        """Start batch upload"""
        items = []
        table_rows = []  # Track actual table row indices

        # Collect data from table using helper
        for row in range(self.table.rowCount()):
            brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
            code_widget = self._get_widget_from_cell(row, 1, LineEdit)
            url_widget = self._get_widget_from_cell(row, 2, LineEdit)

            if not (brand_widget and code_widget and url_widget):
                continue

            brand = brand_widget.currentText()
            code_raw = code_widget.text().strip()
            code = self._normalize_code(code_raw)
            url = url_widget.text().strip()

            if not (brand in self.brands and code and url):
                continue

            desc_widget = self._get_widget_from_cell(row, 3, ComboBox)
            desc = desc_widget.currentText() if desc_widget else ""

            frameset_checkbox = self._get_widget_from_cell(row, 4, CheckBox)
            frameset = frameset_checkbox.isChecked() if frameset_checkbox else False

            disclaimer_checkbox = self._get_widget_from_cell(row, 5, CheckBox)
            disclaimer = disclaimer_checkbox.isChecked() if disclaimer_checkbox else False

            items.append({
                'brand': brand,
                'code': code,
                'url': url,
                'description_name': desc,
                'frameset_only': frameset,
                'append_disclaimer': disclaimer,
            })
            table_rows.append(row)  # Track which table row this item came from

        if not items:
            return

        # If batch includes brands that use external credentials, ensure master is unlocked
        master_password = getattr(self.main, "_unlocked_master_password", None)
        try:
            needs_master = False
            for it in items:
                if it.get('brand') == 'Basso':
                    if not self.main.credential_manager.has_external_credentials('basso'):
                        InfoBar.error(
                            title=self.main.i18n.tr("common.error"),
                            content=self.main.i18n.tr("account.brand_missing", brand='Basso'),
                            orient=Qt.Orientation.Horizontal,
                            isClosable=True,
                            position=InfoBarPosition.TOP,
                            duration=5000,
                            parent=self
                        )
                        return
                    needs_master = True

                if it.get('brand') == 'Lee Cougan':
                    if not self.main.credential_manager.has_external_credentials('leecougan'):
                        InfoBar.error(
                            title=self.main.i18n.tr("common.error"),
                            content=self.main.i18n.tr("account.brand_missing", brand='Lee Cougan'),
                            orient=Qt.Orientation.Horizontal,
                            isClosable=True,
                            position=InfoBarPosition.TOP,
                            duration=5000,
                            parent=self
                        )
                        return
                    needs_master = True

                if needs_master:
                    break
            if needs_master:
                master_password = self.main.get_unlocked_master_password(parent=self)
                if not master_password:
                    InfoBar.error(
                        title=self.main.i18n.tr("common.error"),
                        content=self.main.i18n.tr("master.invalid.content"),
                        orient=Qt.Orientation.Horizontal,
                        isClosable=True,
                        position=InfoBarPosition.TOP,
                        duration=4000,
                        parent=self
                    )
                    return
        except Exception:
            pass

        # Initialize status/error columns for all valid rows
        row_idx = 0
        for row in range(self.table.rowCount()):
            brand_widget = self._get_widget_from_cell(row, 0, ComboBox)
            code_widget = self._get_widget_from_cell(row, 1, LineEdit)
            url_widget = self._get_widget_from_cell(row, 2, LineEdit)

            if not (brand_widget and code_widget and url_widget):
                continue

            brand = brand_widget.currentText()
            code = code_widget.text().strip()
            url = url_widget.text().strip()

            if brand in self.brands and code and url:
                # Clear status and error for this row
                status_item = self.table.item(row, 6)
                if not status_item:
                    status_item = QTableWidgetItem("")
                    status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    self.table.setItem(row, 6, status_item)
                status_item.setText("Pending")

                error_item = self.table.item(row, 7)
                if not error_item:
                    error_item = QTableWidgetItem("")
                    self.table.setItem(row, 7, error_item)
                error_item.setText("")

                row_idx += 1

        # Check if multi-session is enabled
        multi_session_enabled = self.main.settings.get('multi_session_enabled', False)
        browser_count = self.main.settings.get('browser_count', 2)

        # Initialize session manager if multi-session enabled
        if multi_session_enabled:
            if self.session_manager is None or not self.session_manager.is_ready():
                from Managers.BrowserSessionManager import BrowserSessionManager

                # Get browser type from settings
                browser_type = self.main.settings.get('browser_choice', 'Chrome')

                InfoBar.info(
                    title=self.main.i18n.tr("common.info"),
                    content=f"Initializing {browser_count} browser instances...",
                    orient=Qt.Orientation.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                    parent=self
                )

                # Create and initialize session pool
                self.session_manager = BrowserSessionManager(
                    browser_type=browser_type,
                    pool_size=browser_count,
                    logger=self.main.logger
                )

                if not self.session_manager.initialize_pool():
                    InfoBar.error(
                        title=self.main.i18n.tr("common.error"),
                        content="Failed to initialize browser pool. Falling back to single-session mode.",
                        orient=Qt.Orientation.Horizontal,
                        isClosable=True,
                        position=InfoBarPosition.TOP,
                        duration=5000,
                        parent=self
                    )
                    self.session_manager = None
                    multi_session_enabled = False

        # Start worker thread (with table row indices for proper UI updates)
        if multi_session_enabled and self.session_manager and self.session_manager.is_ready():
            # Use parallel worker with session pool
            self.worker = ParallelBatchUploadWorker(
                items,
                self.session_manager,
                self.main,
                master_password=master_password,
                table_row_indices=table_rows
            )
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self.main.logger.log)
            self.worker.done.connect(self._on_done)
            self.worker.session_status.connect(self._on_session_status_update)
            self.worker.start()

            # Show session status if we have the UI element
            if hasattr(self, 'session_status_label'):
                self.session_status_label.show()
        else:
            # Use single-session worker
            self.worker = BatchUploadWorker(items, self.main, master_password=master_password, table_row_indices=table_rows)
            self.worker.row_update.connect(self._on_row_update)
            self.worker.log.connect(self.main.logger.log)
            self.worker.done.connect(self._on_done)
            self.worker.start()

        # Update UI state
        self.start_btn.setEnabled(False)

        # Show progress ring
        if hasattr(self, 'progress_ring'):
            self.progress_ring.start()
            self.progress_ring.show()

        # Start elapsed time timer
        if hasattr(self, 'elapsed_label'):
            self.start_time = __import__('time').time()
            self._elapsed_timer = QTimer()
            self._elapsed_timer.timeout.connect(self._update_elapsed)
            self._elapsed_timer.start(1000)
            self.elapsed_label.show()

        InfoBar.info(
            title=self.main.i18n.tr("batch.run.started.title"),
            content=self.main.i18n.tr("batch.run.started.content", count=len(items)),
            orient=Qt.Orientation.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=3000,
            parent=self
        )

    def _ensure_combo_placeholder_item(
        self,
        combo: ComboBox,
        placeholder_text: str,
        valid_values: list[str],
        force_placeholder: bool = False,
    ) -> None:
        """Ensure a visible, disabled placeholder item at index 0.

        QComboBox placeholderText behavior can be inconsistent across styles/widgets.
        This approach avoids a blank selectable row and prevents defaulting to a real brand.
        """

        combo.blockSignals(True)
        try:
            current = combo.currentText().strip()
            keep_current = (not force_placeholder) and (current in valid_values)

            combo.clear()
            combo.addItem(placeholder_text)
            combo.addItems(valid_values)

            try:
                model = combo.model()
                if isinstance(model, QStandardItemModel):
                    item = model.item(0)
                    if item is not None:
                        item.setEnabled(False)
            except Exception:
                pass

            if keep_current:
                combo.setCurrentText(current)
            else:
                combo.setCurrentIndex(0)
        finally:
            combo.blockSignals(False)

    def _update_table_theme(self):
        """Update table styling based on current theme"""
        is_dark = isDarkTheme()
        table_colors = COMPONENT_COLORS['table']
        # Use slightly-tinted base rows so white inputs don't blend into the table.
        bg_color = table_colors['row_alt_bg_dark'] if is_dark else table_colors['row_alt_bg_light']
        alt_bg = table_colors['row_bg_dark'] if is_dark else table_colors['row_bg_light']
        border_color = table_colors['border_dark'] if is_dark else table_colors['border_light']

        # Header colors: lavender_grey for dark mode, space_indigo for light mode
        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']

        text_color = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']

        from GUI_Qt.styles.theme_config import (
            get_embedded_input_border,
            get_scrollbar_handle_bg,
            get_scrollbar_handle_hover_bg,
            get_selection_bg,
        )

        # Stronger default borders for inputs embedded inside the table.
        # This avoids "border disappears" issues against light table backgrounds.
        embedded_input_border = get_embedded_input_border(is_dark)
        embedded_input_bg = COMPONENT_COLORS['input']['bg_dark'] if is_dark else COMPONENT_COLORS['input']['bg_light']


        self.table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg_color};
                alternate-background-color: {alt_bg};
                border: 1px solid {border_color};
                border-radius: {RADII['md']}px;
                gridline-color: {border_color};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::viewport {{
                background-color: {bg_color};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableView::item:focus, QAbstractItemView::item:focus {{
                outline: none;
            }}
            QTableWidget QWidget:focus {{
                outline: none;
            }}
            QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-bottom: 1px solid {border_color};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
                letter-spacing: 0.5px;
                text-transform: uppercase;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}
            QHeaderView::section {{
                border-right: 1px solid {border_color};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:horizontal:last {{
                border-top-right-radius: {RADII['md']}px;
            }}

            /* Inputs inside table cells: add contrast + prevent gridline overlap */
            QTableWidget QWidget[ubTableCell="true"] LineEdit,
            QTableWidget QWidget[ubTableCell="true"] PasswordLineEdit,
            QTableWidget QWidget[ubTableCell="true"] SearchLineEdit,
            QTableWidget QWidget[ubTableCell="true"] QLineEdit,
            QTableWidget QWidget[ubTableCell="true"] ComboBox,
            QTableWidget QWidget[ubTableCell="true"] QComboBox {{
                background-color: {embedded_input_bg};
                border: 1px solid {embedded_input_border};
                border-bottom: 2px solid {embedded_input_border};
                border-radius: {RADII['sm']}px;
                padding: {PADDINGS['combo']};
                outline: none;
            }}

            /* Custom scrollbar styling */
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}

            QScrollBar:horizontal {{
                background: transparent;
                height: {SIZES['scrollbar_thickness']}px;
                margin: 0px {SPACING['xxs']}px {SPACING['xxs']}px {SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:horizontal {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-width: {SIZES['scrollbar_handle_min']}px;
                margin: {SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:horizontal:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{
                width: 0px;
                border: none;
            }}
            QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{
                background: none;
            }}
        """)

    def _on_breakpoint_changed(self, breakpoint: str):
        """Respond to breakpoint changes - adjust margins and spacing."""
        # Adjust margins based on breakpoint
        margins = get_responsive_margins(breakpoint)
        spacing = get_responsive_spacing(breakpoint)
        if hasattr(self, 'content_widget') and self.content_widget and self.content_widget.layout():
            self.content_widget.layout().setContentsMargins(*margins)
            self.content_widget.layout().setSpacing(spacing)

        # Note: Table already has horizontal scrolling enabled (ScrollPerPixel)
        # Toolbar stays horizontal at all breakpoints for BatchUploadScreen
        # (too many controls to stack vertically effectively)

    def _on_theme_changed(self):
        """Handle theme change event"""
        apply_screen_theme(
            self,
            "BatchUploadScreen",
            scroll=self.scroll,
            content=self.content_widget
        )

        # Update table theme
        self._update_table_theme()
