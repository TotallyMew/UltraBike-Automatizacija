"""Integrated Pimbo-to-Orbea catalogue, report, and table-image workflow."""

from __future__ import annotations

import json
import re
import threading
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence

from PySide6.QtCore import QTimer, QUrl
from PySide6.QtGui import QColor, QDesktopServices
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QTabBar,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    CheckBox,
    ComboBox,
    FluentIcon,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    PillPushButton,
    PlainTextEdit,
    PrimaryPushButton,
    ProgressBar,
    PushButton,
    ScrollArea,
    TitleLabel,
    FlowLayout,
    isDarkTheme,
    qconfig,
)

from GUI_Qt.styles.screen_theme import (
    CARD_MARGINS,
    CARD_SPACING,
    CONTENT_SPACING,
    ICON_TEXT_GAP,
    PAGE_MARGINS,
    PAGE_SPACING,
    ROW_SPACING,
    apply_screen_theme,
    enforce_transparent_labels,
)
from GUI_Qt.styles.theme_config import (
    COLORS,
    COMPONENT_COLORS,
    FONTS,
    PADDINGS,
    RADII,
    get_accent_colors,
    get_selection_bg,
    get_status_text_color,
    get_subtle_border,
    get_subtle_item_hover_bg,
    get_text_color,
    rgba_from_hex,
)
from GUI_Qt.widgets import enable_table_copy
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget


CATALOGUE_SETTING = "orbea_catalogue_path"
OUTPUT_SETTING = "orbea_output_root"
FILTER_SETTING = "orbea_filter_preset"
DESCRIPTION_OUTPUT_SETTING = "orbea_description_output"
PHOTO_OUTPUT_SETTING = "orbea_photo_output"
TABLE_OUTPUT_SETTING = "orbea_table_output"
DIRECT_GEOMETRY_SETTING = "orbea_direct_geometry_images"
DIRECT_SIZE_GUIDE_SETTING = "orbea_direct_size_guide_image"
DIRECT_PRODUCT_PHOTOS_SETTING = "orbea_direct_product_photos"
TABLE_IMAGES_SETTING = "orbea_download_table_images"
PRODUCT_PHOTOS_SETTING = "orbea_download_product_photos"
PREVIEW_LIMIT = 500


from GUI_Qt.orbea.controller import (
    OrbeaWorkflowController, _create_model, _plain, _read,
)
from GUI_Qt.orbea.workers import (
    OrbeaDescriptionWorker, OrbeaExcelSortWorker, OrbeaFilterWorker,
    OrbeaPhotoWorker, OrbeaRunWorker, OrbeaTableImageWorker,
)
from GUI_Qt.orbea.tabs import OrbeaSectionPage, OrbeaSectionTabs

class OrbeaScreen(ResponsiveWidget):
    """End-to-end Orbea automation UI backed by the authenticated Pimbo driver."""

    def __init__(
        self,
        main_window,
        *,
        settings_manager=None,
        service_factory: Callable[[Any], Any] | None = None,
        image_driver_factory: Callable[[], Any] | None = None,
        description_service_factory: Callable[[], Any] | None = None,
        photo_service_factory: Callable[[], Any] | None = None,
        table_image_service_factory: Callable[[], Any] | None = None,
    ):
        super().__init__(main_window)
        self.main = main_window
        self.settings = settings_manager or getattr(main_window, "settings", None)
        self.tr = main_window.i18n.tr
        self.workflow_controller = OrbeaWorkflowController(
            self,
            service_factory=service_factory,
            image_driver_factory=image_driver_factory,
            description_service_factory=description_service_factory,
            photo_service_factory=photo_service_factory,
            table_image_service_factory=table_image_service_factory,
        )
        self._worker: OrbeaRunWorker | None = None
        self._filter_worker: OrbeaFilterWorker | None = None
        self._description_worker: OrbeaDescriptionWorker | None = None
        self._photo_worker: OrbeaPhotoWorker | None = None
        self._table_image_worker: OrbeaTableImageWorker | None = None
        self._excel_sort_worker: OrbeaExcelSortWorker | None = None
        self._owns_browser_lease = False
        self._restoring_filters = False
        self._closing = False
        self._workbook_path: Path | None = None
        self._run_dir: Path | None = None
        self._description_output_dir: Path | None = None
        self._photo_output_dir: Path | None = None
        self._table_output_dir: Path | None = None
        self._run_operation_id: str | None = None
        self._status_buttons: dict[str, PillPushButton] = {}
        self._stock_buttons: dict[str, PillPushButton] = {}
        self._bucket_buttons: dict[str, PillPushButton] = {}
        self._status_group = None
        self._stock_group = None
        self._bucket_group = None
        self._config_widgets: list[QWidget] = []
        self._description_config_widgets: list[QWidget] = []
        self._photo_config_widgets: list[QWidget] = []
        self._table_image_config_widgets: list[QWidget] = []
        self._saved_filter_state = self._load_filter_state()
        self._auto_refreshed_driver_id: int | None = None

        self.setObjectName("OrbeaScreen")
        self._build_ui()
        self._install_default_filters()
        self._load_paths()
        self.retranslate_ui()
        self._update_action_states()
        qconfig.themeChangedFinished.connect(self._on_theme_changed)
        QTimer.singleShot(0, self._auto_refresh_filters)

    def _on_theme_changed(self):
        apply_screen_theme(
            self, "OrbeaScreen", scroll=self._scroll, content=self._container
        )
        self._subtitle.setStyleSheet(
            f"color: {get_text_color(isDarkTheme(), 'secondary')}; "
            "background: transparent; border: none;"
        )
        self._update_table_theme()
        enforce_transparent_labels(self)

    # ------------------------------------------------------------------ UI

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        self._scroll = ScrollArea()
        self._scroll.setWidgetResizable(True)
        root.addWidget(self._scroll)
        self._container = QWidget()
        self._layout = QVBoxLayout(self._container)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(CONTENT_SPACING)
        self._scroll.setWidget(self._container)
        apply_screen_theme(self, "OrbeaScreen", scroll=self._scroll, content=self._container)

        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)
        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self._title = TitleLabel("")
        self._subtitle = CaptionLabel("")
        self._subtitle.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        title_col.addWidget(self._title)
        title_col.addWidget(self._subtitle)
        header.addLayout(title_col)
        header.addStretch()
        self._layout.addLayout(header)

        self._section_tabs = OrbeaSectionTabs(self)
        self._section_keys = self._section_tabs.KEYS
        self._layout.addWidget(self._section_tabs)

        self._setup_page, self._setup_layout = self._section_page()
        self._progress_page, self._progress_layout = self._section_page()
        self._photos_page, self._photos_layout = self._section_page()
        self._descriptions_page, self._descriptions_layout = self._section_page()
        self._results_page, self._results_layout = self._section_page()

        self._build_paths_card()
        self._build_actions_card()
        self._build_filters_card()
        self._build_progress_card()
        self._build_table_image_card()
        self._build_photo_card()
        self._build_description_card()
        self._build_results_table()

        self._setup_layout.addStretch()
        self._section_pages = {
            "setup": self._setup_page,
            "progress": self._progress_page,
            "photos": self._photos_page,
            "descriptions": self._descriptions_page,
            "results": self._results_page,
        }
        for page in self._section_pages.values():
            self._layout.addWidget(page, 1)

        self._section_tabs.keyChanged.connect(self._switch_section)
        self._switch_section("setup")

        enforce_transparent_labels(self)

    @staticmethod
    def _section_page() -> tuple[QWidget, QVBoxLayout]:
        page = OrbeaSectionPage()
        return page, page.content_layout

    def _switch_section(self, route_key: str) -> None:
        route_key = self._section_tabs.select_key(route_key)
        for key, page in self._section_pages.items():
            page.setVisible(key == route_key)
        self._scroll.verticalScrollBar().setValue(0)

    def _card(self) -> tuple[CardWidget, QVBoxLayout]:
        card = CardWidget()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(*CARD_MARGINS)
        layout.setSpacing(CARD_SPACING)
        return card, layout

    def _build_paths_card(self):
        card, layout = self._card()
        self._paths_title = BodyLabel("")
        self._paths_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        layout.addWidget(self._paths_title)
        grid = QGridLayout()
        grid.setHorizontalSpacing(CARD_SPACING)
        grid.setVerticalSpacing(ROW_SPACING)
        grid.setColumnStretch(1, 1)

        self._catalogue_label = BodyLabel("")
        self._catalogue_edit = LineEdit()
        self._catalogue_edit.setClearButtonEnabled(True)
        self._catalogue_edit.editingFinished.connect(self._paths_changed)
        self._catalogue_btn = PushButton(FluentIcon.DOCUMENT, "")
        self._catalogue_btn.clicked.connect(self._browse_catalogue)
        grid.addWidget(self._catalogue_label, 0, 0)
        grid.addWidget(self._catalogue_edit, 0, 1)
        grid.addWidget(self._catalogue_btn, 0, 2)

        self._output_label = BodyLabel("")
        self._output_edit = LineEdit()
        self._output_edit.setClearButtonEnabled(True)
        self._output_edit.editingFinished.connect(self._paths_changed)
        self._output_btn = PushButton(FluentIcon.FOLDER, "")
        self._output_btn.clicked.connect(self._browse_output)
        grid.addWidget(self._output_label, 1, 0)
        grid.addWidget(self._output_edit, 1, 1)
        grid.addWidget(self._output_btn, 1, 2)

        # The Pimbo query is intentionally fixed and is not an actionable
        # setting, so retain it for configuration/tests without spending a
        # full visible form row on a disabled control.
        self._search_label = BodyLabel("", card)
        self._search_edit = LineEdit(card)
        self._search_edit.setText("orbea")
        self._search_edit.setReadOnly(True)
        self._search_edit.setEnabled(False)
        self._search_label.setVisible(False)
        self._search_edit.setVisible(False)
        layout.addLayout(grid)

        self._downloads_label = BodyLabel("")
        self._downloads_label.setVisible(False)
        layout.addWidget(self._downloads_label)
        download_options = FlowLayout()
        download_options.setHorizontalSpacing(CARD_SPACING)
        download_options.setVerticalSpacing(ROW_SPACING)
        self._table_images_check = CheckBox("")
        self._table_images_check.setChecked(False)
        self._table_images_check.setVisible(False)
        self._table_images_check.stateChanged.connect(
            self._download_options_changed
        )
        self._product_photos_check = CheckBox("")
        self._product_photos_check.setChecked(False)
        self._product_photos_check.setVisible(False)
        self._product_photos_check.stateChanged.connect(
            self._download_options_changed
        )
        download_options.addWidget(self._table_images_check)
        download_options.addWidget(self._product_photos_check)
        layout.addLayout(download_options)
        self._downloads_hint = CaptionLabel("")
        self._downloads_hint.setVisible(False)
        self._downloads_hint.setWordWrap(True)
        self._downloads_hint.setStyleSheet(
            f"color: {get_text_color(isDarkTheme(), 'secondary')};"
        )
        layout.addWidget(self._downloads_hint)

        self._setup_layout.addWidget(card)
        self._config_widgets.extend([
            self._catalogue_edit,
            self._catalogue_btn,
            self._output_edit,
            self._output_btn,
            self._table_images_check,
            self._product_photos_check,
        ])

    def _build_filters_card(self):
        card, layout = self._card()
        top = QHBoxLayout()
        self._filters_title = BodyLabel("")
        self._filters_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        self._filter_state_label = CaptionLabel("")
        self._filter_state_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self._refresh_btn = PushButton(FluentIcon.SYNC, "")
        self._refresh_btn.clicked.connect(self.refresh_filter_options)
        top.addWidget(self._filters_title)
        top.addWidget(self._filter_state_label, 1)
        top.addWidget(self._refresh_btn)
        layout.addLayout(top)

        self._status_label = BodyLabel("")
        layout.addWidget(self._status_label)
        self._status_layout = FlowLayout()
        self._status_layout.setHorizontalSpacing(ROW_SPACING)
        self._status_layout.setVerticalSpacing(ROW_SPACING)
        layout.addLayout(self._status_layout)

        grid = QGridLayout()
        grid.setHorizontalSpacing(CARD_SPACING)
        grid.setVerticalSpacing(ROW_SPACING)
        self._family_combo = ComboBox()
        self._category_combo = ComboBox()
        self._source_combo = ComboBox()
        self._locale_combo = ComboBox()
        self._sort_combo = ComboBox()
        self._family_field, self._family_label = self._combo_field(self._family_combo)
        self._category_field, self._category_label = self._combo_field(self._category_combo)
        self._source_field, self._source_label = self._combo_field(self._source_combo)
        self._locale_field, self._locale_label = self._combo_field(self._locale_combo)
        self._sort_field, self._sort_label = self._combo_field(self._sort_combo)
        grid.addWidget(self._family_field, 0, 0)
        grid.addWidget(self._category_field, 0, 1)
        grid.addWidget(self._source_field, 0, 2)
        grid.addWidget(self._locale_field, 1, 0)
        grid.addWidget(self._sort_field, 1, 1)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(2, 1)
        layout.addLayout(grid)

        self._stock_label = BodyLabel("")
        layout.addWidget(self._stock_label)
        self._stock_layout = FlowLayout()
        self._stock_layout.setHorizontalSpacing(ROW_SPACING)
        self._stock_layout.setVerticalSpacing(ROW_SPACING)
        layout.addLayout(self._stock_layout)

        self._bucket_label = BodyLabel("")
        layout.addWidget(self._bucket_label)
        self._bucket_layout = FlowLayout()
        self._bucket_layout.setHorizontalSpacing(ROW_SPACING)
        self._bucket_layout.setVerticalSpacing(ROW_SPACING)
        layout.addLayout(self._bucket_layout)

        self._setup_layout.addWidget(card)
        combos = [
            self._family_combo,
            self._category_combo,
            self._source_combo,
            self._locale_combo,
            self._sort_combo,
        ]
        for combo in combos:
            combo.currentIndexChanged.connect(self._filter_changed)
        self._config_widgets.extend([self._refresh_btn, *combos])

    def _combo_field(self, combo: ComboBox) -> tuple[QWidget, BodyLabel]:
        field = QWidget()
        box = QVBoxLayout(field)
        box.setContentsMargins(0, 0, 0, 0)
        box.setSpacing(4)
        label = BodyLabel("")
        combo.setMinimumWidth(160)
        box.addWidget(label)
        box.addWidget(combo)
        return field, label

    def _build_actions_card(self):
        card, layout = self._card()
        self._actions_title = BodyLabel("")
        self._actions_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        layout.addWidget(self._actions_title)
        actions = QGridLayout()
        actions.setHorizontalSpacing(ROW_SPACING)
        actions.setVerticalSpacing(ROW_SPACING)
        self._start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._start_btn.setObjectName("orbeaPrimaryAction")
        self._start_btn.clicked.connect(self._on_start_stop)
        self._resume_btn = PushButton(FluentIcon.UPDATE, "")
        self._resume_btn.clicked.connect(
            lambda: self._start_run(resume=True, retry_failed=False, require_existing=True)
        )
        self._retry_btn = PushButton(FluentIcon.SYNC, "")
        self._retry_btn.clicked.connect(
            lambda: self._start_run(resume=True, retry_failed=True, require_existing=True)
        )
        self._open_excel_btn = PushButton(FluentIcon.DOCUMENT, "")
        self._open_excel_btn.setEnabled(False)
        self._open_excel_btn.clicked.connect(self._open_excel)
        self._open_folder_btn = PushButton(FluentIcon.FOLDER, "")
        self._open_folder_btn.setEnabled(False)
        self._open_folder_btn.clicked.connect(self._open_folder)
        self._excel_sort_btn = PushButton(FluentIcon.DOCUMENT, "")
        self._excel_sort_btn.clicked.connect(self._sort_existing_excel)
        actions.addWidget(self._start_btn, 0, 0)
        actions.addWidget(self._resume_btn, 0, 1)
        actions.addWidget(self._retry_btn, 0, 2)
        for button in (self._start_btn, self._resume_btn, self._retry_btn):
            button.setMinimumHeight(38)
        for column in range(3):
            actions.setColumnStretch(column, 1)
        layout.addLayout(actions)
        self._setup_layout.addWidget(card)
        self._config_widgets.append(self._excel_sort_btn)

    def _build_progress_card(self):
        card, layout = self._card()
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        status_row = QHBoxLayout()
        self._stage_label = BodyLabel("")
        self._stage_label.setStyleSheet("font-size: 15px; font-weight: 600;")
        self._eta_label = CaptionLabel("")
        self._eta_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        self._progress_stop_btn = PushButton(FluentIcon.CLOSE, "")
        self._progress_stop_btn.setObjectName("orbeaDangerAction")
        self._progress_stop_btn.clicked.connect(self._on_start_stop)
        self._progress_stop_btn.setVisible(False)
        status_row.addWidget(self._stage_label, 1)
        status_row.addWidget(self._eta_label)
        status_row.addWidget(self._progress_stop_btn)
        layout.addLayout(status_row)
        self._progress = ProgressBar()
        self._progress.setRange(0, 100)
        self._progress.setValue(0)
        layout.addWidget(self._progress)
        self._progress_label = CaptionLabel("")
        self._progress_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._progress_label)

        stats = QGridLayout()
        stats.setHorizontalSpacing(CARD_SPACING)
        stats.setVerticalSpacing(ROW_SPACING)
        self._stat_values: dict[str, BodyLabel] = {}
        self._stat_labels: dict[str, CaptionLabel] = {}
        for index, key in enumerate(("scanned", "matched", "review", "images", "unavailable", "errors")):
            widget = QWidget()
            stat_layout = QVBoxLayout(widget)
            stat_layout.setContentsMargins(8, 4, 8, 4)
            stat_layout.setSpacing(2)
            label = CaptionLabel("")
            label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
            value = BodyLabel("0")
            value.setStyleSheet("font-size: 20px; font-weight: 600;")
            stat_layout.addWidget(label)
            stat_layout.addWidget(value)
            stats.addWidget(widget, index // 3, index % 3)
            self._stat_labels[key] = label
            self._stat_values[key] = value
        layout.addLayout(stats)

        self._log = PlainTextEdit()
        self._log.setReadOnly(True)
        self._log.setMinimumHeight(180)
        layout.addWidget(self._log, 1)
        self._progress_layout.addWidget(card, 1)

    def _build_description_card(self):
        card, layout = self._card()

        self._description_title = BodyLabel("")
        self._description_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        self._description_subtitle = CaptionLabel("")
        self._description_subtitle.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._description_title)
        layout.addWidget(self._description_subtitle)

        self._description_urls_label = BodyLabel("")
        self._description_urls_edit = PlainTextEdit()
        self._description_urls_edit.setMinimumHeight(100)
        self._description_urls_edit.setMaximumHeight(170)
        self._description_urls_edit.textChanged.connect(self._update_action_states)
        layout.addWidget(self._description_urls_label)
        layout.addWidget(self._description_urls_edit)

        output_row = QGridLayout()
        output_row.setHorizontalSpacing(CARD_SPACING)
        output_row.setVerticalSpacing(ROW_SPACING)
        output_row.setColumnStretch(0, 1)
        self._description_output_label = BodyLabel("")
        self._description_output_edit = LineEdit()
        self._description_output_edit.setClearButtonEnabled(True)
        self._description_output_edit.editingFinished.connect(self._description_output_changed)
        self._description_output_btn = PushButton(FluentIcon.FOLDER, "")
        self._description_output_btn.clicked.connect(self._browse_description_output)
        output_row.addWidget(self._description_output_label, 0, 0, 1, 2)
        output_row.addWidget(self._description_output_edit, 1, 0)
        output_row.addWidget(self._description_output_btn, 1, 1)
        layout.addLayout(output_row)

        action_row = QVBoxLayout()
        action_row.setSpacing(ROW_SPACING)
        self._description_start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._description_start_btn.setObjectName("orbeaPrimaryAction")
        self._description_start_btn.clicked.connect(self._on_description_start_stop)
        self._description_open_btn = PushButton(FluentIcon.FOLDER, "")
        self._description_open_btn.setEnabled(False)
        self._description_open_btn.clicked.connect(self._open_description_folder)
        action_row.addWidget(self._description_start_btn)
        action_row.addWidget(self._description_open_btn)
        layout.addLayout(action_row)

        self._description_status_label = BodyLabel("")
        self._description_status_label.setStyleSheet("font-size: 14px; font-weight: 600;")
        layout.addWidget(self._description_status_label)
        self._description_progress = ProgressBar()
        self._description_progress.setRange(0, 100)
        self._description_progress.setValue(0)
        layout.addWidget(self._description_progress)
        self._description_progress_label = CaptionLabel("")
        self._description_progress_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._description_progress_label)

        self._description_log = PlainTextEdit()
        self._description_log.setReadOnly(True)
        self._description_log.setMaximumHeight(90)
        layout.addWidget(self._description_log)

        self._description_config_widgets.extend([
            self._description_urls_edit,
            self._description_output_edit,
            self._description_output_btn,
        ])
        self._descriptions_layout.addWidget(card)
        self._descriptions_layout.addStretch()

    def _build_table_image_card(self):
        card, layout = self._card()

        self._table_image_title = BodyLabel("")
        self._table_image_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        self._table_image_subtitle = CaptionLabel("")
        self._table_image_subtitle.setWordWrap(True)
        self._table_image_subtitle.setStyleSheet(
            f"color: {get_text_color(isDarkTheme(), 'secondary')};"
        )
        layout.addWidget(self._table_image_title)
        layout.addWidget(self._table_image_subtitle)

        self._table_image_url_label = BodyLabel("")
        self._table_image_url_edit = PlainTextEdit()
        self._table_image_url_edit.setMinimumHeight(82)
        self._table_image_url_edit.setMaximumHeight(140)
        self._table_image_url_edit.textChanged.connect(
            self._on_table_image_urls_changed
        )
        layout.addWidget(self._table_image_url_label)
        layout.addWidget(self._table_image_url_edit)
        self._table_image_urls_hint = CaptionLabel("")
        self._table_image_urls_hint.setWordWrap(True)
        self._table_image_urls_hint.setStyleSheet(
            f"color: {get_text_color(isDarkTheme(), 'secondary')};"
        )
        layout.addWidget(self._table_image_urls_hint)

        self._table_image_types_label = BodyLabel("")
        layout.addWidget(self._table_image_types_label)
        image_types = FlowLayout()
        image_types.setHorizontalSpacing(CARD_SPACING)
        image_types.setVerticalSpacing(ROW_SPACING)
        self._table_geometry_check = CheckBox("")
        self._table_geometry_check.setChecked(True)
        self._table_size_guide_check = CheckBox("")
        self._table_size_guide_check.setChecked(True)
        self._table_product_photos_check = CheckBox("")
        self._table_product_photos_check.setChecked(False)
        for checkbox in (
            self._table_geometry_check,
            self._table_size_guide_check,
            self._table_product_photos_check,
        ):
            checkbox.stateChanged.connect(self._table_image_options_changed)
            image_types.addWidget(checkbox)
        layout.addLayout(image_types)

        output_row = QGridLayout()
        output_row.setHorizontalSpacing(CARD_SPACING)
        output_row.setVerticalSpacing(ROW_SPACING)
        output_row.setColumnStretch(0, 1)
        self._table_image_output_label = BodyLabel("")
        self._table_image_output_edit = LineEdit()
        self._table_image_output_edit.setClearButtonEnabled(True)
        self._table_image_output_edit.editingFinished.connect(
            self._table_image_output_changed
        )
        self._table_image_output_btn = PushButton(FluentIcon.FOLDER, "")
        self._table_image_output_btn.clicked.connect(
            self._browse_table_image_output
        )
        output_row.addWidget(self._table_image_output_label, 0, 0, 1, 2)
        output_row.addWidget(self._table_image_output_edit, 1, 0)
        output_row.addWidget(self._table_image_output_btn, 1, 1)
        layout.addLayout(output_row)

        action_row = QHBoxLayout()
        action_row.setSpacing(ROW_SPACING)
        self._table_image_start_btn = PrimaryPushButton(FluentIcon.DOWNLOAD, "")
        self._table_image_start_btn.setObjectName("orbeaPrimaryAction")
        self._table_image_start_btn.clicked.connect(
            self._on_table_image_start_stop
        )
        self._table_image_open_btn = PushButton(FluentIcon.FOLDER, "")
        self._table_image_open_btn.setEnabled(False)
        self._table_image_open_btn.clicked.connect(
            self._open_table_image_folder
        )
        action_row.addWidget(self._table_image_start_btn)
        action_row.addWidget(self._table_image_open_btn)
        action_row.addStretch()
        layout.addLayout(action_row)

        self._table_image_status_label = BodyLabel("")
        self._table_image_status_label.setStyleSheet(
            "font-size: 14px; font-weight: 600;"
        )
        layout.addWidget(self._table_image_status_label)
        self._table_image_progress = ProgressBar()
        self._table_image_progress.setRange(0, 100)
        self._table_image_progress.setValue(0)
        layout.addWidget(self._table_image_progress)
        self._table_image_progress_label = CaptionLabel("")
        self._table_image_progress_label.setWordWrap(True)
        self._table_image_progress_label.setStyleSheet(
            f"color: {get_text_color(isDarkTheme(), 'secondary')};"
        )
        layout.addWidget(self._table_image_progress_label)

        self._table_image_log = PlainTextEdit()
        self._table_image_log.setReadOnly(True)
        self._table_image_log.setMinimumHeight(80)
        self._table_image_log.setMaximumHeight(120)
        self._table_image_log.setVisible(False)
        layout.addWidget(self._table_image_log)

        self._table_image_config_widgets.extend(
            [
                self._table_image_url_edit,
                self._table_image_output_edit,
                self._table_image_output_btn,
                self._table_geometry_check,
                self._table_size_guide_check,
                self._table_product_photos_check,
            ]
        )
        self._photos_layout.addWidget(card)

    def _build_photo_card(self):
        card, layout = self._card()
        self._photo_card = card

        self._photo_title = BodyLabel("")
        self._photo_title.setStyleSheet("font-size: 15px; font-weight: 600;")
        self._photo_subtitle = CaptionLabel("")
        self._photo_subtitle.setWordWrap(True)
        self._photo_subtitle.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._photo_title)
        layout.addWidget(self._photo_subtitle)

        self._photo_url_label = BodyLabel("")
        self._photo_url_edit = PlainTextEdit()
        self._photo_url_edit.setMinimumHeight(92)
        self._photo_url_edit.setMaximumHeight(160)
        self._photo_url_edit.textChanged.connect(self._on_photo_urls_changed)
        layout.addWidget(self._photo_url_label)
        layout.addWidget(self._photo_url_edit)
        self._photo_urls_hint = CaptionLabel("")
        self._photo_urls_hint.setWordWrap(True)
        self._photo_urls_hint.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._photo_urls_hint)

        output_row = QGridLayout()
        output_row.setHorizontalSpacing(CARD_SPACING)
        output_row.setVerticalSpacing(ROW_SPACING)
        output_row.setColumnStretch(0, 1)
        self._photo_output_label = BodyLabel("")
        self._photo_output_edit = LineEdit()
        self._photo_output_edit.setClearButtonEnabled(True)
        self._photo_output_edit.editingFinished.connect(self._photo_output_changed)
        self._photo_output_btn = PushButton(FluentIcon.FOLDER, "")
        self._photo_output_btn.clicked.connect(self._browse_photo_output)
        output_row.addWidget(self._photo_output_label, 0, 0, 1, 2)
        output_row.addWidget(self._photo_output_edit, 1, 0)
        output_row.addWidget(self._photo_output_btn, 1, 1)
        layout.addLayout(output_row)

        action_row = QHBoxLayout()
        action_row.setSpacing(ROW_SPACING)
        self._photo_start_btn = PrimaryPushButton(FluentIcon.DOWNLOAD, "")
        self._photo_start_btn.setObjectName("orbeaPrimaryAction")
        self._photo_start_btn.clicked.connect(self._on_photo_start_stop)
        self._photo_open_btn = PushButton(FluentIcon.FOLDER, "")
        self._photo_open_btn.setEnabled(False)
        self._photo_open_btn.clicked.connect(self._open_photo_folder)
        action_row.addWidget(self._photo_start_btn)
        action_row.addWidget(self._photo_open_btn)
        action_row.addStretch()
        layout.addLayout(action_row)

        self._photo_status_label = BodyLabel("")
        self._photo_status_label.setStyleSheet("font-size: 14px; font-weight: 600;")
        layout.addWidget(self._photo_status_label)
        self._photo_progress = ProgressBar()
        self._photo_progress.setRange(0, 100)
        self._photo_progress.setValue(0)
        layout.addWidget(self._photo_progress)
        self._photo_progress_label = CaptionLabel("")
        self._photo_progress_label.setWordWrap(True)
        self._photo_progress_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        layout.addWidget(self._photo_progress_label)

        self._photo_log = PlainTextEdit()
        self._photo_log.setReadOnly(True)
        self._photo_log.setMinimumHeight(150)
        layout.addWidget(self._photo_log, 1)

        self._photo_config_widgets.extend(
            [self._photo_url_edit, self._photo_output_edit, self._photo_output_btn]
        )
        # Product photography is intentionally not part of this workflow.
        # Keep the existing controls constructed for backward compatibility
        # with saved settings and older integrations, but do not expose a
        # second, competing downloader in the table-only interface.
        card.setVisible(False)
        self._photos_layout.addWidget(card)

    def _build_results_table(self):
        self._results_label = CaptionLabel("")
        self._results_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'secondary')};")
        toolbar = QGridLayout()
        toolbar.setHorizontalSpacing(ROW_SPACING)
        toolbar.setVerticalSpacing(ROW_SPACING)
        toolbar.addWidget(self._results_label, 0, 0, 1, 4)
        toolbar.addWidget(self._excel_sort_btn, 1, 1)
        toolbar.addWidget(self._open_excel_btn, 1, 2)
        toolbar.addWidget(self._open_folder_btn, 1, 3)
        toolbar.setColumnStretch(0, 1)
        self._results_layout.addLayout(toolbar)
        self._table = QTableWidget(0, 6)
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        self._table.setShowGrid(False)
        self._table.setWordWrap(False)
        self._table.verticalHeader().setDefaultSectionSize(36)
        self._table.horizontalHeader().setMinimumHeight(42)
        self._table.horizontalHeader().setHighlightSections(False)
        self._table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self._table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self._table.setMinimumHeight(420)
        self._table.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        enable_table_copy(self._table)
        self._results_layout.addWidget(self._table, 1)
        self._update_table_theme()

    # -------------------------------------------------------------- Filters

    def _install_default_filters(self):
        defaults = {
            "statuses": [("Draft", "Draft"), ("In Review", "In Review"), ("Published", "Published"), ("Disabled", "Disabled")],
            "families": [],
            "categories": [],
            "sources": [],
            "stock": [("Any", "Any"), ("In stock", "In stock"), ("Out of stock", "Out of stock")],
            "locales": [("Overall", "Overall"), ("LT", "LT"), ("EN", "EN"), ("LV", "LV"), ("EE", "EE")],
            "buckets": [("<40%", "<40%"), ("40–80%", "40–80%"), ("≥80%", "≥80%"), ("100%", "100%")],
            "sort": [("Recent", "Recent"), ("Least complete", "Least complete"), ("Most complete", "Most complete")],
        }
        self._apply_filter_options(defaults)

    def _normalise_options(self, raw: Any) -> list[tuple[str, Any]]:
        if raw is None:
            return []
        if isinstance(raw, Mapping):
            raw = list(raw.items())
        result: list[tuple[str, Any]] = []
        for item in raw:
            if isinstance(item, (tuple, list)) and len(item) >= 2:
                label, value = item[0], item[1]
            elif isinstance(item, str):
                label = value = item
            else:
                label = _read(item, "label", "name", "text", default="")
                value = _read(item, "value", "id", "key", default=label)
            label = str(label or value or "").strip()
            if label:
                result.append((label, _plain(value)))
        return result

    def _options(self, source: Any, *names: str) -> list[tuple[str, Any]]:
        return self._normalise_options(_read(source, *names, default=[]))

    def _apply_filter_options(self, options: Any):
        state = self._collect_filter_state() if self._status_buttons else dict(self._saved_filter_state)
        self._restoring_filters = True
        try:
            statuses = self._options(options, "statuses", "status_options") or self._normalise_options([
                ("Draft", "Draft"), ("In Review", "In Review"), ("Published", "Published"), ("Disabled", "Disabled")
            ])
            stock = self._options(options, "stock", "stock_options") or self._normalise_options([
                ("Any", "Any"), ("In stock", "In stock"), ("Out of stock", "Out of stock")
            ])
            buckets = self._options(options, "buckets", "completeness_buckets", "bucket_options") or self._normalise_options([
                ("<40%", "<40%"), ("40–80%", "40–80%"), ("≥80%", "≥80%"), ("100%", "100%")
            ])
            self._rebuild_multi_chips(self._status_layout, "_status_buttons", statuses, state.get("statuses", ["Draft"]))
            self._stock_group = self._rebuild_single_chips(self._stock_layout, "_stock_buttons", stock, state.get("stock", "In stock"))
            selected_buckets = state.get("completeness_buckets")
            if selected_buckets is None:
                legacy_bucket = state.get("completeness_bucket")
                selected_buckets = [legacy_bucket] if legacy_bucket and legacy_bucket != "any" else []
            self._rebuild_multi_chips(self._bucket_layout, "_bucket_buttons", buckets, selected_buckets)

            self._populate_combo(self._family_combo, self._options(options, "families", "family_options"), "All families", state.get("family_id"))
            self._populate_combo(self._category_combo, self._options(options, "categories", "category_options"), "All categories", state.get("category_id"))
            self._populate_combo(self._source_combo, self._options(options, "sources", "source_options"), "All sources", state.get("source_id"))
            locales = self._options(options, "locales", "completeness_locales", "locale_options") or self._normalise_options([
                ("Overall", "Overall"), ("LT", "LT"), ("EN", "EN"), ("LV", "LV"), ("EE", "EE")
            ])
            sorts = self._options(options, "sort", "sorts", "sort_options") or self._normalise_options([
                ("Recent", "Recent"), ("Least complete", "Least complete"), ("Most complete", "Most complete")
            ])
            self._populate_combo(self._locale_combo, locales, None, state.get("completeness_locale", "Overall"))
            self._populate_combo(self._sort_combo, sorts, None, state.get("sort", "Recent"))
        finally:
            self._restoring_filters = False
        self._save_filter_state()
        self._style_filter_buttons()

    def _clear_chip_layout(self, layout, attr_name: str):
        buttons = getattr(self, attr_name, {})
        for button in buttons.values():
            layout.removeWidget(button)
            button.deleteLater()
            if button in self._config_widgets:
                self._config_widgets.remove(button)
        setattr(self, attr_name, {})

    def _rebuild_multi_chips(self, layout, attr_name, options, selected):
        self._clear_chip_layout(layout, attr_name)
        selected_keys = {str(_plain(value)).lower() for value in (selected or [])}
        buttons: dict[str, PillPushButton] = {}
        for label, value in options:
            key = str(_plain(value))
            button = PillPushButton()
            button.setText(label)
            button.setCheckable(True)
            button.setProperty("filterValue", value)
            button.setChecked(key.lower() in selected_keys or label.lower() in selected_keys)
            button.toggled.connect(self._filter_changed)
            layout.insertWidget(layout.count(), button)
            buttons[key] = button
            self._config_widgets.append(button)
        setattr(self, attr_name, buttons)

    def _rebuild_single_chips(self, layout, attr_name, options, selected):
        self._clear_chip_layout(layout, attr_name)
        group = QButtonGroup(self)
        group.setExclusive(True)
        selected_key = str(_plain(selected)).lower()
        buttons: dict[str, PillPushButton] = {}
        first = None
        for label, value in options:
            key = str(_plain(value))
            button = PillPushButton()
            button.setText(label)
            button.setCheckable(True)
            button.setProperty("filterValue", value)
            button.toggled.connect(self._filter_changed)
            group.addButton(button)
            layout.insertWidget(layout.count(), button)
            buttons[key] = button
            self._config_widgets.append(button)
            first = first or button
            if key.lower() == selected_key or label.lower() == selected_key:
                button.setChecked(True)
        if group.checkedButton() is None and first is not None:
            first.setChecked(True)
        setattr(self, attr_name, buttons)
        return group

    def _populate_combo(self, combo: ComboBox, options, all_label: str | None, selected):
        combo.clear()
        if all_label is not None:
            combo.addItem(all_label, userData=None)
        for label, value in options:
            if all_label is not None and (value in (None, "") or label.lower().startswith("all ")):
                continue
            combo.addItem(label, userData=value)
        target = str(_plain(selected)) if selected is not None else None
        if target is not None:
            for index in range(combo.count()):
                if (
                    str(_plain(combo.itemData(index))) == target
                    or combo.itemText(index).strip().lower() == target.strip().lower()
                ):
                    combo.setCurrentIndex(index)
                    break
        if combo.count() and combo.currentIndex() < 0:
            combo.setCurrentIndex(0)

    def _checked_value(self, group: QButtonGroup | None, default=None):
        button = group.checkedButton() if group else None
        return _plain(button.property("filterValue")) if button else default

    def _collect_filter_state(self) -> dict[str, Any]:
        return {
            "statuses": [_plain(button.property("filterValue")) for button in self._status_buttons.values() if button.isChecked()],
            "family_id": _plain(self._family_combo.currentData()) if hasattr(self, "_family_combo") else None,
            "category_id": _plain(self._category_combo.currentData()) if hasattr(self, "_category_combo") else None,
            "source_id": _plain(self._source_combo.currentData()) if hasattr(self, "_source_combo") else None,
            "stock": self._checked_value(self._stock_group, "In stock"),
            "completeness_locale": _plain(self._locale_combo.currentData()) if hasattr(self, "_locale_combo") else "Overall",
            "completeness_buckets": [
                _plain(button.property("filterValue"))
                for button in self._bucket_buttons.values()
                if button.isChecked()
            ],
            "sort": _plain(self._sort_combo.currentData()) if hasattr(self, "_sort_combo") else "Recent",
        }

    def _filter_changed(self, *_args):
        if not self._restoring_filters:
            self._save_filter_state()

    def _load_filter_state(self) -> dict[str, Any]:
        raw = self._setting_get(FILTER_SETTING, "")
        try:
            value = json.loads(raw) if raw else {}
            return value if isinstance(value, dict) else {}
        except Exception:
            return {}

    def _save_filter_state(self):
        if self._restoring_filters:
            return
        try:
            self._setting_set(FILTER_SETTING, json.dumps(self._collect_filter_state(), ensure_ascii=False))
        except Exception:
            pass

    # --------------------------------------------------------------- Service

    def _make_service(self, driver):
        return self.workflow_controller.make_service(driver)

    def _create_run_config(self):
        from tools.orbea_automation import OrbeaRunConfig, PimboFilterSpec

        state = self._collect_filter_state()
        filter_values = {
            **state,
            "statuses": tuple(state.get("statuses") or ()),
            "family_id": state.get("family_id") or "",
            "category_id": state.get("category_id") or "",
            "source_id": state.get("source_id") or "",
            "stock": state.get("stock") or "Any",
            "completeness_locale": state.get("completeness_locale") or "Overall",
            "completeness_buckets": tuple(state.get("completeness_buckets") or ()),
            "sort": state.get("sort") or "Recent",
        }
        filter_aliases = {
            "status": "statuses",
            "status_values": "statuses",
            "family": "family_id",
            "family_value": "family_id",
            "category": "category_id",
            "category_value": "category_id",
            "source": "source_id",
            "source_value": "source_id",
            "stock_status": "stock",
            "locale": "completeness_locale",
            "completeness": "completeness_bucket",
            "sort_order": "sort",
        }
        filters = _create_model(PimboFilterSpec, filter_values, filter_aliases)
        config_values = {
            "catalogue_path": Path(self._catalogue_edit.text().strip()),
            "output_root": Path(self._output_edit.text().strip()),
            "filters": filters,
            "all_products": True,
            # Image work has its own URL-driven table downloader. A Pimbo
            # catalogue scan must never start image downloads implicitly.
            "download_images": False,
            "download_product_photos": False,
            "browser_name": str(self._setting_get("browser_choice", "Chrome") or "Chrome").strip().lower(),
        }
        config_aliases = {
            "catalogue": "catalogue_path",
            "catalog_path": "catalogue_path",
            "output_dir": "output_root",
            "filter_spec": "filters",
        }
        return _create_model(OrbeaRunConfig, config_values, config_aliases)

    def _make_description_service(self):
        return self.workflow_controller.make_description_service()

    def _make_photo_service(self):
        return self.workflow_controller.make_photo_service()

    def _make_table_image_service(self):
        return self.workflow_controller.make_table_image_service()

    def _table_image_link_state(
        self,
    ) -> tuple[tuple[str, ...], int, tuple[str, ...], tuple[str, ...]]:
        from tools.orbea_automation import (
            normalize_orbea_table_url,
            unique_orbea_table_urls,
        )

        entries: list[str] = []
        invalid: list[str] = []
        for line in self._table_image_url_edit.toPlainText().splitlines():
            value = line.strip().rstrip(".\"'()[]{}<>")
            if not value:
                continue
            try:
                entries.append(normalize_orbea_table_url(value))
            except (TypeError, ValueError):
                invalid.append(line.strip())
        unique, duplicates = unique_orbea_table_urls(entries)
        return unique, len(duplicates), tuple(invalid), tuple(entries)

    def _table_image_urls(self) -> tuple[str, ...]:
        return self._table_image_link_state()[0]

    def _table_image_selection(self) -> tuple[bool, bool, bool]:
        return (
            self._table_geometry_check.isChecked(),
            self._table_size_guide_check.isChecked(),
            self._table_product_photos_check.isChecked(),
        )

    def _table_image_options_changed(self, _state=None):
        geometry, size_guide, product_photos = self._table_image_selection()
        self._setting_set(DIRECT_GEOMETRY_SETTING, geometry)
        self._setting_set(DIRECT_SIZE_GUIDE_SETTING, size_guide)
        self._setting_set(DIRECT_PRODUCT_PHOTOS_SETTING, product_photos)
        self._update_action_states()

    def _on_table_image_urls_changed(self):
        urls, duplicates, invalid, _entries = self._table_image_link_state()
        if duplicates == 1:
            duplicate_summary = self._t(
                "orbea.tables.urls.duplicate.one", "1 duplicate ignored"
            )
        else:
            duplicate_summary = self._t(
                "orbea.tables.urls.duplicate.many",
                "{count:,} duplicates ignored",
                count=duplicates,
            )
        if invalid:
            summary = self._t(
                "orbea.tables.urls.summary.invalid",
                "Products: {unique:,} • {duplicate_summary} • Invalid lines: {invalid:,}",
                unique=len(urls),
                duplicate_summary=duplicate_summary,
                invalid=len(invalid),
            )
        elif urls:
            summary = self._t(
                "orbea.tables.urls.summary",
                "Products: {unique:,} • {duplicate_summary}",
                unique=len(urls),
                duplicate_summary=duplicate_summary,
            )
        else:
            summary = self._t(
                "orbea.tables.urls.hint",
                "Paste one Orbea product-page URL per line. No Pimbo scan is used.",
            )
        self._table_image_urls_hint.setText(summary)
        self._update_action_states()

    def _photo_link_state(
        self,
    ) -> tuple[tuple[str, ...], int, tuple[str, ...], tuple[str, ...]]:
        from tools.orbea_automation import (
            normalize_orbea_product_url,
            unique_orbea_product_urls,
        )

        entries: list[str] = []
        invalid: list[str] = []
        for line in self._photo_url_edit.toPlainText().splitlines():
            value = line.strip().rstrip(".\"'()[]{}<>")
            if not value:
                continue
            try:
                entries.append(normalize_orbea_product_url(value))
            except (TypeError, ValueError):
                invalid.append(line.strip())
        unique, duplicates = unique_orbea_product_urls(entries)
        return unique, len(duplicates), tuple(invalid), tuple(entries)

    def _photo_urls(self) -> tuple[str, ...]:
        return self._photo_link_state()[0]

    def _photo_url(self) -> str:
        urls = self._photo_urls()
        return urls[0] if urls else ""

    def _on_photo_urls_changed(self):
        urls, duplicates, invalid, _entries = self._photo_link_state()
        duplicate_summary = (
            self._t("orbea.photo.urls.duplicate.one", "1 duplicate ignored")
            if duplicates == 1
            else self._t(
                "orbea.photo.urls.duplicate.many",
                "{count:,} duplicates ignored",
                count=duplicates,
            )
        )
        if invalid:
            summary = self._t(
                "orbea.photo.urls.summary.invalid",
                "Products: {unique:,} • {duplicate_summary} • Invalid lines: {invalid:,}",
                unique=len(urls),
                duplicate_summary=duplicate_summary,
                invalid=len(invalid),
            )
        elif urls:
            summary = self._t(
                "orbea.photo.urls.summary",
                "Products: {unique:,} • {duplicate_summary}",
                unique=len(urls),
                duplicate_summary=duplicate_summary,
            )
        else:
            summary = self._t(
                "orbea.photo.urls.hint",
                "Enter one product URL per line. Duplicate links are ignored automatically.",
            )
        self._photo_urls_hint.setText(summary)
        self._update_action_states()

    def _description_urls(self) -> tuple[str, ...]:
        candidates = re.findall(
            r"https?://[^\s,;]+",
            self._description_urls_edit.toPlainText(),
            flags=re.IGNORECASE,
        )
        urls: list[str] = []
        seen: set[str] = set()
        for candidate in candidates:
            url = candidate.rstrip(".\"'()[]{}<>")
            if not re.match(r"^https?://(?:cms|www)\.orbea\.com/[^?#\s]+", url, re.IGNORECASE):
                continue
            key = url.lower()
            if key not in seen:
                urls.append(url)
                seen.add(key)
        return tuple(urls)

    def _create_description_config(self):
        from tools.orbea_automation import DescriptionRunConfig

        values = {
            "urls": self._description_urls(),
            "output_dir": Path(self._description_output_edit.text().strip()),
            "browser_name": str(self._setting_get("browser_choice", "Chrome") or "Chrome").strip().lower(),
            "show_browser": False,
            "headless": True,
        }
        aliases = {
            "orbea_urls": "urls",
            "input_urls": "urls",
            "output_root": "output_dir",
            "destination": "output_dir",
            "browser": "browser_name",
        }
        return _create_model(DescriptionRunConfig, values, aliases)

    def _auto_refresh_filters(self):
        driver = getattr(self.main, "driver", None)
        if (
            driver is not None
            and id(driver) != self._auto_refreshed_driver_id
            and not self.is_running()
        ):
            self._auto_refreshed_driver_id = id(driver)
            self.refresh_filter_options(show_errors=False)

    def on_activated(self) -> None:
        """Finish deferred browser setup after a login-time screen preload."""
        self._auto_refresh_filters()

    def refresh_filter_options(self, *_args, show_errors=True):
        if self.is_running():
            return
        # ``shutdown()`` is also used during logout. A later login may reuse
        # this lazily-created screen, so a new user action re-enables callbacks.
        self._closing = False
        driver = getattr(self.main, "driver", None)
        if driver is None:
            if show_errors:
                self._warn(self._t("common.error", "Not connected"), self._t("batchdesc.no_session", "Log in to Pimbo first."))
            return
        if not self._acquire_browser():
            self._warn(self._t("orbea.browser.busy.title", "Browser busy"), self._t("orbea.browser.busy", "Another tool is using Pimbo."))
            return
        self._filter_state_label.setText(self._t("orbea.filters.loading", "Loading Pimbo filters…"))
        self._refresh_btn.setEnabled(False)
        self._filter_worker = OrbeaFilterWorker(driver, self._make_service)
        self._filter_worker.loaded.connect(self._filters_loaded)
        self._filter_worker.failed.connect(lambda message: self._filters_failed(message, show_errors))
        self._filter_worker.finished.connect(self._filter_finished)
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._filter_worker, "orbea", "orbea", stage="Loading Pimbo filters"
            )
        self._filter_worker.start()
        self._update_action_states()

    def _filters_loaded(self, options):
        if self._closing:
            return
        self._apply_filter_options(options)
        self._filter_state_label.setText(self._t("orbea.filters.loaded", "Filters loaded from Pimbo"))

    def _filters_failed(self, message: str, show_errors: bool):
        self._filter_state_label.setText(self._t("orbea.filters.defaults", "Using saved/default filters"))
        self._append_log(f"Filter discovery: {message}")
        if show_errors and not self._closing:
            self._warn(self._t("orbea.filters.error.title", "Could not load filters"), message)

    def _filter_finished(self):
        self._filter_worker = None
        self._refresh_btn.setEnabled(True)
        self._release_browser()
        self._update_action_states()

    # --------------------------------------------------------------- Running

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            self._progress_stop_btn.setEnabled(False)
            self._stage_label.setText(self._t("orbea.stopping", "Stopping safely…"))
            return
        # The primary action automatically resumes the newest compatible run.
        self._start_run(resume=True, retry_failed=False)

    def _start_run(
        self,
        *,
        resume: bool,
        retry_failed: bool,
        require_existing: bool = False,
    ):
        if self.is_running() or not self._validate_inputs():
            return
        self._closing = False
        driver = getattr(self.main, "driver", None)
        if driver is None:
            self._warn(self._t("common.error", "Not connected"), self._t("batchdesc.no_session", "Log in to Pimbo first."))
            return
        try:
            config = self._create_run_config()
        except Exception as exc:
            self._error(self._t("orbea.service.error.title", "Orbea service unavailable"), str(exc))
            return
        if require_existing:
            try:
                from tools.orbea_automation import find_latest_compatible_run

                existing = find_latest_compatible_run(
                    config,
                    include_completed_errors=retry_failed,
                )
            except Exception as exc:
                self._error(self._t("orbea.service.error.title", "Orbea service unavailable"), str(exc))
                return
            if existing is None:
                self._warn(
                    self._t("orbea.resume.none.title", "Nothing to resume"),
                    self._t(
                        "orbea.retry.none" if retry_failed else "orbea.resume.none",
                        "No compatible run with retryable errors was found."
                        if retry_failed
                        else "No compatible incomplete run was found.",
                    ),
                )
                return
        if not self._acquire_browser():
            self._warn(self._t("orbea.browser.busy.title", "Browser busy"), self._t("orbea.browser.busy", "Another tool is using Pimbo."))
            return

        self._save_paths()
        self._save_filter_state()
        self._table.setRowCount(0)
        self._workbook_path = None
        self._run_dir = None
        self._open_excel_btn.setEnabled(False)
        self._open_folder_btn.setEnabled(False)
        self._log.clear()
        self._progress.setValue(0)
        self._switch_section("progress")
        self._set_busy(True)
        self._worker = OrbeaRunWorker(
            driver,
            self._make_service,
            config,
            resume=resume,
            retry_failed=retry_failed,
        )
        self._worker.progress_changed.connect(self._on_progress)
        self._worker.log_message.connect(self._append_log)
        self._worker.succeeded.connect(self._on_result)
        self._worker.failed.connect(self._on_run_error)
        self._worker.finished.connect(self._run_thread_finished)
        if hasattr(self.main, "track_worker"):
            operation = self.main.track_worker(
                self._worker,
                "orbea",
                "orbea",
                output_path=str(config.output_root),
                resume_kind="orbea_checkpoint",
                resume_ref=str(config.output_root),
            )
            self._run_operation_id = operation.id
        self._worker.start()

    def _on_progress(self, update):
        stage = str(_plain(_read(update, "stage", "phase", default="Running")) or "Running")
        message = str(_read(update, "message", "detail", default="") or "")
        current = int(_read(update, "current", "done", default=0) or 0)
        total = int(_read(update, "total", default=0) or 0)
        eta = _read(update, "eta_seconds", "eta", default=None)
        self._stage_label.setText(stage.replace("_", " ").title())
        self._progress_label.setText(message or (f"{current:,} / {total:,}" if total else f"{current:,}"))
        self._progress.setValue(max(0, min(100, int(current * 100 / total)))) if total else self._progress.setValue(0)
        self._eta_label.setText(self._format_eta(eta))
        self._update_counts(_read(update, "counts", default={}) or {})

    def _on_result(self, result):
        workbook = _read(result, "workbook_path", "excel_path")
        run_dir = _read(result, "run_dir", "output_dir")
        self._workbook_path = Path(workbook) if workbook else None
        self._run_dir = Path(run_dir) if run_dir else (self._workbook_path.parent if self._workbook_path else None)
        self._update_counts(_read(result, "counts", default={}) or {})
        if self._workbook_path and self._workbook_path.exists():
            self._load_workbook_preview(self._workbook_path)
        cancelled = bool(_read(result, "cancelled", default=False))
        completed = bool(_read(result, "completed", default=not cancelled))
        if self._run_operation_id and hasattr(self.main, "operation_tracker"):
            status = "cancelled" if cancelled else ("succeeded" if completed else "partial")
            checkpoint = _read(result, "checkpoint_path", default="")
            self.main.operation_tracker.update(
                self._run_operation_id,
                resume_ref=str(checkpoint or self._run_dir or ""),
            )
            self.main.operation_tracker.finish(
                self._run_operation_id,
                status,
                output_path=str(self._run_dir or self._workbook_path or ""),
                summary=dict(_read(result, "counts", default={}) or {}),
            )
        if cancelled:
            self._stage_label.setText(self._t("orbea.cancelled", "Stopped — partial report saved"))
        elif completed:
            self._stage_label.setText(self._t("orbea.complete", "Complete"))
            self._progress.setValue(100)
        else:
            self._stage_label.setText(self._t("orbea.incomplete", "Incomplete — ready to resume"))
        self._switch_section("results")

    def _on_run_error(self, message: str):
        if self._run_operation_id and hasattr(self.main, "operation_tracker"):
            self.main.operation_tracker.finish(
                self._run_operation_id, "failed", error_summary=message
            )
        self._stage_label.setText(self._t("orbea.failed", "Run failed — progress was checkpointed"))
        self._append_log(message)
        if not self._closing:
            self._error(self._t("orbea.failed.title", "Orbea automation failed"), message)

    def _run_thread_finished(self):
        self._worker = None
        self._set_busy(False)
        self._release_browser()
        self._open_excel_btn.setEnabled(bool(self._workbook_path and self._workbook_path.exists()))
        self._open_folder_btn.setEnabled(bool(self._run_dir and self._run_dir.exists()))
        self._update_action_states()

    def _sort_existing_excel(self):
        if self.is_running():
            return
        start = (
            self._workbook_path.parent
            if self._workbook_path and self._workbook_path.exists()
            else self._desktop_dir()
        )
        path, _ = QFileDialog.getOpenFileName(
            self,
            self._t("orbea.excel_sort.pick", "Select Orbea match Excel"),
            str(start),
            "Excel (*.xlsx)",
        )
        if not path:
            return

        self._excel_sort_btn.setEnabled(False)
        self._switch_section("progress")
        self._stage_label.setText(
            self._t("orbea.excel_sort.running", "Sorting existing Excel…")
        )
        self._append_log(f"Sorting existing Excel: {path}")
        self._excel_sort_worker = OrbeaExcelSortWorker(Path(path))
        self._excel_sort_worker.succeeded.connect(self._on_excel_sorted)
        self._excel_sort_worker.failed.connect(self._on_excel_sort_error)
        self._excel_sort_worker.finished.connect(self._excel_sort_finished)
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._excel_sort_worker,
                "orbea",
                "orbea",
                output_path=str(Path(path).parent),
            )
        self._excel_sort_worker.start()
        self._update_action_states()

    def _on_excel_sorted(self, path: str):
        self._workbook_path = Path(path)
        self._run_dir = self._workbook_path.parent
        self._load_workbook_preview(self._workbook_path)
        self._open_excel_btn.setEnabled(True)
        self._open_folder_btn.setEnabled(True)
        self._stage_label.setText(
            self._t("orbea.excel_sort.complete", "Excel sorted")
        )
        self._append_log(f"Sorted Excel saved: {self._workbook_path}")
        self._switch_section("results")
        InfoBar.success(
            self._t("orbea.excel_sort.complete", "Excel sorted"),
            self._t(
                "orbea.excel_sort.saved",
                "A clean sorted copy was saved next to the selected file.",
            ),
            parent=self,
            position=InfoBarPosition.TOP,
            duration=4500,
        )

    def _on_excel_sort_error(self, message: str):
        self._stage_label.setText(
            self._t("orbea.excel_sort.failed", "Excel sorting failed")
        )
        self._append_log(message)
        if not self._closing:
            self._error(
                self._t("orbea.excel_sort.failed", "Excel sorting failed"),
                message,
            )

    def _excel_sort_finished(self):
        self._excel_sort_worker = None
        self._update_action_states()

    def _set_busy(self, busy: bool):
        for widget in self._config_widgets:
            widget.setEnabled(not busy)
        self._resume_btn.setEnabled(not busy)
        self._retry_btn.setEnabled(not busy)
        for widget in self._description_config_widgets:
            widget.setEnabled(not busy)
        self._description_start_btn.setEnabled(not busy)
        for widget in self._photo_config_widgets:
            widget.setEnabled(not busy)
        self._photo_start_btn.setEnabled(not busy)
        for widget in self._table_image_config_widgets:
            widget.setEnabled(not busy)
        self._table_image_start_btn.setEnabled(not busy)
        self._progress_stop_btn.setVisible(busy)
        self._progress_stop_btn.setEnabled(busy)
        if busy:
            self._start_btn.setText(self._t("orbea.stop", "Stop"))
            self._start_btn.setIcon(FluentIcon.CLOSE)
            self._start_btn.setEnabled(True)
            self._stage_label.setText(self._t("orbea.starting", "Starting…"))
        else:
            self._start_btn.setText(self._t("orbea.start", "Start Pimbo scan"))
            self._start_btn.setIcon(FluentIcon.PLAY)

    # --------------------------------------------------- Description extractor

    def _on_description_start_stop(self):
        if self._description_worker and self._description_worker.isRunning():
            self._description_worker.request_stop()
            self._description_start_btn.setEnabled(False)
            self._description_status_label.setText(
                self._t("orbea.description.stopping", "Stopping safely…")
            )
            return
        if self.is_running() or not self._validate_description_inputs():
            return
        self._closing = False
        try:
            config = self._create_description_config()
        except Exception as exc:
            self._error(
                self._t("orbea.description.service_error.title", "Description extractor unavailable"),
                str(exc),
            )
            return

        self._save_description_output()
        self._description_output_dir = Path(self._description_output_edit.text().strip())
        self._description_open_btn.setEnabled(False)
        self._description_log.clear()
        self._description_progress.setValue(0)
        self._set_description_busy(True)
        self._description_worker = OrbeaDescriptionWorker(self._make_description_service, config)
        self._description_worker.progress_changed.connect(self._on_description_progress)
        self._description_worker.log_message.connect(self._append_description_log)
        self._description_worker.succeeded.connect(self._on_description_result)
        self._description_worker.failed.connect(self._on_description_error)
        self._description_worker.finished.connect(self._description_thread_finished)
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._description_worker,
                "orbea",
                "orbea",
                output_path=str(self._description_output_dir),
            )
        self._description_worker.start()

    def _validate_description_inputs(self) -> bool:
        if not self._description_urls():
            self._warn(
                self._t("orbea.description.urls_invalid.title", "Orbea URL required"),
                self._t(
                    "orbea.description.urls_invalid",
                    "Paste one or more cms.orbea.com model URLs containing /m/.",
                ),
            )
            return False
        if not self._description_output_edit.text().strip():
            self._warn(
                self._t("orbea.description.output_invalid.title", "Output folder required"),
                self._t("orbea.description.output_invalid", "Choose where extracted text should be saved."),
            )
            return False
        return True

    def _on_description_progress(self, update):
        status = str(_read(update, "status", "stage", default="Extracting") or "Extracting")
        current = int(_read(update, "current", "done", default=0) or 0)
        total = int(_read(update, "total", default=0) or 0)
        message = str(_read(update, "message", "url", default="") or "")
        succeeded = int(_read(update, "succeeded", default=0) or 0)
        failed = int(_read(update, "failed", default=0) or 0)
        self._description_status_label.setText(status.replace("_", " ").title())
        if message:
            self._description_progress_label.setText(message)
        elif total:
            self._description_progress_label.setText(
                self._t(
                    "orbea.description.progress",
                    "{current:,} / {total:,} URLs • {succeeded:,} saved • {failed:,} failed",
                    current=current,
                    total=total,
                    succeeded=succeeded,
                    failed=failed,
                )
            )
        self._description_progress.setValue(
            max(0, min(100, int(current * 100 / total))) if total else 0
        )

    def _on_description_result(self, result):
        output_dir = _read(result, "output_dir", "output_root", default=None)
        if output_dir:
            self._description_output_dir = Path(output_dir)
        files = _read(result, "files", "written_files", "paths", default=()) or ()
        succeeded = int(_read(result, "succeeded", default=len(files)) or 0)
        failures = _read(result, "failures", default=()) or ()
        cancelled = bool(_read(result, "cancelled", default=False))
        if cancelled:
            self._description_status_label.setText(
                self._t("orbea.description.stopped", "Stopped — completed text files were kept")
            )
        else:
            self._description_status_label.setText(
                self._t("orbea.description.complete", "Description extraction complete")
            )
            self._description_progress.setValue(100)
        self._description_progress_label.setText(
            self._t(
                "orbea.description.result",
                "{succeeded:,} saved • {failed:,} failed",
                succeeded=succeeded,
                failed=len(failures),
            )
        )

    def _on_description_error(self, message: str):
        self._description_status_label.setText(
            self._t("orbea.description.failed", "Description extraction failed")
        )
        self._append_description_log(message)
        if not self._closing:
            self._error(
                self._t("orbea.description.failed.title", "Description extraction failed"),
                message,
            )

    def _description_thread_finished(self):
        self._description_worker = None
        self._set_description_busy(False)
        self._description_open_btn.setEnabled(
            bool(self._description_output_dir and self._description_output_dir.exists())
        )
        self._update_action_states()

    def _set_description_busy(self, busy: bool):
        for widget in self._config_widgets:
            widget.setEnabled(not busy)
        self._start_btn.setEnabled(not busy)
        self._resume_btn.setEnabled(not busy)
        self._retry_btn.setEnabled(not busy)
        for widget in self._description_config_widgets:
            widget.setEnabled(not busy)
        for widget in self._photo_config_widgets:
            widget.setEnabled(not busy)
        self._photo_start_btn.setEnabled(not busy)
        for widget in self._table_image_config_widgets:
            widget.setEnabled(not busy)
        self._table_image_start_btn.setEnabled(not busy)
        if busy:
            self._description_start_btn.setText(self._t("orbea.description.stop", "Stop"))
            self._description_start_btn.setIcon(FluentIcon.CLOSE)
            self._description_start_btn.setEnabled(True)
            self._description_status_label.setText(
                self._t("orbea.description.starting", "Starting description extraction…")
            )
        else:
            self._description_start_btn.setText(
                self._t("orbea.description.extract", "Extract descriptions")
            )
            self._description_start_btn.setIcon(FluentIcon.PLAY)

    # ---------------------------------------------------- Direct table downloader

    def _on_table_image_start_stop(self):
        if self._table_image_worker and self._table_image_worker.isRunning():
            self._table_image_worker.request_stop()
            self._table_image_start_btn.setEnabled(False)
            self._table_image_status_label.setText(
                self._t("orbea.tables.stopping", "Stopping safely…")
            )
            return
        if self.is_running() or not self._validate_table_image_inputs():
            return

        self._closing = False
        _unique, _duplicates, _invalid, entries = self._table_image_link_state()
        geometry, size_guide, product_photos = self._table_image_selection()
        output_dir = Path(self._table_image_output_edit.text().strip())
        self._save_table_image_output()
        self._table_output_dir = output_dir
        self._table_image_open_btn.setEnabled(False)
        self._table_image_log.clear()
        self._table_image_log.setVisible(True)
        self._table_image_progress.setValue(0)
        self._set_table_image_busy(True)
        self._table_image_worker = OrbeaTableImageWorker(
            self._make_table_image_service,
            entries,
            output_dir,
            download_geometry=geometry,
            download_size_guide=size_guide,
            download_product_photos=product_photos,
        )
        self._table_image_worker.progress_changed.connect(
            self._on_table_image_progress
        )
        self._table_image_worker.log_message.connect(self._append_table_image_log)
        self._table_image_worker.succeeded.connect(self._on_table_image_result)
        self._table_image_worker.failed.connect(self._on_table_image_error)
        self._table_image_worker.finished.connect(self._table_image_thread_finished)
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._table_image_worker,
                "orbea",
                "orbea",
                total=len(entries),
                output_path=str(output_dir),
            )
        self._table_image_worker.start()

    def _validate_table_image_inputs(self) -> bool:
        urls, _duplicates, invalid, _entries = self._table_image_link_state()
        if not any(self._table_image_selection()):
            self._warn(
                self._t(
                    "orbea.tables.selection_invalid.title",
                    "Choose what to download",
                ),
                self._t(
                    "orbea.tables.selection_invalid",
                    "Select geometry, the CM size guide, product photos, or any combination.",
                ),
            )
            return False
        if invalid:
            self._warn(
                self._t(
                    "orbea.tables.urls_invalid.title",
                    "Some product links are invalid",
                ),
                self._t(
                    "orbea.tables.urls_invalid",
                    "Fix or remove {count:,} invalid lines. Enter one Orbea product URL per line.",
                    count=len(invalid),
                ),
            )
            return False
        if not urls:
            self._warn(
                self._t(
                    "orbea.tables.url_invalid.title",
                    "Orbea product URLs required",
                ),
                self._t(
                    "orbea.tables.url_invalid",
                    "Paste one or more public Orbea bicycle product-page URLs.",
                ),
            )
            return False
        if not self._table_image_output_edit.text().strip():
            self._warn(
                self._t(
                    "orbea.tables.output_invalid.title",
                    "Output folder required",
                ),
                self._t(
                    "orbea.tables.output_invalid",
                    "Choose where the selected Orbea images should be saved.",
                ),
            )
            return False
        return True

    def _on_table_image_progress(self, update):
        status = str(_read(update, "status", default="downloading") or "downloading")
        current = int(_read(update, "current", default=0) or 0)
        total = int(_read(update, "total", default=0) or 0)
        message = str(_read(update, "message", default="") or "")
        status_fallbacks = {
            "opening_page": "Opening product page…",
            "saved": "Images saved",
            "partial": "Some images could not be saved",
            "downloading": "Downloading images…",
            "product_photos": "Downloading product photos…",
        }
        self._table_image_status_label.setText(
            self._t(
                f"orbea.tables.status.{status}",
                status_fallbacks.get(status, status.replace("_", " ").title()),
            )
        )
        self._table_image_progress.setValue(
            max(0, min(100, int(current * 100 / total))) if total else 0
        )
        self._table_image_progress_label.setText(
            message
            or self._t(
                "orbea.tables.progress",
                "{current:,} / {total:,} products",
                current=current,
                total=total,
            )
        )

    def _on_table_image_result(self, result):
        output_dir = _read(result, "output_dir", default=None)
        if output_dir:
            self._table_output_dir = Path(output_dir)
        files = _read(result, "files", default=()) or ()
        failures = _read(result, "failures", default=()) or ()
        unavailable = _read(result, "unavailable", default=()) or ()
        products = int(_read(result, "products", default=0) or 0)
        duplicates = int(_read(result, "duplicates", default=0) or 0)
        photo_variants = int(_read(result, "photo_variants", default=0) or 0)
        photo_views = int(_read(result, "photo_views", default=0) or 0)
        cancelled = bool(_read(result, "cancelled", default=False))
        if cancelled:
            self._table_image_status_label.setText(
                self._t(
                    "orbea.tables.stopped",
                    "Stopped — completed images were kept",
                )
            )
        elif failures:
            self._table_image_status_label.setText(
                self._t(
                    "orbea.tables.partial",
                    "Completed with some failed images",
                )
            )
            self._table_image_progress.setValue(100)
        else:
            self._table_image_status_label.setText(
                self._t("orbea.tables.complete", "Image download complete")
            )
            self._table_image_progress.setValue(100)
        self._table_image_progress_label.setText(
            self._t(
                "orbea.tables.result",
                "Products {products:,} • Images saved {saved:,} • Photo colours {photo_variants:,} • Photo views {photo_views:,} • Duplicates ignored {duplicates:,} • Unavailable {unavailable:,} • Failed {failed:,}",
                products=products,
                saved=len(files),
                photo_variants=photo_variants,
                photo_views=photo_views,
                duplicates=duplicates,
                unavailable=len(unavailable),
                failed=len(failures),
            )
        )
        if failures and not cancelled:
            self._warn(
                self._t(
                    "orbea.tables.partial.title",
                    "Some images were not downloaded",
                ),
                self._t(
                    "orbea.tables.partial.detail",
                    "Successful images were kept. Check the activity log for details.",
                ),
            )

    def _on_table_image_error(self, message: str):
        self._table_image_status_label.setText(
            self._t("orbea.tables.failed", "Image download failed")
        )
        self._append_table_image_log(message)
        if not self._closing:
            self._error(
                self._t("orbea.tables.failed.title", "Image download failed"),
                message,
            )

    def _table_image_thread_finished(self):
        self._table_image_worker = None
        self._set_table_image_busy(False)
        self._table_image_open_btn.setEnabled(
            bool(self._table_output_dir and self._table_output_dir.exists())
        )
        self._update_action_states()

    def _set_table_image_busy(self, busy: bool):
        for group in (
            self._config_widgets,
            self._description_config_widgets,
            self._photo_config_widgets,
            self._table_image_config_widgets,
        ):
            for widget in group:
                widget.setEnabled(not busy)
        self._start_btn.setEnabled(not busy)
        self._resume_btn.setEnabled(not busy)
        self._retry_btn.setEnabled(not busy)
        self._description_start_btn.setEnabled(not busy)
        self._photo_start_btn.setEnabled(not busy)
        if busy:
            self._table_image_start_btn.setText(
                self._t("orbea.tables.stop", "Stop")
            )
            self._table_image_start_btn.setIcon(FluentIcon.CLOSE)
            self._table_image_start_btn.setEnabled(True)
            self._table_image_status_label.setText(
                self._t("orbea.tables.starting", "Opening Orbea product pages…")
            )
        else:
            self._table_image_start_btn.setText(
                self._t("orbea.tables.download", "Download selected")
            )
            self._table_image_start_btn.setIcon(FluentIcon.DOWNLOAD)

    # --------------------------------------------------------- Photo downloader

    def _on_photo_start_stop(self):
        if self._photo_worker and self._photo_worker.isRunning():
            self._photo_worker.request_stop()
            self._photo_start_btn.setEnabled(False)
            self._photo_status_label.setText(
                self._t("orbea.photo.stopping", "Stopping safely…")
            )
            return
        if self.is_running() or not self._validate_photo_inputs():
            return

        self._closing = False
        _unique, _duplicates, _invalid, entries = self._photo_link_state()
        output_dir = Path(self._photo_output_edit.text().strip())
        self._save_photo_output()
        self._photo_output_dir = output_dir
        self._photo_open_btn.setEnabled(False)
        self._photo_log.clear()
        self._photo_progress.setValue(0)
        self._set_photo_busy(True)
        self._photo_worker = OrbeaPhotoWorker(
            self._make_photo_service, entries, output_dir
        )
        self._photo_worker.progress_changed.connect(self._on_photo_progress)
        self._photo_worker.log_message.connect(self._append_photo_log)
        self._photo_worker.succeeded.connect(self._on_photo_result)
        self._photo_worker.failed.connect(self._on_photo_error)
        self._photo_worker.finished.connect(self._photo_thread_finished)
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._photo_worker,
                "orbea",
                "orbea",
                total=len(entries),
                output_path=str(output_dir),
            )
        self._photo_worker.start()

    def _validate_photo_inputs(self) -> bool:
        urls, _duplicates, invalid, _entries = self._photo_link_state()
        if invalid:
            self._warn(
                self._t("orbea.photo.urls_invalid.title", "Some product links are invalid"),
                self._t(
                    "orbea.photo.urls_invalid",
                    "Fix or remove {count:,} invalid lines. Enter one Orbea product URL per line.",
                    count=len(invalid),
                ),
            )
            return False
        if not urls:
            self._warn(
                self._t("orbea.photo.url_invalid.title", "Orbea product URLs required"),
                self._t(
                    "orbea.photo.url_invalid",
                    "Paste one or more public cms.orbea.com bicycle product URLs.",
                ),
            )
            return False
        if not self._photo_output_edit.text().strip():
            self._warn(
                self._t("orbea.photo.output_invalid.title", "Output folder required"),
                self._t(
                    "orbea.photo.output_invalid",
                    "Choose where product photo folders should be saved.",
                ),
            )
            return False
        return True

    def _on_photo_progress(self, update):
        status = str(_read(update, "status", default="downloading") or "downloading")
        current = int(_read(update, "current", default=0) or 0)
        total = int(_read(update, "total", default=0) or 0)
        succeeded = int(_read(update, "succeeded", default=0) or 0)
        failed = int(_read(update, "failed", default=0) or 0)
        message = str(_read(update, "message", default="") or "")
        self._photo_status_label.setText(status.replace("_", " ").title())
        self._photo_progress.setValue(
            max(0, min(100, int(current * 100 / total))) if total else 0
        )
        self._photo_progress_label.setText(
            message
            or self._t(
                "orbea.photo.progress",
                "{current:,} / {total:,} images • {succeeded:,} saved • {failed:,} failed",
                current=current,
                total=total,
                succeeded=succeeded,
                failed=failed,
            )
        )

    def _on_photo_result(self, result):
        product_results = _read(result, "product_results", default=()) or ()
        product_dir = _read(result, "product_dir", default=None)
        if not product_dir and len(product_results) == 1:
            product_dir = _read(product_results[0], "product_dir", default=None)
        if not product_dir:
            product_dir = _read(result, "output_dir", default=None)
        if product_dir:
            self._photo_output_dir = Path(product_dir)
        files = _read(result, "files", default=()) or ()
        failures = _read(result, "failures", default=()) or ()
        unavailable = _read(result, "unavailable", default=()) or ()
        variants = int(_read(result, "variants", default=0) or 0)
        products = int(_read(result, "products", default=1 if files else 0) or 0)
        duplicates = int(_read(result, "duplicates", default=0) or 0)
        duplicate_summary = (
            self._t("orbea.photo.urls.duplicate.one", "1 duplicate ignored")
            if duplicates == 1
            else self._t(
                "orbea.photo.urls.duplicate.many",
                "{count:,} duplicates ignored",
                count=duplicates,
            )
        )
        cancelled = bool(_read(result, "cancelled", default=False))
        if cancelled:
            self._photo_status_label.setText(
                self._t("orbea.photo.stopped", "Stopped — completed photos were kept")
            )
        elif failures:
            self._photo_status_label.setText(
                self._t("orbea.photo.partial", "Completed with some failed photos")
            )
            self._photo_progress.setValue(100)
        else:
            self._photo_status_label.setText(
                self._t("orbea.photo.complete", "Photo download complete")
            )
            self._photo_progress.setValue(100)
        self._photo_progress_label.setText(
            self._t(
                "orbea.photo.result",
                "Products {products:,} • Colours {variants:,} • Photos saved {saved:,} • {duplicate_summary} • Unavailable {unavailable:,} • Failed {failed:,}",
                products=products,
                variants=variants,
                saved=len(files),
                duplicate_summary=duplicate_summary,
                unavailable=len(unavailable),
                failed=len(failures),
            )
        )
        if failures and not cancelled:
            self._warn(
                self._t("orbea.photo.partial.title", "Some photos were not downloaded"),
                self._t(
                    "orbea.photo.partial.detail",
                    "The successful photos were kept. Check the activity log for details.",
                ),
            )

    def _on_photo_error(self, message: str):
        self._photo_status_label.setText(
            self._t("orbea.photo.failed", "Photo download failed")
        )
        self._append_photo_log(message)
        if not self._closing:
            self._error(
                self._t("orbea.photo.failed.title", "Photo download failed"),
                message,
            )

    def _photo_thread_finished(self):
        self._photo_worker = None
        self._set_photo_busy(False)
        self._photo_open_btn.setEnabled(
            bool(self._photo_output_dir and self._photo_output_dir.exists())
        )
        self._update_action_states()

    def _set_photo_busy(self, busy: bool):
        for widget in self._config_widgets:
            widget.setEnabled(not busy)
        for widget in self._description_config_widgets:
            widget.setEnabled(not busy)
        self._start_btn.setEnabled(not busy)
        self._resume_btn.setEnabled(not busy)
        self._retry_btn.setEnabled(not busy)
        self._description_start_btn.setEnabled(not busy)
        for widget in self._photo_config_widgets:
            widget.setEnabled(not busy)
        for widget in self._table_image_config_widgets:
            widget.setEnabled(not busy)
        self._table_image_start_btn.setEnabled(not busy)
        if busy:
            self._photo_start_btn.setText(self._t("orbea.photo.stop", "Stop"))
            self._photo_start_btn.setIcon(FluentIcon.CLOSE)
            self._photo_start_btn.setEnabled(True)
            self._photo_status_label.setText(
                self._t("orbea.photo.starting", "Reading product colours…")
            )
        else:
            self._photo_start_btn.setText(
                self._t("orbea.photo.download", "Download all colours")
            )
            self._photo_start_btn.setIcon(FluentIcon.DOWNLOAD)

    def _update_counts(self, counts: Any):
        aliases = {
            "scanned": ("scanned", "products_scanned", "processed"),
            "matched": ("matched", "matches", "code_matches"),
            "review": ("review", "review_count", "needs_review"),
            "images": ("images", "images_downloaded", "downloaded"),
            "unavailable": ("unavailable", "not_available", "tables_not_available"),
            "errors": ("errors", "error_count", "transient_errors"),
        }
        for key, names in aliases.items():
            value = _read(counts, *names, default=None)
            if value is not None:
                self._stat_values[key].setText(str(value))

    # -------------------------------------------------------------- Results

    def _load_workbook_preview(self, path: Path):
        rows: list[list[str]] = []
        total_rows = 0
        try:
            import openpyxl

            workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for sheet_name, review_sheet in (("Matches", False), ("Review", True)):
                if sheet_name not in workbook.sheetnames:
                    continue
                sheet = workbook[sheet_name]
                iterator = sheet.iter_rows(values_only=True)
                headers = [str(value or "").strip() for value in next(iterator, ())]
                lookup = {name.lower(): index for index, name in enumerate(headers)}
                for source in iterator:
                    total_rows += 1
                    if len(rows) >= PREVIEW_LIMIT:
                        continue
                    rows.append([
                        self._cell(source, lookup, "variant sku", "pimbo sku", "sku"),
                        self._cell(source, lookup, "pimbo product", "product", "product title"),
                        self._cell(source, lookup, "match method", "status", "reason") or ("Review" if review_sheet else "Match"),
                        self._cell(source, lookup, "orbea url", "catalogue url"),
                        self._cell(source, lookup, "geometry status", "geometry", "geometry image"),
                        self._cell(source, lookup, "size guide status", "size status", "size guide", "size image"),
                    ])
            workbook.close()
        except Exception as exc:
            self._append_log(f"Could not preview workbook: {exc}")
            return

        self._table.setUpdatesEnabled(False)
        self._table.setRowCount(len(rows))
        for row_index, values in enumerate(rows):
            for column, value in enumerate(values):
                item = QTableWidgetItem(value)
                if column == 2:
                    lower = value.lower()
                    status_role = "success" if "code" in lower or lower == "match" else "warning"
                    item.setForeground(QColor(get_status_text_color(status_role, isDarkTheme())))
                self._table.setItem(row_index, column, item)
        self._table.setUpdatesEnabled(True)
        self._results_label.setText(
            self._t("orbea.results.preview", "Showing {shown:,} of {total:,} rows. Excel contains the complete report.", shown=len(rows), total=total_rows)
        )

    @staticmethod
    def _cell(row, lookup: dict[str, int], *names: str) -> str:
        for name in names:
            index = lookup.get(name.lower())
            if index is not None and index < len(row):
                return str(row[index] or "")
        return ""

    # --------------------------------------------------------------- Paths

    def _load_paths(self):
        catalogue = str(self._setting_get(CATALOGUE_SETTING, "") or "").strip()
        output = str(self._setting_get(OUTPUT_SETTING, "") or "").strip()
        description_output = str(
            self._setting_get(DESCRIPTION_OUTPUT_SETTING, "") or ""
        ).strip()
        photo_output = str(self._setting_get(PHOTO_OUTPUT_SETTING, "") or "").strip()
        table_output = str(self._setting_get(TABLE_OUTPUT_SETTING, "") or "").strip()
        if not catalogue:
            catalogue = self._detect_catalogue()
        if not output:
            output = str(self._desktop_dir() / "UltraBike Orbea Runs")
        if not description_output:
            description_output = str(self._desktop_dir() / "UltraBike Orbea Descriptions")
        if not photo_output:
            photo_output = str(self._desktop_dir() / "UltraBike Orbea Photos")
        if not table_output:
            table_output = str(self._desktop_dir() / "UltraBike Orbea Downloads")
        self._catalogue_edit.setText(catalogue)
        self._output_edit.setText(output)
        self._description_output_edit.setText(description_output)
        self._photo_output_edit.setText(photo_output)
        self._table_image_output_edit.setText(table_output)
        self._table_geometry_check.setChecked(
            self._setting_bool(DIRECT_GEOMETRY_SETTING, True)
        )
        self._table_size_guide_check.setChecked(
            self._setting_bool(DIRECT_SIZE_GUIDE_SETTING, True)
        )
        self._table_product_photos_check.setChecked(
            self._setting_bool(DIRECT_PRODUCT_PHOTOS_SETTING, False)
        )
        # Migrate older saved choices to the new explicit table-only flow.
        self._table_images_check.setChecked(False)
        self._product_photos_check.setChecked(False)
        self._setting_set(TABLE_IMAGES_SETTING, False)
        self._setting_set(PRODUCT_PHOTOS_SETTING, False)
        candidate = Path(description_output)
        self._description_output_dir = candidate if candidate.exists() else None
        photo_candidate = Path(photo_output)
        self._photo_output_dir = photo_candidate if photo_candidate.exists() else None
        self._photo_open_btn.setEnabled(bool(self._photo_output_dir))
        table_candidate = Path(table_output)
        self._table_output_dir = table_candidate if table_candidate.exists() else None
        self._table_image_open_btn.setEnabled(bool(self._table_output_dir))

    def _detect_catalogue(self) -> str:
        candidates = [
            self._desktop_dir() / "Orbea-Scraper" / "data" / "output" / "orbea_bicycle_catalogue.xlsx",
            Path.cwd() / "Orbea-Scraper" / "data" / "output" / "orbea_bicycle_catalogue.xlsx",
            Path(__file__).resolve().parents[2] / "data" / "output" / "orbea_bicycle_catalogue.xlsx",
        ]
        return str(next((path for path in candidates if path.is_file()), ""))

    @staticmethod
    def _desktop_dir() -> Path:
        candidates = [Path.home() / "Desktop", Path.home() / "OneDrive" / "Desktop"]
        return next((path for path in candidates if path.exists()), candidates[0])

    def _browse_catalogue(self):
        path, _ = QFileDialog.getOpenFileName(self, self._t("orbea.catalogue.pick", "Select Orbea catalogue"), self._catalogue_edit.text(), "Excel (*.xlsx)")
        if path:
            self._catalogue_edit.setText(path)
            self._paths_changed()

    def _browse_output(self):
        path = QFileDialog.getExistingDirectory(self, self._t("orbea.output.pick", "Select output folder"), self._output_edit.text())
        if path:
            self._output_edit.setText(path)
            self._paths_changed()

    def _browse_description_output(self):
        path = QFileDialog.getExistingDirectory(
            self,
            self._t("orbea.description.output.pick", "Select description output folder"),
            self._description_output_edit.text(),
        )
        if path:
            self._description_output_edit.setText(path)
            self._description_output_changed()

    def _browse_photo_output(self):
        path = QFileDialog.getExistingDirectory(
            self,
            self._t("orbea.photo.output.pick", "Select photo output folder"),
            self._photo_output_edit.text(),
        )
        if path:
            self._photo_output_edit.setText(path)
            self._photo_output_changed()

    def _browse_table_image_output(self):
        path = QFileDialog.getExistingDirectory(
            self,
            self._t(
                "orbea.tables.output.pick",
                "Select table-image output folder",
            ),
            self._table_image_output_edit.text(),
        )
        if path:
            self._table_image_output_edit.setText(path)
            self._table_image_output_changed()

    def _description_output_changed(self):
        self._save_description_output()
        path = Path(self._description_output_edit.text().strip())
        self._description_output_dir = path if path.exists() else None
        self._description_open_btn.setEnabled(bool(self._description_output_dir))
        self._update_action_states()

    def _photo_output_changed(self):
        self._save_photo_output()
        value = self._photo_output_edit.text().strip()
        path = Path(value) if value else None
        self._photo_output_dir = path if path is not None and path.exists() else None
        self._photo_open_btn.setEnabled(bool(self._photo_output_dir))
        self._update_action_states()

    def _table_image_output_changed(self):
        self._save_table_image_output()
        value = self._table_image_output_edit.text().strip()
        path = Path(value) if value else None
        self._table_output_dir = path if path is not None and path.exists() else None
        self._table_image_open_btn.setEnabled(bool(self._table_output_dir))
        self._update_action_states()

    def _paths_changed(self):
        self._save_paths()
        self._update_action_states()

    def _download_options_changed(self, _state=None):
        # Legacy hidden controls may still be changed by an older integration;
        # keep catalogue scans image-free regardless of stale UI state.
        self._setting_set(TABLE_IMAGES_SETTING, False)
        self._setting_set(PRODUCT_PHOTOS_SETTING, False)
        self._update_action_states()

    def _save_paths(self):
        self._setting_set(CATALOGUE_SETTING, self._catalogue_edit.text().strip())
        self._setting_set(OUTPUT_SETTING, self._output_edit.text().strip())

    def _save_description_output(self):
        self._setting_set(
            DESCRIPTION_OUTPUT_SETTING,
            self._description_output_edit.text().strip(),
        )

    def _save_photo_output(self):
        self._setting_set(PHOTO_OUTPUT_SETTING, self._photo_output_edit.text().strip())

    def _save_table_image_output(self):
        self._setting_set(
            TABLE_OUTPUT_SETTING,
            self._table_image_output_edit.text().strip(),
        )

    def _validate_inputs(self) -> bool:
        catalogue = Path(self._catalogue_edit.text().strip())
        if not catalogue.is_file() or catalogue.suffix.lower() != ".xlsx":
            self._warn(self._t("orbea.catalogue.invalid.title", "Catalogue required"), self._t("orbea.catalogue.invalid", "Choose the Orbea catalogue .xlsx file."))
            return False
        if not self._output_edit.text().strip():
            self._warn(self._t("orbea.output.invalid.title", "Output folder required"), self._t("orbea.output.invalid", "Choose where Orbea run folders should be saved."))
            return False
        return True

    def _open_excel(self):
        if self._workbook_path and self._workbook_path.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._workbook_path)))

    def _open_folder(self):
        if self._run_dir and self._run_dir.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._run_dir)))

    def _open_description_folder(self):
        if self._description_output_dir and self._description_output_dir.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._description_output_dir)))

    def _open_photo_folder(self):
        if self._photo_output_dir and self._photo_output_dir.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._photo_output_dir)))

    def _open_table_image_folder(self):
        if self._table_output_dir and self._table_output_dir.exists():
            QDesktopServices.openUrl(QUrl.fromLocalFile(str(self._table_output_dir)))

    # ---------------------------------------------------------- App lifecycle

    def _acquire_browser(self) -> bool:
        acquired = self.workflow_controller.acquire_browser()
        self._owns_browser_lease = self.workflow_controller.owns_browser_lease
        return acquired

    def _release_browser(self):
        self.workflow_controller.release_browser()
        self._owns_browser_lease = self.workflow_controller.owns_browser_lease

    def is_running(self) -> bool:
        return bool(
            (self._worker and self._worker.isRunning())
            or (self._filter_worker and self._filter_worker.isRunning())
            or (self._description_worker and self._description_worker.isRunning())
            or (self._photo_worker and self._photo_worker.isRunning())
            or (self._table_image_worker and self._table_image_worker.isRunning())
            or (self._excel_sort_worker and self._excel_sort_worker.isRunning())
        )

    def shutdown(self, wait_ms: int = 5000) -> bool:
        """Request a checkpointed stop and wait briefly; never terminate threads."""
        self._closing = True
        workers = [
            worker
            for worker in (
                self._worker,
                self._filter_worker,
                self._description_worker,
                self._photo_worker,
                self._table_image_worker,
                self._excel_sort_worker,
            )
            if worker and worker.isRunning()
        ]
        for worker in workers:
            worker.request_stop()
        remaining = max(0, int(wait_ms))
        for worker in workers:
            if not worker.isRunning():
                continue
            slice_ms = remaining if len(workers) == 1 else max(1, remaining // len(workers))
            if not worker.wait(slice_ms):
                self._closing = False
                return False
            remaining = max(0, remaining - slice_ms)
        self._release_browser()
        return True

    # ------------------------------------------------------------- Utilities

    def _setting_get(self, key: str, default=None):
        try:
            return self.settings.get(key, default) if self.settings is not None else default
        except Exception:
            return default

    def _setting_bool(self, key: str, default: bool) -> bool:
        value = self._setting_get(key, default)
        if isinstance(value, str):
            normalized = value.strip().casefold()
            if normalized in {"true", "1", "yes", "on"}:
                return True
            if normalized in {"false", "0", "no", "off", ""}:
                return False
        return bool(value)

    def _setting_set(self, key: str, value):
        try:
            if self.settings is not None:
                self.settings.set(key, value)
        except Exception:
            pass

    def _update_action_states(self):
        main_running = bool(self._worker and self._worker.isRunning())
        filter_running = bool(self._filter_worker and self._filter_worker.isRunning())
        excel_sort_running = self._excel_sort_worker is not None
        description_running = bool(
            self._description_worker and self._description_worker.isRunning()
        )
        photo_running = bool(self._photo_worker and self._photo_worker.isRunning())
        table_running = bool(
            self._table_image_worker and self._table_image_worker.isRunning()
        )
        if excel_sort_running:
            for group in (
                self._config_widgets,
                self._description_config_widgets,
                self._photo_config_widgets,
                self._table_image_config_widgets,
            ):
                for widget in group:
                    widget.setEnabled(False)
            for button in (
                self._start_btn,
                self._resume_btn,
                self._retry_btn,
                self._description_start_btn,
                self._photo_start_btn,
                self._table_image_start_btn,
            ):
                button.setEnabled(False)
            return
        if main_running:
            for group in (
                self._description_config_widgets,
                self._photo_config_widgets,
                self._table_image_config_widgets,
            ):
                for widget in group:
                    widget.setEnabled(False)
            self._description_start_btn.setEnabled(False)
            self._photo_start_btn.setEnabled(False)
            self._table_image_start_btn.setEnabled(False)
            return
        if description_running:
            for widget in self._config_widgets:
                widget.setEnabled(False)
            for widget in self._description_config_widgets:
                widget.setEnabled(False)
            for widget in self._photo_config_widgets:
                widget.setEnabled(False)
            for widget in self._table_image_config_widgets:
                widget.setEnabled(False)
            self._start_btn.setEnabled(False)
            self._resume_btn.setEnabled(False)
            self._retry_btn.setEnabled(False)
            self._description_start_btn.setEnabled(True)
            self._photo_start_btn.setEnabled(False)
            self._table_image_start_btn.setEnabled(False)
            return
        if photo_running:
            for group in (
                self._config_widgets,
                self._description_config_widgets,
                self._photo_config_widgets,
                self._table_image_config_widgets,
            ):
                for widget in group:
                    widget.setEnabled(False)
            self._start_btn.setEnabled(False)
            self._resume_btn.setEnabled(False)
            self._retry_btn.setEnabled(False)
            self._description_start_btn.setEnabled(False)
            self._photo_start_btn.setEnabled(True)
            self._table_image_start_btn.setEnabled(False)
            return
        if table_running:
            for group in (
                self._config_widgets,
                self._description_config_widgets,
                self._photo_config_widgets,
                self._table_image_config_widgets,
            ):
                for widget in group:
                    widget.setEnabled(False)
            self._start_btn.setEnabled(False)
            self._resume_btn.setEnabled(False)
            self._retry_btn.setEnabled(False)
            self._description_start_btn.setEnabled(False)
            self._photo_start_btn.setEnabled(False)
            self._table_image_start_btn.setEnabled(True)
            return

        for widget in self._config_widgets:
            widget.setEnabled(not filter_running)
        for widget in self._description_config_widgets:
            widget.setEnabled(not filter_running)
        for widget in self._photo_config_widgets:
            widget.setEnabled(not filter_running)
        for widget in self._table_image_config_widgets:
            widget.setEnabled(not filter_running)
        valid = bool(
            getattr(self.main, "driver", None) is not None
            and Path(self._catalogue_edit.text().strip()).is_file()
            and self._output_edit.text().strip()
            and not filter_running
        )
        self._start_btn.setEnabled(valid)
        self._resume_btn.setEnabled(valid)
        self._retry_btn.setEnabled(valid)
        self._description_start_btn.setEnabled(
            bool(
                not filter_running
                and self._description_urls()
                and self._description_output_edit.text().strip()
            )
        )
        photo_urls, _duplicates, photo_invalid, _entries = self._photo_link_state()
        self._photo_start_btn.setEnabled(
            bool(
                not filter_running
                and photo_urls
                and not photo_invalid
                and self._photo_output_edit.text().strip()
            )
        )
        table_urls, _duplicates, table_invalid, _entries = (
            self._table_image_link_state()
        )
        self._table_image_start_btn.setEnabled(
            bool(
                not filter_running
                and table_urls
                and not table_invalid
                and any(self._table_image_selection())
                and self._table_image_output_edit.text().strip()
            )
        )

    @staticmethod
    def _format_eta(seconds: Any) -> str:
        try:
            seconds = max(0, int(float(seconds)))
        except (TypeError, ValueError):
            return "ETA —"
        minutes, secs = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        return f"ETA {hours}h {minutes:02d}m" if hours else f"ETA {minutes}m {secs:02d}s"

    def _append_log(self, message: str):
        if message:
            self._log.appendPlainText(str(message))

    def _append_description_log(self, message: str):
        if message:
            self._description_log.appendPlainText(str(message))

    def _append_photo_log(self, message: str):
        if message:
            self._photo_log.appendPlainText(str(message))

    def _append_table_image_log(self, message: str):
        if message:
            self._table_image_log.appendPlainText(str(message))

    def _t(self, key: str, fallback: str, **kwargs) -> str:
        value = self.tr(key, **kwargs)
        if value == key:
            try:
                return fallback.format(**kwargs)
            except Exception:
                return fallback
        return value

    def _warn(self, title: str, message: str):
        InfoBar.warning(title, message, parent=self, position=InfoBarPosition.TOP, duration=4500)

    def _error(self, title: str, message: str):
        InfoBar.error(title, message, parent=self, position=InfoBarPosition.TOP, duration=6000)

    def _update_table_theme(self):
        dark = isDarkTheme()
        table = COMPONENT_COLORS["table"]
        bg = table["row_bg_dark"] if dark else table["row_bg_light"]
        alt = table["row_alt_bg_dark"] if dark else table["row_alt_bg_light"]
        border = table["border_dark"] if dark else table["border_light"]
        text = COLORS["text_primary_dark"] if dark else COLORS["text_primary_light"]
        muted = COLORS["text_secondary_dark"] if dark else COLORS["text_secondary_light"]
        disabled_text = COLORS["text_disabled_dark"] if dark else COLORS["text_disabled_light"]
        header_bg = table["header_bg_dark" if dark else "header_bg_light"]
        header_text = table["header_text_dark" if dark else "header_text_light"]
        accent_colors = get_accent_colors(dark)
        accent = accent_colors["base"]
        accent_text = accent_colors["text"]
        accent_hover = accent_colors["hover"]
        accent_pressed = accent_colors["pressed"]
        accent_soft = rgba_from_hex(accent, 0.16 if dark else 0.07)
        outline = get_subtle_border(dark)
        hover = get_subtle_item_hover_bg(dark)
        disabled_bg = COLORS["disabled_surface_dark"] if dark else COLORS["disabled_surface_light"]
        danger = COLORS["error_text_dark"] if dark else COLORS["error_text_light"]

        self._section_tabs.setStyleSheet(f"""
            QTabBar::tab {{
                background: transparent;
                color: {disabled_text};
                padding: {PADDINGS['tab']};
                border: none;
                border-bottom: 3px solid transparent;
            }}
            QTabBar::tab:hover {{
                background: {hover};
                color: {text};
            }}
            QTabBar::tab:selected {{
                background: {accent_soft};
                color: {accent};
                border-bottom: 3px solid {accent};
                font-weight: 600;
            }}
        """)
        primary_style = f"""
            PrimaryPushButton {{
                background-color: {accent};
                border: 1px solid {accent};
                color: {accent_text};
                font-weight: 600;
                border-radius: {RADII['md']}px;
            }}
            PrimaryPushButton:hover {{
                background-color: {accent_hover};
                border-color: {accent_hover};
            }}
            PrimaryPushButton:pressed {{
                background-color: {accent_pressed};
                border-color: {accent_pressed};
            }}
            PrimaryPushButton:disabled {{
                background-color: {disabled_bg};
                border-color: {border};
                color: {muted};
            }}
        """
        for button in (
            self._start_btn,
            self._description_start_btn,
            self._photo_start_btn,
            self._table_image_start_btn,
        ):
            button.setStyleSheet(primary_style)
        self._progress_stop_btn.setStyleSheet(f"""
            PushButton {{
                background: transparent;
                border: 1px solid {rgba_from_hex(danger, 0.65)};
                color: {danger};
                border-radius: {RADII['sm']}px;
            }}
            PushButton:hover {{ background: {rgba_from_hex(danger, 0.10)}; }}
        """)
        self._table.setStyleSheet(f"""
            QTableWidget {{ background: {bg}; alternate-background-color: {alt}; color: {text}; border: 1px solid {border}; border-radius: {RADII['md']}px; gridline-color: transparent; }}
            QTableWidget::viewport {{ background: {bg}; border-radius: {RADII['md']}px; }}
            QTableWidget::item {{ padding: {PADDINGS['table_cell']}; border: none; border-bottom: 1px solid {border}; }}
            QTableWidget::item:hover {{ background: {hover}; }}
            QTableWidget::item:selected {{ background: {get_selection_bg(dark)}; color: {text}; }}
            QHeaderView::section {{ background: {header_bg}; color: {header_text}; padding: {PADDINGS['table_header']}; border: none; font-weight: 600; font-size: {FONTS['size_body_sm']}; }}
        """)
        self._style_filter_buttons()

    def _style_filter_buttons(self) -> None:
        if not hasattr(self, "_status_buttons"):
            return
        dark = isDarkTheme()
        accent = COLORS["lavender_grey"] if dark else COLORS["space_indigo"]
        accent_text = COLORS["space_indigo"] if dark else COLORS["text_white"]
        text = COLORS["text_primary_dark"] if dark else COLORS["text_primary_light"]
        outline = get_subtle_border(dark)
        hover = get_subtle_item_hover_bg(dark)
        style = f"""
            PillPushButton {{
                background: transparent;
                border: 1px solid {outline};
                color: {text};
                border-radius: 14px;
            }}
            PillPushButton:hover {{ background: {hover}; }}
            PillPushButton:checked {{
                background: {accent};
                border-color: {accent};
                color: {accent_text};
                font-weight: 600;
            }}
        """
        for collection in (
            self._status_buttons,
            self._stock_buttons,
            self._bucket_buttons,
        ):
            for button in collection.values():
                button.setStyleSheet(style)

    def retranslate_ui(self):
        self.tr = self.main.i18n.tr
        self._title.setText(self._t("orbea.title", "Orbea Automation"))
        self._subtitle.setText(
            self._t(
                "orbea.subtitle",
                "Match filtered Pimbo products and choose which Orbea images to download.",
            )
        )
        self._section_tabs.setTabText(0, self._t("orbea.tab.setup", "Setup"))
        self._section_tabs.setTabText(1, self._t("orbea.tab.progress", "Progress"))
        self._section_tabs.setTabText(
            2, self._t("orbea.tab.photos", "Image downloads")
        )
        self._section_tabs.setTabText(3, self._t("orbea.tab.descriptions", "Descriptions"))
        self._section_tabs.setTabText(4, self._t("orbea.tab.results", "Results"))
        self._paths_title.setText(self._t("orbea.paths", "Catalogue and output"))
        self._catalogue_label.setText(self._t("orbea.catalogue", "Catalogue"))
        self._output_label.setText(self._t("orbea.output", "Output folder"))
        self._downloads_label.setText(
            self._t("orbea.downloads", "Include in catalogue run")
        )
        self._table_images_check.setText(
            self._t(
                "orbea.downloads.tables", "Geometry + CM size tables"
            )
        )
        self._product_photos_check.setText(
            self._t(
                "orbea.downloads.photos", "Product photos (all colours)"
            )
        )
        self._downloads_hint.setText(
            self._t(
                "orbea.downloads.hint",
                "This option belongs to the Pimbo catalogue scan. For URLs you paste yourself, use Image downloads.",
            )
        )
        self._search_label.setText(self._t("orbea.search", "Fixed Pimbo search"))
        self._catalogue_btn.setText(self._t("common.browse", "Browse"))
        self._output_btn.setText(self._t("common.browse", "Browse"))
        self._filters_title.setText(self._t("orbea.filters", "Pimbo filters"))
        self._refresh_btn.setText(self._t("orbea.filters.refresh", "Refresh filters"))
        self._status_label.setText(self._t("orbea.filters.status", "Status (select any)"))
        self._family_label.setText(self._t("orbea.filters.family", "Family"))
        self._category_label.setText(self._t("orbea.filters.category", "Category"))
        self._source_label.setText(self._t("orbea.filters.source", "Source"))
        self._locale_label.setText(self._t("orbea.filters.locale", "Completeness locale"))
        self._sort_label.setText(self._t("orbea.filters.sort", "Sort"))
        self._stock_label.setText(self._t("orbea.filters.stock", "Stock"))
        self._bucket_label.setText(self._t("orbea.filters.completeness", "Completeness"))
        self._actions_title.setText(
            self._t("orbea.actions", "Pimbo catalogue scan")
        )
        running = bool(self._worker and self._worker.isRunning())
        self._start_btn.setText(
            self._t("orbea.stop", "Stop")
            if running
            else self._t("orbea.start", "Start Pimbo scan")
        )
        self._progress_stop_btn.setText(self._t("orbea.stop", "Stop"))
        self._resume_btn.setText(self._t("orbea.resume", "Resume latest"))
        self._retry_btn.setText(self._t("orbea.retry", "Retry failed"))
        self._excel_sort_btn.setText(
            self._t("orbea.excel_sort", "Sort existing Excel")
        )
        self._excel_sort_btn.setToolTip(
            self._t(
                "orbea.excel_sort.tooltip",
                "Create a clean five-column copy sorted by Catalogue Model.",
            )
        )
        self._open_excel_btn.setText(self._t("orbea.open_excel", "Open Excel"))
        self._open_folder_btn.setText(self._t("orbea.open_folder", "Open folder"))
        self._table_image_title.setText(
            self._t("orbea.tables.title", "Orbea image downloader")
        )
        self._table_image_subtitle.setText(
            self._t(
                "orbea.tables.subtitle",
                "Choose geometry tables, the CM size guide, product photos, or any combination. Pimbo is not scanned.",
            )
        )
        self._table_image_url_label.setText(
            self._t("orbea.tables.url", "Orbea product URLs")
        )
        self._table_image_url_edit.setPlaceholderText(
            self._t(
                "orbea.tables.url.placeholder",
                "One product URL per line\nhttps://www.orbea.com/en-be/onna-20",
            )
        )
        self._table_image_types_label.setText(
            self._t("orbea.tables.types", "What to download")
        )
        self._table_geometry_check.setText(
            self._t(
                "orbea.tables.type.geometry",
                "Geometry — every frame size",
            )
        )
        self._table_size_guide_check.setText(
            self._t("orbea.tables.type.size_guide", "CM size guide")
        )
        self._table_product_photos_check.setText(
            self._t(
                "orbea.tables.type.product_photos",
                "Product photos — every colour and view",
            )
        )
        self._table_image_output_label.setText(
            self._t("orbea.tables.output", "Image output folder")
        )
        self._table_image_output_btn.setText(self._t("common.browse", "Browse"))
        table_running = bool(
            self._table_image_worker and self._table_image_worker.isRunning()
        )
        self._table_image_start_btn.setText(
            self._t("orbea.tables.stop", "Stop")
            if table_running
            else self._t("orbea.tables.download", "Download selected")
        )
        self._table_image_open_btn.setText(
            self._t("orbea.tables.open_folder", "Open download folder")
        )
        if not self._table_image_status_label.text():
            self._table_image_status_label.setText(
                self._t("orbea.tables.ready", "Orbea downloader ready")
            )
        if not self._table_image_progress_label.text():
            self._table_image_progress_label.setText(
                self._t(
                    "orbea.tables.ready.detail",
                    "Paste Orbea product pages, choose image types, then download.",
                )
            )
        self._on_table_image_urls_changed()
        self._photo_title.setText(
            self._t("orbea.photo.title", "Orbea product photos")
        )
        self._photo_subtitle.setText(
            self._t(
                "orbea.photo.subtitle",
                "Download every official colour as full-resolution images from Orbea’s product configurator.",
            )
        )
        self._photo_url_label.setText(
            self._t("orbea.photo.url", "Orbea product URLs")
        )
        self._photo_url_edit.setPlaceholderText(
            self._t(
                "orbea.photo.url.placeholder",
                "One product URL per line\nhttps://cms.orbea.com/en-au/kimu-27-h20",
            )
        )
        self._photo_output_label.setText(
            self._t("orbea.photo.output", "Photo output folder")
        )
        self._photo_output_btn.setText(self._t("common.browse", "Browse"))
        photo_running = bool(self._photo_worker and self._photo_worker.isRunning())
        self._photo_start_btn.setText(
            self._t("orbea.photo.stop", "Stop")
            if photo_running
            else self._t("orbea.photo.download", "Download all colours")
        )
        self._photo_open_btn.setText(
            self._t("orbea.photo.open_folder", "Open photos folder")
        )
        if not self._photo_status_label.text():
            self._photo_status_label.setText(
                self._t("orbea.photo.ready", "Photo downloader ready")
            )
        if not self._photo_progress_label.text():
            self._photo_progress_label.setText(
                self._t(
                    "orbea.photo.ready.detail",
                    "Paste one or more Orbea product URLs to download every published colour.",
                )
            )
        self._on_photo_urls_changed()
        self._description_title.setText(
            self._t("orbea.description.title", "Description extractor")
        )
        self._description_subtitle.setText(
            self._t(
                "orbea.description.subtitle",
                "Paste Orbea model URLs to save all visible, expanded, and carousel description text.",
            )
        )
        self._description_urls_label.setText(
            self._t("orbea.description.urls", "Orbea model URLs")
        )
        self._description_urls_edit.setPlaceholderText(
            self._t(
                "orbea.description.urls.placeholder",
                "One URL per line, for example:\nhttps://cms.orbea.com/en-au/m/kemen-adv",
            )
        )
        self._description_output_label.setText(
            self._t("orbea.description.output", "Description output folder")
        )
        self._description_output_btn.setText(self._t("common.browse", "Browse"))
        description_running = bool(
            self._description_worker and self._description_worker.isRunning()
        )
        self._description_start_btn.setText(
            self._t("orbea.description.stop", "Stop")
            if description_running
            else self._t("orbea.description.extract", "Extract descriptions")
        )
        self._description_open_btn.setText(
            self._t("orbea.description.open_folder", "Open descriptions folder")
        )
        if not self._description_status_label.text():
            self._description_status_label.setText(
                self._t("orbea.description.ready", "Description extractor ready")
            )
        if not self._description_progress_label.text():
            self._description_progress_label.setText(
                self._t(
                    "orbea.description.ready.detail",
                    "Paste one or more /m/ URLs, then extract.",
                )
            )
        if not self._stage_label.text():
            self._stage_label.setText(self._t("orbea.ready", "Ready"))
        if not self._eta_label.text():
            self._eta_label.setText("ETA —")
        if not self._progress_label.text():
            self._progress_label.setText(self._t("orbea.ready.detail", "Choose filters, then start the complete workflow."))
        stat_text = {
            "scanned": self._t("orbea.stat.scanned", "Scanned"),
            "matched": self._t("orbea.stat.matched", "Matched"),
            "review": self._t("orbea.stat.review", "Review"),
            "images": self._t("orbea.stat.images", "Table images"),
            "unavailable": self._t("orbea.stat.unavailable", "Not available"),
            "errors": self._t("orbea.stat.errors", "Errors"),
        }
        for key, text in stat_text.items():
            self._stat_labels[key].setText(text)
        self._table.setHorizontalHeaderLabels([
            self._t("orbea.col.sku", "Variant SKU"),
            self._t("orbea.col.product", "Pimbo product"),
            self._t("orbea.col.match", "Match"),
            self._t("orbea.col.url", "Orbea URL"),
            self._t("orbea.col.geometry", "Geometry"),
            self._t("orbea.col.size", "Size guide"),
        ])
        if not self._results_label.text():
            self._results_label.setText(self._t("orbea.results.empty", "Results will appear here; Excel always contains the complete run."))
        self._update_table_theme()
