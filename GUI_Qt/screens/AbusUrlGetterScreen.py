"""GUI_Qt/screens/AbusUrlGetterScreen.py

Looks up ABUS product URLs from UltraBike or bare ABUS product codes.
"""

from __future__ import annotations

import os
import re
import time
from datetime import datetime
from urllib.parse import quote_plus, urljoin

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Signal
from PySide6.QtGui import QColor
from PySide6.QtWidgets import QFileDialog, QHeaderView, QTableWidgetItem

from qfluentwidgets import FluentIcon, InfoBar, InfoBarPosition, isDarkTheme

from Config.Selectors import AbusSelectors
from GUI_Qt.screens.CastelliUrlGetterScreen import CastelliUrlGetterScreen
from GUI_Qt.widgets import show_file_saved_bar
from GUI_Qt.styles.theme_config import get_text_color


def extract_abus_code(raw_code: str) -> str:
    """Extract the bare ABUS code from UB-prefixed or plain values."""
    value = (raw_code or "").strip()
    ub_match = re.search(r"\bUB-([A-Za-z0-9]+)", value, re.IGNORECASE)
    if ub_match:
        return ub_match.group(1)
    token_match = re.search(r"[A-Za-z0-9]{4,}", value)
    return token_match.group(0) if token_match else value


class AbusUrlGetterWorker(QThread):
    row_update = Signal(int, str, str)
    progress_update = Signal(int, int)
    done = Signal(int, int)
    log = Signal(str)

    def __init__(self, driver, rows: list[dict[str, str]]):
        super().__init__()
        self.driver = driver
        self.rows = rows
        self._stop = False

    def request_stop(self):
        self._stop = True

    def run(self):
        found = 0
        errors = 0

        for idx, row in enumerate(self.rows):
            if self._stop:
                self.row_update.emit(idx, "Skipped", "")
                continue

            search_code = row["search_code"]
            self.progress_update.emit(idx + 1, len(self.rows))
            self.row_update.emit(idx, "Searching", "")
            self.log.emit(f"Searching {search_code}")

            try:
                urls = self._lookup_urls(search_code)
                if urls:
                    found += 1
                    status = "Found" if len(urls) == 1 else "Multiple results"
                    self.row_update.emit(idx, status, "\n".join(urls))
                else:
                    errors += 1
                    self.row_update.emit(idx, "Not found", "")
            except Exception as e:
                errors += 1
                self.row_update.emit(idx, f"Error: {e}", "")

        self.done.emit(found, errors)

    def _lookup_urls(self, search_code: str) -> list[str]:
        search_url = AbusSelectors.SEARCH_URL.format(query=quote_plus(search_code))
        self.log.emit(f"Opening {search_url}")
        self._navigate_to(search_url)
        self._accept_cookies()
        urls = self._wait_for_search_results(search_code)
        if urls or self._has_no_results():
            return urls

        self._navigate_to(AbusSelectors.BASE_URL)
        self._accept_cookies()
        return self._search_with_page_ui(search_code)

    def _navigate_to(self, url: str):
        try:
            self.driver.switch_to.alert.accept()
            time.sleep(0.2)
        except Exception:
            pass

        try:
            self.driver.get(url)
        except Exception:
            self.driver.execute_script("window.location.href = arguments[0];", url)

    def _accept_cookies(self) -> bool:
        for locator in AbusSelectors.COOKIE_BUTTON_CANDIDATES:
            try:
                for btn in self.driver.find_elements(*locator):
                    if btn.is_displayed() and btn.is_enabled():
                        self.driver.execute_script("arguments[0].click();", btn)
                        time.sleep(0.2)
                        return True
            except Exception:
                continue
        return False

    def _wait_for_search_results(self, search_code: str) -> list[str]:
        from selenium.webdriver.support.ui import WebDriverWait

        try:
            WebDriverWait(self.driver, 8).until(
                lambda d: self._collect_matching_urls(search_code) or self._has_no_results()
            )
        except Exception:
            pass
        return self._collect_matching_urls(search_code)

    def _search_with_page_ui(self, search_code: str) -> list[str]:
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        input_el = None

        for locator in AbusSelectors.SEARCH_INPUT_CANDIDATES:
            try:
                input_el = WebDriverWait(self.driver, 2).until(EC.visibility_of_element_located(locator))
                break
            except Exception:
                continue

        if input_el is None:
            for locator in AbusSelectors.SEARCH_OPEN_CANDIDATES:
                try:
                    opener = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable(locator))
                    self.driver.execute_script("arguments[0].click();", opener)
                    time.sleep(0.3)
                    for input_locator in AbusSelectors.SEARCH_INPUT_CANDIDATES:
                        try:
                            input_el = WebDriverWait(self.driver, 2).until(
                                EC.visibility_of_element_located(input_locator)
                            )
                            break
                        except Exception:
                            continue
                    if input_el is not None:
                        break
                except Exception:
                    continue

        if input_el is None:
            return []

        try:
            input_el.clear()
        except Exception:
            pass
        input_el.send_keys(search_code)
        time.sleep(0.2)

        clicked_show_all = False
        for locator in AbusSelectors.SHOW_ALL_RESULTS:
            try:
                button = WebDriverWait(self.driver, 2).until(EC.element_to_be_clickable(locator))
                self.driver.execute_script("arguments[0].click();", button)
                clicked_show_all = True
                break
            except Exception:
                continue

        if not clicked_show_all:
            input_el.send_keys(Keys.ENTER)

        try:
            WebDriverWait(self.driver, 6).until(
                lambda d: self._collect_matching_urls(search_code) or self._has_no_results()
            )
        except Exception:
            pass

        return self._collect_matching_urls(search_code)

    def _collect_matching_urls(self, search_code: str) -> list[str]:
        seen = set()
        exact_matches = []
        fallback_matches = []
        lowered_code = search_code.lower()

        try:
            cards = self.driver.find_elements(*AbusSelectors.PRODUCT_CARD)
        except Exception:
            cards = []

        if not cards:
            try:
                cards = self.driver.find_elements(*AbusSelectors.PRODUCT_LINK)
            except Exception:
                cards = []

        for card in cards:
            try:
                href = (card.get_attribute("href") or "").strip()
                text = (card.text or "").strip()
                title = (card.get_attribute("title") or "").strip()
                outer_html = (card.get_attribute("outerHTML") or "").strip()
            except Exception:
                continue

            if not href:
                continue

            url = urljoin(AbusSelectors.BASE_URL, href)
            if url in seen:
                continue

            seen.add(url)
            fallback_matches.append(url)

            haystack = " ".join([url, text, title, outer_html]).lower()
            if lowered_code in haystack or f"/product/{lowered_code}" in url.lower():
                exact_matches.append(url)

        if exact_matches:
            return exact_matches
        if len(fallback_matches) == 1:
            return fallback_matches

        return fallback_matches

    def _has_no_results(self) -> bool:
        try:
            return any(el.is_displayed() for el in self.driver.find_elements(*AbusSelectors.NO_RESULTS))
        except Exception:
            return False


class AbusUrlGetterScreen(CastelliUrlGetterScreen):
    def __init__(self, main_window):
        super().__init__(main_window)
        self._worker: AbusUrlGetterWorker | None = None

    def _build_ui(self):
        super()._build_ui()
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.setColumnWidth(1, 170)
        self._table.setColumnWidth(2, 140)
        self._table.setColumnWidth(3, 130)

    def retranslate_ui(self):
        self._title.setText(self.tr("abusurl.title"))
        self._subtitle.setText(self.tr("abusurl.subtitle"))
        self._browse_btn.setText(self.tr("abusurl.browse"))
        self._start_btn.setText(self.tr("abusurl.start"))
        self._export_btn.setText(self.tr("abusurl.export"))
        self._table.setHorizontalHeaderLabels([
            self.tr("abusurl.col.index"),
            self.tr("abusurl.col.original_code"),
            self.tr("abusurl.col.search_code"),
            self.tr("abusurl.col.status"),
            self.tr("abusurl.col.url"),
        ])

    def _load_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            rows: list[dict[str, str]] = []

            for excel_row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
                value = excel_row[0]
                if value is None:
                    continue
                original_code = str(value).strip()
                if not original_code:
                    continue
                if original_code.strip().lower() in {"code", "product code", "original code"}:
                    continue
                rows.append({
                    "original_code": original_code,
                    "search_code": extract_abus_code(original_code),
                })
            wb.close()

            if not rows:
                InfoBar.warning(
                    self.tr("common.error"),
                    self.tr("abusurl.no_codes"),
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                )
                return

            self._rows = rows
            self._file_path = path
            self._file_label.setText(os.path.basename(path))
            self._table.setRowCount(len(rows))

            for idx, row in enumerate(rows):
                self._table.setItem(idx, 0, QTableWidgetItem(str(idx + 1)))
                self._table.setItem(idx, 1, QTableWidgetItem(row["original_code"]))
                self._table.setItem(idx, 2, QTableWidgetItem(row["search_code"]))
                status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
                status.setForeground(QColor(get_text_color(isDarkTheme(), "secondary")))
                self._table.setItem(idx, 3, status)
                self._table.setItem(idx, 4, QTableWidgetItem(""))

            self._start_btn.setEnabled(True)
            self._export_btn.setEnabled(False)
            self._drop_zone.setVisible(False)
            self._progress_label.setText(self.tr("abusurl.loaded", count=len(rows)))
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        if not self._rows:
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        for idx in range(self._table.rowCount()):
            status = QTableWidgetItem(self.tr("batchdesc.status.pending"))
            status.setForeground(QColor(get_text_color(isDarkTheme(), "secondary")))
            self._table.setItem(idx, 3, status)
            self._table.setItem(idx, 4, QTableWidgetItem(""))

        self._start_btn.setText(self.tr("abusurl.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._browse_btn.setEnabled(False)
        self._export_btn.setEnabled(False)

        self._worker = AbusUrlGetterWorker(driver, self._rows)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[AbusUrlGetter] {msg}"))
        self._worker.start()

    def _on_progress(self, current: int, total: int):
        self._progress_label.setText(self.tr("abusurl.progress", current=current, total=total))

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("abusurl.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._browse_btn.setEnabled(True)
        self._export_btn.setEnabled(True)

        total = found + errors
        text = self.tr("abusurl.done", found=found, errors=errors, total=total)
        self._progress_label.setText(text)
        InfoBar.success(
            self.tr("abusurl.done_title"),
            text,
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("abusurl.export"),
            f"abus_urls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ABUS URLs"

            headers = ["#", "Original Code", "Search Code", "Status", "URL"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(5):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 100

            wb.save(path)
            show_file_saved_bar(self, self.tr("common.success"), self.tr("abusurl.exported"), path)
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )
