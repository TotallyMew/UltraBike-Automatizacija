"""
History Screen - Fluent Design System
Space Indigo/Lavender Grey color scheme with FluentIcons
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFileDialog, QLabel
)
from PySide6.QtCore import Qt, QUrl
from PySide6.QtGui import QPixmap, QIcon, QDesktopServices, QCursor
from qfluentwidgets import (
    ComboBox, PushButton, TransparentToolButton, FluentIcon,
    BodyLabel, TitleLabel, StrongBodyLabel, CardWidget, CaptionLabel,
    InfoBar, InfoBarPosition, ScrollArea, isDarkTheme, IconWidget, qconfig
)
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from GUI_Qt.styles.theme_config import COLORS, FONTS
from PySide6.QtWidgets import QSizePolicy


class ClickableLabel(BodyLabel):
    """Clickable label that opens URLs in browser"""

    def __init__(self, text, url, parent=None):
        super().__init__(parent)
        self.setText(text)
        self.url = url
        self.setCursor(QCursor(Qt.CursorShape.PointingHandCursor))
        self.setWordWrap(False)

    def mousePressEvent(self, event):
        """Open URL in default browser when clicked"""
        if event.button() == Qt.MouseButton.LeftButton:
            QDesktopServices.openUrl(QUrl(self.url))
        super().mousePressEvent(event)

    def enterEvent(self, event):
        """Add underline on hover"""
        current_style = self.styleSheet()
        if 'text-decoration' not in current_style:
            self.setStyleSheet(current_style + " text-decoration: underline;")
        super().enterEvent(event)

    def leaveEvent(self, event):
        """Remove underline when not hovering"""
        current_style = self.styleSheet()
        self.setStyleSheet(current_style.replace(" text-decoration: underline;", ""))
        super().leaveEvent(event)


class StatCard(CardWidget):
    """Statistics card with FluentIcon"""

    def __init__(self, label, value, icon, parent=None):
        super().__init__(parent)
        self.setBorderRadius(8)
        self.setFixedHeight(110)

        layout = QHBoxLayout(self)
        layout.setSpacing(16)
        layout.setContentsMargins(20, 20, 20, 20)

        # Icon on the left
        icon_widget = IconWidget(icon)
        icon_widget.setFixedSize(40, 40)
        layout.addWidget(icon_widget)

        # Value and label on the right
        text_layout = QVBoxLayout()
        text_layout.setSpacing(4)

        self.value_label = TitleLabel(value)
        self.value_label.setAlignment(Qt.AlignmentFlag.AlignLeft)

        self.label_label = CaptionLabel(label)
        self.label_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.label_label.setAlignment(Qt.AlignmentFlag.AlignLeft)

        text_layout.addWidget(self.value_label)
        text_layout.addWidget(self.label_label)
        text_layout.addStretch()

        layout.addLayout(text_layout, 1)


from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QSizePolicy
from PySide6.QtCore import Qt

class HistoryItemCard(CardWidget):
    """Single history item card with status indicator - improved layout"""

    def __init__(self, row, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self._expanded = False
        self.setBorderRadius(8)
        self.setMinimumHeight(160)  # Increased to prevent clipping

        self._details_json = None
        try:
            self._details_json = row['details_json'] if 'details_json' in row.keys() else None
        except Exception:
            self._details_json = None

        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)

        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(32, 24, 32, 24)

        # --- TOP ROW ---
        top_row = QHBoxLayout()
        top_row.setSpacing(24)

        # STATUS ICON
        is_success = row['status'] == 'success'
        status_icon = FluentIcon.ACCEPT if is_success else FluentIcon.CANCEL
        status_color = COLORS['success'] if is_success else COLORS['error']

        icon_widget = IconWidget(status_icon)
        icon_widget.setFixedSize(32, 32)
        icon_widget.setStyleSheet(f"color: {status_color}; background: transparent;")
        top_row.addWidget(icon_widget, 0, Qt.AlignmentFlag.AlignVCenter)

        # LEFT: Brand / Product Code
        left_section = QVBoxLayout()
        left_section.setSpacing(2)

        title = StrongBodyLabel(row['brand'])
        title.setStyleSheet(f"font-size: 15px; font-weight: 600; color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']};")

        product_code = BodyLabel(row['product_code'])
        product_code.setStyleSheet(f"font-size: 12px; color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['space_indigo']};")

        left_section.addWidget(title)
        left_section.addWidget(product_code)

        left_widget = QWidget()
        left_widget.setLayout(left_section)
        left_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        top_row.addWidget(left_widget, 2)  # Stretch factor 2

        # VERTICAL SEPARATOR
        sep1 = QWidget()
        sep1.setFixedWidth(1)
        sep1.setStyleSheet(f"background-color: {COLORS['border_dark'] if isDarkTheme() else COLORS['border_light']};")
        top_row.addWidget(sep1)

        # METRICS
        metrics_layout = QHBoxLayout()
        metrics_layout.setSpacing(24)

        for label_text, value in [
            (self.tr("history.metric.duration"), f"{row['duration_seconds']:.1f}s" if row['duration_seconds'] else self.tr("common.na")),
            (self.tr("history.metric.features"), str(row['features_uploaded']) if 'features_uploaded' in row.keys() and row['features_uploaded'] else "0"),
            (self.tr("history.metric.images"), str(row['images_uploaded']) if 'images_uploaded' in row.keys() and row['images_uploaded'] else "0"),
        ]:
            v_layout = QVBoxLayout()
            v_layout.setSpacing(2)
            v_layout.setContentsMargins(0, 0, 0, 0)

            label = CaptionLabel(label_text)
            label.setStyleSheet(f"font-size: 10px; text-transform: uppercase; color: {COLORS['text_tertiary']};")
            value_label = StrongBodyLabel(value)
            value_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']};")
            value_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)

            v_layout.addWidget(label, 0, Qt.AlignmentFlag.AlignHCenter)
            v_layout.addWidget(value_label, 1, Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)

            container = QWidget()
            container.setLayout(v_layout)
            container.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
            metrics_layout.addWidget(container)

        metrics_widget = QWidget()
        metrics_widget.setLayout(metrics_layout)
        metrics_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        top_row.addWidget(metrics_widget, 3)  # Stretch factor 3 to give metrics more room

        # VERTICAL SEPARATOR
        sep2 = QWidget()
        sep2.setFixedWidth(1)
        sep2.setStyleSheet(f"background-color: {COLORS['border_dark'] if isDarkTheme() else COLORS['border_light']};")
        top_row.addWidget(sep2)

        # RIGHT: Timestamp / Failed stage
        right_layout = QVBoxLayout()
        right_layout.setSpacing(2)
        right_layout.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        timestamp = CaptionLabel(str(row['processed_at']))
        timestamp.setStyleSheet(f"font-size: 11px; color: {COLORS['text_tertiary']};")
        right_layout.addWidget(timestamp)

        if not is_success and 'failed_stage' in row.keys() and row['failed_stage']:
            failed_label = CaptionLabel(self.tr("history.failed_prefix", stage=row['failed_stage']))
            failed_label.setStyleSheet(f"font-size: 11px; color: {COLORS['error']}; font-weight: 500;")
            failed_label.setAlignment(Qt.AlignmentFlag.AlignRight)
            right_layout.addWidget(failed_label)

        # Expand details toggle (always visible so users can discover it)
        self._details_btn = None
        try:
            self._details_btn = TransparentToolButton(FluentIcon.DOWN, self)
            self._details_btn.setFixedSize(28, 28)
            self._details_btn.setToolTip(self.tr("history.details"))
            self._details_btn.clicked.connect(self._toggle_details)
            right_layout.addWidget(self._details_btn, 0, Qt.AlignmentFlag.AlignRight)
        except Exception:
            self._details_btn = None

        right_widget = QWidget()
        right_widget.setLayout(right_layout)
        right_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        top_row.addWidget(right_widget, 1)  # Stretch factor 1

        main_layout.addLayout(top_row)

        # --- Details section (collapsed by default) ---
        self._details_card = CardWidget()
        self._details_card.setVisible(False)
        self._details_card.setStyleSheet(
            f"background-color: {'rgba(141,153,174,0.10)' if isDarkTheme() else 'rgba(43,45,66,0.04)'};"
            f"border: 1px solid {COLORS['border_dark'] if isDarkTheme() else COLORS['border_light']};"
        )
        details_layout = QVBoxLayout(self._details_card)
        details_layout.setContentsMargins(12, 10, 12, 10)
        details_layout.setSpacing(6)

        self._details_label = BodyLabel("")
        self._details_label.setWordWrap(True)
        self._details_label.setStyleSheet(
            f"font-size: 11px; color: {COLORS['text_primary_dark'] if isDarkTheme() else COLORS['text_primary_light']};"
        )
        details_layout.addWidget(self._details_label)
        main_layout.addWidget(self._details_card)

        # --- BOTTOM ROW: URL/Code ---
        url_or_code = row['url_or_code'] if 'url_or_code' in row.keys() else None
        if url_or_code:
            main_layout.addSpacing(12)
            divider = QWidget()
            divider.setFixedHeight(1)
            divider.setStyleSheet(f"background-color: {COLORS['border_dark'] if isDarkTheme() else COLORS['border_light']};")
            main_layout.addWidget(divider)
            main_layout.addSpacing(12)

            url_layout = QHBoxLayout()
            url_layout.setSpacing(10)

            url_icon = IconWidget(FluentIcon.LINK)
            url_icon.setFixedSize(16, 16)
            url_icon.setStyleSheet(f"color: {COLORS['lavender_grey']};")

            display_text = str(url_or_code)
            if len(display_text) > 120:
                display_text = display_text[:117] + "..."

            url_label = ClickableLabel(display_text, str(url_or_code))
            url_label.setStyleSheet(f"font-size: 12px; color: {COLORS['lavender_grey']}; font-family: {FONTS['family_mono']};")

            url_layout.addWidget(url_icon)
            url_layout.addWidget(url_label)
            url_layout.addStretch()
            main_layout.addLayout(url_layout)

        # --- Error message ---
        if not is_success and 'error_message' in row.keys() and row['error_message']:
            error_card = CardWidget()
            error_card.setStyleSheet(f"background-color: {'rgba(200,29,37,0.08)' if isDarkTheme() else 'rgba(200,29,37,0.05)'}; border: 1px solid {COLORS['error']};")
            error_layout = QHBoxLayout(error_card)
            error_layout.setContentsMargins(12, 8, 12, 8)

            error_icon = IconWidget(FluentIcon.INFO)
            error_icon.setFixedSize(16, 16)
            error_icon.setStyleSheet(f"color: {COLORS['error']};")

            error_label = BodyLabel(self.tr("history.error_prefix", error=row['error_message']))
            error_label.setStyleSheet(f"font-size: 11px; color: {COLORS['error']};")
            error_label.setWordWrap(True)

            error_layout.addWidget(error_icon)
            error_layout.addWidget(error_label, 1)
            main_layout.addWidget(error_card)

    def _toggle_details(self):
        self._expanded = not self._expanded
        self._details_card.setVisible(self._expanded)

        if self._details_btn is not None:
            try:
                self._details_btn.setIcon(FluentIcon.UP if self._expanded else FluentIcon.DOWN)
            except Exception:
                pass

        if self._expanded:
            self._details_label.setText(self._format_details(self._details_json))

    def _format_details(self, details_json: str) -> str:
        if not details_json:
            return self.tr("history.details.empty")
        try:
            import json
            data = json.loads(details_json) or {}
        except Exception:
            return str(details_json)

        lines = []
        skipped = data.get("skipped_features")
        if isinstance(skipped, list) and skipped:
            # Show up to a reasonable number; keep UI compact
            shown = skipped[:30]
            items = []
            for item in shown:
                try:
                    items.append(str(item.get("key") or ""))
                except Exception:
                    continue
            items = [s for s in items if s]
            if items:
                lines.append(self.tr("history.details.skipped_features") + ":")
                lines.append("\n".join(f"- {k}" for k in items))
                if len(skipped) > len(shown):
                    lines.append(self.tr("history.details.more", count=len(skipped) - len(shown)))

        if not lines:
            return self.tr("history.details.empty")

        return "\n".join(lines)


class HistoryScreen(QWidget):
    """History screen with filtering, statistics, and Excel export"""

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self._FILTER_ALL = "__all__"
        self.current_page = 0
        self.items_per_page = 10
        self.all_history = []
        self.filtered_history = []
        self._init_ui()
        self.refresh_history()

        # Connect to theme change signal
        qconfig.themeChangedFinished.connect(self._on_theme_changed)

    def _init_ui(self):
        """Initialize UI with Fluent Design"""
        # Apply background color based on theme
        is_dark = isDarkTheme()
        bg_color = COLORS['space_indigo'] if is_dark else COLORS['platinum']

        # Force the background color
        self.setStyleSheet(f"""
            HistoryScreen {{
                background-color: {bg_color};
                font-family: {FONTS['family']};
            }}
        """)
        self.setAutoFillBackground(True)

        # Main layout
        layout = QVBoxLayout(self)
        layout.setContentsMargins(40, 20, 40, 20)  # Fluent standard: 40px sides
        layout.setSpacing(20)

        # === HEADER SECTION ===
        header = QHBoxLayout()
        self.title_label = TitleLabel("")
        header.addWidget(self.title_label)
        header.addStretch()
        layout.addLayout(header)

        # === STATISTICS CARDS ===
        stats_container = QHBoxLayout()
        stats_container.setSpacing(16)

        # Using FluentIcons for stats
        self.total_stat = StatCard("", "0", FluentIcon.FOLDER)
        self.success_stat = StatCard("", "0%", FluentIcon.ACCEPT)
        self.duration_stat = StatCard("", "0s", FluentIcon.HISTORY)
        self.today_stat = StatCard("", "0", FluentIcon.CALENDAR)

        stats_container.addWidget(self.total_stat, 1)
        stats_container.addWidget(self.success_stat, 1)
        stats_container.addWidget(self.duration_stat, 1)
        stats_container.addWidget(self.today_stat, 1)

        layout.addLayout(stats_container)

        # === TOOLBAR SECTION ===
        toolbar_card = CardWidget()
        toolbar_card.setBorderRadius(8)
        toolbar_layout = QHBoxLayout(toolbar_card)
        toolbar_layout.setContentsMargins(24, 20, 24, 20)  # Fluent standard card padding
        toolbar_layout.setSpacing(16)  # Fluent 4px increment

        # Brand filter
        self.brand_label = BodyLabel("")
        brand_label = self.brand_label
        brand_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.brand_filter = ComboBox()
        self.brand_filter.setMinimumWidth(140)
        self.brand_filter.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.brand_filter.currentTextChanged.connect(self.refresh_history)

        # Status filter
        self.status_label = BodyLabel("")
        status_label = self.status_label
        status_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.status_filter = ComboBox()
        self.status_filter.setMinimumWidth(120)
        self.status_filter.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.status_filter.currentTextChanged.connect(self.refresh_history)

        # Date filter
        self.period_label = BodyLabel("")
        date_label = self.period_label
        date_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self.date_filter = ComboBox()
        self.date_filter.setMinimumWidth(140)
        self.date_filter.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.date_filter.currentTextChanged.connect(self.refresh_history)

        # Refresh button
        self.refresh_button = TransparentToolButton(FluentIcon.SYNC, self)
        self.refresh_button.clicked.connect(self.refresh_history)

        # Export button
        self.export_button = PushButton("")
        self.export_button.setIcon(FluentIcon.DOCUMENT)
        self.export_button.clicked.connect(self.export_to_excel)

        toolbar_layout.addWidget(brand_label)
        toolbar_layout.addWidget(self.brand_filter)
        toolbar_layout.addWidget(status_label)
        toolbar_layout.addWidget(self.status_filter)
        toolbar_layout.addWidget(date_label)
        toolbar_layout.addWidget(self.date_filter)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(self.refresh_button)
        toolbar_layout.addWidget(self.export_button)

        layout.addWidget(toolbar_card)

        # === HISTORY LIST ===
        scroll = ScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background: transparent;
            }
            QScrollArea > QWidget > QWidget {
                background: transparent;
            }
        """)

        # CRITICAL: Prevent viewport from clipping child widgets
        viewport = scroll.viewport()
        if viewport:
            viewport.setAutoFillBackground(False)
            viewport.setAttribute(Qt.WidgetAttribute.WA_OpaquePaintEvent, False)

        self.history_container = QWidget()
        self.history_container.setStyleSheet("background: transparent;")
        self.history_layout = QVBoxLayout(self.history_container)
        self.history_layout.setSpacing(16)  # Spacing between cards
        self.history_layout.setContentsMargins(5, 15, 5, 15)  # Generous margins to prevent any clipping
        self.history_layout.addStretch()

        # No centering wrapper - let history items expand naturally
        scroll.setWidget(self.history_container)

        # Add spacing between toolbar and scroll area
        layout.addSpacing(16)
        layout.addWidget(scroll, 1)

        # === PAGINATION CONTROLS ===
        pagination_card = CardWidget()
        pagination_card.setBorderRadius(8)
        pagination_layout = QHBoxLayout(pagination_card)
        pagination_layout.setContentsMargins(20, 12, 20, 12)
        pagination_layout.setSpacing(16)

        self.prev_btn = TransparentToolButton(FluentIcon.LEFT_ARROW, self)
        self.prev_btn.setFixedSize(32, 32)
        self.prev_btn.clicked.connect(self._prev_page)
        self.prev_btn.setEnabled(False)

        self.page_label = BodyLabel("")
        self.page_label.setStyleSheet(f"color: {COLORS['text_secondary']}; font-weight: 500;")

        self.next_btn = TransparentToolButton(FluentIcon.RIGHT_ARROW, self)
        self.next_btn.setFixedSize(32, 32)
        self.next_btn.clicked.connect(self._next_page)
        self.next_btn.setEnabled(False)

        self.items_per_page_combo = ComboBox()
        self.items_per_page_combo.addItems(["10", "20", "50", "100"])
        self.items_per_page_combo.setCurrentText("10")
        self.items_per_page_combo.setFixedWidth(80)
        self.items_per_page_combo.currentTextChanged.connect(self._on_items_per_page_changed)

        self.items_label = CaptionLabel("")
        items_label = self.items_label
        items_label.setStyleSheet(f"color: {COLORS['text_tertiary']};")

        pagination_layout.addWidget(self.prev_btn)
        pagination_layout.addWidget(self.page_label)
        pagination_layout.addWidget(self.next_btn)
        pagination_layout.addStretch()
        pagination_layout.addWidget(self.items_per_page_combo)
        pagination_layout.addWidget(items_label)

        layout.addWidget(pagination_card)

        self._populate_filters()
        self.retranslate_ui()

    def _populate_filters(self):
        """Populate filter combos with translated labels while keeping stable userData."""
        tr = self.main.i18n.tr

        # Preserve selection by userData
        current_brand = self.brand_filter.currentData()
        current_status = self.status_filter.currentData()
        current_period = self.date_filter.currentData()

        self.brand_filter.blockSignals(True)
        self.brand_filter.clear()
        self.brand_filter.addItem(tr("common.all"), self._FILTER_ALL)
        for brand in ["KROSS", "Pinarello", "Basso", "Factor", "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan"]:
            self.brand_filter.addItem(brand, brand)
        if current_brand is not None:
            idx = self.brand_filter.findData(current_brand)
            if idx >= 0:
                self.brand_filter.setCurrentIndex(idx)
        self.brand_filter.blockSignals(False)

        self.status_filter.blockSignals(True)
        self.status_filter.clear()
        self.status_filter.addItem(tr("common.all"), self._FILTER_ALL)
        self.status_filter.addItem(tr("common.success"), "success")
        self.status_filter.addItem(tr("common.failed"), "failed")
        if current_status is not None:
            idx = self.status_filter.findData(current_status)
            if idx >= 0:
                self.status_filter.setCurrentIndex(idx)
        self.status_filter.blockSignals(False)

        self.date_filter.blockSignals(True)
        self.date_filter.clear()
        self.date_filter.addItem(tr("common.all_time"), self._FILTER_ALL)
        self.date_filter.addItem(tr("common.today"), "today")
        self.date_filter.addItem(tr("common.last_7_days"), "last_7")
        self.date_filter.addItem(tr("common.last_30_days"), "last_30")
        if current_period is not None:
            idx = self.date_filter.findData(current_period)
            if idx >= 0:
                self.date_filter.setCurrentIndex(idx)
        self.date_filter.blockSignals(False)

    def retranslate_ui(self):
        tr = self.main.i18n.tr

        self.title_label.setText(tr("history.title"))
        self.total_stat.label_label.setText(tr("history.stats.total"))
        self.success_stat.label_label.setText(tr("history.stats.success_rate"))
        self.duration_stat.label_label.setText(tr("history.stats.avg_duration"))
        self.today_stat.label_label.setText(tr("history.stats.today"))

        self.brand_label.setText(f"{tr('common.brand')}:" if not tr('common.brand').endswith(":") else tr('common.brand'))
        self.status_label.setText(f"{tr('common.status')}:" if not tr('common.status').endswith(":") else tr('common.status'))
        self.period_label.setText(f"{tr('common.period')}:" if not tr('common.period').endswith(":") else tr('common.period'))

        self.refresh_button.setToolTip(tr("history.refresh"))
        self.export_button.setText(tr("common.export_excel"))
        self.prev_btn.setToolTip(tr("common.previous_page"))
        self.next_btn.setToolTip(tr("common.next_page"))
        self.items_label.setText(tr("common.items_per_page"))

        # Rebuild filter item text for current language
        self._populate_filters()

        # Re-render to update page label + no-data text + item cards
        self._render_page()

    def refresh_history(self):
        """Refresh history list based on filters"""
        # Database already has sqlite3.Row factory set
        cursor = self.main.db.conn.cursor()

        # Build query with filters
        query = """
            SELECT brand, product_code, url_or_code, status, duration_seconds,
                   features_uploaded, images_uploaded, error_message,
                   failed_stage, details_json, processed_at
            FROM processing_history
            WHERE 1=1
        """
        params = []

        # Brand filter (case insensitive)
        brand = self.brand_filter.currentData()
        if brand and brand != self._FILTER_ALL:
            query += " AND UPPER(brand) = UPPER(?)"
            params.append(brand)

        # Status filter
        status_value = self.status_filter.currentData()
        if status_value and status_value != self._FILTER_ALL:
            query += " AND status = ?"
            params.append(status_value)

        # Date filter
        date_range = self.date_filter.currentData()
        if date_range and date_range != self._FILTER_ALL:
            if date_range == "today":
                date_threshold = datetime.now().strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) = ?"
                params.append(date_threshold)
            elif date_range == "last_7":
                date_threshold = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) >= ?"
                params.append(date_threshold)
            elif date_range == "last_30":
                date_threshold = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                query += " AND DATE(processed_at) >= ?"
                params.append(date_threshold)

        query += " ORDER BY processed_at DESC"

        # Execute query and store all results
        self.all_history = cursor.execute(query, params).fetchall()
        self.filtered_history = self.all_history

        # Reset to first page
        self.current_page = 0
        self._render_page()

        # Update statistics
        self._update_stats()

    def _render_page(self):
        """Render current page of history items"""
        # Clear existing items (keep the stretch at the end)
        while self.history_layout.count() > 1:
            item = self.history_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Calculate pagination
        total_items = len(self.filtered_history)
        total_pages = max(1, (total_items + self.items_per_page - 1) // self.items_per_page)
        start_idx = self.current_page * self.items_per_page
        end_idx = min(start_idx + self.items_per_page, total_items)

        # Update pagination controls
        self.page_label.setText(self.main.i18n.tr("history.page", current=self.current_page + 1, total=total_pages))
        self.prev_btn.setEnabled(self.current_page > 0)
        self.next_btn.setEnabled(self.current_page < total_pages - 1)

        # Add history items for current page
        if self.filtered_history:
            page_items = self.filtered_history[start_idx:end_idx]
            for row in page_items:
                card = HistoryItemCard(row, self.main.i18n.tr)
                self.history_layout.insertWidget(self.history_layout.count() - 1, card)
        else:
            no_data = BodyLabel(self.main.i18n.tr("history.no_records"))
            no_data.setStyleSheet(f"color: {COLORS['text_tertiary']}; padding: 40px 20px;")
            no_data.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.history_layout.insertWidget(0, no_data)

    def _prev_page(self):
        """Go to previous page"""
        if self.current_page > 0:
            self.current_page -= 1
            self._render_page()

    def _next_page(self):
        """Go to next page"""
        total_pages = max(1, (len(self.filtered_history) + self.items_per_page - 1) // self.items_per_page)
        if self.current_page < total_pages - 1:
            self.current_page += 1
            self._render_page()

    def _on_items_per_page_changed(self, text):
        """Handle items per page change"""
        self.items_per_page = int(text)
        self.current_page = 0
        self._render_page()

    def _update_stats(self):
        """Update statistics cards"""
        cursor = self.main.db.conn.cursor()

        # Total uploads
        total = cursor.execute("SELECT COUNT(*) FROM processing_history").fetchone()[0]
        self.total_stat.value_label.setText(str(total))

        # Success rate
        success = cursor.execute(
            "SELECT COUNT(*) FROM processing_history WHERE status = 'success'"
        ).fetchone()[0]
        success_rate = (success / total * 100) if total > 0 else 0
        self.success_stat.value_label.setText(f"{success_rate:.1f}%")

        # Average duration
        avg_duration = cursor.execute(
            "SELECT AVG(duration_seconds) FROM processing_history WHERE status = 'success'"
        ).fetchone()[0] or 0
        self.duration_stat.value_label.setText(f"{avg_duration:.1f}s")

        # Today's uploads
        today = datetime.now().strftime('%Y-%m-%d')
        today_count = cursor.execute(
            "SELECT COUNT(*) FROM processing_history WHERE DATE(processed_at) = ?",
            (today,)
        ).fetchone()[0]
        self.today_stat.value_label.setText(str(today_count))

    def export_to_excel(self):
        """Export history to Excel file"""
        try:
            # Get filtered data
            cursor = self.main.db.conn.cursor()

            query = """
                SELECT brand, product_code, url_or_code, status, duration_seconds,
                       features_uploaded, images_uploaded, error_message, processed_at
                FROM processing_history
                WHERE 1=1
            """
            params = []

            # Apply same filters (case insensitive for brand)
            brand = self.brand_filter.currentData()
            if brand and brand != self._FILTER_ALL:
                query += " AND UPPER(brand) = UPPER(?)"
                params.append(brand)

            status_value = self.status_filter.currentData()
            if status_value and status_value != self._FILTER_ALL:
                query += " AND status = ?"
                params.append(status_value)

            date_range = self.date_filter.currentData()
            if date_range and date_range != self._FILTER_ALL:
                if date_range == "today":
                    date_threshold = datetime.now().strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) = ?"
                    params.append(date_threshold)
                elif date_range == "last_7":
                    date_threshold = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) >= ?"
                    params.append(date_threshold)
                elif date_range == "last_30":
                    date_threshold = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                    query += " AND DATE(processed_at) >= ?"
                    params.append(date_threshold)

            query += " ORDER BY processed_at DESC"

            history = cursor.execute(query, params).fetchall()

            if not history:
                InfoBar.warning(
                    title=self.main.i18n.tr("history.no_data.title"),
                    content=self.main.i18n.tr("history.no_data.content"),
                    parent=self,
                    position=InfoBarPosition.TOP
                )
                return

            # Create Excel file
            file_dialog = QFileDialog()
            file_path, _ = file_dialog.getSaveFileName(
                self,
                self.main.i18n.tr("history.export.title"),
                f"history_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                self.main.i18n.tr("history.excel.filter")
            )

            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = self.main.i18n.tr("history.sheet.title")

            # Headers with Space Indigo background and Lavender Grey text
            headers = ["Brand", "Product Code", "URL/Code", "Status", "Duration (s)",
                      "Features", "Images", "Error", "Processed At"]
            ws.append(headers)

            header_fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
            header_font = Font(bold=True, color="8D99AE")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Data rows with color-coded status
            for row in history:
                ws.append(row)

                # Color code status column (column D)
                try:
                    status_cell = ws.cell(ws.max_row, 4)
                    if row[3] == "success":  # Status column
                        # Success: Emerald Green (#10B981)
                        status_cell.font = Font(color="10B981", bold=True)
                    else:
                        # Failed: Flag Red (#C81D25)
                        status_cell.font = Font(color="C81D25", bold=True)
                except:
                    pass

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)

            InfoBar.success(
                title=self.main.i18n.tr("history.export.success.title"),
                content=self.main.i18n.tr("history.export.success.content", path=file_path),
                parent=self,
                position=InfoBarPosition.TOP
            )

        except Exception as ex:
            InfoBar.error(
                title=self.main.i18n.tr("history.export.failed.title"),
                content=str(ex),
                parent=self,
                position=InfoBarPosition.TOP
            )

    def _on_theme_changed(self):
        """Handle theme change event"""
        is_dark = isDarkTheme()
        bg_color = COLORS['space_indigo'] if is_dark else COLORS['platinum']

        self.setStyleSheet(f"""
            HistoryScreen {{
                background-color: {bg_color};
                font-family: {FONTS['family']};
            }}
        """)
