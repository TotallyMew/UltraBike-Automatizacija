"""Earnings tracker UI: entries, timer, goals, forecasts, and exports."""

from __future__ import annotations

import math
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from PySide6.QtCore import QDate, QEvent, QPoint, QRectF, QSize, Qt, QTimer, QUrl
from PySide6.QtGui import QBrush, QColor, QFont, QPainter, QPen, QTextCharFormat
from PySide6.QtMultimedia import QSoundEffect
from PySide6.QtWidgets import (
    QAbstractItemView,
    QAbstractSpinBox,
    QCalendarWidget,
    QCheckBox,
    QDateEdit,
    QDateTimeEdit,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLayout,
    QMessageBox,
    QProgressBar,
    QSizePolicy,
    QSpinBox,
    QStackedWidget,
    QTabBar,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTableView,
    QToolButton,
    QToolTip,
    QVBoxLayout,
    QWidget,
)
from qfluentwidgets import (
    Action,
    BodyLabel,
    CaptionLabel,
    CheckBox,
    ComboBox,
    DateEdit as FluentDateEdit,
    DoubleSpinBox as FluentDoubleSpinBox,
    FluentIcon,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    MessageBox,
    PrimaryPushButton,
    PushButton,
    RoundMenu,
    ScrollArea,
    SpinBox as FluentSpinBox,
    StrongBodyLabel,
    TitleLabel,
    isDarkTheme,
    qconfig,
)

from GUI_Qt.styles.screen_theme import apply_screen_theme
from GUI_Qt.styles.theme_config import (
    COLORS,
    COMPONENT_COLORS,
    FONTS,
    PADDINGS,
    RADII,
    SIZES,
    SPACING,
    get_dialog_button_style,
    get_dialog_danger_button_style,
    get_dialog_section_style,
    get_dialog_table_style,
    get_calendar_popup_style,
    get_form_dialog_style,
    get_form_input_style,
    get_accent_colors,
    get_selection_bg,
    get_status_text_color,
    get_subtle_border,
    get_subtle_item_hover_bg,
    get_surface_color,
    get_text_color,
    rgba_from_hex,
)
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from Managers.EarningsManager import (
    ActiveGoalError,
    ActiveSessionError,
    EarningsManager,
    GoalStatus,
    ProductType,
    QuestKind,
)


from GUI_Qt.earnings.presentation import (
    PRODUCT_TYPES, duration, goal_level_state, goal_progress_state, local_datetime, money,
)
from GUI_Qt.earnings.widgets import (
    ActivityHeatmap, ActivityLegend, AnimatedSubmitButton, BatchProgressTicks, EarningsBurstBadge,
    EarningsChart, FluentCalendarWidget, GoalMilestoneBar, MetricCard, ProjectionMetric,
    PerformanceTargetWidget, QuestCelebrationOverlay, QuestProgressWidget,
    apply_earnings_datetime_theme, configure_earnings_datetime_edit,
)
from GUI_Qt.earnings.dialogs import (
    BrandManagerDialog, BrandNameDialog, BulkEarningEditDialog,
    EarningEntryDialog, EarningsSettingsDialog, GoalAdjustmentDialog, GoalDialog,
    QuestPickerDialog, SessionRecapDialog, UploadEarningsDialog,
)
from Utilities.ResourcePaths import resource_path

class EarningsScreen(ResponsiveWidget):
    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.service: EarningsManager = getattr(
            main_window, "earnings_manager", EarningsManager(main_window.db, main_window.settings)
        )
        self._entries: list[dict[str, Any]] = []
        self._sessions_count = 0
        self._last_expired_session = None
        self._engagement_settings = self.service.engagement_settings()
        self._quest_observed_session_id: int | None = None
        self._quest_observed_checkpoints = 0
        self._quest_observed_complete = False
        self._quest_observed_bonus_complete = False
        self._init_ui()
        self._success_sound = QSoundEffect(self)
        self._success_sound.setSource(
            QUrl.fromLocalFile(str(resource_path("Assets/Sounds/earnings_success.wav")))
        )
        self._success_sound.setVolume(self.service.celebration_sound_volume() / 100.0)
        self._tick_timer = QTimer(self)
        self._tick_timer.timeout.connect(self._timer_tick)
        self._tick_timer.start(500)
        qconfig.themeChangedFinished.connect(self._apply_theme)
        self.refresh_all()
        self.retranslate_ui()

    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        self.scroll = ScrollArea(self)
        self.scroll.setWidgetResizable(True)
        root.addWidget(self.scroll)
        self.content = QWidget()
        self.scroll.setWidget(self.content)
        layout = QVBoxLayout(self.content)
        layout.setContentsMargins(24, 20, 28, 28)
        layout.setSpacing(10)

        self.header_layout = QGridLayout()
        self.header_layout.setHorizontalSpacing(8)
        self.header_layout.setVerticalSpacing(8)
        self.title = TitleLabel("Earnings")
        self.brands_button = PushButton("Brands")
        self.brands_button.setObjectName("earningsHeaderAction")
        self.brands_button.clicked.connect(self._manage_brands)
        self.settings_button = PushButton("Settings")
        self.settings_button.setObjectName("earningsHeaderAction")
        self.settings_button.clicked.connect(self._settings)
        self.export_button = PushButton("Export")
        self.export_button.setObjectName("earningsHeaderAction")
        self.export_menu = RoundMenu(parent=self.export_button)
        self.export_menu.setObjectName("earningsExportMenu")
        self.export_menu.setAccessibleName("Export options")
        self.export_menu.setMinimumWidth(248)
        self.export_filtered_action = Action(
            FluentIcon.FILTER, "Export filtered records", self.export_menu
        )
        self.export_all_action = Action(
            FluentIcon.SAVE_AS, "Export all records", self.export_menu
        )
        self.export_filtered_action.setStatusTip("Export only the records shown by the current filters")
        self.export_all_action.setStatusTip("Export the complete earnings history")
        self.export_filtered_action.triggered.connect(lambda: self._export(True))
        self.export_all_action.triggered.connect(lambda: self._export(False))
        self.export_menu.addAction(self.export_filtered_action)
        self.export_menu.addAction(self.export_all_action)
        self.export_button.setMenu(self.export_menu)
        self.header_action_buttons = (self.brands_button, self.settings_button, self.export_button)
        for button in self.header_action_buttons:
            button.setMinimumHeight(36)
            button.setMaximumHeight(36)
        self._arrange_header(compact=False)
        layout.addLayout(self.header_layout)

        self.section_tabs = QTabBar(self)
        self.section_tabs.setObjectName("earningsTabs")
        self.section_tabs.setDrawBase(False)
        self.section_tabs.setExpanding(False)
        self.section_tabs.setMovable(False)
        self.section_keys = ("logging", "analytics", "history")
        self.section_tabs.addTab("Logging")
        self.section_tabs.addTab("Analytics")
        self.section_tabs.addTab("History")
        layout.addWidget(self.section_tabs)

        self.logging_page = QWidget()
        self.logging_page.setObjectName("earningsLoggingPage")
        logging_layout = QVBoxLayout(self.logging_page)
        logging_layout.setContentsMargins(0, 0, 0, 0)
        logging_layout.setSpacing(10)

        self.metrics_layout = QGridLayout()
        self.metrics_layout.setSpacing(8)
        self.metric_today = MetricCard("Today")
        self.metric_week = MetricCard("This week")
        self.metric_all = MetricCard("All time")
        self.metric_rate = MetricCard("Effective hourly rate")
        self.metric_cards = (self.metric_today, self.metric_week, self.metric_all, self.metric_rate)
        self._arrange_metrics(compact=False)
        logging_layout.addLayout(self.metrics_layout)

        self.goal_quest_panel = self._goal_quest_card()
        logging_layout.addWidget(self.goal_quest_panel)

        self.live_tools_layout = QGridLayout()
        self.live_tools_layout.setSpacing(10)
        self.entry_panel = self._entry_card()
        self.timer_panel = self._timer_card()
        self._arrange_live_tools(compact=False)
        logging_layout.addLayout(self.live_tools_layout)
        logging_layout.addStretch()
        layout.addWidget(self.logging_page)

        self.analytics_page = QWidget()
        self.analytics_page.setObjectName("earningsAnalyticsPage")
        analytics_layout = QVBoxLayout(self.analytics_page)
        analytics_layout.setContentsMargins(0, 0, 0, 0)
        analytics_layout.setSpacing(10)

        self.analytics_metrics_layout = QGridLayout()
        self.analytics_metrics_layout.setSpacing(8)
        self.analytics_total = MetricCard("Total earnings")
        self.analytics_products = MetricCard("Products")
        self.analytics_hours = MetricCard("Tracked hours")
        self.analytics_rate = MetricCard("Effective hourly rate")
        self.analytics_metric_cards = (
            self.analytics_total,
            self.analytics_products,
            self.analytics_hours,
            self.analytics_rate,
        )
        self._arrange_analytics_metrics(compact=False)
        analytics_layout.addLayout(self.analytics_metrics_layout)
        self.projection_panel = self._projection_card()
        analytics_layout.addWidget(self.projection_panel)
        analytics_layout.addWidget(self._analytics_card())

        self.analytics_config_layout = QGridLayout()
        self.analytics_config_layout.setSpacing(10)
        self.goal_panel = self._goal_card()
        self.performance_panel = self._performance_card()
        self._arrange_analytics_config(compact=False)
        analytics_layout.addLayout(self.analytics_config_layout)
        analytics_layout.addStretch()
        self.analytics_page.hide()
        layout.addWidget(self.analytics_page)

        self.history_page = QWidget()
        self.history_page.setObjectName("earningsHistoryPage")
        history_layout = QVBoxLayout(self.history_page)
        history_layout.setContentsMargins(0, 0, 0, 0)
        history_layout.setSpacing(8)
        self.history_panel = self._history_card()
        history_layout.addWidget(self.history_panel, 1)
        self.history_page.hide()
        layout.addWidget(self.history_page)

        self.section_tabs.currentChanged.connect(
            lambda index: self._switch_section(self.section_keys[index])
        )
        self.section_tabs.setCurrentIndex(0)
        QTimer.singleShot(0, self.sku_input.setFocus)
        self._apply_theme()

    @staticmethod
    def _surface() -> QWidget:
        surface = QWidget()
        surface.setObjectName("earningsSurface")
        surface.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        return surface

    def _switch_section(self, route_key: str) -> None:
        if route_key not in self.section_keys:
            route_key = "logging"
        self.logging_page.setVisible(route_key == "logging")
        self.analytics_page.setVisible(route_key == "analytics")
        self.history_page.setVisible(route_key == "history")
        target_index = self.section_keys.index(route_key)
        if self.section_tabs.currentIndex() != target_index:
            self.section_tabs.setCurrentIndex(target_index)
        self.scroll.verticalScrollBar().setValue(0)
        if route_key == "analytics":
            self._refresh_analytics()
        elif route_key == "history":
            self._refresh_entries()
            self._refresh_sessions()
        else:
            QTimer.singleShot(0, self.sku_input.setFocus)

    @staticmethod
    def _clear_grid(layout: QGridLayout) -> None:
        while layout.count():
            layout.takeAt(0)

    def _arrange_metrics(self, *, compact: bool) -> None:
        self._clear_grid(self.metrics_layout)
        columns = 2 if compact else 4
        for index, card in enumerate(self.metric_cards):
            self.metrics_layout.addWidget(card, index // columns, index % columns)
        for column in range(4):
            self.metrics_layout.setColumnStretch(column, 1 if column < columns else 0)

    def _arrange_analytics_metrics(self, *, compact: bool) -> None:
        self._clear_grid(self.analytics_metrics_layout)
        columns = 2 if compact else 4
        for index, card in enumerate(self.analytics_metric_cards):
            self.analytics_metrics_layout.addWidget(card, index // columns, index % columns)
        for column in range(4):
            self.analytics_metrics_layout.setColumnStretch(column, 1 if column < columns else 0)

    def _arrange_projection_metrics(self, *, compact: bool) -> None:
        self._clear_grid(self.projection_metrics_layout)
        columns = 2 if compact else 4
        for index, metric in enumerate(self.projection_metrics):
            self.projection_metrics_layout.addWidget(metric, index // columns, index % columns)
        for column in range(4):
            self.projection_metrics_layout.setColumnStretch(column, 1 if column < columns else 0)

    def _arrange_analytics_config(self, *, compact: bool) -> None:
        self._clear_grid(self.analytics_config_layout)
        if compact:
            self.analytics_config_layout.addWidget(self.goal_panel, 0, 0)
            self.analytics_config_layout.addWidget(self.performance_panel, 1, 0)
            self.analytics_config_layout.setColumnStretch(0, 1)
            self.analytics_config_layout.setColumnStretch(1, 0)
        else:
            self.analytics_config_layout.addWidget(self.goal_panel, 0, 0)
            self.analytics_config_layout.addWidget(self.performance_panel, 0, 1)
            self.analytics_config_layout.setColumnStretch(0, 1)
            self.analytics_config_layout.setColumnStretch(1, 1)

    def _arrange_header(self, *, compact: bool) -> None:
        self._clear_grid(self.header_layout)
        for column in range(4):
            self.header_layout.setColumnStretch(column, 0)
        if compact:
            self.header_layout.addWidget(self.title, 0, 0, 1, 3)
            for column, button in enumerate(self.header_action_buttons):
                self.header_layout.addWidget(button, 1, column)
                self.header_layout.setColumnStretch(column, 1)
        else:
            self.header_layout.addWidget(self.title, 0, 0)
            self.header_layout.setColumnStretch(0, 1)
            for offset, button in enumerate(self.header_action_buttons, start=1):
                self.header_layout.addWidget(button, 0, offset)

    def _arrange_live_tools(self, *, compact: bool) -> None:
        self._clear_grid(self.live_tools_layout)
        if compact:
            self.live_tools_layout.addWidget(self.entry_panel, 0, 0)
            self.live_tools_layout.addWidget(self.timer_panel, 1, 0)
            self.live_tools_layout.setColumnStretch(0, 1)
            self.live_tools_layout.setColumnStretch(1, 0)
        else:
            self.live_tools_layout.addWidget(self.entry_panel, 0, 0)
            self.live_tools_layout.addWidget(self.timer_panel, 0, 1)
            self.live_tools_layout.setColumnStretch(0, 3)
            self.live_tools_layout.setColumnStretch(1, 2)

    def _arrange_goal_quest_header(self, *, compact: bool) -> None:
        self._clear_grid(self.goal_quest_header_layout)
        for column in range(6):
            self.goal_quest_header_layout.setColumnStretch(column, 0)
        if compact:
            self.goal_quest_header_layout.addWidget(self.goal_quest_title, 0, 0, 1, 2)
            self.goal_quest_header_layout.addWidget(self.goal_quest_level, 0, 2)
            self.goal_quest_header_layout.addWidget(self.goal_quest_create, 1, 0)
            self.goal_quest_header_layout.addWidget(self.goal_quest_adjust, 1, 1)
            self.goal_quest_header_layout.addWidget(self.goal_quest_view, 1, 2)
            for column in range(3):
                self.goal_quest_header_layout.setColumnStretch(column, 1)
        else:
            self.goal_quest_header_layout.addWidget(self.goal_quest_title, 0, 0)
            self.goal_quest_header_layout.addWidget(self.goal_quest_level, 0, 1)
            self.goal_quest_header_layout.setColumnStretch(2, 1)
            self.goal_quest_header_layout.addWidget(self.goal_quest_create, 0, 3)
            self.goal_quest_header_layout.addWidget(self.goal_quest_adjust, 0, 4)
            self.goal_quest_header_layout.addWidget(self.goal_quest_view, 0, 5)

    def _on_breakpoint_changed(self, breakpoint: str):
        compact = breakpoint in ("xs", "sm")
        self._update_header_labels(compact=compact)
        self._arrange_header(compact=compact)
        self._arrange_metrics(compact=compact)
        self._arrange_analytics_metrics(compact=compact)
        self._arrange_projection_metrics(compact=compact)
        self._arrange_analytics_config(compact=compact)
        self._arrange_goal_quest_header(compact=compact)
        self._arrange_live_tools(compact=compact)
        self._arrange_history_toolbar(compact=compact)
        margins = (16, 14, 18, 20) if compact else (24, 20, 28, 28)
        self.content.layout().setContentsMargins(*margins)

    def _update_header_labels(self, *, compact: bool | None = None) -> None:
        if compact is None:
            compact = self.get_current_breakpoint() in ("xs", "sm")
        tr = self.main.i18n.tr
        labels = (
            (self.brands_button, tr("earnings.brands")),
            (self.settings_button, tr("earnings.settings")),
            (self.export_button, tr("earnings.export")),
        )
        for button, label in labels:
            button.setAccessibleName(label)
            button.setToolTip("")
            button.setText(label)
            button.setMinimumWidth(0)
            button.setMaximumWidth(16777215)

    def _entry_card(self):
        card = self._surface()
        grid = QGridLayout(card)
        grid.setContentsMargins(14, 10, 14, 12)
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(7)
        self.entry_heading = StrongBodyLabel("Add earning")
        grid.addWidget(self.entry_heading, 0, 0, 1, 2)
        self.sku_input = LineEdit()
        self.sku_input.setPlaceholderText("SKU (required)")
        self.name_input = LineEdit()
        self.name_input.setPlaceholderText("Optional product name")
        self.brand_input = ComboBox()
        self.type_input = ComboBox()
        for key, label in PRODUCT_TYPES:
            self.type_input.addItem(label, userData=key)
        self.date_input = QDateTimeEdit(datetime.now())
        configure_earnings_datetime_edit(self.date_input)
        self.add_button = AnimatedSubmitButton("Add earning")
        self.add_button.setObjectName("earningsPrimaryAction")
        self.add_button.clicked.connect(self._add_entry)
        self.sku_input.returnPressed.connect(self._add_entry)
        self.sku_label = CaptionLabel("SKU *")
        self.name_label = CaptionLabel("Product name")
        self.brand_label = CaptionLabel("Brand")
        self.type_label = CaptionLabel("Product type")
        self.date_label = CaptionLabel("Date and time")

        def field(label: QLabel, control: QWidget) -> QWidget:
            wrapper = QWidget()
            wrapper.setObjectName("earningsField")
            wrapper.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            wrapper.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
            box = QVBoxLayout(wrapper)
            box.setContentsMargins(0, 0, 0, 0)
            box.setSpacing(5)
            box.addWidget(label)
            box.addWidget(control)
            return wrapper

        grid.addWidget(field(self.sku_label, self.sku_input), 1, 0)
        grid.addWidget(field(self.name_label, self.name_input), 1, 1)
        grid.addWidget(field(self.brand_label, self.brand_input), 2, 0)
        grid.addWidget(field(self.type_label, self.type_input), 2, 1)
        grid.addWidget(field(self.date_label, self.date_input), 3, 0)
        button_wrapper = QWidget()
        button_wrapper.setObjectName("earningsField")
        button_wrapper.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        button_layout = QVBoxLayout(button_wrapper)
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(5)
        button_layout.addSpacing(self.date_label.sizeHint().height())
        button_layout.addWidget(self.add_button, 0, Qt.AlignmentFlag.AlignRight)
        self.batch_counter = BatchProgressTicks()
        button_layout.addWidget(self.batch_counter, 0, Qt.AlignmentFlag.AlignRight)
        grid.addWidget(button_wrapper, 3, 1)
        self.earning_burst = EarningsBurstBadge(card)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        return card

    def _timer_card(self):
        card = self._surface()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 10, 14, 12)
        layout.setSpacing(6)
        row = QHBoxLayout()
        self.timer_heading = StrongBodyLabel("Work timer")
        row.addWidget(self.timer_heading)
        row.addStretch()
        self.timer_mode = ComboBox()
        self.timer_mode.addItem("Stopwatch", userData="stopwatch")
        self.timer_mode.addItem("Countdown", userData="countdown")
        self.timer_mode.currentIndexChanged.connect(self._timer_mode_changed)
        row.addWidget(self.timer_mode)
        layout.addLayout(row)
        self.timer_display = QLabel("00:00:00")
        self.timer_display.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.timer_display.setFont(QFont(FONTS["family"], 30, QFont.Weight.Bold))
        layout.addWidget(self.timer_display)
        self.session_earnings = StrongBodyLabel("Session earnings: €0.00")
        self.session_earnings.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.session_earnings.setObjectName("sessionEarnings")
        layout.addWidget(self.session_earnings)
        self.session_products = CaptionLabel("SKUs logged this session: 0")
        self.session_products.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.session_products)
        self.timer_status = CaptionLabel("Ready")
        self.timer_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.timer_status)
        self.quest_progress = QuestProgressWidget(card)
        layout.addWidget(self.quest_progress)
        countdown = QHBoxLayout()
        self.countdown_minutes = QSpinBox()
        self.countdown_minutes.setRange(1, 24 * 60)
        self.countdown_minutes.setValue(25)
        self.countdown_minutes.setSuffix(" min")
        countdown.addWidget(self.countdown_minutes)
        for value in (25, 45, 60):
            button = PushButton(str(value))
            button.setToolTip(f"{value} minutes")
            button.clicked.connect(lambda _=False, minutes=value: self.countdown_minutes.setValue(minutes))
            countdown.addWidget(button)
        self.countdown_row = QWidget()
        self.countdown_row.setLayout(countdown)
        self.countdown_row.setVisible(False)
        layout.addWidget(self.countdown_row)
        actions = QHBoxLayout()
        actions.setSpacing(8)
        actions.addStretch()
        self.timer_start = PrimaryPushButton("Start", self)
        self.timer_pause = PrimaryPushButton("Pause", self)
        self.timer_finish = PushButton("Finish session", self)
        self.timer_reset = PushButton("Discard", self)
        self.timer_start.setObjectName("timerStart")
        self.timer_pause.setObjectName("timerStart")
        self.timer_finish.setObjectName("timerSecondary")
        self.timer_reset.setObjectName("earningsDangerAction")
        for button in (self.timer_start, self.timer_pause, self.timer_finish, self.timer_reset):
            button.setFixedHeight(36)
        self.timer_start.setMinimumWidth(96)
        self.timer_pause.setMinimumWidth(96)
        self.timer_finish.setMinimumWidth(132)
        self.timer_reset.setMinimumWidth(84)
        self.timer_start.clicked.connect(self._start_or_resume)
        self.timer_pause.clicked.connect(self._pause_timer)
        self.timer_finish.clicked.connect(self._finish_timer)
        self.timer_reset.clicked.connect(self._reset_timer)
        actions.addWidget(self.timer_start)
        actions.addWidget(self.timer_pause)
        actions.addWidget(self.timer_finish)
        actions.addWidget(self.timer_reset)
        actions.addStretch()
        layout.addLayout(actions)
        self.quest_celebration = QuestCelebrationOverlay(card)
        self._set_timer_action_labels("ready")
        return card

    def _goal_quest_card(self) -> QWidget:
        card = self._surface()
        card.setObjectName("earningsGoalQuest")
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)

        layout = QVBoxLayout(card)
        layout.setContentsMargins(18, 14, 18, 15)
        layout.setSpacing(SPACING["sm"])
        layout.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)

        self.goal_quest_header_layout = QGridLayout()
        self.goal_quest_header_layout.setHorizontalSpacing(SPACING["sm"])
        self.goal_quest_header_layout.setVerticalSpacing(SPACING["sm"])
        self.goal_quest_title = StrongBodyLabel("Current money goal")
        self.goal_quest_title.setWordWrap(True)
        self.goal_quest_level = CaptionLabel("")
        self.goal_quest_level.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )

        self.goal_quest_create = PrimaryPushButton("Create goal")
        self.goal_quest_adjust = PushButton("Adjust progress")
        self.goal_quest_view = PushButton("View forecast")
        self.goal_quest_create.setObjectName("earningsPrimaryAction")
        self.goal_quest_adjust.setObjectName("earningsSecondaryAction")
        self.goal_quest_view.setObjectName("earningsSecondaryAction")
        for button in (
            self.goal_quest_create,
            self.goal_quest_adjust,
            self.goal_quest_view,
        ):
            button.setFixedHeight(34)
            button.setMinimumWidth(104)
        self.goal_quest_create.clicked.connect(self._create_goal)
        self.goal_quest_adjust.clicked.connect(self._add_goal_adjustment)
        self.goal_quest_view.clicked.connect(lambda: self._switch_section("analytics"))
        self._arrange_goal_quest_header(compact=False)
        layout.addLayout(self.goal_quest_header_layout)

        self.goal_quest_empty = BodyLabel("")
        self.goal_quest_empty.setWordWrap(True)
        layout.addWidget(self.goal_quest_empty)

        self.goal_quest_details = QWidget()
        self.goal_quest_details.setObjectName("earningsGoalQuestDetails")
        details = QVBoxLayout(self.goal_quest_details)
        details.setContentsMargins(0, 0, 0, 0)
        details.setSpacing(SPACING["sm"])
        details.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)

        value_row = QHBoxLayout()
        value_row.setSpacing(SPACING["sm"])
        self.goal_quest_value = StrongBodyLabel("")
        self.goal_quest_value.setWordWrap(True)
        self.goal_quest_percentage = StrongBodyLabel("")
        self.goal_quest_percentage.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        self.goal_quest_percentage.setMinimumWidth(64)
        value_row.addWidget(self.goal_quest_value, 1)
        value_row.addWidget(self.goal_quest_percentage)
        details.addLayout(value_row)

        self.goal_quest_progress = GoalMilestoneBar()
        self.goal_quest_progress.setFixedHeight(14)
        details.addWidget(self.goal_quest_progress)

        footer = QHBoxLayout()
        footer.setSpacing(SPACING["sm"])
        self.goal_quest_next = CaptionLabel("")
        self.goal_quest_next.setWordWrap(True)
        self.goal_quest_remaining = CaptionLabel("")
        self.goal_quest_remaining.setWordWrap(True)
        self.goal_quest_remaining.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        footer.addWidget(self.goal_quest_next, 1)
        footer.addWidget(self.goal_quest_remaining, 1)
        details.addLayout(footer)
        layout.addWidget(self.goal_quest_details)
        return card

    def _projection_card(self):
        card = self._surface()
        card.setToolTip(
            "Estimates use your measured effective hourly rate and configured work schedule. "
            "They do not account for holidays, breaks, or changes in pace."
        )
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(4)
        self.projection_heading = StrongBodyLabel("Income at your current pace")
        self.projection_basis = CaptionLabel("")
        self.projection_basis.setWordWrap(True)
        layout.addWidget(self.projection_heading)
        layout.addWidget(self.projection_basis)
        self.projection_metrics_layout = QGridLayout()
        self.projection_metrics_layout.setHorizontalSpacing(12)
        self.projection_metrics_layout.setVerticalSpacing(6)
        self.projection_day = ProjectionMetric("Work day")
        self.projection_week = ProjectionMetric("Work week")
        self.projection_month = ProjectionMetric("Average month")
        self.projection_year = ProjectionMetric("Work year")
        self.projection_metrics = (
            self.projection_day,
            self.projection_week,
            self.projection_month,
            self.projection_year,
        )
        self._arrange_projection_metrics(compact=False)
        layout.addLayout(self.projection_metrics_layout)
        return card

    def _goal_card(self):
        card = self._surface()
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(0)
        layout.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)
        header = QHBoxLayout()
        self.goal_heading = StrongBodyLabel("Money goal and forecast")
        header.addWidget(self.goal_heading)
        header.addStretch()
        self.goal_create = PrimaryPushButton("Create goal")
        self.goal_create.setObjectName("earningsPrimaryAction")
        self.goal_create.setFixedSize(128, 34)
        self.goal_edit = PushButton("Edit")
        self.goal_adjust = PushButton("Add progress")
        self.goal_archive = PushButton("Archive")
        self.goal_adjust.setObjectName("earningsSecondaryAction")
        self.goal_edit.setObjectName("earningsSecondaryAction")
        self.goal_archive.setObjectName("earningsSecondaryAction")
        self.goal_edit.setFixedHeight(34)
        self.goal_edit.setMinimumWidth(70)
        self.goal_adjust.setFixedHeight(34)
        self.goal_adjust.setMinimumWidth(108)
        self.goal_archive.setFixedHeight(34)
        self.goal_archive.setMinimumWidth(82)
        self.goal_create.clicked.connect(self._create_goal)
        self.goal_adjust.clicked.connect(self._add_goal_adjustment)
        self.goal_edit.clicked.connect(self._edit_goal)
        self.goal_archive.clicked.connect(self._archive_goal)
        header.addWidget(self.goal_create)
        header.addWidget(self.goal_adjust)
        header.addWidget(self.goal_edit)
        header.addWidget(self.goal_archive)
        layout.addLayout(header)

        self.goal_empty = BodyLabel("Create a money goal to see your forecast.")
        self.goal_empty.setWordWrap(True)
        self.goal_empty.setContentsMargins(0, SPACING["sm"], 0, 0)
        layout.addWidget(self.goal_empty)

        self.goal_details = QWidget()
        self.goal_details.setObjectName("earningsGoalDetails")
        self.goal_details.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum
        )
        details = QVBoxLayout(self.goal_details)
        details.setContentsMargins(0, SPACING["md"], 0, 0)
        details.setSpacing(0)
        details.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)

        summary = QHBoxLayout()
        summary.setSpacing(SPACING["sm"])
        self.goal_title = BodyLabel("")
        self.goal_title.setWordWrap(True)
        self.goal_percentage = BodyLabel("")
        self.goal_percentage.setAlignment(
            Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
        )
        self.goal_percentage.setMinimumWidth(64)
        summary.addWidget(self.goal_title, 1)
        summary.addWidget(self.goal_percentage)
        details.addLayout(summary)

        details.addSpacing(SPACING["xs"])
        self.goal_remaining = CaptionLabel("")
        self.goal_remaining.setWordWrap(True)
        details.addWidget(self.goal_remaining)

        details.addSpacing(SPACING["sm"])
        self.goal_progress = QProgressBar()
        self.goal_progress.setRange(0, 1000)
        self.goal_progress.setTextVisible(False)
        self.goal_progress.setFixedHeight(12)
        details.addWidget(self.goal_progress)

        self.goal_adjustment_block = QWidget()
        self.goal_adjustment_block.setObjectName("earningsGoalAdjustment")
        self.goal_adjustment_block.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum
        )
        adjustment_layout = QVBoxLayout(self.goal_adjustment_block)
        adjustment_layout.setContentsMargins(0, SPACING["sm"], 0, 0)
        adjustment_layout.setSpacing(SPACING["xs"])
        adjustment_layout.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)
        self.goal_adjustment_summary = CaptionLabel("")
        self.goal_adjustment_summary.setWordWrap(True)
        self.goal_adjustment_note = CaptionLabel("")
        self.goal_adjustment_note.setWordWrap(True)
        adjustment_layout.addWidget(self.goal_adjustment_summary)
        adjustment_layout.addWidget(self.goal_adjustment_note)
        details.addWidget(self.goal_adjustment_block)

        self.goal_forecast_block = QWidget()
        self.goal_forecast_block.setObjectName("earningsGoalForecast")
        self.goal_forecast_block.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Minimum
        )
        forecast_layout = QVBoxLayout(self.goal_forecast_block)
        forecast_layout.setContentsMargins(0, SPACING["md"], 0, 0)
        forecast_layout.setSpacing(SPACING["xs"])
        forecast_layout.setSizeConstraint(QLayout.SizeConstraint.SetMinimumSize)
        self.goal_forecast = BodyLabel("")
        self.goal_forecast.setWordWrap(True)
        self.goal_forecast_detail = CaptionLabel("")
        self.goal_forecast_detail.setWordWrap(True)
        forecast_layout.addWidget(self.goal_forecast)
        forecast_layout.addWidget(self.goal_forecast_detail)

        self.goal_deadline = CaptionLabel("")
        self.goal_deadline.setWordWrap(True)
        self.goal_deadline.setContentsMargins(0, SPACING["xs"], 0, 0)
        forecast_layout.addWidget(self.goal_deadline)
        details.addWidget(self.goal_forecast_block)
        layout.addWidget(self.goal_details)
        self._goal_is_complete = False
        return card

    def _analytics_card(self):
        card = self._surface()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 10, 14, 12)
        layout.setSpacing(7)
        header = QHBoxLayout()
        self.analytics_heading = StrongBodyLabel("Earnings trend")
        header.addWidget(self.analytics_heading)
        header.addStretch()
        self.period = ComboBox()
        for label, key in (("Hourly", "hourly"), ("Daily", "daily"), ("Weekly", "weekly"),
                           ("Monthly", "monthly"), ("Quarterly", "quarterly"),
                           ("Yearly", "yearly"), ("All time", "all_time")):
            self.period.addItem(label, userData=key)
        self.period.setCurrentIndex(1)
        self.period.currentIndexChanged.connect(self._refresh_analytics)
        header.addWidget(self.period)
        layout.addLayout(header)
        self.chart = EarningsChart()
        self.analytics_empty = QWidget()
        self.analytics_empty.setObjectName("earningsEmptyState")
        self.analytics_empty.setMinimumHeight(118)
        self.analytics_empty.setMaximumHeight(132)
        empty_layout = QVBoxLayout(self.analytics_empty)
        empty_layout.setContentsMargins(12, 10, 12, 10)
        empty_layout.setSpacing(5)
        empty_layout.addStretch()
        self.analytics_empty_title = StrongBodyLabel("No earnings recorded yet")
        self.analytics_empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.analytics_empty_body = CaptionLabel(
            "Add your first earning to start seeing trends, hourly rate, brands, and product-type breakdowns."
        )
        self.analytics_empty_body.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        self.analytics_empty_body.setWordWrap(True)
        self.analytics_empty_body.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.analytics_empty_add = PrimaryPushButton("Add earning")
        self.analytics_empty_add.setObjectName("earningsPrimaryAction")
        self.analytics_empty_add.setFixedSize(140, 34)
        self.analytics_empty_add.clicked.connect(self._go_to_logging)
        empty_layout.addWidget(self.analytics_empty_title)
        empty_layout.addWidget(self.analytics_empty_body)
        empty_layout.addWidget(self.analytics_empty_add, 0, Qt.AlignmentFlag.AlignCenter)
        empty_layout.addStretch()
        layout.addWidget(self.analytics_empty)
        layout.addWidget(self.chart)
        breakdown = QHBoxLayout()
        self.type_breakdown = CaptionLabel("")
        self.type_breakdown.setWordWrap(True)
        self.brand_breakdown = CaptionLabel("")
        self.brand_breakdown.setWordWrap(True)
        breakdown.addWidget(self.type_breakdown, 1)
        breakdown.addWidget(self.brand_breakdown, 1)
        layout.addLayout(breakdown)
        return card

    def _performance_card(self):
        card = self._surface()
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 16)
        layout.setSpacing(12)
        header = QHBoxLayout()
        self.performance_heading = StrongBodyLabel("Daily and weekly targets")
        header.addWidget(self.performance_heading)
        header.addStretch()
        self.streak_label = CaptionLabel("")
        self.streak_label.setObjectName("earningsStreakBadge")
        self.streak_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.streak_label.setMinimumHeight(26)
        header.addWidget(self.streak_label)
        layout.addLayout(header)
        grid = QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(8)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        self.performance_widgets = {}
        definitions = (
            ("daily_money", "Daily earnings", "earnings.performance.daily_money"),
            ("weekly_money", "Weekly earnings", "earnings.performance.weekly_money"),
            ("daily_time", "Daily work time", "earnings.performance.daily_time"),
            ("weekly_time", "Weekly work time", "earnings.performance.weekly_time"),
        )
        self.performance_label_keys = {}
        for index, (key, label, label_key) in enumerate(definitions):
            target = PerformanceTargetWidget(label)
            grid.addWidget(target, index // 2, index % 2)
            self.performance_widgets[key] = target
            self.performance_label_keys[key] = label_key
        layout.addLayout(grid)
        self.no_performance_targets = CaptionLabel("Set targets in Earnings settings to see progress and streaks.")
        self.no_performance_targets.setWordWrap(True)
        layout.addWidget(self.no_performance_targets)

        self.activity_block = QWidget()
        self.activity_block.setObjectName("earningsActivityBlock")
        self.activity_block.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        activity_layout = QVBoxLayout(self.activity_block)
        activity_layout.setContentsMargins(12, 11, 12, 12)
        activity_layout.setSpacing(7)
        heatmap_header = QHBoxLayout()
        self.activity_heatmap_heading = CaptionLabel("Product activity · last 30 days")
        self.activity_heatmap_heading.setObjectName("activityHeatmapHeading")
        heatmap_header.addWidget(self.activity_heatmap_heading)
        heatmap_header.addStretch()
        self.activity_heatmap_legend = ActivityLegend(self.main.i18n.tr)
        heatmap_header.addWidget(self.activity_heatmap_legend)
        activity_layout.addLayout(heatmap_header)
        self.activity_heatmap_explanation = CaptionLabel(
            "Each square is one day. Darker means more SKUs logged."
        )
        self.activity_heatmap_explanation.setWordWrap(True)
        activity_layout.addWidget(self.activity_heatmap_explanation)
        self.activity_heatmap = ActivityHeatmap(self.main.i18n.tr)
        activity_layout.addWidget(self.activity_heatmap)

        self.activity_summary = QWidget()
        self.activity_summary.setObjectName("earningsActivitySummary")
        summary_layout = QGridLayout(self.activity_summary)
        summary_layout.setContentsMargins(0, 2, 0, 0)
        summary_layout.setHorizontalSpacing(8)
        summary_layout.setVerticalSpacing(2)
        self.activity_summary_widgets = []
        for column in range(3):
            metric = QWidget()
            metric.setObjectName("earningsActivityMetric")
            metric.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
            metric_layout = QVBoxLayout(metric)
            metric_layout.setContentsMargins(10, 8, 10, 8)
            metric_layout.setSpacing(1)
            value = QLabel("0")
            value.setObjectName("earningsActivityValue")
            caption = CaptionLabel("")
            caption.setObjectName("earningsActivityCaption")
            metric_layout.addWidget(value)
            metric_layout.addWidget(caption)
            summary_layout.addWidget(metric, 0, column)
            summary_layout.setColumnStretch(column, 1)
            self.activity_summary_widgets.append((value, caption))
        activity_layout.addWidget(self.activity_summary)
        layout.addWidget(self.activity_block)
        return card

    def _history_card(self):
        card = QWidget()
        card.setObjectName("earningsFlatSection")
        card.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        layout = QVBoxLayout(card)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)
        heading_row = QHBoxLayout()
        self.history_heading = StrongBodyLabel("History")
        self.filtered_total = StrongBodyLabel("")
        heading_row.addWidget(self.history_heading)
        heading_row.addStretch()
        heading_row.addWidget(self.filtered_total)
        self.history_heading_row = heading_row
        layout.addLayout(heading_row)

        self.search = LineEdit()
        self.search.setPlaceholderText("Search SKU, name, or brand")
        self.search.textChanged.connect(self._refresh_entries)
        self.filter_brand = ComboBox()
        self.filter_brand.currentIndexChanged.connect(self._refresh_entries)
        self.filter_type = ComboBox()
        self.filter_type.addItem("All types", userData=None)
        for key, label in PRODUCT_TYPES:
            self.filter_type.addItem(label, userData=key)
        self.filter_type.currentIndexChanged.connect(self._refresh_entries)
        self.filter_source = ComboBox()
        self.filter_source.addItem("All sources", userData=None)
        self.filter_source.addItem("Manual", userData="manual")
        self.filter_source.addItem("Regular upload", userData="regular_upload")
        self.filter_source.addItem("Batch upload", userData="batch_upload")
        self.filter_source.currentIndexChanged.connect(self._refresh_entries)
        self.filter_date = ComboBox()
        self.filter_date.addItem("All dates", userData=None)
        self.filter_date.addItem("Today", userData=1)
        self.filter_date.addItem("Last 7 days", userData=7)
        self.filter_date.addItem("Last 30 days", userData=30)
        self.filter_date.currentIndexChanged.connect(self._refresh_entries)
        self.clear_filters_button = PushButton("Clear filters")
        self.clear_filters_button.setObjectName("earningsSecondaryAction")
        self.clear_filters_button.clicked.connect(self._clear_history_filters)
        self.clear_filters_button.setVisible(False)
        self.edit_entry_button = PushButton("Edit")
        self.edit_entry_button.setObjectName("earningsSecondaryAction")
        self.edit_entry_button.setToolTip("Edit selected earning")
        self.edit_entry_button.clicked.connect(self._edit_selected_entry)
        self.edit_entry_button.setVisible(False)
        self.bulk_edit_button = PushButton("Edit selected")
        self.bulk_edit_button.setObjectName("earningsSecondaryAction")
        self.bulk_edit_button.setToolTip(
            "Apply the same brand, product type, or date to selected earnings"
        )
        self.bulk_edit_button.clicked.connect(self._bulk_edit_selected_entries)
        self.bulk_edit_button.setVisible(False)
        self.history_heading_row.insertWidget(2, self.bulk_edit_button)
        self.delete_entry_button = PushButton("Delete")
        self.delete_entry_button.setObjectName("earningsDangerAction")
        self.delete_entry_button.setToolTip("Delete selected earning")
        self.delete_entry_button.clicked.connect(self._delete_selected_entry)
        self.delete_entry_button.setVisible(False)
        for combo in (self.filter_brand, self.filter_type, self.filter_source, self.filter_date):
            combo.setMinimumWidth(90)
            combo.setMaximumWidth(150)
            combo.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
        self.edit_entry_button.setMinimumWidth(72)
        self.edit_entry_button.setMaximumWidth(110)
        self.edit_entry_button.setFixedHeight(34)
        self.bulk_edit_button.setMinimumWidth(112)
        self.bulk_edit_button.setMaximumWidth(170)
        self.bulk_edit_button.setFixedHeight(34)
        self.delete_entry_button.setMinimumWidth(78)
        self.delete_entry_button.setMaximumWidth(110)
        self.delete_entry_button.setFixedHeight(34)
        self.clear_filters_button.setFixedHeight(34)

        self.history_toolbar = QWidget()
        self.history_toolbar.setObjectName("earningsToolbar")
        self.history_toolbar_layout = QGridLayout(self.history_toolbar)
        self.history_toolbar_layout.setContentsMargins(0, 0, 0, 0)
        self.history_toolbar_layout.setHorizontalSpacing(8)
        self.history_toolbar_layout.setVerticalSpacing(8)
        self.history_toolbar_widgets = (
            self.search,
            self.filter_brand,
            self.filter_type,
            self.filter_source,
            self.filter_date,
            self.clear_filters_button,
            self.edit_entry_button,
            self.delete_entry_button,
        )
        self._arrange_history_toolbar(compact=False)
        layout.addWidget(self.history_toolbar)
        self.selection_hint = CaptionLabel("")
        self.selection_hint.setWordWrap(True)
        layout.addWidget(self.selection_hint)

        self.history_tabs = QTabWidget()
        self.history_tabs.setObjectName("earningsHistoryTabs")
        self.entries_table = QTableWidget(0, 10)
        self.entries_table.setObjectName("earningsEntriesTable")
        self.entries_table.setHorizontalHeaderLabels(
            ["#", "Date", "SKU", "Name", "Brand", "Type", "Source", "Earning", "Session", "ID"]
        )
        self.entries_table.setColumnHidden(9, True)
        self.entries_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.entries_table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.entries_table.setAccessibleName("Earnings history")
        self.entries_table.setAccessibleDescription(
            "Use Control-click or Shift-click to select multiple earnings; Control+A selects all"
        )
        self.entries_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.entries_table.setAlternatingRowColors(True)
        self.entries_table.setShowGrid(False)
        self.entries_table.setWordWrap(False)
        self.entries_table.verticalHeader().setVisible(False)
        self.entries_table.verticalHeader().setDefaultSectionSize(36)
        self.entries_table.horizontalHeader().setMinimumHeight(42)
        self.entries_table.horizontalHeader().setHighlightSections(False)
        self.entries_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.entries_table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.entries_table.setSortingEnabled(True)
        self.entries_table.doubleClicked.connect(self._edit_selected_entry)
        self.entries_table.itemSelectionChanged.connect(self._update_history_actions)
        entries_header = self.entries_table.horizontalHeader()
        entries_header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )
        self.entries_table.horizontalHeaderItem(0).setTextAlignment(
            Qt.AlignmentFlag.AlignCenter
        )
        for column in range(9):
            entries_header.setSectionResizeMode(column, QHeaderView.ResizeMode.Interactive)
        entries_header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        for column, width in {
            0: 54,
            1: 150,
            2: 120,
            4: 120,
            5: 100,
            6: 125,
            7: 100,
            8: 110,
        }.items():
            self.entries_table.setColumnWidth(column, width)

        self.entries_empty_state = QWidget()
        self.entries_empty_state.setObjectName("earningsEmptyState")
        empty_layout = QVBoxLayout(self.entries_empty_state)
        empty_layout.setContentsMargins(16, 18, 16, 18)
        empty_layout.setSpacing(5)
        empty_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.entries_empty_title = StrongBodyLabel("No earnings yet")
        self.entries_empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.entries_empty_body = CaptionLabel("Your logged products will appear here.")
        self.entries_empty_body.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.entries_empty_add = PrimaryPushButton("Add earning")
        self.entries_empty_add.setObjectName("earningsPrimaryAction")
        self.entries_empty_add.setFixedSize(140, 34)
        self.entries_empty_add.clicked.connect(self._go_to_logging)
        empty_layout.addWidget(self.entries_empty_title)
        empty_layout.addWidget(self.entries_empty_body)
        empty_layout.addWidget(self.entries_empty_add, 0, Qt.AlignmentFlag.AlignCenter)

        self.entries_stack = QStackedWidget()
        self.entries_stack.addWidget(self.entries_table)
        self.entries_stack.addWidget(self.entries_empty_state)
        entry_page = QWidget()
        entry_layout = QVBoxLayout(entry_page)
        entry_layout.setContentsMargins(0, 8, 0, 0)
        entry_layout.addWidget(self.entries_stack)
        self.history_tabs.addTab(entry_page, "Earnings")
        self.sessions_table = QTableWidget(0, 8)
        self.sessions_table.setObjectName("earningsSessionsTable")
        self.sessions_table.setHorizontalHeaderLabels(
            ["Started", "Mode", "Status", "Target", "Worked", "Products", "Earned", "€/hour"]
        )
        self.sessions_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.sessions_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.sessions_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.sessions_table.setAlternatingRowColors(True)
        self.sessions_table.setShowGrid(False)
        self.sessions_table.setWordWrap(False)
        self.sessions_table.verticalHeader().setVisible(False)
        self.sessions_table.verticalHeader().setDefaultSectionSize(36)
        self.sessions_table.horizontalHeader().setMinimumHeight(42)
        self.sessions_table.horizontalHeader().setHighlightSections(False)
        self.sessions_table.setHorizontalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        self.sessions_table.setVerticalScrollMode(QAbstractItemView.ScrollMode.ScrollPerPixel)
        sessions_header = self.sessions_table.horizontalHeader()
        sessions_header.setDefaultAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )
        sessions_header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for column in range(1, 8):
            sessions_header.setSectionResizeMode(column, QHeaderView.ResizeMode.ResizeToContents)

        self.sessions_empty_state = QWidget()
        self.sessions_empty_state.setObjectName("earningsEmptyState")
        sessions_empty_layout = QVBoxLayout(self.sessions_empty_state)
        sessions_empty_layout.setContentsMargins(16, 18, 16, 18)
        sessions_empty_layout.setSpacing(5)
        sessions_empty_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sessions_empty_title = StrongBodyLabel("No work sessions yet")
        self.sessions_empty_title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sessions_empty_body = CaptionLabel("Start the timer to track your effective hourly rate.")
        self.sessions_empty_body.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sessions_empty_add = PrimaryPushButton("Open timer")
        self.sessions_empty_add.setObjectName("earningsPrimaryAction")
        self.sessions_empty_add.setFixedSize(140, 34)
        self.sessions_empty_add.clicked.connect(self._go_to_logging)
        sessions_empty_layout.addWidget(self.sessions_empty_title)
        sessions_empty_layout.addWidget(self.sessions_empty_body)
        sessions_empty_layout.addWidget(self.sessions_empty_add, 0, Qt.AlignmentFlag.AlignCenter)

        self.sessions_stack = QStackedWidget()
        self.sessions_stack.addWidget(self.sessions_table)
        self.sessions_stack.addWidget(self.sessions_empty_state)
        sessions_page = QWidget()
        sessions_layout = QVBoxLayout(sessions_page)
        sessions_layout.setContentsMargins(0, 8, 0, 0)
        sessions_layout.addWidget(self.sessions_stack)
        self.history_tabs.addTab(sessions_page, "Work sessions")
        self.history_tabs.setMinimumHeight(420)
        self.history_tabs.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        layout.addWidget(self.history_tabs, 1)
        return card

    def _arrange_history_toolbar(self, *, compact: bool) -> None:
        layout = getattr(self, "history_toolbar_layout", None)
        if layout is None:
            return
        self._clear_grid(layout)
        if compact:
            layout.addWidget(self.search, 0, 0, 1, 3)
            layout.addWidget(self.clear_filters_button, 0, 3)
            layout.addWidget(self.edit_entry_button, 0, 4)
            layout.addWidget(self.delete_entry_button, 0, 5)
            layout.addWidget(self.filter_brand, 1, 0)
            layout.addWidget(self.filter_type, 1, 1)
            layout.addWidget(self.filter_source, 1, 2)
            layout.addWidget(self.filter_date, 1, 3, 1, 3)
            layout.setColumnStretch(0, 1)
            layout.setColumnStretch(1, 1)
            layout.setColumnStretch(2, 1)
            layout.setColumnStretch(3, 1)
            layout.setColumnStretch(4, 1)
            layout.setColumnStretch(5, 1)
        else:
            for column, widget in enumerate(self.history_toolbar_widgets):
                layout.addWidget(widget, 0, column)
            layout.setColumnStretch(0, 2)
            for column in range(1, len(self.history_toolbar_widgets)):
                layout.setColumnStretch(column, 0)

    def _go_to_logging(self) -> None:
        self._switch_section("logging")
        QTimer.singleShot(0, self.sku_input.setFocus)

    def _clear_history_filters(self) -> None:
        controls = (self.search, self.filter_brand, self.filter_type, self.filter_source, self.filter_date)
        for control in controls:
            control.blockSignals(True)
        self.search.clear()
        for combo in (self.filter_brand, self.filter_type, self.filter_source, self.filter_date):
            combo.setCurrentIndex(0)
        for control in controls:
            control.blockSignals(False)
        self._refresh_entries()

    def _history_filters_active(self) -> bool:
        return bool(
            self.search.text().strip()
            or self.filter_brand.currentData() is not None
            or self.filter_type.currentData() is not None
            or self.filter_source.currentData() is not None
            or self.filter_date.currentData() is not None
        )

    def _update_history_actions(self) -> None:
        selected_count = len(self._selected_entries())
        single = selected_count == 1
        multiple = selected_count > 1
        self.edit_entry_button.setVisible(single)
        self.delete_entry_button.setVisible(single)
        self.bulk_edit_button.setVisible(multiple)
        tr = self.main.i18n.tr
        self.selection_hint.setText(
            tr("earnings.history.selection.count", count=selected_count)
            if selected_count
            else tr("earnings.history.selection.hint")
        )

    # ------------------------------------------------------------- refresh/UI
    def refresh_all(self):
        self._reload_brands()
        self._refresh_metrics()
        self._refresh_goal()
        self._refresh_performance()
        self._refresh_analytics()
        self._refresh_entries()
        self._refresh_sessions()
        self._timer_tick()

    def _reload_brands(self):
        current = self.brand_input.currentData() if self.brand_input.count() else None
        self.brand_input.clear()
        self.brand_input.addItem("No brand", userData=None)
        self.filter_brand.blockSignals(True)
        self.filter_brand.clear()
        self.filter_brand.addItem("All brands", userData=None)
        for brand in self.service.list_brands():
            self.brand_input.addItem(brand["name"], userData=brand["id"])
            self.filter_brand.addItem(brand["name"], userData=brand["id"])
        index = self.brand_input.findData(current)
        self.brand_input.setCurrentIndex(max(0, index))
        self.filter_brand.blockSignals(False)

    def _refresh_metrics(self):
        values = self.service.summary()

        def activity_line(count: int, seconds: int, comparison: str = "") -> str:
            if not count and not seconds:
                return "No activity yet"
            parts = []
            if count:
                parts.append(f"{count} product{'s' if count != 1 else ''}")
            if seconds:
                parts.append(duration(seconds))
            if comparison:
                parts.append(comparison)
            return " • ".join(parts)

        today_delta = self._comparison(values["today_cents"], values["yesterday_cents"], "yesterday")
        week_delta = self._comparison(values["week_cents"], values["previous_week_cents"], "last week")
        self.metric_today.set_values(
            money(values["today_cents"]),
            activity_line(values["today_count"], values["today_seconds"], today_delta),
        )
        self.metric_week.set_values(
            money(values["week_cents"]),
            activity_line(values["week_count"], values["week_seconds"], week_delta),
        )
        self.metric_all.set_values(
            money(values["all_cents"]),
            activity_line(values["all_count"], values["all_seconds"]),
        )
        if values["effective_hourly_cents"] is None:
            rate_subtitle = "Start the timer before logging earnings"
        elif values["untimed_count"]:
            rate_subtitle = (
                f"Timed earnings only | {values['untimed_count']:,} untimed excluded"
            )
        else:
            rate_subtitle = "Timed earnings per tracked hour"
        self.metric_rate.set_values(money(values["effective_hourly_cents"]), rate_subtitle)
        self.analytics_total.set_values(money(values["all_cents"]), "Across all recorded earnings")
        self.analytics_products.set_values(f"{values['all_count']:,}", "Paid products recorded")
        self.analytics_hours.set_values(f"{values['all_seconds'] / 3600:.1f} h", "Tracked work time")
        self.analytics_rate.set_values(money(values["effective_hourly_cents"]), rate_subtitle)
        self._refresh_projections(values)

    def _refresh_projections(self, summary: dict[str, Any] | None = None) -> None:
        summary = summary or self.service.summary()
        values = self.service.income_projections(
            effective_hourly_cents=summary["effective_hourly_cents"]
        )
        tr = self.main.i18n.tr

        def hours(value: float) -> str:
            return f"{value:,.0f}" if float(value).is_integer() else f"{value:,.1f}"

        day_hours = hours(values["day_hours"])
        week_hours = hours(values["week_hours"])
        month_hours = hours(values["month_hours"])
        year_hours = hours(values["year_hours"])
        if values["effective_hourly_cents"] is None:
            self.projection_basis.setText(
                tr(
                    "earnings.projection.no_data",
                    day_hours=day_hours,
                    days=values["workdays_per_week"],
                )
            )
        else:
            self.projection_basis.setText(
                tr(
                    "earnings.projection.basis",
                    rate=money(values["effective_hourly_cents"]),
                    tracked_hours=f"{values['tracked_seconds'] / 3600:,.1f}",
                    day_hours=day_hours,
                    days=values["workdays_per_week"],
                )
            )
        self.projection_day.set_values(
            money(values["day_cents"]),
            tr("earnings.projection.day.subtitle", hours=day_hours),
        )
        self.projection_week.set_values(
            money(values["week_cents"]),
            tr(
                "earnings.projection.week.subtitle",
                days=values["workdays_per_week"],
                hours=week_hours,
            ),
        )
        self.projection_month.set_values(
            money(values["month_cents"]),
            tr("earnings.projection.month.subtitle", hours=month_hours),
        )
        self.projection_year.set_values(
            money(values["year_cents"]),
            tr("earnings.projection.year.subtitle", hours=year_hours),
        )

    @staticmethod
    def _comparison(current: int, previous: int, label: str) -> str:
        if previous <= 0:
            return ""
        delta = (current - previous) * 100.0 / previous
        arrow = "↑" if delta > 0 else ("↓" if delta < 0 else "→")
        return f"{arrow} {abs(delta):.0f}% vs {label}"

    def _refresh_goal(self):
        tr = self.main.i18n.tr
        forecast = self.service.goal_forecast()
        editable = forecast is not None
        if forecast is None:
            goals = self.service.list_goals()
            latest = goals[0] if goals else None
            if latest and latest.get("status") == GoalStatus.COMPLETED.value:
                forecast = self.service.goal_forecast(int(latest["id"]))

        has_goal = forecast is not None
        self.goal_create.setVisible(not editable)
        self.goal_adjust.setVisible(editable)
        self.goal_edit.setVisible(editable)
        self.goal_archive.setVisible(editable)
        self.goal_empty.setVisible(not has_goal)
        self.goal_details.setVisible(has_goal)
        if not has_goal:
            self.goal_panel.setMinimumHeight(92)
            self.goal_panel.setMaximumHeight(96)
            self.goal_empty.setText(tr("earnings.goal.empty"))
            self.goal_title.setText("")
            self.goal_percentage.setText("")
            self.goal_remaining.setText("")
            self.goal_adjustment_summary.setText("")
            self.goal_adjustment_note.setText("")
            self.goal_adjustment_block.setVisible(False)
            self.goal_forecast.setText("")
            self.goal_forecast_detail.setText("")
            self.goal_deadline.setText("")
            self._goal_is_complete = False

            self.goal_quest_title.setText(tr("earnings.goal.quest.create.title"))
            self.goal_quest_level.setVisible(False)
            self.goal_quest_create.setVisible(True)
            self.goal_quest_create.setText(tr("earnings.goal.quest.create.action"))
            self.goal_quest_adjust.setVisible(False)
            self.goal_quest_view.setVisible(False)
            self.goal_quest_empty.setText(tr("earnings.goal.quest.create.body"))
            self.goal_quest_empty.setVisible(True)
            self.goal_quest_details.setVisible(False)
            self.goal_quest_progress.set_progress(0, 0)
            self._apply_goal_progress_theme()
            return

        self.goal_panel.setMinimumHeight(0)
        self.goal_panel.setMaximumHeight(16777215)
        goal = forecast["goal"]
        earned = int(forecast.get("earned_cents") or 0)
        target = int(goal.get("target_cents") or 0)
        state = goal_progress_state(earned, target)
        level_state = goal_level_state(earned, target)
        percent = float(state["percent"])
        percent_text = f"{percent:.0f}%" if abs(percent - round(percent)) < 0.05 else f"{percent:.1f}%"

        self.goal_title.setText(
            tr("earnings.goal.progress", current=money(earned), target=money(target))
        )
        self.goal_percentage.setText(percent_text)
        if state["above_cents"]:
            self.goal_remaining.setText(
                tr("earnings.goal.above", amount=money(state["above_cents"]))
            )
        elif state["reached"]:
            self.goal_remaining.setText(tr("earnings.goal.reached"))
        else:
            self.goal_remaining.setText(
                tr("earnings.goal.remaining", amount=money(state["remaining_cents"]))
            )
        self.goal_progress.setValue(int(round(float(state["visual_percent"]) * 10)))

        self.goal_quest_title.setText(tr("earnings.goal.quest.title"))
        self.goal_quest_level.setText(
            tr(
                "earnings.goal.quest.level",
                level=level_state["level"],
                levels=level_state["levels"],
            )
        )
        self.goal_quest_level.setVisible(True)
        self.goal_quest_create.setVisible(not editable)
        self.goal_quest_create.setText(tr("earnings.goal.quest.new.action"))
        self.goal_quest_adjust.setVisible(editable)
        self.goal_quest_view.setVisible(True)
        self.goal_quest_empty.setVisible(False)
        self.goal_quest_details.setVisible(True)
        self.goal_quest_value.setText(
            tr("earnings.goal.progress", current=money(earned), target=money(target))
        )
        self.goal_quest_percentage.setText(percent_text)
        self.goal_quest_progress.set_progress(earned, target)
        if level_state["complete"]:
            self.goal_quest_next.setText(tr("earnings.goal.quest.complete"))
            self.goal_quest_remaining.setText(
                tr("earnings.goal.above", amount=money(state["above_cents"]))
                if state["above_cents"]
                else ""
            )
        else:
            self.goal_quest_next.setText(
                tr(
                    "earnings.goal.quest.next",
                    amount=money(level_state["next_level_cents"]),
                    level=level_state["next_level"],
                )
            )
            self.goal_quest_remaining.setText(
                tr(
                    "earnings.goal.quest.remaining",
                    amount=money(state["remaining_cents"]),
                )
            )

        adjustment_cents = int(forecast.get("adjustment_cents") or 0)
        adjustments = (
            self.service.list_goal_adjustments(int(goal["id"]))
            if adjustment_cents > 0
            else []
        )
        latest_note = next(
            (str(item.get("note") or "").strip() for item in adjustments if item.get("note")),
            "",
        )
        self.goal_adjustment_block.setVisible(adjustment_cents > 0)
        self.goal_adjustment_summary.setText(
            self.main.i18n.tr(
                "earnings.goal.adjust.summary",
                amount=money(adjustment_cents),
            )
            if adjustment_cents > 0
            else ""
        )
        self.goal_adjustment_note.setVisible(bool(latest_note))
        self.goal_adjustment_note.setText(
            self.main.i18n.tr("earnings.goal.adjust.note.latest", note=latest_note)
            if latest_note
            else ""
        )

        self._goal_is_complete = bool(state["reached"])
        self.goal_forecast_block.setVisible(not self._goal_is_complete)
        likely = (
            tr("earnings.goal.forecast.products", count=forecast["likely_products"])
            if forecast["likely_products"] is not None
            else tr("earnings.goal.forecast.products.unknown")
        )
        hours = (
            tr(
                "earnings.goal.forecast.hours",
                hours=f"{forecast['estimated_hours']:.1f}",
            )
            if forecast["estimated_hours"] is not None
            else tr("earnings.goal.forecast.hours.unknown")
        )
        self.goal_forecast.setText(
            tr("earnings.goal.forecast.likely", products=likely, hours=hours)
        )
        basis = tr(
            "earnings.goal.forecast.basis.last_30_days"
            if forecast["product_basis"] == "last_30_days"
            else "earnings.goal.forecast.basis.all"
        )
        payout_basis = tr(
            "earnings.goal.forecast.payout.recent"
            if forecast["product_basis"] == "last_30_days"
            else "earnings.goal.forecast.payout.recorded"
        )
        self.goal_forecast_detail.setText(
            tr(
                "earnings.goal.forecast.range",
                optimistic=forecast["optimistic_products"],
                conservative=forecast["conservative_products"],
                payout_basis=payout_basis,
                sample=forecast["product_sample"],
                basis=basis,
            )
        )
        deadline = forecast.get("deadline")
        if not deadline:
            self.goal_deadline.setText("")
            self.goal_deadline.setVisible(False)
        elif deadline["overdue"]:
            self.goal_deadline.setText(
                tr("earnings.goal.forecast.deadline.overdue", date=deadline["date"])
            )
            self.goal_deadline.setVisible(True)
        else:
            pace_parts = [
                tr(
                    "earnings.goal.forecast.pace.money",
                    daily=money(deadline["cents_per_day"]),
                    weekly=money(deadline["cents_per_week"]),
                )
            ]
            if deadline["products_per_day"] is not None:
                pace_parts.append(
                    tr(
                        "earnings.goal.forecast.pace.products",
                        products=f"{deadline['products_per_day']:.1f}",
                    )
                )
            if deadline["hours_per_week"] is not None:
                pace_parts.append(
                    tr(
                        "earnings.goal.forecast.pace.hours",
                        hours=f"{deadline['hours_per_week']:.1f}",
                    )
                )
            self.goal_deadline.setText(
                tr(
                    "earnings.goal.forecast.deadline",
                    date=deadline["date"],
                    pace=" • ".join(pace_parts),
                )
            )
            self.goal_deadline.setVisible(True)
        self._apply_goal_progress_theme()
        self.goal_adjustment_block.updateGeometry()
        self.goal_forecast_block.updateGeometry()
        self.goal_details.updateGeometry()
        self.goal_panel.updateGeometry()
        self.goal_quest_details.updateGeometry()
        self.goal_quest_panel.updateGeometry()
        self.analytics_config_layout.invalidate()

    def _refresh_analytics(self):
        chart_data = self.service.trend_data(self.period.currentData() or "daily")
        has_data = any(
            int(item.get("cents", 0)) > 0 or int(item.get("count", 0)) > 0
            for item in chart_data
        )
        self.chart.set_data(chart_data if has_data else [])
        self.chart.setVisible(has_data)
        self.analytics_empty.setVisible(not has_data)
        self.type_breakdown.setVisible(has_data)
        self.brand_breakdown.setVisible(has_data)
        type_rows = self.service.type_breakdown()
        type_text = "Product types\n" + ("\n".join(
            f"{row['product_type'].title()}: {money(row['cents'])} ({row['count']})" for row in type_rows
        ) or "No data")
        brand_rows = self.service.brand_breakdown()
        brand_text = "Top brands\n" + ("\n".join(
            f"{row['brand']}: {money(row['cents'])} ({row['count']})" for row in brand_rows
        ) or "No data")
        self.type_breakdown.setText(type_text)
        self.brand_breakdown.setText(brand_text)

    def _refresh_performance(self):
        values = self.service.performance_progress()
        activity = self.service.trend_data("daily")
        self.activity_heatmap.set_data(activity)
        rows = {
            "daily_money": (values["daily_earned_cents"], values["daily_earning_goal_cents"], money),
            "weekly_money": (values["weekly_earned_cents"], values["weekly_earning_goal_cents"], money),
            "daily_time": (values["daily_work_minutes"], values["daily_work_goal_minutes"], lambda value: f"{float(value):.0f} min"),
            "weekly_time": (values["weekly_work_minutes"], values["weekly_work_goal_minutes"], lambda value: f"{float(value):.0f} min"),
        }
        any_enabled = False
        for key, (current, target, formatter) in rows.items():
            widget = self.performance_widgets[key]
            enabled = target > 0
            widget.setVisible(enabled)
            if enabled:
                any_enabled = True
                ratio = float(current) / float(target)
                widget.set_progress(
                    formatter(current),
                    formatter(target),
                    ratio,
                )
        self.no_performance_targets.setVisible(not any_enabled)
        self.performance_panel.setMinimumHeight(210)
        self.performance_panel.setMaximumHeight(16777215)
        streak = int(values["streak"])
        if any_enabled:
            self.streak_label.setText(
                self.main.i18n.tr(
                    "earnings.performance.streak",
                    count=streak,
                )
                if streak > 0
                else self.main.i18n.tr("earnings.performance.streak_start")
            )
        else:
            self.streak_label.setText("")

        counts = [max(0, int(item.get("count", 0))) for item in activity[-30:]]
        active_days = sum(1 for count in counts if count > 0)
        total_skus = sum(counts)
        best_day = max(counts, default=0)
        for (value_label, _caption), value in zip(
            self.activity_summary_widgets,
            (active_days, total_skus, best_day),
        ):
            value_label.setText(str(value))

    def _refresh_entries(self):
        if not hasattr(self, "entries_table"):
            return
        days = self.filter_date.currentData()
        start = datetime.now().astimezone() - timedelta(days=int(days)) if days else None
        if days == 1:
            local_now = datetime.now().astimezone()
            start = local_now.replace(hour=0, minute=0, second=0, microsecond=0)
        self._entries = self.service.list_entries(
            search=self.search.text(), brand_id=self.filter_brand.currentData(),
            product_type=self.filter_type.currentData(), source=self.filter_source.currentData(),
            start=start,
        )
        self.entries_table.setSortingEnabled(False)
        self.entries_table.clearSelection()
        self.entries_table.setRowCount(len(self._entries))
        for row, entry in enumerate(self._entries):
            values = (
                row + 1, local_datetime(entry["earned_at"]), entry["sku"], entry.get("product_name") or "",
                entry.get("brand_name") or "", entry["product_type"].title(),
                entry["source"].replace("_", " ").title(), money(entry["payout_cents"]),
                str(entry.get("session_id") or ""), str(entry["id"]),
            )
            for column, value in enumerate(values):
                item = QTableWidgetItem()
                item.setData(
                    Qt.ItemDataRole.DisplayRole,
                    int(value) if column == 0 else str(value),
                )
                item.setData(Qt.ItemDataRole.UserRole, entry)
                if value:
                    item.setToolTip(str(value))
                if column == 0:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                elif column in (7, 8):
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                    )
                self.entries_table.setItem(row, column, item)
        self.entries_table.setSortingEnabled(True)
        has_rows = bool(self._entries)
        self.entries_stack.setCurrentIndex(0 if has_rows else 1)
        self.entries_stack.setMaximumHeight(16777215 if has_rows else 210)
        filters_active = self._history_filters_active()
        tr = self.main.i18n.tr
        self.clear_filters_button.setVisible(filters_active)
        self.entries_empty_title.setText(
            tr("earnings.history.empty.filtered.title")
            if filters_active
            else tr("earnings.history.empty.title")
        )
        self.entries_empty_body.setText(
            tr("earnings.history.empty.filtered.body")
            if filters_active
            else tr("earnings.history.empty.body")
        )
        self.entries_empty_add.setVisible(not filters_active)
        self._update_history_actions()
        total = sum(int(row["payout_cents"]) for row in self._entries)
        self.filtered_total.setText(f"{len(self._entries)} records • {money(total)}")
        self._update_history_height()

    def _refresh_sessions(self):
        sessions = self.service.list_sessions()
        self._sessions_count = len(sessions)
        self.sessions_table.setRowCount(len(sessions))
        for row, session in enumerate(sessions):
            target = duration(session["target_seconds"]) if session["target_seconds"] else "—"
            values = (
                local_datetime(session["started_at"]), session["mode"].title(), session["status"].title(),
                target, duration(session["elapsed_seconds"]), session["product_count"],
                money(session["earned_cents"]), money(session["hourly_cents"]),
            )
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if value:
                    item.setToolTip(str(value))
                if column >= 3:
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter
                    )
                self.sessions_table.setItem(row, column, item)
        has_sessions = bool(sessions)
        self.sessions_stack.setCurrentIndex(0 if has_sessions else 1)
        self.sessions_stack.setMaximumHeight(16777215 if has_sessions else 210)
        self._update_history_height()

    def _update_history_height(self) -> None:
        if not hasattr(self, "history_tabs"):
            return
        empty = not self._entries and self._sessions_count == 0
        self.history_tabs.setMaximumHeight(280 if empty else 16777215)

    # --------------------------------------------------------------- entries
    def _add_entry(self):
        sku = self.sku_input.text().strip()
        if not sku:
            InfoBar.warning(title="SKU required", content="Enter an SKU before adding the earning.", parent=self, position=InfoBarPosition.TOP)
            return
        if self.service.duplicate_sku_count(sku):
            dialog = MessageBox("Duplicate SKU", f"{sku} already exists. Add another paid record anyway?", self)
            dialog.yesButton.setText("Add anyway")
            dialog.cancelButton.setText("Cancel")
            if not dialog.exec():
                return
        try:
            payout_cents = self.service.get_rate_cents(self.type_input.currentData())
            self.service.create_entry(
                sku, self.type_input.currentData(), product_name=self.name_input.text(),
                brand_id=self.brand_input.currentData(), earned_at=self.date_input.dateTime().toPython(),
            )
            self.sku_input.clear()
            self.name_input.clear()
            self.date_input.setDateTime(datetime.now())
            self.refresh_all()
            self.batch_counter.advance()
            self.add_button.animate_success()
            self.earning_burst.show_amount(payout_cents, self.add_button)
            InfoBar.success(title="Earning added", content=f"{sku} was added to your earnings.", parent=self, position=InfoBarPosition.TOP, duration=2500)
        except Exception as error:
            InfoBar.error(title="Could not add earning", content=str(error), parent=self, position=InfoBarPosition.TOP)

    def _selected_entry(self):
        entries = self._selected_entries()
        return entries[0] if len(entries) == 1 else None

    def _selected_entries(self) -> list[dict[str, Any]]:
        selection = self.entries_table.selectionModel()
        if selection is None:
            return []
        entries: list[dict[str, Any]] = []
        for index in sorted(selection.selectedRows(0), key=lambda value: value.row()):
            item = self.entries_table.item(index.row(), 0)
            entry = item.data(Qt.ItemDataRole.UserRole) if item is not None else None
            if isinstance(entry, dict):
                entries.append(entry)
        return entries

    def _edit_selected_entry(self):
        entry = self._selected_entry()
        if not entry:
            return
        dialog = EarningEntryDialog(self.service, entry, self)
        if dialog.exec():
            try:
                self.service.update_entry(entry["id"], **dialog.values())
                self.refresh_all()
            except Exception as error:
                QMessageBox.warning(self, "Could not update earning", str(error))

    def _bulk_edit_selected_entries(self):
        entries = self._selected_entries()
        if len(entries) < 2:
            return
        dialog = BulkEarningEditDialog(
            self.service,
            len(entries),
            self,
            translate=self.main.i18n.tr,
        )
        if not dialog.exec():
            return
        try:
            updated = self.service.bulk_update_entries(
                (entry["id"] for entry in entries), **dialog.values()
            )
            self.refresh_all()
            InfoBar.success(
                title=self.main.i18n.tr("earnings.bulk.complete.title"),
                content=self.main.i18n.tr(
                    "earnings.bulk.complete", count=updated
                ),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
        except Exception as error:
            InfoBar.error(
                title=self.main.i18n.tr("earnings.bulk.failed"),
                content=str(error),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _delete_selected_entry(self):
        entry = self._selected_entry()
        if not entry:
            return
        dialog = MessageBox("Delete earning?", f"Delete {entry['sku']} and {money(entry['payout_cents'])} from your history?", self)
        dialog.yesButton.setText("Delete")
        if dialog.exec():
            self.service.delete_entry(entry["id"])
            self.refresh_all()

    # ---------------------------------------------------------------- timer
    def _timer_mode_changed(self):
        self.countdown_row.setVisible(self.timer_mode.currentData() == "countdown")

    def _set_timer_action_labels(self, state: str) -> None:
        tr = self.main.i18n.tr
        start_label = {
            "ready": tr("earnings.timer.start"),
            "paused": tr("earnings.timer.resume"),
            "running": tr("earnings.timer.running"),
            "overtime": tr("earnings.timer.overtime"),
        }.get(state, tr("earnings.timer.start"))
        controls = (
            (self.timer_start, start_label),
            (self.timer_pause, tr("earnings.timer.pause")),
            (self.timer_finish, tr("earnings.timer.finish")),
            (self.timer_reset, tr("earnings.timer.discard")),
        )
        for button, label in controls:
            button.setText(label)
            button.setToolTip(label)
            button.setAccessibleName(label)

    def _quest_value_text(self, kind: str, value: int) -> str:
        tr = self.main.i18n.tr
        if kind == QuestKind.SKU.value:
            return tr("earnings.quest.value.sku", value=int(value))
        if kind == QuestKind.EARNINGS.value:
            return money(int(value))
        minutes = int(value) / 60.0
        rendered = f"{minutes:.0f}" if minutes.is_integer() else f"{minutes:.1f}"
        return tr("earnings.quest.value.minutes", value=rendered)

    def _render_quest_progress(self, progress: dict[str, Any] | None) -> None:
        if progress is None:
            self.quest_progress.clear()
            return
        tr = self.main.i18n.tr
        kind = progress["kind"]
        current = self._quest_value_text(kind, progress["current_value"])
        target = self._quest_value_text(kind, progress["target_value"])
        if progress["complete"]:
            badge = tr("earnings.quest.complete.badge")
            if progress.get("bonus_complete"):
                detail = tr("earnings.quest.bonus.complete")
            else:
                detail = tr(
                    "earnings.quest.bonus.next",
                    current=current,
                    value=self._quest_value_text(kind, progress["bonus_target_value"]),
                )
        else:
            badge = tr(
                "earnings.quest.percent", percent=int(round(progress["percent"]))
            )
            next_value = self._quest_value_text(
                kind, progress["next_checkpoint_value"]
            )
            detail = tr(
                "earnings.quest.progress.detail",
                current=current,
                target=target,
                next=next_value,
            )
        title = tr(f"earnings.quest.{kind}.title")
        animations = self._engagement_settings["animations_enabled"]
        self.quest_progress.set_state(
            title=title,
            badge=badge,
            detail=detail,
            percent=progress["percent"],
            animated=animations,
            complete=progress["complete"],
        )

    def _observe_quest_progress(
        self, session_id: int, progress: dict[str, Any] | None
    ) -> None:
        if progress is None:
            self._quest_observed_session_id = session_id
            self._quest_observed_checkpoints = 0
            self._quest_observed_complete = False
            self._quest_observed_bonus_complete = False
            return
        reached = int(progress["reached_checkpoints"])
        complete = bool(progress["complete"])
        bonus_complete = bool(progress.get("bonus_complete"))
        if self._quest_observed_session_id != session_id:
            # A refresh or app restart silently adopts persisted state so a
            # celebration can never replay for work already observed.
            self._quest_observed_session_id = session_id
            self._quest_observed_checkpoints = reached
            self._quest_observed_complete = complete
            self._quest_observed_bonus_complete = bonus_complete
            return

        settings = self._engagement_settings
        completed_now = complete and not self._quest_observed_complete
        checkpoint_now = reached > self._quest_observed_checkpoints
        bonus_now = bonus_complete and not self._quest_observed_bonus_complete
        if completed_now:
            if settings["animations_enabled"]:
                self.quest_celebration.celebrate(
                    completion=True,
                    badge_text=self.main.i18n.tr("earnings.quest.complete.badge"),
                )
            if settings["sound_enabled"]:
                self._success_sound.play()
        elif (checkpoint_now or bonus_now) and settings["animations_enabled"]:
            self.quest_celebration.celebrate(completion=False)
        self._quest_observed_checkpoints = reached
        self._quest_observed_complete = complete
        self._quest_observed_bonus_complete = bonus_complete

    def _timer_tick(self):
        snapshot = self.service.timer_snapshot()
        active = snapshot is not None
        self.timer_mode.setEnabled(not active)
        self.countdown_minutes.setEnabled(not active)
        if not snapshot:
            self.timer_display.setText("00:00:00")
            self.session_earnings.setText(
                self.main.i18n.tr("earnings.timer.session_earnings", amount=money(0))
            )
            self.session_products.setText(
                self.main.i18n.tr("earnings.timer.session_products", count=0)
            )
            self.timer_status.setText("Ready")
            self.quest_progress.clear()
            self._quest_observed_session_id = None
            self._quest_observed_checkpoints = 0
            self._quest_observed_complete = False
            self._quest_observed_bonus_complete = False
            self._set_timer_action_labels("ready")
            self.timer_start.setVisible(True)
            self.timer_pause.setVisible(False)
            self.timer_finish.setVisible(False)
            self.timer_reset.setVisible(False)
            self.timer_start.setEnabled(True)
            return
        session = self.service.get_session(snapshot.id)
        self._render_quest_progress(session["quest_progress"])
        self._observe_quest_progress(snapshot.id, session["quest_progress"])
        self.session_earnings.setText(
            self.main.i18n.tr(
                "earnings.timer.session_earnings", amount=money(session["earned_cents"])
            )
        )
        self.session_products.setText(
            self.main.i18n.tr(
                "earnings.timer.session_products", count=session["product_count"]
            )
        )
        shown = snapshot.remaining_seconds if snapshot.mode == "countdown" and not snapshot.allow_overtime else snapshot.elapsed_seconds
        self.timer_display.setText(duration(shown))
        self.timer_status.setText(
            "Overtime" if snapshot.allow_overtime and snapshot.status == "running" else snapshot.status.title()
        )
        finished_countdown = snapshot.mode == "countdown" and snapshot.remaining_seconds == 0 and not snapshot.allow_overtime
        action_state = "overtime" if finished_countdown else ("paused" if snapshot.status == "paused" else "running")
        self._set_timer_action_labels(action_state)
        self.timer_start.setVisible(snapshot.status == "paused")
        self.timer_pause.setVisible(snapshot.status == "running")
        self.timer_finish.setVisible(True)
        self.timer_reset.setVisible(True)
        if snapshot.expired and self._last_expired_session != snapshot.id:
            self._last_expired_session = snapshot.id
            InfoBar.info(title="Countdown complete", content="The timer paused at zero. Finish the session or start overtime.", parent=self, position=InfoBarPosition.TOP, duration=7000)
            self._refresh_sessions()

    def _start_or_resume(self):
        snapshot = self.service.timer_snapshot()
        try:
            if snapshot is None:
                dialog = QuestPickerDialog(
                    self.service,
                    self,
                    translate=self.main.i18n.tr,
                )
                if not dialog.exec():
                    return
                quest_kind, quest_target = dialog.values()
                target = self.countdown_minutes.value() * 60 if self.timer_mode.currentData() == "countdown" else None
                session_id = self.service.start_session(
                    self.timer_mode.currentData(),
                    target,
                    quest_kind=quest_kind,
                    quest_target_value=quest_target,
                )
                if quest_kind == QuestKind.FOCUS.value:
                    focus_index = self.timer_mode.findData("countdown")
                    if focus_index >= 0:
                        self.timer_mode.setCurrentIndex(focus_index)
                # New sessions begin at zero and may celebrate future progress.
                self._quest_observed_session_id = session_id
                self._quest_observed_checkpoints = 0
                self._quest_observed_complete = False
                self._quest_observed_bonus_complete = False
            else:
                overtime = snapshot.mode == "countdown" and snapshot.remaining_seconds == 0
                self.service.resume_session(overtime=overtime)
            self._timer_tick()
        except (ActiveSessionError, ValueError) as error:
            QMessageBox.warning(self, "Timer", str(error))

    def _pause_timer(self):
        self.service.pause_session()
        self._timer_tick()
        self._refresh_sessions()

    def _finish_timer(self):
        finished = self.service.finish_session()
        recap = self.service.session_recap(finished["id"])
        self.refresh_all()
        dialog = SessionRecapDialog(
            recap,
            self,
            translate=self.main.i18n.tr,
        )
        dialog.exec()
        if dialog.start_another:
            self._start_or_resume()

    def _reset_timer(self):
        dialog = MessageBox("Discard session?", "The work session will be removed. Linked earnings will remain and be detached.", self)
        dialog.yesButton.setText("Discard")
        if dialog.exec():
            self.service.reset_session()
            self.refresh_all()

    # ---------------------------------------------------------------- goals
    def _create_goal(self):
        dialog = GoalDialog(parent=self)
        if not dialog.exec():
            return
        target, deadline = dialog.values()
        try:
            self.service.create_goal(target, deadline)
            self.refresh_all()
        except ActiveGoalError:
            replace = MessageBox("Replace active goal?", "Archive the current goal and start this new one?", self)
            replace.yesButton.setText("Archive and replace")
            if replace.exec():
                self.service.create_goal(target, deadline, replace_status=GoalStatus.ARCHIVED)
                self.refresh_all()
        except Exception as error:
            QMessageBox.warning(self, "Could not create goal", str(error))

    def _edit_goal(self):
        goal = self.service.active_goal()
        if not goal:
            return
        dialog = GoalDialog(goal, self)
        if dialog.exec():
            target, deadline = dialog.values()
            self.service.update_goal(goal["id"], target, deadline)
            self.refresh_all()

    def _add_goal_adjustment(self):
        goal = self.service.active_goal()
        if not goal:
            return
        current_progress_cents = int(
            self.service.goal_progress(int(goal["id"]))["earned_cents"]
        )
        dialog = GoalAdjustmentDialog(
            self,
            translate=self.main.i18n.tr,
            current_progress_cents=current_progress_cents,
        )
        if not dialog.exec():
            return
        mode, entered_cents, note = dialog.values()
        try:
            if mode == GoalAdjustmentDialog.SET_TOTAL:
                _adjustment_id, amount_cents = self.service.add_goal_adjustment_to_total(
                    int(goal["id"]), entered_cents, note
                )
            else:
                amount_cents = entered_cents
                self.service.add_goal_adjustment(int(goal["id"]), amount_cents, note)
            self.refresh_all()
            InfoBar.success(
                title=self.main.i18n.tr("earnings.goal.adjust.done.title"),
                content=self.main.i18n.tr(
                    "earnings.goal.adjust.done.content",
                    amount=money(amount_cents),
                ),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=2500,
            )
        except Exception as error:
            QMessageBox.warning(
                self,
                self.main.i18n.tr("earnings.goal.adjust.failed"),
                str(error),
            )

    def _archive_goal(self):
        goal = self.service.active_goal()
        if not goal:
            return
        dialog = MessageBox("Archive goal?", "Its current result will be preserved in goal history.", self)
        dialog.yesButton.setText("Archive")
        if dialog.exec():
            self.service.close_goal(goal["id"], status=GoalStatus.ARCHIVED)
            self.refresh_all()

    # --------------------------------------------------------------- dialogs
    def _manage_brands(self):
        BrandManagerDialog(self.service, self).exec()
        self._reload_brands()

    def _settings(self):
        dialog = EarningsSettingsDialog(
            self.service,
            self,
            translate=self.main.i18n.tr,
        )
        if dialog.exec():
            dialog.save()
            self._engagement_settings = self.service.engagement_settings()
            self._success_sound.setVolume(
                self.service.celebration_sound_volume() / 100.0
            )
            self.refresh_all()

    def prompt_upload_entries(self, items: list[dict[str, Any]]) -> int:
        dialog = UploadEarningsDialog(self.service, items, self)
        if not dialog.items or not dialog.exec():
            return 0
        try:
            count = dialog.save_entries()
            self.refresh_all()
            InfoBar.success(title="Earnings added", content=f"Added {count} saved product(s).", parent=self, position=InfoBarPosition.TOP, duration=3500)
            return count
        except Exception as error:
            QMessageBox.warning(self, "Could not add saved products", str(error))
            return 0

    # ---------------------------------------------------------------- export
    def _export(self, filtered: bool):
        default = str(Path.home() / "Desktop" / f"earnings_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "Export earnings", default, "Excel workbook (*.xlsx)")
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        entries = self._entries if filtered else self.service.list_entries()
        try:
            self._build_workbook(path, entries)
            InfoBar.success(title="Export complete", content=path, parent=self, position=InfoBarPosition.TOP, duration=5000)
        except Exception as error:
            InfoBar.error(title="Export failed", content=str(error), parent=self, position=InfoBarPosition.TOP)

    def _build_workbook(self, path: str, entries: list[dict[str, Any]]):
        workbook = openpyxl.Workbook()
        header_fill = PatternFill("solid", fgColor="2B2D42")
        header_font = Font(color="FFFFFF", bold=True)

        def sheet(name, headers):
            reuse_blank = (
                len(workbook.sheetnames) == 1
                and workbook.active.max_row == 1
                and workbook.active["A1"].value is None
            )
            ws = workbook.active if reuse_blank else workbook.create_sheet()
            ws.title = name
            if reuse_blank:
                for column, header in enumerate(headers, start=1):
                    ws.cell(1, column, header)
            else:
                ws.append(headers)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
            return ws

        ws = sheet("Earnings", ["Date", "SKU", "Name", "Brand", "Product Type", "Source", "Earning", "Session ID"])
        for row in entries:
            ws.append([local_datetime(row["earned_at"]), row["sku"], row.get("product_name") or "", row.get("brand_name") or "",
                       row["product_type"].title(), row["source"], row["payout_cents"] / 100, row.get("session_id")])
            ws.cell(ws.max_row, 7).number_format = '€0.00'

        ws_sessions = sheet("Work Sessions", ["Started", "Completed", "Mode", "Status", "Target Hours", "Worked Hours", "Products", "Earned", "Effective €/Hour"])
        for row in self.service.list_sessions():
            ws_sessions.append([local_datetime(row["started_at"]), local_datetime(row.get("completed_at")), row["mode"], row["status"],
                                (row["target_seconds"] or 0) / 3600, row["elapsed_seconds"] / 3600,
                                row["product_count"], row["earned_cents"] / 100,
                                row["hourly_cents"] / 100 if row["hourly_cents"] is not None else None])
            ws_sessions.cell(ws_sessions.max_row, 8).number_format = '€0.00'
            ws_sessions.cell(ws_sessions.max_row, 9).number_format = '€0.00'

        ws_goals = sheet("Goal History", ["Started", "Deadline", "Status", "Target", "Completed", "Final Progress", "Goal-only Adjustments", "Products", "Tracked Hours"])
        for row in self.service.list_goals():
            ws_goals.append([local_datetime(row["started_at"]), row.get("deadline_date"), row["status"], row["target_cents"] / 100,
                             local_datetime(row.get("completed_at")), (row.get("final_earned_cents") or 0) / 100,
                             (row.get("adjustment_cents") or 0) / 100,
                             row.get("final_product_count") or 0, (row.get("final_tracked_seconds") or 0) / 3600])
            ws_goals.cell(ws_goals.max_row, 4).number_format = '€0.00'
            ws_goals.cell(ws_goals.max_row, 6).number_format = '€0.00'
            ws_goals.cell(ws_goals.max_row, 7).number_format = '€0.00'

        values = self.service.summary()
        ws_summary = sheet("Summary", ["Metric", "Value"])
        ws_summary.append(["Today earnings", values["today_cents"] / 100])
        ws_summary.append(["This week earnings", values["week_cents"] / 100])
        ws_summary.append(["All-time earnings", values["all_cents"] / 100])
        ws_summary.append(["All-time products", values["all_count"]])
        ws_summary.append(["Tracked hours", values["all_seconds"] / 3600])
        ws_summary.append(["Timed earnings used for rate", values["timed_cents"] / 100])
        ws_summary.append(["Untimed earnings excluded", values["untimed_cents"] / 100])
        ws_summary.append(["Effective €/hour", values["effective_hourly_cents"] / 100 if values["effective_hourly_cents"] is not None else None])
        projections = self.service.income_projections(
            effective_hourly_cents=values["effective_hourly_cents"]
        )
        ws_summary.append(["Normal workday projection", projections["day_cents"] / 100 if projections["day_cents"] is not None else None])
        ws_summary.append(["Work-week projection", projections["week_cents"] / 100 if projections["week_cents"] is not None else None])
        ws_summary.append(["Average-month projection", projections["month_cents"] / 100 if projections["month_cents"] is not None else None])
        ws_summary.append(["Work-year projection", projections["year_cents"] / 100 if projections["year_cents"] is not None else None])
        ws_summary.append(["Normal workday hours", projections["day_hours"]])
        ws_summary.append(["Workdays per week", projections["workdays_per_week"]])
        for row in range(2, 5):
            ws_summary.cell(row, 2).number_format = '€0.00'
        for row in range(7, 12):
            ws_summary.cell(row, 2).number_format = '€0.00'
        for ws_current in workbook.worksheets:
            for column in ws_current.columns:
                letter = column[0].column_letter
                ws_current.column_dimensions[letter].width = min(45, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        workbook.save(path)

    def _apply_goal_progress_theme(self) -> None:
        dark = isDarkTheme()
        complete = bool(getattr(self, "_goal_is_complete", False))
        accent = COLORS["lavender_grey" if dark else "space_indigo"]
        track = (
            rgba_from_hex(COLORS["lavender_grey"], 0.18)
            if dark
            else COLORS["platinum"]
        )
        fill = COLORS["success"] if complete else COLORS["lavender_grey"]
        quest_fill = COLORS["success"] if complete else accent
        quest_background = "#292D40" if dark else "#F7F6FB"
        quest_track = rgba_from_hex(accent, 0.18 if dark else 0.12)
        quest_border = rgba_from_hex(accent, 0.52 if dark else 0.32)
        primary = get_text_color(dark, "primary")
        secondary = get_text_color(dark, "secondary")
        tertiary = get_text_color(dark, "tertiary")
        status = get_status_text_color("success", dark) if complete else secondary

        self.goal_quest_panel.setStyleSheet(f"""
            QWidget#earningsGoalQuest {{
                background-color: {quest_background};
                border: 1px solid {quest_border};
                border-radius: 12px;
            }}
        """)
        self.goal_quest_details.setStyleSheet(
            "QWidget#earningsGoalQuestDetails { background: transparent; border: none; }"
        )
        self.goal_quest_progress.setStyleSheet(f"""
            QProgressBar {{
                background-color: {quest_track};
                border: none;
                border-radius: {RADII['sm']}px;
                padding: 0;
            }}
            QProgressBar::chunk {{
                background-color: {quest_fill};
                border: none;
                border-radius: {RADII['sm']}px;
                margin: 0;
            }}
        """)
        self.goal_quest_title.setStyleSheet(
            f"color: {primary}; font-size: 16px; font-weight: 650; "
            "background: transparent; border: none;"
        )
        self.goal_quest_level.setStyleSheet(
            f"color: {quest_fill}; background-color: {quest_track}; "
            "border: none; border-radius: 8px; padding: 4px 9px; font-weight: 600;"
        )
        self.goal_quest_value.setStyleSheet(
            f"color: {primary}; font-size: 15px; font-weight: 600; "
            "background: transparent; border: none;"
        )
        self.goal_quest_percentage.setStyleSheet(
            f"color: {quest_fill}; font-size: 15px; font-weight: 700; "
            "background: transparent; border: none;"
        )
        self.goal_quest_next.setStyleSheet(
            f"color: {quest_fill}; font-weight: 600; background: transparent; border: none;"
        )
        self.goal_quest_remaining.setStyleSheet(
            f"color: {secondary}; background: transparent; border: none;"
        )
        self.goal_quest_empty.setStyleSheet(
            f"color: {secondary}; background: transparent; border: none;"
        )

        self.goal_details.setStyleSheet(
            "QWidget#earningsGoalDetails { background: transparent; border: none; }"
        )
        self.goal_forecast_block.setStyleSheet(
            "QWidget#earningsGoalForecast { background: transparent; border: none; }"
        )
        self.goal_adjustment_block.setStyleSheet(
            "QWidget#earningsGoalAdjustment { background: transparent; border: none; }"
        )
        self.goal_progress.setStyleSheet(f"""
            QProgressBar {{
                background-color: {track};
                border: none;
                border-radius: {RADII['sm']}px;
                padding: 0;
            }}
            QProgressBar::chunk {{
                background-color: {fill};
                border: none;
                border-radius: {RADII['sm']}px;
                margin: 0;
            }}
        """)
        self.goal_title.setStyleSheet(
            f"color: {primary}; font-weight: 600; background: transparent; border: none;"
        )
        self.goal_percentage.setStyleSheet(
            f"color: {primary}; font-weight: 600; background: transparent; border: none;"
        )
        self.goal_remaining.setStyleSheet(
            f"color: {status}; background: transparent; border: none;"
        )
        self.goal_adjustment_summary.setStyleSheet(
            f"color: {tertiary}; background: transparent; border: none;"
        )
        self.goal_adjustment_note.setStyleSheet(
            f"color: {secondary}; background: transparent; border: none;"
        )
        self.goal_forecast.setStyleSheet(
            f"color: {primary}; font-weight: 500; background: transparent; border: none;"
        )
        self.goal_forecast_detail.setStyleSheet(
            f"color: {tertiary}; background: transparent; border: none;"
        )
        self.goal_deadline.setStyleSheet(
            f"color: {secondary}; background: transparent; border: none;"
        )

    def _apply_theme(self):
        apply_screen_theme(self, "EarningsScreen", scroll=self.scroll, content=self.content)
        dark = isDarkTheme()
        canvas = get_surface_color(dark, "canvas")
        surface = get_surface_color(dark)
        metric_surface = get_surface_color(dark, "alternate") if dark else surface
        empty_surface = get_surface_color(dark, "alternate")
        text = COLORS["text_primary_dark" if dark else "text_primary_light"]
        muted = COLORS["text_secondary_dark" if dark else "text_secondary_light"]
        outline = get_subtle_border(dark)
        card_border = COMPONENT_COLORS["card"]["border_dark" if dark else "border_light"]
        secondary_hover = get_subtle_item_hover_bg(dark)
        danger = COLORS["error_text_dark" if dark else "error_text_light"]
        accent_colors = get_accent_colors(dark)
        accent = accent_colors["base"]
        accent_text = accent_colors["text"]
        accent_hover = accent_colors["hover"]
        accent_pressed = accent_colors["pressed"]
        accent_soft = rgba_from_hex(
            COLORS["lavender_grey" if dark else "space_indigo"],
            0.16 if dark else 0.07,
        )
        disabled_bg = COLORS["bg_alt_dark"] if dark else "#E6E9ED"
        table_colors = COMPONENT_COLORS["table"]
        table_bg = table_colors["row_bg_dark" if dark else "row_bg_light"]
        table_alt = table_colors["row_alt_bg_dark" if dark else "row_alt_bg_light"]
        table_border = table_colors["border_dark" if dark else "border_light"]
        table_header = table_colors["header_bg_dark" if dark else "header_bg_light"]
        table_header_text = table_colors["header_text_dark" if dark else "header_text_light"]
        self.setStyleSheet(
            self.styleSheet()
            + f"""
            EarningsScreen QWidget#earningsSurface,
            EarningsScreen QWidget#earningsMetric {{
                background-color: {surface};
                border: 1px solid {card_border};
                border-radius: 12px;
            }}
            EarningsScreen QWidget#earningsMetric {{
                background-color: {metric_surface};
            }}
            EarningsScreen QWidget#earningsTargetMetric {{
                background-color: {empty_surface};
                border: 1px solid {outline};
                border-radius: 10px;
            }}
            EarningsScreen QWidget#earningsTargetMetric QLabel {{
                background: transparent;
                border: none;
            }}
            EarningsScreen QLabel#earningsTargetTitle {{
                color: {muted};
                font-weight: 600;
            }}
            EarningsScreen QLabel#earningsTargetCurrent {{
                color: {text};
                font-size: 19px;
                font-weight: 700;
            }}
            EarningsScreen QLabel#earningsTargetTotal {{
                color: {muted};
            }}
            EarningsScreen QLabel#earningsTargetPercentage {{
                color: {accent};
                background-color: {accent_soft};
                border: 1px solid {accent};
                border-radius: 8px;
                padding: 2px 7px;
                font-weight: 700;
            }}
            EarningsScreen QProgressBar#earningsTargetProgress {{
                background-color: {disabled_bg};
                border: none;
                border-radius: 4px;
            }}
            EarningsScreen QProgressBar#earningsTargetProgress::chunk {{
                background-color: {accent};
                border: none;
                border-radius: 4px;
            }}
            EarningsScreen QLabel#earningsStreakBadge {{
                color: {accent};
                background-color: {accent_soft};
                border: 1px solid {accent};
                border-radius: 9px;
                padding: 3px 10px;
                font-weight: 600;
            }}
            EarningsScreen QWidget#earningsActivityBlock {{
                background-color: {accent_soft};
                border: 1px solid {outline};
                border-radius: 10px;
            }}
            EarningsScreen QWidget#earningsActivityBlock QLabel {{
                background: transparent;
                border: none;
            }}
            EarningsScreen QLabel#activityHeatmapHeading {{
                color: {text};
                font-weight: 650;
            }}
            EarningsScreen QWidget#earningsActivityMetric {{
                background-color: {surface};
                border: 1px solid {outline};
                border-radius: 8px;
            }}
            EarningsScreen QLabel#earningsActivityValue {{
                color: {text};
                font-size: 20px;
                font-weight: 700;
            }}
            EarningsScreen QLabel#earningsActivityCaption {{
                color: {muted};
            }}
            EarningsScreen PrimaryPushButton#earningsPrimaryAction,
            EarningsScreen PrimaryPushButton#timerStart {{
                background-color: {accent};
                border: 1px solid {accent};
                color: {accent_text};
                font-weight: 600;
                border-radius: 8px;
            }}
            EarningsScreen PrimaryPushButton#earningsPrimaryAction:hover,
            EarningsScreen PrimaryPushButton#timerStart:hover {{
                background-color: {accent_hover};
                border-color: {accent_hover};
            }}
            EarningsScreen PrimaryPushButton#earningsPrimaryAction:pressed,
            EarningsScreen PrimaryPushButton#timerStart:pressed {{
                background-color: {accent_pressed};
                border-color: {accent_pressed};
            }}
            EarningsScreen PrimaryPushButton#earningsPrimaryAction:disabled,
            EarningsScreen PrimaryPushButton#timerStart:disabled {{
                background-color: {disabled_bg};
                border-color: {table_border};
                color: {muted};
            }}
            EarningsScreen PushButton#earningsSecondaryAction {{
                background: transparent;
                border: 1px solid {outline};
                color: {text};
                border-radius: 7px;
            }}
            EarningsScreen PushButton#earningsSecondaryAction:hover {{
                background-color: {secondary_hover};
                border-color: {accent};
            }}
            EarningsScreen PushButton#earningsHeaderAction {{
                background-color: {accent_soft};
                border: 1px solid {accent};
                color: {accent};
                border-radius: 7px;
            }}
            EarningsScreen PushButton#earningsHeaderAction:hover {{
                background-color: {accent};
                border-color: {accent};
                color: {accent_text};
            }}
            EarningsScreen PushButton#earningsDangerAction {{
                background: transparent;
                border: 1px solid {outline};
                color: {danger};
            }}
            EarningsScreen PushButton#earningsDangerAction:hover {{
                background-color: {'#442A32' if dark else '#FFF1F0'};
                border-color: {danger};
            }}
            EarningsScreen PushButton#timerSecondary {{
                background: transparent;
                border: 1px solid {outline};
                border-radius: 7px;
            }}
            EarningsScreen PushButton#timerSecondary:hover {{
                background-color: {secondary_hover};
                border-color: {accent};
            }}
            EarningsScreen QWidget#earningsEmptyState {{
                background-color: {empty_surface};
                border: 1px solid {card_border};
                border-radius: 10px;
            }}
            EarningsScreen QTabWidget#earningsHistoryTabs::pane {{
                border: 1px solid {table_border};
                border-radius: 8px;
                background-color: {table_bg};
                top: -1px;
            }}
            EarningsScreen QTabWidget#earningsHistoryTabs QTabBar::tab {{
                background-color: transparent;
                color: {muted};
                border: none;
                border-bottom: 2px solid transparent;
                padding: 8px 14px;
                margin-right: 2px;
            }}
            EarningsScreen QTabWidget#earningsHistoryTabs QTabBar::tab:hover {{
                background-color: {accent_soft};
                color: {text};
            }}
            EarningsScreen QTabWidget#earningsHistoryTabs QTabBar::tab:selected {{
                background-color: {table_bg};
                color: {accent};
                border-bottom: 2px solid {accent};
                font-weight: 600;
            }}
            EarningsScreen QTableWidget {{
                background-color: {table_bg};
                alternate-background-color: {table_alt};
                color: {text};
                border: none;
                border-radius: 8px;
                gridline-color: transparent;
                selection-background-color: {get_selection_bg()};
                selection-color: {text};
            }}
            EarningsScreen QTableWidget::viewport {{
                background-color: {table_bg};
                border-radius: 8px;
            }}
            EarningsScreen QTableWidget::item {{
                border: none;
                border-bottom: 1px solid {table_border};
                padding: 7px 10px;
            }}
            EarningsScreen QTableWidget::item:hover {{
                background-color: {secondary_hover};
            }}
            EarningsScreen QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text};
            }}
            EarningsScreen QHeaderView::section {{
                background-color: {table_header};
                color: {table_header_text};
                border: none;
                border-right: 1px solid {outline};
                padding: 10px 12px;
                font-weight: 600;
            }}
            EarningsScreen QTableCornerButton::section {{
                background-color: {table_header};
                border: none;
            }}
            """
        )
        self.content.setStyleSheet(f"background-color: {canvas};")
        self.scroll.setStyleSheet(
            f"QScrollArea {{ border: none; background-color: {canvas}; }} "
            f"QScrollArea QWidget#qt_scrollarea_viewport {{ background-color: {canvas}; }}"
        )
        self.section_tabs.setStyleSheet(f"""
            QTabBar {{ background: transparent; border: none; }}
            QTabBar::tab {{
                background: transparent;
                color: {muted};
                border: none;
                border-bottom: 3px solid transparent;
                padding: 9px 20px 8px 20px;
                margin-right: 4px;
                font-weight: 500;
            }}
            QTabBar::tab:hover {{ background-color: {secondary_hover}; color: {text}; }}
            QTabBar::tab:selected {{
                background-color: {accent_soft};
                color: {accent};
                border-bottom: 3px solid {accent};
                font-weight: 600;
            }}
        """)
        surface_style = f"""
            QWidget#earningsSurface {{
                background-color: {surface};
                border: 1px solid {card_border};
                border-radius: 12px;
            }}
        """
        metric_style = f"""
            QWidget#earningsMetric {{
                background-color: {metric_surface};
                border: 1px solid {card_border};
                border-radius: 12px;
            }}
        """
        for widget in self.findChildren(QWidget, "earningsSurface"):
            widget.setStyleSheet(surface_style)
        for widget in self.findChildren(QWidget, "earningsMetric"):
            widget.setStyleSheet(metric_style)
        for widget in self.findChildren(QWidget, "earningsField"):
            widget.setStyleSheet("QWidget#earningsField { background: transparent; border: none; }")
        for widget in self.findChildren(QWidget, "earningsEmptyState"):
            widget.setStyleSheet(
                f"QWidget#earningsEmptyState {{ background-color: {empty_surface}; border: 1px solid {card_border}; border-radius: 10px; }}"
            )
        self.session_earnings.setStyleSheet(
            f"color: {COLORS['success_text_dark' if dark else 'success_text_light']}; "
            "font-size: 15px; font-weight: 700; background: transparent; border: none;"
        )

        primary_push_style = f"""
            PrimaryPushButton {{
                background-color: {accent};
                border: 1px solid {accent};
                border-radius: 8px;
                color: {accent_text};
                font-weight: 600;
                padding-left: 16px;
                padding-right: 16px;
            }}
            PrimaryPushButton:hover {{ background-color: {accent_hover}; border-color: {accent_hover}; }}
            PrimaryPushButton:pressed {{ background-color: {accent_pressed}; border-color: {accent_pressed}; }}
            PrimaryPushButton:disabled {{ background-color: {disabled_bg}; border-color: {table_border}; color: {muted}; }}
        """
        for button in (
            self.add_button,
            self.goal_create,
            self.goal_quest_create,
            self.analytics_empty_add,
            self.entries_empty_add,
            self.sessions_empty_add,
            self.timer_start,
            self.timer_pause,
        ):
            button.setStyleSheet(primary_push_style)
        secondary_style = f"""
            PushButton {{ background: transparent; border: 1px solid {outline}; border-radius: 7px; color: {text}; padding-left: 12px; padding-right: 12px; }}
            PushButton:hover {{ background-color: {secondary_hover}; border-color: {accent}; }}
        """
        for button in (
            self.goal_edit,
            self.goal_adjust,
            self.goal_archive,
            self.goal_quest_adjust,
            self.goal_quest_view,
            self.clear_filters_button,
            self.edit_entry_button,
            self.bulk_edit_button,
            self.timer_finish,
        ):
            button.setStyleSheet(secondary_style)
        danger_style = f"""
            PushButton {{ background: transparent; border: 1px solid {outline}; border-radius: 7px; color: {danger}; padding-left: 12px; padding-right: 12px; }}
            PushButton:hover {{ background-color: {'#442A32' if dark else '#FFF1F0'}; border-color: {danger}; }}
        """
        self.delete_entry_button.setStyleSheet(danger_style)
        self.timer_reset.setStyleSheet(danger_style)
        header_style = f"""
            PushButton {{ background-color: {accent_soft}; border: 1px solid {accent}; border-radius: 7px; color: {accent}; padding-left: 12px; padding-right: 12px; }}
            PushButton:hover {{ background-color: {accent}; border-color: {accent}; color: {accent_text}; }}
        """
        for button in (self.brands_button, self.settings_button, self.export_button):
            button.setStyleSheet(header_style)
        apply_earnings_datetime_theme(self.date_input)
        quest_surface = "#222C2A" if dark else "#F3FAF7"
        quest_border = "#355C52" if dark else "#B8DFD0"
        quest_track = "#303748" if dark else "#DDE6E2"
        quest_fill = COLORS["success"]
        self.quest_progress.setStyleSheet(
            f"""
            QWidget#earningsQuestProgress {{
                background-color: {quest_surface};
                border: 1px solid {quest_border};
                border-radius: 9px;
            }}
            QWidget#earningsQuestProgress QLabel {{
                background: transparent;
                border: none;
            }}
            QWidget#earningsQuestProgress QProgressBar {{
                background-color: {quest_track};
                border: none;
                border-radius: 6px;
            }}
            QWidget#earningsQuestProgress QProgressBar::chunk {{
                background-color: {quest_fill};
                border: none;
                border-radius: 6px;
            }}
            """
        )
        self._apply_goal_progress_theme()
        self.chart.update()

    def retranslate_ui(self):
        tr = self.main.i18n.tr
        self.title.setText(tr("earnings.title"))
        self._update_header_labels()
        self.export_filtered_action.setText(tr("earnings.export.filtered"))
        self.export_all_action.setText(tr("earnings.export.all"))
        self.section_tabs.setTabText(0, tr("earnings.tab.logging"))
        self.section_tabs.setTabText(1, tr("earnings.tab.analytics"))
        self.section_tabs.setTabText(2, tr("earnings.tab.history"))
        self.metric_today.title.setText(tr("earnings.metric.today"))
        self.metric_week.title.setText(tr("earnings.metric.week"))
        self.metric_all.title.setText(tr("earnings.metric.all"))
        self.metric_rate.title.setText(tr("earnings.metric.rate"))
        self.analytics_total.title.setText(tr("earnings.metric.total_earnings"))
        self.analytics_products.title.setText(tr("earnings.metric.products"))
        self.analytics_hours.title.setText(tr("earnings.metric.hours"))
        self.analytics_rate.title.setText(tr("earnings.metric.rate"))
        self.entry_heading.setText(tr("earnings.add.title"))
        self.add_button.setText(tr("earnings.add.action"))
        self.timer_heading.setText(tr("earnings.timer.title"))
        self.goal_heading.setText(tr("earnings.goal.title"))
        self.goal_create.setText(tr("earnings.goal.quest.create.action"))
        self.goal_edit.setText(tr("earnings.goal.edit.action"))
        self.goal_adjust.setText(tr("earnings.goal.adjust.action"))
        self.goal_archive.setText(tr("earnings.goal.archive.action"))
        self.goal_quest_adjust.setText(tr("earnings.goal.adjust.action"))
        self.goal_quest_view.setText(tr("earnings.goal.quest.view"))
        self.performance_heading.setText(tr("earnings.performance.title"))
        for key, target in self.performance_widgets.items():
            target.set_title(tr(self.performance_label_keys[key]))
        self.no_performance_targets.setText(tr("earnings.performance.empty"))
        self.activity_heatmap_heading.setText(tr("earnings.activity_30_days"))
        self.activity_heatmap_explanation.setText(tr("earnings.activity_legend"))
        self.activity_heatmap_legend.retranslate_ui()
        self.activity_heatmap.retranslate_ui()
        for (_value, caption), key in zip(
            self.activity_summary_widgets,
            (
                "earnings.activity.active_days",
                "earnings.activity.total_skus",
                "earnings.activity.best_day",
            ),
        ):
            caption.setText(tr(key))
        self.analytics_heading.setText(tr("earnings.analytics.title"))
        self.projection_heading.setText(tr("earnings.projection.title"))
        self.projection_day.title.setText(tr("earnings.projection.day"))
        self.projection_week.title.setText(tr("earnings.projection.week"))
        self.projection_month.title.setText(tr("earnings.projection.month"))
        self.projection_year.title.setText(tr("earnings.projection.year"))
        self.history_heading.setText(tr("earnings.history.title"))
        self.sku_label.setText(tr("earnings.field.sku"))
        self.name_label.setText(tr("earnings.field.name"))
        self.brand_label.setText(tr("earnings.field.brand"))
        self.type_label.setText(tr("earnings.field.type"))
        self.date_label.setText(tr("earnings.field.date"))
        self.edit_entry_button.setText(tr("earnings.history.edit"))
        self.bulk_edit_button.setText(tr("earnings.history.bulk_edit"))
        self.bulk_edit_button.setToolTip(tr("earnings.history.bulk_edit.tooltip"))
        self.delete_entry_button.setText(tr("earnings.history.delete"))
        self.clear_filters_button.setText(tr("earnings.history.clear_filters"))
        self.analytics_empty_title.setText(tr("earnings.analytics.empty.title"))
        self.analytics_empty_body.setText(tr("earnings.analytics.empty.body"))
        self.analytics_empty_add.setText(tr("earnings.add.action"))
        self.entries_empty_title.setText(tr("earnings.history.empty.title"))
        self.entries_empty_body.setText(tr("earnings.history.empty.body"))
        self.entries_empty_add.setText(tr("earnings.add.action"))
        self.sessions_empty_title.setText(tr("earnings.sessions.empty.title"))
        self.sessions_empty_body.setText(tr("earnings.sessions.empty.body"))
        self.sessions_empty_add.setText(tr("earnings.sessions.empty.action"))
        self.history_tabs.setTabText(0, tr("earnings.history.earnings"))
        self.history_tabs.setTabText(1, tr("earnings.history.sessions"))
        self.sku_input.setPlaceholderText(tr("earnings.sku.placeholder"))
        self.name_input.setPlaceholderText(tr("earnings.name.placeholder"))
        self.search.setPlaceholderText(tr("earnings.search.placeholder"))
        self._update_history_actions()
        self._refresh_projections()
        self._refresh_goal()
        self._timer_tick()
