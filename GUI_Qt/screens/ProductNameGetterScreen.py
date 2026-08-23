"""Collect every product name from all product-list pages."""

from __future__ import annotations

import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import Signal
from PySide6.QtWidgets import QFileDialog
from qfluentwidgets import FluentIcon, InfoBar, InfoBarPosition

from GUI_Qt.screens.CodeGetterScreen import CodeGetterScreen, CodeGetterWorker
from GUI_Qt.widgets import show_file_saved_bar


class ProductNameGetterWorker(CodeGetterWorker):
    row_found = Signal(int, str, str, str, str)  # index, name, code, page, status

    def _process_row(self, page_no: int, row_index: int):
        display_index = self._found + self._errors + 1
        page_label = str(page_no)

        try:
            _, title, visible_code = self._get_row_snapshot(row_index)
            self.progress_update.emit(display_index, f"Page {page_no}, row {row_index + 1}")

            status = "Found" if title else "Missing name"
            if title:
                self._found += 1
            else:
                self._errors += 1

            self.row_found.emit(display_index, title, visible_code, page_label, status)
            self.log.emit(f"{display_index}: {title or '(missing)'}")
        except Exception as e:
            self._errors += 1
            self.row_found.emit(display_index, "", "", page_label, f"Error: {e}")
            self.log.emit(f"{display_index}: error {e}")


class ProductNameGetterScreen(CodeGetterScreen):
    def __init__(self, main_window):
        super().__init__(main_window)
        self._worker: ProductNameGetterWorker | None = None

    @staticmethod
    def _sort_key(name: str) -> tuple[str, str]:
        value = (name or "").strip()
        lowered = value.casefold()

        if lowered.startswith("orbea "):
            tokens = value.split()
            model_tokens = tokens[1:]
            for index, token in enumerate(model_tokens):
                if re.fullmatch(r"20\d{2}", token):
                    model_tokens = model_tokens[:index]
                    break

            model_key = " ".join(model_tokens).casefold().strip()
            return ("orbea", model_key or lowered)

        return (lowered, lowered)

    @staticmethod
    def _orbea_model_info(name: str) -> tuple[str, str]:
        value = (name or "").strip()
        lowered = value.casefold()

        if not lowered.startswith("orbea "):
            return "", ""

        tokens = value.split()
        model_tokens = tokens[1:]
        for index, token in enumerate(model_tokens):
            if re.fullmatch(r"20\d{2}", token):
                model_tokens = model_tokens[:index]
                break

        model = " ".join(model_tokens).strip()
        model_key = model.casefold()
        return model, model_key

    def _build_ui(self):
        super()._build_ui()
        self._table.setColumnWidth(1, 420)
        self._table.setColumnWidth(2, 180)

    def retranslate_ui(self):
        self._title.setText(self.tr("productnamegetter.title"))
        self._subtitle.setText(self.tr("productnamegetter.subtitle"))
        self._start_btn.setText(self.tr("productnamegetter.start"))
        self._export_btn.setText(self.tr("productnamegetter.export"))
        self._status_label.setText(self.tr("productnamegetter.ready"))
        self._table.setHorizontalHeaderLabels([
            self.tr("productnamegetter.col.index"),
            self.tr("productnamegetter.col.name"),
            self.tr("productnamegetter.col.code"),
            self.tr("productnamegetter.col.page"),
            self.tr("productnamegetter.col.status"),
        ])

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        self._table.setRowCount(0)
        self._export_btn.setEnabled(False)
        self._start_btn.setText(self.tr("productnamegetter.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._status_label.setText(self.tr("productnamegetter.running"))

        self._worker = ProductNameGetterWorker(driver)
        self._worker.row_found.connect(self._on_row_found)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[ProductNameGetter] {msg}"))
        if hasattr(self.main, "track_worker"):
            self.main.track_worker(
                self._worker, "name_scanner", "product_name_getter"
            )
        self._worker.start()

    def _on_progress(self, current: int, label: str):
        self._progress_label.setText(
            self.tr("productnamegetter.progress", current=current, label=label)
        )

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("productnamegetter.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._export_btn.setEnabled(self._table.rowCount() > 0)

        total = found + errors
        done_text = self.tr(
            "productnamegetter.done",
            found=found,
            errors=errors,
            total=total,
        )
        self._status_label.setText(done_text)
        self._progress_label.setText(done_text)

        InfoBar.success(
            self.tr("productnamegetter.done_title"),
            done_text,
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self,
            self.tr("productnamegetter.export"),
            f"product_names_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)",
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Product Names"

            headers = ["#", "Name", "Code", "Page", "Status"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid",
            )
            group_fill = PatternFill(
                start_color="D9EAF7",
                end_color="D9EAF7",
                fill_type="solid",
            )
            group_font = Font(bold=True, color="1F1F1F")

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            rows = []
            for row in range(self._table.rowCount()):
                values = []
                for col in range(5):
                    item = self._table.item(row, col)
                    values.append(item.text() if item else "")
                rows.append(values)

            # Group Orbea products by model frequency so the most common models
            # appear first. Preserve relative ordering within each model group.
            from collections import defaultdict

            model_buckets: dict[str, list[list[str]]] = defaultdict(list)
            other_rows: list[list[str]] = []

            for values in rows:
                name = values[1]
                model_name, model_key = self._orbea_model_info(name)
                if model_key:
                    model_buckets[model_key].append(values)
                else:
                    other_rows.append(values)

            # Order models by descending frequency (most common first)
            ordered_models = sorted(
                model_buckets.items(), key=lambda kv: len(kv[1]), reverse=True
            )

            # Build final ordered rows: all Orbea rows ordered by model frequency,
            # then the remaining non-Orbea rows (sorted by name for stability).
            ordered_rows: list[list[str]] = []
            for model_key, group in ordered_models:
                ordered_rows.extend(group)

            other_rows.sort(key=lambda r: (r[1] or "").casefold())
            ordered_rows.extend(other_rows)

            # Write rows to worksheet without complex merged headers.
            current_row = 2
            for values in ordered_rows:
                for col_index, value in enumerate(values, start=1):
                    ws.cell(row=current_row, column=col_index, value=value)
                current_row += 1

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 70
            ws.column_dimensions["C"].width = 24
            ws.column_dimensions["D"].width = 10
            ws.column_dimensions["E"].width = 22

            wb.save(path)
            show_file_saved_bar(
                self,
                self.tr("common.success"),
                self.tr("productnamegetter.exported"),
                path,
            )
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )
