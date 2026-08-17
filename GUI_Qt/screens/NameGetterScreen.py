"""GUI_Qt/screens/NameGetterScreen.py

Batch product name retriever.

Reads product codes from an Excel file, searches each code on the
pim.bo product list via the External ID filter, and extracts the
product name from the results table (td.cell-name).

No navigation into individual product pages is needed — names are
read directly from the list view, making this very fast.
"""

from __future__ import annotations

import os
import time
import threading
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Qt, Signal
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent
from PySide6.QtWidgets import (
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    FluentIcon,
    IconWidget,
    InfoBar,
    InfoBarPosition,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    ComboBox,
    TitleLabel,
    isDarkTheme,
)

from GUI_Qt.widgets import enable_table_copy, show_file_saved_bar
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import (
    COLORS, COMPONENT_COLORS, FONTS, RADII, PADDINGS, SIZES, SPACING as THEME_SPACING,
    get_status_text_color, get_text_color,
)
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS, PAGE_SPACING, ICON_TEXT_GAP, ROW_SPACING, TOOLBAR_MARGINS,
    CARD_SPACING, CONTENT_SPACING, SPACING,
    apply_screen_theme, enforce_transparent_labels,
)
from Config.Selectors import ProductListSelectors
from Managers.PimboProductEditor import PIMBO_LOGIN_URL, PimboProductEditor


# ---------------------------------------------------------------------------
# Drop zone (reused pattern)
# ---------------------------------------------------------------------------

class DropZoneWidget(QWidget):
    """Drag & drop zone for a single .xlsx file."""

    file_dropped = Signal(str)

    def __init__(self, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(PAGE_SPACING)

        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES['icon_huge'], SIZES['icon_huge'])

        self.title_label = BodyLabel("")
        self.title_label.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']}; color: {COLORS['lavender_grey']};")

        self.subtitle_label = CaptionLabel("")
        self.subtitle_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        layout.addStretch()
        layout.addWidget(icon, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.title_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.subtitle_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addStretch()

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._apply_style(False)
        self.retranslate_ui()

    def retranslate_ui(self):
        self.title_label.setText(self.tr("batch.drop.title"))
        self.subtitle_label.setText(self.tr("batch.drop.subtitle"))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1 and urls[0].toLocalFile().endswith('.xlsx'):
                event.acceptProposedAction()
                self._apply_style(True)

    def dragLeaveEvent(self, event):
        self._apply_style(False)

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith('.xlsx'):
            self.file_dropped.emit(files[0])
        self._apply_style(False)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.file_dropped.emit("__browse__")

    def _apply_style(self, drag_active: bool):
        is_dark = isDarkTheme()
        border = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        dashed = COLORS['lavender_grey']
        from GUI_Qt.styles.theme_config import get_hover_bg
        hover_bg = get_hover_bg(is_dark)

        if drag_active:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: {hover_bg};
                    border: 2px solid {border};
                    border-radius: {RADII['lg']}px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: transparent;
                    border: 2px dashed {dashed};
                    border-radius: {RADII['lg']}px;
                }}
                DropZoneWidget:hover {{
                    border-color: {border};
                    background-color: {hover_bg};
                }}
            """)


# ---------------------------------------------------------------------------
# Background worker
# ---------------------------------------------------------------------------

class NameGetterWorker(QThread):
    """Searches each code on the product list and extracts the name."""

    row_update = Signal(int, str, str)      # row_index, status, product_name
    log = Signal(str)
    done = Signal(int, int)                 # found, errors
    progress_update = Signal(int, int, float, float)  # current, total, speed, eta

    def __init__(self, main_window, codes: list[str], browser_count: int = 1):
        super().__init__()
        self.main = main_window
        self.codes = codes
        self.browser_count = browser_count
        self._stop = False

        self._lock = threading.Lock()
        self._work_index = 0
        self._found = 0
        self._errors = 0
        self._processed_count = 0
        self._processed_times: list[float] = []
        self._seen: dict[str, tuple[str, str]] = {}  # code_upper -> (status, name)

        self._session_manager = None

    def request_stop(self):
        self._stop = True

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _ensure_products_page(driver) -> bool:
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        if "/dashboard/products/" in (driver.current_url or ""):
            if PimboProductEditor(driver).is_dirty():
                raise RuntimeError(
                    "Atidaryta forma turi neišsaugotų pakeitimų; vardų tikrinimas jos neatmes."
                )
        try:
            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.NAV_PRODUCTS)
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
            )
        except Exception:
            try:
                driver.get(ProductListSelectors.URL)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
                )
            except Exception:
                pass
        return False

    def _login_session(self, driver) -> bool:
        from Config.LoginConfig.LoginHandler import LoginHandler
        email, password = self.main.credential_manager.get_saved_credentials()
        if not email or not password:
            return False
        login_handler = LoginHandler(driver, self.main.credential_manager, logger=self.main.logger)
        return login_handler.attempt_login(email, password)

    # -- progress ------------------------------------------------------------

    def _emit_progress(self):
        with self._lock:
            current = self._processed_count
            times = list(self._processed_times)
        total = len(self.codes)
        if not times:
            return
        avg = sum(times) / len(times)
        speed = 1.0 / avg if avg > 0 else 0
        remaining = total - current
        eta = remaining * avg
        self.progress_update.emit(current, total, speed, eta)

    # -- search & extract name -----------------------------------------------

    def _setup_filter(self, driver) -> None:
        """Wait until the dashboard product search field is ready."""
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        driver.execute_script("window.scrollTo(0, 0);")
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(ProductListSelectors.SEARCH_NAME)
        )

    def _search_and_get_name(self, code: str, driver) -> str | None:
        """Search for a code using the dashboard search and return the name.

        Returns the product name string, or None if not found.
        """
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.common.exceptions import TimeoutException

        search_input = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable(ProductListSelectors.SEARCH_NAME)
        )
        previous_rows = tuple(
            row.text for row in driver.find_elements(*ProductListSelectors.PRODUCT_ROWS)
        )
        search_input.send_keys(Keys.CONTROL, "a")
        search_input.send_keys(code)
        search_input.send_keys(Keys.ENTER)

        def exact_result_loaded(current_driver):
            exact = current_driver.find_elements(*ProductListSelectors.product_row_by_code(code))
            current_rows = tuple(
                row.text for row in current_driver.find_elements(*ProductListSelectors.PRODUCT_ROWS)
            )
            return exact[0] if exact else (False if current_rows == previous_rows else False)

        # Try to find the product row matching this code
        try:
            row = WebDriverWait(driver, 8).until(exact_result_loaded)
        except TimeoutException:
            return None

        try:
            name_element = row.find_element(
                *ProductListSelectors.product_link_by_brand("")
            )
            return name_element.text.strip()
        except Exception:
            return None

    # -- single-item check ---------------------------------------------------

    def _check_one(self, idx: int, code: str, driver):
        tr = self.main.i18n.tr
        total = len(self.codes)
        code_upper = code.strip().upper()

        # Duplicate detection
        is_dup = False
        with self._lock:
            if code_upper in self._seen:
                prev_status, prev_name = self._seen[code_upper]
                self._processed_count += 1
                is_dup = True
        if is_dup:
            self.row_update.emit(idx, tr("namegetter.status.duplicate"), prev_name)
            self.log.emit(f"[{idx + 1}/{total}] {code} — {tr('namegetter.status.duplicate')}")
            self._emit_progress()
            return

        t0 = time.perf_counter()
        self.row_update.emit(idx, tr("batchdesc.status.running"), "")
        self.log.emit(f"[{idx + 1}/{total}] {tr('namegetter.log.searching', code=code)}")

        try:
            # Search and extract name (filter is already set up)
            name = self._search_and_get_name(code, driver)

            if name:
                status = tr("namegetter.status.found")
                with self._lock:
                    self._seen[code_upper] = (status, name)
                    self._found += 1
                self.row_update.emit(idx, status, name)
                self.log.emit(f"[{idx + 1}/{total}] {code} — {name}")
            else:
                status = tr("namegetter.status.not_found")
                with self._lock:
                    self._seen[code_upper] = (status, "")
                    self._errors += 1
                self.row_update.emit(idx, status, "")
                self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")

        except Exception as e:
            status = f"{tr('namegetter.status.error')}: {e}"
            with self._lock:
                self._seen[code_upper] = (status, "")
                self._errors += 1
            self.row_update.emit(idx, status, "")
            self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")

        elapsed = time.perf_counter() - t0
        with self._lock:
            self._processed_times.append(elapsed)
            self._processed_count += 1
        self._emit_progress()

    # -- worker thread -------------------------------------------------------

    def _worker_thread(self, driver):
        # One-time setup: navigate to product list and open the filter
        self._ensure_products_page(driver)
        self._setup_filter(driver)

        while not self._stop:
            with self._lock:
                if self._work_index >= len(self.codes):
                    break
                idx = self._work_index
                self._work_index += 1
            code = self.codes[idx]
            self._check_one(idx, code, driver)

    # -- main entry ----------------------------------------------------------

    def run(self):
        tr = self.main.i18n.tr
        total = len(self.codes)

        main_driver = getattr(self.main, "driver", None)
        if main_driver is None:
            for i in range(total):
                self.row_update.emit(i, tr("namegetter.status.error"), tr("batchdesc.no_session"))
            self.done.emit(0, total)
            return

        extra_drivers = []

        if self.browser_count > 1:
            from Managers.BrowserSessionManager import BrowserSessionManager

            pool_size = self.browser_count - 1
            self.log.emit(tr("namegetter.log.init_browsers", count=self.browser_count))

            browser_type = self.main.settings.get('browser_choice', 'Chrome')
            self._session_manager = BrowserSessionManager(
                browser_type=browser_type,
                pool_size=pool_size,
                logger=self.main.logger,
            )

            if self._session_manager.initialize_pool():
                for session in self._session_manager.sessions:
                    if self._stop:
                        break
                    self.log.emit(f"Logging in browser {session.session_id + 2}...")
                    session.driver.get(PIMBO_LOGIN_URL)
                    if self._login_session(session.driver):
                        extra_drivers.append(session.driver)
                        self.log.emit(f"Browser {session.session_id + 2} logged in")
                    else:
                        self.log.emit(f"Browser {session.session_id + 2} login failed, skipping")
            else:
                self.log.emit("Failed to create browser pool, using single browser")

        all_drivers = [main_driver] + extra_drivers
        actual_count = len(all_drivers)
        self.log.emit(f"Starting name getter with {actual_count} browser(s), {total} codes")

        if actual_count == 1:
            self._worker_thread(main_driver)
        else:
            threads = []
            for driver in all_drivers:
                t = threading.Thread(target=self._worker_thread, args=(driver,), daemon=True)
                threads.append(t)
                t.start()
            for t in threads:
                t.join()

        # Mark remaining as skipped if stopped early
        if self._stop:
            for j in range(self._work_index, total):
                self.row_update.emit(j, tr("namegetter.status.skipped"), "")

        if self._session_manager:
            self._session_manager.shutdown_all()
            self._session_manager = None

        self.done.emit(self._found, self._errors)


# ---------------------------------------------------------------------------
# Main screen
# ---------------------------------------------------------------------------

class NameGetterScreen(ResponsiveWidget):

    def __init__(self, main_window):
        super().__init__(main_window)
        self.main = main_window
        self.tr = main_window.i18n.tr
        self._worker: NameGetterWorker | None = None
        self._codes: list[str] = []
        self._file_path: str | None = None

        self._build_ui()
        self.retranslate_ui()

    # -- UI construction -----------------------------------------------------

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        self._scroll = ScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        root.addWidget(self._scroll)

        self._container = QWidget()
        self._layout = QVBoxLayout(self._container)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(CONTENT_SPACING)
        self._scroll.setWidget(self._container)

        apply_screen_theme(
            self,
            "NameGetterScreen",
            scroll=self._scroll,
            content=self._container,
        )

        # -- Header ----------------------------------------------------------
        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)
        icon = IconWidget(FluentIcon.SEARCH)
        icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        header.addWidget(icon)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self._title = TitleLabel("")
        self._subtitle = CaptionLabel("")
        self._subtitle.setStyleSheet(f"color: {COLORS['text_secondary']};")
        title_col.addWidget(self._title)
        title_col.addWidget(self._subtitle)
        header.addLayout(title_col)
        header.addStretch()
        self._layout.addLayout(header)

        # -- Toolbar card ----------------------------------------------------
        toolbar_card = CardWidget()
        toolbar_layout = QGridLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self._browse_btn = PushButton(FluentIcon.FOLDER, "")
        self._browse_btn.clicked.connect(self._browse_file)

        self._template_btn = PushButton(FluentIcon.DOWNLOAD, "")
        self._template_btn.clicked.connect(self._download_template)

        self._file_label = CaptionLabel("")
        self._file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        # Browser count selector
        self._browser_label = BodyLabel("")
        self._browser_combo = ComboBox()
        self._browser_combo.addItems(["1", "2", "3", "4"])
        self._browser_combo.setCurrentIndex(0)
        self._browser_combo.setFixedWidth(80)

        self._start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._start_btn.setEnabled(False)
        self._start_btn.clicked.connect(self._on_start_stop)

        self._export_btn = PushButton(FluentIcon.SAVE, "")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_results)

        toolbar_layout.addWidget(self._browse_btn, 0, 0)
        toolbar_layout.addWidget(self._template_btn, 0, 1)
        toolbar_layout.addWidget(self._file_label, 1, 0, 1, 2)
        toolbar_layout.addWidget(self._browser_label, 2, 0)
        toolbar_layout.addWidget(self._browser_combo, 2, 1)
        toolbar_layout.addWidget(self._export_btn, 3, 0)
        toolbar_layout.addWidget(self._start_btn, 3, 1)
        toolbar_layout.setColumnStretch(0, 1)
        toolbar_layout.setColumnStretch(1, 1)
        self._layout.addWidget(toolbar_card)

        # -- Drop zone -------------------------------------------------------
        self._drop_zone = DropZoneWidget(self.tr)
        self._drop_zone.setMinimumHeight(120)
        self._drop_zone.file_dropped.connect(self._on_file_dropped)
        self._layout.addWidget(self._drop_zone)

        # -- Progress label --------------------------------------------------
        self._progress_label = CaptionLabel("")
        self._progress_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self._layout.addWidget(self._progress_label)

        # -- Results table ---------------------------------------------------
        self._table = QTableWidget(0, 4)
        self._table.setHorizontalHeaderLabels(["#", "Code", "Status", "Name"])
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Interactive)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._table.setColumnWidth(1, 160)
        self._table.setColumnWidth(2, 130)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setMinimumHeight(200)
        enable_table_copy(self._table)
        self._layout.addWidget(self._table, 1)

        self._update_table_theme()
        enforce_transparent_labels(self)

    # -- Table theme ---------------------------------------------------------

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        tc = COMPONENT_COLORS['table']
        bg = tc['row_alt_bg_dark'] if is_dark else tc['row_alt_bg_light']
        alt_bg = tc['row_bg_dark'] if is_dark else tc['row_bg_light']
        border = tc['border_dark'] if is_dark else tc['border_light']
        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']
        text_color = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']

        from GUI_Qt.styles.theme_config import (
            get_scrollbar_handle_bg, get_scrollbar_handle_hover_bg, get_selection_bg,
        )

        self._table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg};
                alternate-background-color: {alt_bg};
                border: 1px solid {border};
                border-radius: {RADII['md']}px;
                gridline-color: {border};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::viewport {{
                background-color: {bg};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-bottom: 1px solid {border};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
                letter-spacing: 0.5px;
                text-transform: uppercase;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}
            QHeaderView::section {{
                border-right: 1px solid {border};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {THEME_SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}
        """)

    # -- File handling -------------------------------------------------------

    def _on_file_dropped(self, path: str):
        if path == "__browse__":
            self._browse_file()
            return
        self._load_file(path)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("batch.file.select_excel.title"),
            "", "Excel files (*.xlsx);;All files (*.*)"
        )
        if path:
            self._load_file(path)

    def _load_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            codes = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                val = row[0]
                if val is not None:
                    code = str(val).strip()
                    if code:
                        if not code.upper().startswith("UB-"):
                            code = f"UB-{code}"
                        codes.append(code)
            wb.close()

            if not codes:
                InfoBar.warning(
                    self.tr("common.error"),
                    self.tr("namegetter.no_codes"),
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                )
                return

            self._codes = codes
            self._file_path = path
            self._file_label.setText(os.path.basename(path))

            # Populate table
            self._table.setRowCount(len(codes))
            for i, code in enumerate(codes):
                self._table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self._table.setItem(i, 1, QTableWidgetItem(code))
                item = QTableWidgetItem(self.tr("batchdesc.status.pending"))
                item.setForeground(QColor(COLORS['text_secondary']))
                self._table.setItem(i, 2, item)
                self._table.setItem(i, 3, QTableWidgetItem(""))

            self._start_btn.setEnabled(True)
            self._export_btn.setEnabled(False)
            self._progress_label.setText(f"{len(codes)} {self.tr('namegetter.codes_loaded')}")
            self._drop_zone.setVisible(False)

        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("batch.template.save.title"),
            "name_getter_template.xlsx", "Excel files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Codes"
            ws["A1"] = "Code"
            ws["A1"].font = Font(bold=True)
            ws.column_dimensions["A"].width = 25
            wb.save(path)
            show_file_saved_bar(
                self,
                self.tr("common.success"),
                self.tr("batch.template.saved"),
                path,
            )
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    # -- Start / Stop --------------------------------------------------------

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        if not self._codes:
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        # Reset table statuses
        for i in range(self._table.rowCount()):
            item = QTableWidgetItem(self.tr("batchdesc.status.pending"))
            item.setForeground(QColor(COLORS['text_secondary']))
            self._table.setItem(i, 2, item)
            self._table.setItem(i, 3, QTableWidgetItem(""))

        self._start_btn.setText(self.tr("namegetter.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._browse_btn.setEnabled(False)
        self._export_btn.setEnabled(False)
        self._browser_combo.setEnabled(False)

        browser_count = int(self._browser_combo.currentText())

        self._worker = NameGetterWorker(self.main, self._codes, browser_count)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[NameGetter] {msg}"))
        self._worker.start()

    def _on_row_update(self, row: int, status: str, name: str):
        tr = self.tr
        item = QTableWidgetItem(status)

        if status == tr("namegetter.status.found"):
            item.setForeground(QColor(get_status_text_color("success", isDarkTheme())))
        elif status == tr("namegetter.status.duplicate"):
            item.setForeground(QColor(get_status_text_color("warning", isDarkTheme())))
        elif status == tr("batchdesc.status.running"):
            item.setForeground(QColor(get_text_color(isDarkTheme(), "secondary")))
        elif status == tr("batchdesc.status.pending"):
            pass
        else:
            item.setForeground(QColor(get_status_text_color("error", isDarkTheme())))

        self._table.setItem(row, 2, item)

        if name:
            self._table.setItem(row, 3, QTableWidgetItem(name))

        self._table.scrollToItem(item)

    def _on_progress(self, current: int, total: int, speed: float, eta: float):
        mins = int(eta // 60)
        secs = int(eta % 60)
        self._progress_label.setText(
            f"{current}/{total} — {speed:.1f} items/s — ETA: {mins}m {secs}s"
        )

    def _on_done(self, found: int, errors: int):
        self._start_btn.setText(self.tr("namegetter.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._browse_btn.setEnabled(True)
        self._export_btn.setEnabled(True)
        self._browser_combo.setEnabled(True)

        total = found + errors
        self._progress_label.setText(
            self.tr("namegetter.done", found=found, errors=errors, total=total)
        )

        InfoBar.success(
            self.tr("namegetter.done_title"),
            self.tr("namegetter.done", found=found, errors=errors, total=total),
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    # -- Export --------------------------------------------------------------

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("namegetter.export"),
            f"product_names_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            headers = ["#", "Code", "Status", "Name"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(4):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 60

            wb.save(path)
            show_file_saved_bar(
                self,
                self.tr("common.success"),
                self.tr("batch.template.saved"),
                path,
            )
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    # -- i18n ----------------------------------------------------------------

    def retranslate_ui(self):
        self.tr = self.main.i18n.tr
        self._title.setText(self.tr("namegetter.title"))
        self._subtitle.setText(self.tr("namegetter.subtitle"))
        self._browse_btn.setText(self.tr("batch.browse_excel"))
        self._template_btn.setText(self.tr("batch.download_template"))
        self._file_label.setText(self.tr("batch.no_file") if not self._file_path else os.path.basename(self._file_path))
        self._export_btn.setText(self.tr("namegetter.export"))
        self._browser_label.setText(self.tr("namegetter.browsers"))

        if self._worker and self._worker.isRunning():
            self._start_btn.setText(self.tr("namegetter.stop"))
        else:
            self._start_btn.setText(self.tr("namegetter.start"))

        self._table.setHorizontalHeaderLabels([
            self.tr("namegetter.col.index"),
            self.tr("namegetter.col.code"),
            self.tr("namegetter.col.status"),
            self.tr("namegetter.col.name"),
        ])

        self._drop_zone.tr = self.tr
        self._drop_zone.retranslate_ui()
