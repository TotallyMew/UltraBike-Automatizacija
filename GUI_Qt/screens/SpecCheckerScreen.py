"""GUI_Qt/screens/SpecCheckerScreen.py

Batch specification template checker.

Reads product codes from an Excel file, navigates to each product in the
pim.bo admin panel, and checks whether the specification template has been
applied (i.e. spec rows exist on the Specifications tab).

Supports multi-browser parallel checking via BrowserSessionManager.
"""

from __future__ import annotations

import os
import time
import threading
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QThread, Qt, Signal
from PySide6.QtGui import QColor, QDragEnterEvent, QDropEvent
from PySide6.QtWidgets import (
    QFileDialog,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    FluentIcon,
    IconWidget,
    InfoBar,
    InfoBarPosition,
    PrimaryPushButton,
    PushButton,
    ScrollArea,
    ComboBox,
    TitleLabel,
    isDarkTheme,
)

from GUI_Qt.widgets import enable_table_copy, show_file_saved_bar
from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.styles.theme_config import (
    COLORS, COMPONENT_COLORS, FONTS, RADII, PADDINGS, SIZES, SPACING as THEME_SPACING,
)
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS, PAGE_SPACING, ICON_TEXT_GAP, ROW_SPACING, TOOLBAR_MARGINS,
    CARD_SPACING, CONTENT_SPACING, SPACING,
    apply_screen_theme, enforce_transparent_labels,
    get_responsive_margins, get_responsive_spacing,
)
from Config.Selectors import ProductListSelectors
from Managers.PimboProductEditor import PIMBO_LOGIN_URL, PimAutomationError, PimboProductEditor
from Utilities.ProductNavigationHandler import ProductNavigationHandler


# ---------------------------------------------------------------------------
# Drop zone (reused pattern from BatchDescriptionsScreen)
# ---------------------------------------------------------------------------

class DropZoneWidget(QWidget):
    """Drag & drop zone for a single .xlsx file."""

    file_dropped = Signal(str)

    def __init__(self, tr, parent=None):
        super().__init__(parent)
        self.tr = tr
        self.setAcceptDrops(True)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.setSpacing(PAGE_SPACING)

        icon = IconWidget(FluentIcon.DOCUMENT)
        icon.setFixedSize(SIZES['icon_huge'], SIZES['icon_huge'])

        self.title_label = BodyLabel("")
        self.title_label.setStyleSheet(f"font-size: {FONTS['size_subtitle_2']}; color: {COLORS['lavender_grey']};")

        self.subtitle_label = CaptionLabel("")
        self.subtitle_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        layout.addStretch()
        layout.addWidget(icon, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.title_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.subtitle_label, 0, Qt.AlignmentFlag.AlignCenter)
        layout.addStretch()

        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._apply_style(False)
        self.retranslate_ui()

    def retranslate_ui(self):
        self.title_label.setText(self.tr("batch.drop.title"))
        self.subtitle_label.setText(self.tr("batch.drop.subtitle"))

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1 and urls[0].toLocalFile().endswith('.xlsx'):
                event.acceptProposedAction()
                self._apply_style(True)

    def dragLeaveEvent(self, event):
        self._apply_style(False)

    def dropEvent(self, event: QDropEvent):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith('.xlsx'):
            self.file_dropped.emit(files[0])
        self._apply_style(False)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.file_dropped.emit("__browse__")

    def _apply_style(self, drag_active: bool):
        is_dark = isDarkTheme()
        border = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        dashed = COLORS['lavender_grey']
        from GUI_Qt.styles.theme_config import get_hover_bg
        hover_bg = get_hover_bg(is_dark)

        if drag_active:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: {hover_bg};
                    border: 2px solid {border};
                    border-radius: {RADII['lg']}px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                DropZoneWidget {{
                    background-color: transparent;
                    border: 2px dashed {dashed};
                    border-radius: {RADII['lg']}px;
                }}
                DropZoneWidget:hover {{
                    border-color: {border};
                    background-color: {hover_bg};
                }}
            """)


# ---------------------------------------------------------------------------
# Background worker — parallel multi-browser
# ---------------------------------------------------------------------------

class SpecCheckerWorker(QThread):
    row_update = Signal(int, str, str)   # row_index, status, notes
    log = Signal(str)
    done = Signal(int, int, int)         # has_specs, no_specs, errors
    progress_update = Signal(int, int, float, float)  # current, total, speed, eta

    def __init__(self, main_window, codes: list[str], browser_count: int = 1):
        super().__init__()
        self.main = main_window
        self.codes = codes
        self.browser_count = browser_count
        self._stop = False

        # Thread-safe shared state
        self._lock = threading.Lock()
        self._work_index = 0
        self._has_specs = 0
        self._no_specs = 0
        self._errors = 0
        self._processed_count = 0
        self._processed_times: list[float] = []
        self._seen: dict[str, tuple[str, str]] = {}  # code_upper -> (status, notes)

        self._session_manager = None

    def request_stop(self):
        self._stop = True

    # -- helpers -------------------------------------------------------------

    @staticmethod
    def _ensure_products_page(driver) -> bool:
        """Navigate read-only; never discard an already dirty product form."""
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait

        if "/dashboard/products/" in (driver.current_url or ""):
            if PimboProductEditor(driver).is_dirty():
                raise RuntimeError(
                    "Atidaryta forma turi neišsaugotų pakeitimų; tikrinimas jos neatmes."
                )
        try:
            products_link = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.NAV_PRODUCTS)
            )
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", products_link)
            products_link.click()

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
            )
        except Exception:
            try:
                driver.get(ProductListSelectors.URL)
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(ProductListSelectors.PRODUCT_TABLE)
                )
            except Exception:
                pass
        return False

    def _login_session(self, driver) -> bool:
        """Log a bare browser session into pim.bo."""
        from Config.LoginConfig.LoginHandler import LoginHandler

        email, password = self.main.credential_manager.get_saved_credentials()
        if not email or not password:
            return False

        login_handler = LoginHandler(driver, self.main.credential_manager, logger=self.main.logger)
        return login_handler.attempt_login(email, password)

    # -- progress ------------------------------------------------------------

    def _emit_progress(self):
        with self._lock:
            current = self._processed_count
            times = list(self._processed_times)
        total = len(self.codes)
        if not times:
            return
        avg = sum(times) / len(times)
        speed = 1.0 / avg if avg > 0 else 0
        remaining = total - current
        eta = remaining * avg
        self.progress_update.emit(current, total, speed, eta)

    # -- single-item check (called from each worker thread) ------------------

    def _check_one(self, idx: int, code: str, driver, nav):
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.common.exceptions import TimeoutException

        tr = self.main.i18n.tr
        total = len(self.codes)
        code_upper = code.strip().upper()

        # Duplicate detection
        is_dup = False
        with self._lock:
            if code_upper in self._seen:
                prev_status, prev_notes = self._seen[code_upper]
                self._processed_count += 1
                is_dup = True
        if is_dup:
            self.row_update.emit(idx, tr("speccheck.status.duplicate"), prev_status)
            self.log.emit(f"[{idx + 1}/{total}] {code} — {tr('speccheck.status.duplicate')}")
            self._emit_progress()
            return

        t0 = time.perf_counter()
        self.row_update.emit(idx, tr("batchdesc.status.running"), "")
        self.log.emit(f"[{idx + 1}/{total}] {tr('speccheck.log.checking', code=code)}")

        notes_parts = []

        try:
            # 1. Go back to product list
            clicked_leave = self._ensure_products_page(driver)
            if clicked_leave:
                notes_parts.append(tr("speccheck.note.left_unsaved"))

            # 2. Navigate to product
            try:
                nav.navigate_to_product("", code)
            except ValueError:
                status = tr("speccheck.status.not_found")
                notes = " / ".join(notes_parts) if notes_parts else ""
                with self._lock:
                    self._seen[code_upper] = (status, notes)
                    self._errors += 1
                self.row_update.emit(idx, status, notes)
                self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")
                elapsed = time.perf_counter() - t0
                with self._lock:
                    self._processed_times.append(elapsed)
                    self._processed_count += 1
                self._emit_progress()
                return

            # 3. Open the current Specifications section semantically.
            try:
                editor = PimboProductEditor(driver, self.main.logger)
                editor.open_section("specifications")
            except (PimAutomationError, TimeoutException):
                status = tr("speccheck.status.no_specs_tab")
                notes = " / ".join(notes_parts) if notes_parts else ""
                with self._lock:
                    self._seen[code_upper] = (status, notes)
                    self._errors += 1
                self.row_update.emit(idx, status, notes)
                self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")
                elapsed = time.perf_counter() - t0
                with self._lock:
                    self._processed_times.append(elapsed)
                    self._processed_count += 1
                self._emit_progress()
                return

            # 4. Read existing Family-provided rows.  No templates are applied.
            if editor.specification_field_count() > 0:
                status = tr("speccheck.status.has_specs")
                with self._lock:
                    self._has_specs += 1
            else:
                status = tr("speccheck.status.no_specs")
                with self._lock:
                    self._no_specs += 1

            notes = " / ".join(notes_parts) if notes_parts else ""
            with self._lock:
                self._seen[code_upper] = (status, notes)
            self.row_update.emit(idx, status, notes)
            self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")

        except Exception as e:
            status = f"{tr('speccheck.status.error')}: {e}"
            notes = " / ".join(notes_parts) if notes_parts else ""
            with self._lock:
                self._seen[code_upper] = (status, notes)
                self._errors += 1
            self.row_update.emit(idx, status, notes)
            self.log.emit(f"[{idx + 1}/{total}] {code} — {status}")

        elapsed = time.perf_counter() - t0
        with self._lock:
            self._processed_times.append(elapsed)
            self._processed_count += 1
        self._emit_progress()

    # -- worker thread (one per browser) -------------------------------------

    def _worker_thread(self, driver):
        tr = self.main.i18n.tr
        nav = ProductNavigationHandler(driver, logger=self.main.logger)
        nav.set_retry_callback(lambda op: False)

        while not self._stop:
            with self._lock:
                if self._work_index >= len(self.codes):
                    break
                idx = self._work_index
                self._work_index += 1

            code = self.codes[idx]
            self._check_one(idx, code, driver, nav)

    # -- main entry ----------------------------------------------------------

    def run(self):
        tr = self.main.i18n.tr
        total = len(self.codes)

        main_driver = getattr(self.main, "driver", None)
        if main_driver is None:
            for i in range(total):
                self.row_update.emit(i, tr("speccheck.status.error"), tr("batchdesc.no_session"))
            self.done.emit(0, 0, total)
            return

        extra_drivers = []

        if self.browser_count > 1:
            # Create extra browser sessions and log them in
            from Managers.BrowserSessionManager import BrowserSessionManager

            pool_size = self.browser_count - 1  # main driver is one of them
            self.log.emit(tr("speccheck.log.init_browsers", count=self.browser_count))

            browser_type = self.main.settings.get('browser_choice', 'Chrome')
            self._session_manager = BrowserSessionManager(
                browser_type=browser_type,
                pool_size=pool_size,
                logger=self.main.logger,
            )

            if self._session_manager.initialize_pool():
                # Log in each extra session
                for session in self._session_manager.sessions:
                    if self._stop:
                        break
                    self.log.emit(f"Logging in browser {session.session_id + 2}...")
                    session.driver.get(PIMBO_LOGIN_URL)
                    if self._login_session(session.driver):
                        extra_drivers.append(session.driver)
                        self.log.emit(f"Browser {session.session_id + 2} logged in")
                    else:
                        self.log.emit(f"Browser {session.session_id + 2} login failed, skipping")
            else:
                self.log.emit("Failed to create browser pool, using single browser")

        # All drivers: main + extras
        all_drivers = [main_driver] + extra_drivers
        actual_count = len(all_drivers)
        self.log.emit(f"Starting spec check with {actual_count} browser(s), {total} codes")

        if actual_count == 1:
            # Single-threaded — run directly
            self._worker_thread(main_driver)
        else:
            # Multi-threaded
            threads = []
            for driver in all_drivers:
                t = threading.Thread(target=self._worker_thread, args=(driver,), daemon=True)
                threads.append(t)
                t.start()

            for t in threads:
                t.join()

        # Mark any remaining (if stopped early) as skipped
        if self._stop:
            for j in range(self._work_index, total):
                self.row_update.emit(j, tr("speccheck.status.skipped"), "")

        # Shutdown extra browsers
        if self._session_manager:
            self._session_manager.shutdown_all()
            self._session_manager = None

        self.done.emit(self._has_specs, self._no_specs, self._errors)


# ---------------------------------------------------------------------------
# Main screen
# ---------------------------------------------------------------------------

class SpecCheckerScreen(ResponsiveWidget):

    def __init__(self, main_window):
        super().__init__(main_window)
        self.main = main_window
        self.tr = main_window.i18n.tr
        self._worker: SpecCheckerWorker | None = None
        self._codes: list[str] = []
        self._file_path: str | None = None

        self._build_ui()
        self.retranslate_ui()

    # -- UI construction -----------------------------------------------------

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(*PAGE_MARGINS)
        root.setSpacing(PAGE_SPACING)

        self._scroll = ScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        root.addWidget(self._scroll)

        self._container = QWidget()
        self._layout = QVBoxLayout(self._container)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(CONTENT_SPACING)
        self._scroll.setWidget(self._container)

        apply_screen_theme(
            self,
            "SpecCheckerScreen",
            scroll=self._scroll,
            content=self._container,
        )

        # -- Header ----------------------------------------------------------
        header = QHBoxLayout()
        header.setSpacing(ICON_TEXT_GAP)
        icon = IconWidget(FluentIcon.CHECKBOX)
        icon.setFixedSize(SIZES['icon_lg'], SIZES['icon_lg'])
        header.addWidget(icon)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        self._title = TitleLabel("")
        self._subtitle = CaptionLabel("")
        self._subtitle.setStyleSheet(f"color: {COLORS['text_secondary']};")
        title_col.addWidget(self._title)
        title_col.addWidget(self._subtitle)
        header.addLayout(title_col)
        header.addStretch()
        self._layout.addLayout(header)

        # -- Toolbar card ----------------------------------------------------
        toolbar_card = CardWidget()
        toolbar_layout = QHBoxLayout(toolbar_card)
        toolbar_layout.setContentsMargins(*TOOLBAR_MARGINS)
        toolbar_layout.setSpacing(CARD_SPACING)

        self._browse_btn = PushButton(FluentIcon.FOLDER, "")
        self._browse_btn.clicked.connect(self._browse_file)

        self._template_btn = PushButton(FluentIcon.DOWNLOAD, "")
        self._template_btn.clicked.connect(self._download_template)

        self._file_label = CaptionLabel("")
        self._file_label.setStyleSheet(f"color: {COLORS['text_secondary']};")

        # Browser count selector
        self._browser_label = BodyLabel("")
        self._browser_combo = ComboBox()
        self._browser_combo.addItems(["1", "2", "3", "4"])
        self._browser_combo.setCurrentIndex(0)
        self._browser_combo.setFixedWidth(80)

        self._start_btn = PrimaryPushButton(FluentIcon.PLAY, "")
        self._start_btn.setEnabled(False)
        self._start_btn.clicked.connect(self._on_start_stop)

        self._export_btn = PushButton(FluentIcon.SAVE, "")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_results)

        toolbar_layout.addWidget(self._browse_btn)
        toolbar_layout.addWidget(self._template_btn)
        toolbar_layout.addWidget(self._file_label, 1)
        toolbar_layout.addWidget(self._browser_label)
        toolbar_layout.addWidget(self._browser_combo)
        toolbar_layout.addWidget(self._export_btn)
        toolbar_layout.addWidget(self._start_btn)
        self._layout.addWidget(toolbar_card)

        # -- Drop zone -------------------------------------------------------
        self._drop_zone = DropZoneWidget(self.tr)
        self._drop_zone.setMinimumHeight(120)
        self._drop_zone.file_dropped.connect(self._on_file_dropped)
        self._layout.addWidget(self._drop_zone)

        # -- Progress label --------------------------------------------------
        self._progress_label = CaptionLabel("")
        self._progress_label.setStyleSheet(f"color: {COLORS['text_secondary']};")
        self._layout.addWidget(self._progress_label)

        # -- Results table ---------------------------------------------------
        self._table = QTableWidget(0, 4)
        self._table.setHorizontalHeaderLabels(["#", "Code", "Status", "Notes"])
        self._table.horizontalHeader().setStretchLastSection(True)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self._table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self._table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self._table.verticalHeader().setVisible(False)
        self._table.setMinimumHeight(200)
        enable_table_copy(self._table)
        self._layout.addWidget(self._table, 1)

        self._update_table_theme()
        enforce_transparent_labels(self)

    # -- Table theme ---------------------------------------------------------

    def _update_table_theme(self):
        is_dark = isDarkTheme()
        tc = COMPONENT_COLORS['table']
        bg = tc['row_alt_bg_dark'] if is_dark else tc['row_alt_bg_light']
        alt_bg = tc['row_bg_dark'] if is_dark else tc['row_bg_light']
        border = tc['border_dark'] if is_dark else tc['border_light']
        header_bg = COLORS['lavender_grey'] if is_dark else COLORS['space_indigo']
        header_text = COLORS['space_indigo'] if is_dark else COLORS['text_white']
        text_color = COLORS['text_primary_dark'] if is_dark else COLORS['text_primary_light']

        from GUI_Qt.styles.theme_config import (
            get_scrollbar_handle_bg, get_scrollbar_handle_hover_bg, get_selection_bg,
        )

        self._table.setStyleSheet(f"""
            QTableWidget {{
                background-color: {bg};
                alternate-background-color: {alt_bg};
                border: 1px solid {border};
                border-radius: {RADII['md']}px;
                gridline-color: {border};
                selection-background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QTableWidget::viewport {{
                background-color: {bg};
                border-bottom-left-radius: {RADII['md']}px;
                border-bottom-right-radius: {RADII['md']}px;
            }}
            QAbstractScrollArea::corner {{
                background: transparent;
            }}
            QTableWidget::item {{
                padding: {PADDINGS['table_cell']};
                color: {text_color};
                border: none;
            }}
            QTableWidget::item:focus {{
                outline: none;
            }}
            QTableWidget:focus {{
                outline: none;
            }}
            QTableWidget::item:selected {{
                background-color: {get_selection_bg()};
                color: {text_color};
            }}
            QHeaderView::section {{
                background-color: {header_bg};
                color: {header_text};
                padding: {PADDINGS['table_header']};
                border: none;
                border-bottom: 1px solid {border};
                font-weight: 600;
                font-size: {FONTS['size_body_sm']};
                letter-spacing: 0.5px;
                text-transform: uppercase;
            }}
            QHeaderView {{
                background: transparent;
                border: none;
            }}
            QHeaderView::section {{
                border-right: 1px solid {border};
            }}
            QHeaderView::section:last {{
                border-right: none;
            }}
            QHeaderView::section:first {{
                border-top-left-radius: {RADII['md']}px;
            }}
            QHeaderView::section:last {{
                border-top-right-radius: {RADII['md']}px;
            }}
            QScrollBar:vertical {{
                background: transparent;
                width: {SIZES['scrollbar_thickness']}px;
                margin: {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px {THEME_SPACING['xxs']}px 0px;
                border: none;
            }}
            QScrollBar::handle:vertical {{
                background: {get_scrollbar_handle_bg(is_dark)};
                border-radius: {RADII['sm']}px;
                min-height: {SIZES['scrollbar_handle_min']}px;
                margin: 0px {THEME_SPACING['xxs']}px;
                border: none;
            }}
            QScrollBar::handle:vertical:hover {{
                background: {get_scrollbar_handle_hover_bg(is_dark)};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
                border: none;
            }}
            QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{
                background: none;
            }}
        """)

    # -- File handling -------------------------------------------------------

    def _on_file_dropped(self, path: str):
        if path == "__browse__":
            self._browse_file()
            return
        self._load_file(path)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, self.tr("batch.file.select_excel.title"),
            "", "Excel files (*.xlsx);;All files (*.*)"
        )
        if path:
            self._load_file(path)

    def _load_file(self, path: str):
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            codes = []
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                val = row[0]
                if val is not None:
                    code = str(val).strip()
                    if code:
                        if not code.upper().startswith("UB-"):
                            code = f"UB-{code}"
                        codes.append(code)
            wb.close()

            if not codes:
                InfoBar.warning(
                    self.tr("common.error"),
                    self.tr("speccheck.no_codes"),
                    parent=self,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                )
                return

            self._codes = codes
            self._file_path = path
            self._file_label.setText(os.path.basename(path))

            # Populate table
            self._table.setRowCount(len(codes))
            for i, code in enumerate(codes):
                self._table.setItem(i, 0, QTableWidgetItem(str(i + 1)))
                self._table.setItem(i, 1, QTableWidgetItem(code))
                item = QTableWidgetItem(self.tr("batchdesc.status.pending"))
                item.setForeground(QColor(COLORS['text_secondary']))
                self._table.setItem(i, 2, item)
                self._table.setItem(i, 3, QTableWidgetItem(""))

            self._start_btn.setEnabled(True)
            self._export_btn.setEnabled(False)
            self._progress_label.setText(f"{len(codes)} {self.tr('speccheck.codes_loaded')}")
            self._drop_zone.setVisible(False)

        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("batch.template.save.title"),
            "spec_checker_template.xlsx", "Excel files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Codes"
            ws["A1"] = "Code"
            ws["A1"].font = Font(bold=True)
            ws.column_dimensions["A"].width = 25
            wb.save(path)
            show_file_saved_bar(
                self,
                self.tr("common.success"),
                self.tr("batch.template.saved"),
                path,
            )
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    # -- Start / Stop --------------------------------------------------------

    def _on_start_stop(self):
        if self._worker and self._worker.isRunning():
            self._worker.request_stop()
            self._start_btn.setEnabled(False)
            return

        if not self._codes:
            return

        driver = getattr(self.main, "driver", None)
        if driver is None:
            InfoBar.warning(
                self.tr("common.error"),
                self.tr("batchdesc.no_session"),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=3000,
            )
            return

        # Reset table statuses
        for i in range(self._table.rowCount()):
            item = QTableWidgetItem(self.tr("batchdesc.status.pending"))
            item.setForeground(QColor(COLORS['text_secondary']))
            self._table.setItem(i, 2, item)
            self._table.setItem(i, 3, QTableWidgetItem(""))

        self._start_btn.setText(self.tr("speccheck.stop"))
        self._start_btn.setIcon(FluentIcon.CLOSE)
        self._browse_btn.setEnabled(False)
        self._export_btn.setEnabled(False)
        self._browser_combo.setEnabled(False)

        browser_count = int(self._browser_combo.currentText())

        self._worker = SpecCheckerWorker(self.main, self._codes, browser_count)
        self._worker.row_update.connect(self._on_row_update)
        self._worker.progress_update.connect(self._on_progress)
        self._worker.done.connect(self._on_done)
        self._worker.log.connect(lambda msg: print(f"[SpecChecker] {msg}"))
        self._worker.start()

    def _on_row_update(self, row: int, status: str, notes: str):
        tr = self.tr
        item = QTableWidgetItem(status)

        if status == tr("speccheck.status.has_specs"):
            item.setForeground(QColor("#4CAF50"))
        elif status == tr("speccheck.status.duplicate"):
            item.setForeground(QColor("#FF9800"))
        elif status == tr("batchdesc.status.running"):
            item.setForeground(QColor(COLORS.get('lavender_grey', '#9E9E9E')))
        elif status == tr("batchdesc.status.pending"):
            pass
        else:
            # No specs, not found, no specs tab, unexpected errors — all red
            item.setForeground(QColor("#F44336"))

        self._table.setItem(row, 2, item)

        # Notes column
        if notes:
            notes_item = QTableWidgetItem(notes)
            notes_item.setForeground(QColor("#FF9800"))
            self._table.setItem(row, 3, notes_item)

        self._table.scrollToItem(item)

    def _on_progress(self, current: int, total: int, speed: float, eta: float):
        mins = int(eta // 60)
        secs = int(eta % 60)
        self._progress_label.setText(
            f"{current}/{total} — {speed:.1f} items/s — ETA: {mins}m {secs}s"
        )

    def _on_done(self, has_specs: int, no_specs: int, errors: int):
        self._start_btn.setText(self.tr("speccheck.start"))
        self._start_btn.setIcon(FluentIcon.PLAY)
        self._start_btn.setEnabled(True)
        self._browse_btn.setEnabled(True)
        self._export_btn.setEnabled(True)
        self._browser_combo.setEnabled(True)

        total = has_specs + no_specs + errors
        self._progress_label.setText(
            self.tr("speccheck.done", has=has_specs, no=no_specs, errors=errors, total=total)
        )

        InfoBar.success(
            self.tr("speccheck.done_title"),
            self.tr("speccheck.done", has=has_specs, no=no_specs, errors=errors, total=total),
            parent=self,
            position=InfoBarPosition.TOP,
            duration=5000,
        )

    # -- Export --------------------------------------------------------------

    def _export_results(self):
        if self._table.rowCount() == 0:
            return

        path, _ = QFileDialog.getSaveFileName(
            self, self.tr("speccheck.export"),
            f"spec_check_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx)"
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"

            headers = ["#", "Code", "Status", "Notes"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            for row in range(self._table.rowCount()):
                for col in range(4):
                    item = self._table.item(row, col)
                    ws.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 20
            ws.column_dimensions["D"].width = 30

            wb.save(path)
            show_file_saved_bar(
                self,
                self.tr("common.success"),
                self.tr("batch.template.saved"),
                path,
            )
        except Exception as e:
            InfoBar.error(
                self.tr("common.error"),
                str(e),
                parent=self,
                position=InfoBarPosition.TOP,
                duration=5000,
            )

    # -- i18n ----------------------------------------------------------------

    def retranslate_ui(self):
        self.tr = self.main.i18n.tr
        self._title.setText(self.tr("speccheck.title"))
        self._subtitle.setText(self.tr("speccheck.subtitle"))
        self._browse_btn.setText(self.tr("batch.browse_excel"))
        self._template_btn.setText(self.tr("batch.download_template"))
        self._file_label.setText(self.tr("batch.no_file") if not self._file_path else os.path.basename(self._file_path))
        self._export_btn.setText(self.tr("speccheck.export"))
        self._browser_label.setText(self.tr("speccheck.browsers"))

        if self._worker and self._worker.isRunning():
            self._start_btn.setText(self.tr("speccheck.stop"))
        else:
            self._start_btn.setText(self.tr("speccheck.start"))

        self._table.setHorizontalHeaderLabels([
            self.tr("speccheck.col.index"),
            self.tr("speccheck.col.code"),
            self.tr("speccheck.col.status"),
            self.tr("speccheck.col.notes"),
        ])

        self._drop_zone.tr = self.tr
        self._drop_zone.retranslate_ui()
