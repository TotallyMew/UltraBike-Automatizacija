"""
Batch Upload Dialog
Modal dialog for uploading multiple products via manual entry or Excel import
"""

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QWidget, QScrollArea,
    QPushButton, QTabWidget, QFileDialog, QLabel, QSizePolicy
)
from PySide6.QtCore import Qt, Signal
from qfluentwidgets import (
    LineEdit, ComboBox, CheckBox, PrimaryPushButton, PushButton,
    BodyLabel, StrongBodyLabel, TransparentToolButton, FluentIcon,
    MessageBox, InfoBar, InfoBarPosition, isDarkTheme, qconfig
)
import openpyxl
from openpyxl.styles import Font, PatternFill
from Managers.DescriptionManager import DescriptionManager

from GUI_Qt.styles.theme_config import get_text_color, get_status_text_color, get_status_row_style, FONTS, SPACING, SIZES
from GUI_Qt.styles.screen_theme import CARD_SPACING, CONTENT_SPACING, ROW_SPACING, MICRO_SPACING
from GUI_Qt.widgets import show_file_saved_bar


class BatchRowWidget(QWidget):
    """Single row for batch upload entry"""
    deleted = Signal(object)  # Emits self when deleted
    changed = Signal()  # Emits when any value changes

    @staticmethod
    def _normalize_code(raw: str) -> str:
        """Normalize product code to always include UB- prefix.

        Mirrors BatchTitles behavior.
        """
        code = (raw or "").strip()
        if not code:
            return ""
        if code.upper().startswith("UB-"):
            return code
        return f"UB-{code}"

    def __init__(self, brands, descriptions, tr, parent=None):
        super().__init__(parent)
        self.brands = brands
        self.descriptions = descriptions
        self._tr = tr
        self._init_ui()

    def retranslate_ui(self):
        self.brand_combo.setPlaceholderText(self._tr("batch.table.brand"))
        self.code_field.setPlaceholderText(self._tr("upload.code.placeholder"))
        self.url_field.setPlaceholderText(self._tr("upload.url.placeholder"))
        self.desc_combo.setPlaceholderText(self._tr("batch.optional"))
        self.frameset_check.setToolTip(self._tr("upload.frameset.tip"))
        self.disclaimer_check.setToolTip(self._tr("upload.disclaimer.tip"))
        self._delete_btn.setToolTip(self._tr("batch.row.remove.tip"))

    def _init_ui(self):
        layout = QHBoxLayout(self)
        layout.setSpacing(ROW_SPACING)
        layout.setContentsMargins(0, 0, 0, 0)

        # Brand dropdown
        self.brand_combo = ComboBox()
        self.brand_combo.addItems(self.brands)
        self.brand_combo.setMinimumWidth(SIZES['field_min_width_sm'])
        self.brand_combo.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.brand_combo.currentTextChanged.connect(self._on_brand_change)
        self.brand_combo.currentTextChanged.connect(self.changed.emit)

        # Product code
        self.code_field = LineEdit()
        self.code_field.setMinimumWidth(SIZES['field_min_width_sm'])
        self.code_field.setSizePolicy(QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        self.code_field.textChanged.connect(self.changed.emit)

        # URL/Code
        self.url_field = LineEdit()
        self.url_field.setMinimumWidth(SIZES['field_min_width_lg'])
        self.url_field.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.url_field.textChanged.connect(self.changed.emit)

        # Description
        self.desc_combo = ComboBox()
        self.desc_combo.addItem("")  # Empty option
        self.desc_combo.addItems(self.descriptions)
        self.desc_combo.setMinimumWidth(SIZES['field_min_width_md'])
        self.desc_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.desc_combo.currentTextChanged.connect(self.changed.emit)

        # Frameset checkbox (hidden by default)
        self.frameset_check = CheckBox()
        self.frameset_check.setVisible(False)
        self.frameset_check.setMinimumWidth(SIZES['check_col_width'])
        self.frameset_check.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.frameset_check.stateChanged.connect(self.changed.emit)

        # Disclaimer checkbox
        self.disclaimer_check = CheckBox()
        self.disclaimer_check.setMinimumWidth(SIZES['check_col_width_sm'])
        self.disclaimer_check.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        self.disclaimer_check.stateChanged.connect(self.changed.emit)

        # Delete button
        self._delete_btn = TransparentToolButton(FluentIcon.DELETE, self)
        self._delete_btn.clicked.connect(lambda: self.deleted.emit(self))
        self._delete_btn.setFixedWidth(SIZES['icon_lg'])

        layout.addWidget(self.brand_combo, 0)
        layout.addWidget(self.code_field, 0)
        layout.addWidget(self.url_field, 2)
        layout.addWidget(self.desc_combo, 1)
        layout.addWidget(self.frameset_check, 0)
        layout.addWidget(self.disclaimer_check, 0)
        layout.addWidget(self._delete_btn, 0)

        self.retranslate_ui()

    def _on_brand_change(self, brand):
        """Show/hide frameset checkbox based on brand"""
        is_pinarello = brand == "Pinarello"
        self.frameset_check.setVisible(is_pinarello)
        if not is_pinarello:
            self.frameset_check.setChecked(False)

    def is_valid(self):
        """Check if row has required data"""
        return bool(
            self.brand_combo.currentText() and
            self.code_field.text().strip() and
            self.url_field.text().strip()
        )

    def get_data(self):
        """Get row data as dict"""
        return {
            'brand': self.brand_combo.currentText(),
            'code': self._normalize_code(self.code_field.text()),
            'url': self.url_field.text().strip(),
            'description_name': self.desc_combo.currentText(),
            'frameset_only': self.frameset_check.isChecked(),
            'append_disclaimer': self.disclaimer_check.isChecked()
        }

    def set_data(self, data):
        """Set row data from dict"""
        if 'brand' in data:
            self.brand_combo.setCurrentText(data['brand'])
        if 'code' in data:
            self.code_field.setText(data['code'])
        if 'url' in data:
            self.url_field.setText(data['url'])
        if 'description_name' in data and data['description_name']:
            self.desc_combo.setCurrentText(data['description_name'])
        if 'frameset_only' in data:
            self.frameset_check.setChecked(data['frameset_only'])
        if 'append_disclaimer' in data:
            self.disclaimer_check.setChecked(data['append_disclaimer'])


class BatchUploadDialog(QDialog):
    """Batch upload dialog with manual and Excel entry"""
    batch_started = Signal(list)  # Emits list of items to process

    def __init__(self, main_window, parent=None):
        super().__init__(parent)
        self.main = main_window
        self.resize(1100, 650)

        self._tr = self.main.i18n.tr

        # Load descriptions
        desc_manager = DescriptionManager(main_window.db)
        self.descriptions = [d['name'] for d in desc_manager.list_descriptions()]

        # Brands list
        self.brands = [
            "KROSS", "Pinarello", "Basso", "Factor",
            "TREK", "Rondo", "Octane", "Rascal", "Lee Cougan"
        ]

        self._init_ui()

        try:
            qconfig.themeChangedFinished.connect(self._on_theme_changed)
        except Exception:
            pass

        if hasattr(self.main, "i18n") and hasattr(self.main.i18n, "languageChanged"):
            self.main.i18n.languageChanged.connect(self.retranslate_ui)

        self.retranslate_ui()

        self._apply_theme()

    def _on_theme_changed(self):
        self._apply_theme()

        # If Excel preview is visible, re-apply row/state styling.
        if hasattr(self, 'excel_preview_widget') and self.excel_preview_widget.isVisible():
            self._validate_excel()

    def _apply_theme(self):
        is_dark = isDarkTheme()

        if hasattr(self, 'excel_filename_label'):
            self.excel_filename_label.setStyleSheet(f"color: {get_text_color(is_dark, 'tertiary')};")

        if hasattr(self, '_empty_text2'):
            self._empty_text2.setStyleSheet(f"color: {get_text_color(is_dark, 'tertiary')};")

    def retranslate_ui(self):
        self.setWindowTitle(self._tr("batch.title"))
        self._title_label.setText(self._tr("batch.title"))

        self.tabs.setTabText(0, self._tr("batch.mode.manual"))
        self.tabs.setTabText(1, self._tr("batch.mode.excel"))

        self._manual_header_brand.setText(self._tr("batch.table.brand"))
        self._manual_header_code.setText(self._tr("batch.table.code"))
        self._manual_header_url.setText(self._tr("batch.table.url"))
        self._manual_header_desc.setText(self._tr("batch.table.desc"))

        self._frameset_label.setText(self._tr("batch.table.frameset"))
        self._frameset_label.setToolTip(self._tr("upload.frameset.tip"))
        self.frameset_select_all.setToolTip(self._tr("batch.bulk.toggle.tip"))

        self._disc_label.setText(self._tr("batch.table.disclaimer"))
        self._disc_label.setToolTip(self._tr("upload.disclaimer.tip"))
        self.disc_select_all.setToolTip(self._tr("batch.bulk.toggle.tip"))

        self._add_row_btn.setText(self._tr("batch.add_row"))
        self._clear_btn.setText(self._tr("batch.clear_all"))

        self._browse_btn.setText(self._tr("batch.browse_excel"))
        self._template_btn.setText(self._tr("batch.download_template"))
        self.excel_filename_label.setText(self._tr("batch.no_file"))

        self._empty_text1.setText(self._tr("batch.drop.title"))
        self._empty_text2.setText(self._tr("batch.drop.subtitle"))

        self._preview_title.setText(self._tr("batch.preview.title"))

        self._cancel_btn.setText(self._tr("common.cancel"))
        self.start_btn.setText(self._tr("batch.start"))

        # Retranslate existing row widgets
        for i in range(self.manual_rows_layout.count() - 1):
            widget = self.manual_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget):
                widget.retranslate_ui()
        for i in range(self.excel_rows_layout.count() - 1):
            widget = self.excel_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget):
                widget.retranslate_ui()

    def _init_ui(self):
        """Initialize UI"""
        layout = QVBoxLayout(self)
        layout.setSpacing(CARD_SPACING)

        # Title
        self._title_label = StrongBodyLabel("")
        self._title_label.setStyleSheet(f"font-size: {FONTS['size_subtitle_1']};")
        layout.addWidget(self._title_label)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.addTab(self._create_manual_tab(), "")
        self.tabs.addTab(self._create_excel_tab(), "")
        layout.addWidget(self.tabs)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        self._cancel_btn = PushButton("")
        self._cancel_btn.clicked.connect(self.reject)

        self.start_btn = PrimaryPushButton("")
        self.start_btn.setIcon(FluentIcon.PLAY)
        self.start_btn.setEnabled(False)
        self.start_btn.clicked.connect(self._handle_start)

        button_layout.addWidget(self._cancel_btn)
        button_layout.addWidget(self.start_btn)
        layout.addLayout(button_layout)

    def _create_manual_tab(self):
        """Create manual entry tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(CONTENT_SPACING)

        # Header row
        header = QWidget()
        header_layout = QHBoxLayout(header)
        header_layout.setSpacing(ROW_SPACING)
        header_layout.setContentsMargins(0, 0, 0, 0)

        self._manual_header_brand = BodyLabel("")
        self._manual_header_code = BodyLabel("")
        self._manual_header_url = BodyLabel("")
        self._manual_header_desc = BodyLabel("")
        header_layout.addWidget(self._manual_header_brand)
        header_layout.addWidget(self._manual_header_code, 0, Qt.AlignmentFlag.AlignLeft)
        header_layout.addWidget(self._manual_header_url, 0, Qt.AlignmentFlag.AlignLeft)
        header_layout.addWidget(self._manual_header_desc, 0, Qt.AlignmentFlag.AlignLeft)

        # Frameset select-all
        frameset_col = QWidget()
        frameset_col_layout = QVBoxLayout(frameset_col)
        frameset_col_layout.setSpacing(MICRO_SPACING)
        frameset_col_layout.setContentsMargins(0, 0, 0, 0)
        self._frameset_label = BodyLabel("")
        self.frameset_select_all = CheckBox()
        self.frameset_select_all.stateChanged.connect(self._toggle_all_framesets)
        frameset_col_layout.addWidget(self._frameset_label)
        frameset_col_layout.addWidget(self.frameset_select_all)
        frameset_col.setFixedWidth(SIZES['check_col_width'])
        header_layout.addWidget(frameset_col)

        # Disclaimer select-all
        disc_col = QWidget()
        disc_col_layout = QVBoxLayout(disc_col)
        disc_col_layout.setSpacing(MICRO_SPACING)
        disc_col_layout.setContentsMargins(0, 0, 0, 0)
        self._disc_label = BodyLabel("")
        self.disc_select_all = CheckBox()
        self.disc_select_all.stateChanged.connect(self._toggle_all_disclaimers)
        disc_col_layout.addWidget(self._disc_label)
        disc_col_layout.addWidget(self.disc_select_all)
        disc_col.setFixedWidth(SIZES['check_col_width_sm'])
        header_layout.addWidget(disc_col)

        header_layout.addSpacing(SPACING["xxl"])  # Space for delete button
        header_layout.addStretch()

        layout.addWidget(header)

        # Scroll area for rows
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        rows_container = QWidget()
        self.manual_rows_layout = QVBoxLayout(rows_container)
        self.manual_rows_layout.setSpacing(ROW_SPACING)
        self.manual_rows_layout.addStretch()

        scroll.setWidget(rows_container)
        layout.addWidget(scroll, 1)

        # Add initial rows
        for _ in range(10):
            self._add_manual_row()

        # Action buttons
        actions = QHBoxLayout()
        self._add_row_btn = PushButton("")
        self._add_row_btn.setIcon(FluentIcon.ADD)
        self._add_row_btn.clicked.connect(self._add_manual_row)

        self._clear_btn = PushButton("")
        self._clear_btn.setIcon(FluentIcon.DELETE)
        self._clear_btn.clicked.connect(self._clear_manual_rows)

        actions.addWidget(self._add_row_btn)
        actions.addWidget(self._clear_btn)
        actions.addStretch()
        layout.addLayout(actions)

        return widget

    def _create_excel_tab(self):
        """Create Excel upload tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setSpacing(CONTENT_SPACING)

        # File selection
        file_row = QHBoxLayout()
        self._browse_btn = PushButton("")
        self._browse_btn.setIcon(FluentIcon.FOLDER)
        self._browse_btn.clicked.connect(self._browse_excel)

        self._template_btn = PushButton("")
        self._template_btn.setIcon(FluentIcon.DOWNLOAD)
        self._template_btn.clicked.connect(self._download_template)

        file_row.addWidget(self._browse_btn)
        file_row.addWidget(self._template_btn)
        file_row.addStretch()
        layout.addLayout(file_row)

        self.excel_filename_label = BodyLabel("")
        self.excel_filename_label.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'tertiary')};")
        layout.addWidget(self.excel_filename_label)

        # Empty state
        self.excel_empty_state = QWidget()
        empty_layout = QVBoxLayout(self.excel_empty_state)
        empty_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_icon = BodyLabel("📤")
        empty_icon.setStyleSheet("font-size: 64px;")
        empty_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_text1 = BodyLabel("")
        self._empty_text1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_text2 = BodyLabel("")
        self._empty_text2.setStyleSheet(f"color: {get_text_color(isDarkTheme(), 'tertiary')};")
        self._empty_text2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        empty_layout.addWidget(empty_icon)
        empty_layout.addWidget(self._empty_text1)
        empty_layout.addWidget(self._empty_text2)
        layout.addWidget(self.excel_empty_state, 1)

        # Preview (hidden initially)
        self.excel_preview_widget = QWidget()
        self.excel_preview_widget.setVisible(False)
        preview_layout = QVBoxLayout(self.excel_preview_widget)
        preview_layout.setSpacing(ROW_SPACING)

        self._preview_title = StrongBodyLabel("")
        preview_layout.addWidget(self._preview_title)

        # Preview scroll
        preview_scroll = QScrollArea()
        preview_scroll.setWidgetResizable(True)
        preview_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        preview_container = QWidget()
        self.excel_rows_layout = QVBoxLayout(preview_container)
        self.excel_rows_layout.setSpacing(ROW_SPACING)
        self.excel_rows_layout.addStretch()

        preview_scroll.setWidget(preview_container)
        preview_layout.addWidget(preview_scroll, 1)

        self.excel_status_label = BodyLabel("")
        preview_layout.addWidget(self.excel_status_label)

        layout.addWidget(self.excel_preview_widget, 1)

        return widget

    def _add_manual_row(self):
        """Add a manual entry row"""
        row = BatchRowWidget(self.brands, self.descriptions, self._tr, self)
        row.deleted.connect(self._remove_row)
        row.changed.connect(self._validate_manual)
        self.manual_rows_layout.insertWidget(self.manual_rows_layout.count() - 1, row)
        self._validate_manual()

    def _remove_row(self, row):
        """Remove a row"""
        self.manual_rows_layout.removeWidget(row)
        row.deleteLater()
        self._validate_manual()

    def _clear_manual_rows(self):
        """Clear all manual rows"""
        while self.manual_rows_layout.count() > 1:  # Keep the stretch
            item = self.manual_rows_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Add 10 new rows
        for _ in range(10):
            self._add_manual_row()

    def _toggle_all_framesets(self, state):
        """Toggle all frameset checkboxes"""
        checked = state == Qt.CheckState.Checked.value
        for i in range(self.manual_rows_layout.count() - 1):
            widget = self.manual_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget):
                if widget.brand_combo.currentText() == "Pinarello":
                    widget.frameset_check.setChecked(checked)

    def _toggle_all_disclaimers(self, state):
        """Toggle all disclaimer checkboxes"""
        checked = state == Qt.CheckState.Checked.value
        for i in range(self.manual_rows_layout.count() - 1):
            widget = self.manual_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget):
                widget.disclaimer_check.setChecked(checked)

    def _validate_manual(self):
        """Validate manual entry rows"""
        valid_count = 0
        for i in range(self.manual_rows_layout.count() - 1):
            widget = self.manual_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget) and widget.is_valid():
                valid_count += 1

        # Only update button if it exists (avoid error during initialization)
        if hasattr(self, 'start_btn'):
            self.start_btn.setEnabled(valid_count > 0)

    def _browse_excel(self):
        """Browse for Excel file"""
        filename, _ = QFileDialog.getOpenFileName(
            self,
            self._tr("batch.file.select_excel.title"),
            "",
            self._tr("batch.file.select_excel.filter")
        )

        if filename:
            self.excel_filename_label.setText(filename)
            self._load_excel_preview(filename)

    def _load_excel_preview(self, filename):
        """Load and preview Excel file"""
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active

            # Clear existing preview
            while self.excel_rows_layout.count() > 1:
                item = self.excel_rows_layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

            # Find header row
            header_row = None
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5), start=1):
                values = [str(c.value).lower() if c.value else "" for c in row]
                if any("brand" in v for v in values):
                    header_row = idx
                    break

            if not header_row:
                self.excel_status_label.setText(self._tr("batch.excel.header_missing"))
                self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('warning')};")
                self.start_btn.setEnabled(False)
                wb.close()
                return

            valid_count = 0
            invalid_count = 0

            # Process rows
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(row):
                    continue

                brand = str(row[0]) if len(row) > 0 and row[0] else ""
                code = str(row[1]) if len(row) > 1 and row[1] else ""
                url = str(row[2]) if len(row) > 2 and row[2] else ""
                desc = str(row[3]).strip() if len(row) > 3 and row[3] and str(row[3]).strip() else ""

                # Parse frameset (column 4)
                frameset_raw = row[4] if len(row) > 4 and row[4] else False
                if isinstance(frameset_raw, str):
                    frameset = frameset_raw.strip().lower() in ("yes", "true", "1", "frameset")
                else:
                    frameset = bool(frameset_raw)

                # Parse disclaimer (column 5)
                disclaimer_raw = row[5] if len(row) > 5 and row[5] else False
                if isinstance(disclaimer_raw, str):
                    disclaimer = disclaimer_raw.strip().lower() in ("yes", "true", "1")
                else:
                    disclaimer = bool(disclaimer_raw)

                # Create row widget
                row_widget = BatchRowWidget(self.brands, self.descriptions, self._tr, self)
                row_widget.set_data({
                    'brand': brand,
                    'code': code,
                    'url': url,
                    'description_name': desc,
                    'frameset_only': frameset,
                    'append_disclaimer': disclaimer
                })
                row_widget.changed.connect(self._validate_excel)

                if row_widget.is_valid():
                    valid_count += 1
                    row_widget.setStyleSheet(get_status_row_style(isDarkTheme(), 'valid'))
                else:
                    invalid_count += 1
                    row_widget.setStyleSheet(get_status_row_style(isDarkTheme(), 'invalid'))

                self.excel_rows_layout.insertWidget(self.excel_rows_layout.count() - 1, row_widget)

            # Update status
            if invalid_count > 0:
                self.excel_status_label.setText(self._tr("batch.excel.rows_valid_invalid", valid=valid_count, invalid=invalid_count))
                self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('warning')};")
                self.start_btn.setEnabled(False)
            else:
                self.excel_status_label.setText(self._tr("batch.excel.rows_valid_ready", valid=valid_count))
                self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('success')};")
                self.start_btn.setEnabled(valid_count > 0)

            # Show preview, hide empty state
            self.excel_empty_state.setVisible(False)
            self.excel_preview_widget.setVisible(True)

            wb.close()

        except Exception as ex:
            self.excel_status_label.setText(self._tr("batch.excel.error", error=str(ex)))
            self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('error')};")
            self.start_btn.setEnabled(False)

    def _validate_excel(self):
        """Validate Excel preview rows"""
        valid_count = 0
        invalid_count = 0

        for i in range(self.excel_rows_layout.count() - 1):
            widget = self.excel_rows_layout.itemAt(i).widget()
            if isinstance(widget, BatchRowWidget):
                if widget.is_valid():
                    valid_count += 1
                    widget.setStyleSheet(get_status_row_style(isDarkTheme(), 'valid'))
                else:
                    invalid_count += 1
                    widget.setStyleSheet(get_status_row_style(isDarkTheme(), 'invalid'))

        # Only update if widgets exist
        if hasattr(self, 'excel_status_label'):
            if invalid_count > 0:
                self.excel_status_label.setText(
                    self._tr("batch.excel.rows_valid_invalid", valid=valid_count, invalid=invalid_count)
                )
                self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('warning')};")
            else:
                self.excel_status_label.setText(self._tr("batch.excel.rows_valid_ready", valid=valid_count))
                self.excel_status_label.setStyleSheet(f"color: {get_status_text_color('success')};")

        if hasattr(self, 'start_btn'):
            if invalid_count > 0:
                self.start_btn.setEnabled(False)
            else:
                self.start_btn.setEnabled(valid_count > 0)

    def _download_template(self):
        """Download Excel template"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            self._tr("batch.template.save_as.title"),
            self._tr("batch.template.default_name"),
            self._tr("batch.template.filter")
        )

        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = self._tr("batch.title")

            # Header
            ws.append(["Brand", "Product Code", "URL or Code", "Description", "Frameset", "Append Disclaimer"])

            # Examples
            ws.append(["KROSS", "UB-1234", "https://kross.pl/product1", "Mountain Bike", "No", "Yes"])
            ws.append(["Pinarello", "UB-5678", "https://pinarello.com/product2", "Road Bike", "Yes", "No"])

            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

            wb.save(filename)
            wb.close()

            show_file_saved_bar(
                self,
                self._tr("common.success"),
                self._tr("batch.template.saved"),
                filename,
            )

        except Exception as ex:
            InfoBar.error(
                title=self._tr("common.error"),
                content=self._tr("batch.template.save_failed", error=str(ex)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self
            )

    def _handle_start(self):
        """Handle start batch upload"""
        items = []

        if self.tabs.currentIndex() == 0:  # Manual entry
            for i in range(self.manual_rows_layout.count() - 1):
                widget = self.manual_rows_layout.itemAt(i).widget()
                if isinstance(widget, BatchRowWidget) and widget.is_valid():
                    items.append(widget.get_data())
        else:  # Excel
            for i in range(self.excel_rows_layout.count() - 1):
                widget = self.excel_rows_layout.itemAt(i).widget()
                if isinstance(widget, BatchRowWidget) and widget.is_valid():
                    items.append(widget.get_data())

        if items:
            self.batch_started.emit(items)
            self.accept()
