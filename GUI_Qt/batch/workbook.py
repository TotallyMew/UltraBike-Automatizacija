"""GUI_Qt/screens/UnifiedBatchScreen.py

Unified batch operations screen — replaces BatchUploadScreen,
BatchDescriptionsScreen and BatchTitlesScreen with a single
customisable table whose columns adapt to the selected operation.
"""

from __future__ import annotations

import json
import os
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from PySide6.QtCore import QEvent, QTimer, Qt, Signal
from PySide6.QtGui import QColor, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from qfluentwidgets import (
    BodyLabel,
    CaptionLabel,
    CardWidget,
    CheckBox,
    ComboBox,
    FluentIcon,
    Flyout,
    FlyoutViewBase,
    IconWidget,
    IndeterminateProgressRing,
    InfoBar,
    InfoBarPosition,
    LineEdit,
    MessageBox,
    PillPushButton,
    PrimaryPushButton,
    PushButton,
    StrongBodyLabel,
    TitleLabel,
    TransparentPushButton,
    TransparentToolButton,
    isDarkTheme,
    qconfig,
)

from GUI_Qt.widgets.ResponsiveWidget import ResponsiveWidget
from GUI_Qt.widgets.DropZoneWidget import DropZoneWidget
from GUI_Qt.widgets.DragFillHandle import DragFillHandle
from GUI_Qt.widgets.TieredHeaderWidget import TieredHeaderWidget
from GUI_Qt.widgets.FrozenColumnTableWidget import FrozenColumnOverlay
from GUI_Qt.widgets.FilterableComboBox import FilterableComboBox
from GUI_Qt.widgets import enable_table_copy
from GUI_Qt.components.dialogs import DestructiveActionDialog
from GUI_Qt.screens.batch_inspector import BatchInspectorPanel
from GUI_Qt.styles.theme_config import (
    COLORS,
    COMPONENT_COLORS,
    FONTS,
    PADDINGS,
    RADII,
    SIZES,
    SPACING,
)
from GUI_Qt.styles.screen_theme import (
    PAGE_MARGINS,
    PAGE_SPACING,
    ICON_TEXT_GAP,
    ROW_SPACING,
    TOOLBAR_MARGINS,
    CONTENT_SPACING,
    TABLE_CELL_MARGINS,
    apply_screen_theme,
    enforce_transparent_labels,
    get_responsive_margins,
    get_responsive_spacing,
)
from GUI_Qt.screens.batch_columns import (
    ATTR_COL_MAP,
    COL,
    COLUMN_GROUPS,
    COLUMNS,
    NUM_COLUMNS,
    PRESETS,
)
from GUI_Qt.screens.batch_strategies import STRATEGIES
from GUI_Qt.screens.UploadScreen import ATTRIBUTE_DEFINITIONS
from GUI_Qt.dialogs.AttributeOptionsDialog import (
    AttributeOptionsDialog,
    load_attribute_options,
)
from Managers.DescriptionManager import DescriptionManager
from Utilities.ExcelHandler import ExcelHandler

_SETTINGS_KEY_COLUMNS = "batch_visible_columns"

# Tuple of combo-like widget types for isinstance checks
_COMBO_TYPES = (ComboBox, FilterableComboBox)


# ---------------------------------------------------------------------------
# Cell focus filter — syncs table selection when a cell widget gets focus
# ---------------------------------------------------------------------------

class BatchWorkbookIO:
    def _handle_file_drop(self, path: str):
        if path == "__browse__":
            self._browse_excel()
        else:
            self._load_excel(path)

    def _browse_excel(self):
        filename, _ = QFileDialog.getOpenFileName(
            self,
            self.main.i18n.tr("batch.file.select_excel.title"),
            "",
            self.main.i18n.tr("batch.file.select_excel.filter"),
        )
        if filename:
            self._load_excel(filename)

    def _load_excel(self, filename: str):
        tr = self.main.i18n.tr
        strategy = self._infer_strategy()
        try:
            wb = openpyxl.load_workbook(filename, read_only=True)
            ws = wb.active

            # Find header row
            header_row = None
            for row_idx, row_data in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
                if row_data and any(
                    str(c).strip().lower() == "brand"
                    for c in row_data
                    if c is not None
                ):
                    header_row = row_idx
                    break

            if header_row is None:
                raise Exception(tr("batch.excel.header_missing_in_file"))

            data_rows = []
            for row_data in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if any(c is not None and str(c).strip() for c in row_data):
                    data_rows.append(row_data)

            wb.close()

            if not data_rows:
                raise Exception(tr("batch.excel.no_data"))

            self.table.setRowCount(len(data_rows))
            for table_row, excel_row in enumerate(data_rows):
                self._setup_table_row(table_row)
                parsed = strategy.parse_excel_row(excel_row, tr)
                self._populate_row_from_dict(table_row, parsed)

            self.excel_file_label.setText(os.path.basename(filename))
            self.excel_file_label.setToolTip(filename)
            self.table_card.show()
            self.drop_zone.hide()
            self._validate()

            InfoBar.success(
                title=tr("common.success"),
                content=tr("batch.excel.loaded", count=len(data_rows)),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self,
            )
        except Exception as ex:
            InfoBar.error(
                title=tr("common.error"),
                content=str(ex),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=5000,
                parent=self,
            )

    def _populate_row_from_dict(self, row: int, data: dict):
        """Fill row widgets from a parsed Excel dict."""
        brand = data.get("brand", "")
        brand_w = self._get_inner_widget(row, COL["brand"])
        if brand_w and isinstance(brand_w, _COMBO_TYPES) and brand:
            idx = brand_w.findText(brand)
            if idx >= 0:
                brand_w.setCurrentIndex(idx)

        code = data.get("code", "")
        code_w = self._get_inner_widget(row, COL["code"])
        if code_w and isinstance(code_w, LineEdit):
            code_w.setText(code)

        url = data.get("url", "")
        url_w = self._get_inner_widget(row, COL["url"])
        if url_w and isinstance(url_w, LineEdit) and url:
            url_w.setText(url)

        desc = data.get("description", "")
        desc_w = self._get_inner_widget(row, COL["description"])
        if desc_w and isinstance(desc_w, _COMBO_TYPES) and desc:
            idx = desc_w.findText(desc)
            if idx >= 0:
                desc_w.setCurrentIndex(idx)

        fs = data.get("frameset", False)
        fs_w = self._get_inner_widget(row, COL["frameset"])
        if fs_w and isinstance(fs_w, CheckBox):
            fs_w.setChecked(bool(fs))

        disc = data.get("disclaimer", False)
        disc_w = self._get_inner_widget(row, COL["disclaimer"])
        if disc_w and isinstance(disc_w, CheckBox):
            disc_w.setChecked(bool(disc))

        # Attributes
        for defn in ATTRIBUTE_DEFINITIONS:
            val = data.get(f"attr_{defn['name']}", "")
            if not val:
                continue
            col_idx = ATTR_COL_MAP.get(defn["name"])
            if col_idx is None:
                continue
            combo = self._get_inner_widget(row, col_idx)
            if combo and isinstance(combo, _COMBO_TYPES):
                idx = combo.findText(val)
                if idx >= 0:
                    combo.setCurrentIndex(idx)

        # Titles
        for key in ("title_lt", "title_en"):
            val = data.get(key, "")
            if not val:
                continue
            w = self._get_inner_widget(row, COL[key])
            if w and isinstance(w, LineEdit):
                w.setText(val)

    def _download_template(self):
        tr = self.main.i18n.tr
        strategy = self._infer_strategy()
        headers = strategy.get_template_headers(tr)
        example = strategy.get_template_example(tr)

        default_name = f"batch_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(
            self, tr("batch.template.save_title"), default_name, "Excel (*.xlsx)"
        )
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Template"

            # Header styling
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2B2D42", end_color="2B2D42", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Example row
            for col_idx, val in enumerate(example, start=1):
                ws.cell(row=2, column=col_idx, value=val)

            # Auto-width
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

            wb.save(filename)
            InfoBar.success(
                title=tr("common.success"),
                content=tr("batch.template.saved"),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=3000,
                parent=self,
            )
        except Exception as ex:
            InfoBar.error(
                title=tr("common.error"),
                content=str(ex),
                orient=Qt.Orientation.Horizontal,
                isClosable=True,
                position=InfoBarPosition.TOP,
                duration=5000,
                parent=self,
            )

    def _export_table_data(self):
        tr = self.main.i18n.tr
        strategy = self._infer_strategy()
        headers = strategy.get_template_headers(tr) + [tr("batch.table.status"), tr("batch.table.error")]

        rows_data = []
        for row in range(self.table.rowCount()):
            brand_w = self._get_inner_widget(row, COL["brand"])
            if not brand_w or not isinstance(brand_w, _COMBO_TYPES) or not brand_w.currentText():
                continue
            row_vals = []
            # Use strategy headers to determine which columns to export
            # Simpler: export all non-hidden, non-locked columns + status + error
            for i, cdef in enumerate(COLUMNS):
                if self.table.isColumnHidden(i) and not cdef.lockable:
                    continue
                if cdef.key in ("delete",):
                    continue
                w = self._get_inner_widget(row, i)
                if w is None:
                    row_vals.append("")
                elif isinstance(w, _COMBO_TYPES):
                    row_vals.append(w.currentText())
                elif isinstance(w, LineEdit):
                    row_vals.append(w.text())
                elif isinstance(w, CheckBox):
                    row_vals.append("Yes" if w.isChecked() else "No")
                elif isinstance(w, BodyLabel):
                    row_vals.append(w.text())
                else:
                    row_vals.append("")
            rows_data.append(row_vals)

        if not rows_data:
            return

        default_name = f"batch_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename, _ = QFileDialog.getSaveFileName(
            self, tr("batch.export.title"), default_name, "Excel (*.xlsx)"
        )
        if filename:
            try:
                # Build actual headers from visible columns
                vis_headers = []
                for i, cdef in enumerate(COLUMNS):
                    if self.table.isColumnHidden(i) and not cdef.lockable:
                        continue
                    if cdef.key == "delete":
                        continue
                    text = tr(cdef.header_i18n) if "." in cdef.header_i18n else cdef.header_i18n
                    vis_headers.append(text or cdef.key)
                ExcelHandler.export_table_data(vis_headers, rows_data, filename, "Batch Data")
                InfoBar.success(
                    title=tr("common.success"),
                    content=tr("batch.export.saved"),
                    orient=Qt.Orientation.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=3000,
                    parent=self,
                )
            except Exception as ex:
                InfoBar.error(
                    title=tr("common.error"),
                    content=str(ex),
                    orient=Qt.Orientation.Horizontal,
                    isClosable=True,
                    position=InfoBarPosition.TOP,
                    duration=5000,
                    parent=self,
                )

    # --------------------------------------------------- Batch start / stop
