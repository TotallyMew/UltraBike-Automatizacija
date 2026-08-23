from __future__ import annotations

import os
from datetime import datetime, timezone


os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication, QWidget
from openpyxl import load_workbook

from Database.DatabaseManager import DatabaseManager
from Database.SettingsManager import SettingsManager
from GUI_Qt.i18n import translate
from GUI_Qt.screens.AnalyticsScreen import AnalyticsScreen
from Managers.EarningsManager import EarningsManager


UTC = timezone.utc


class _I18n:
    @staticmethod
    def tr(key, **values):
        return translate("en", key, **values)


class _Main(QWidget):
    def __init__(self, database):
        super().__init__()
        self.db = database
        self.i18n = _I18n()
        self.opened = []

    def open_route(self, key):
        self.opened.append(key)


def test_dashboard_includes_manual_and_app_linked_earnings(tmp_path, monkeypatch):
    app = QApplication.instance() or QApplication([])
    db = DatabaseManager(":memory:")
    try:
        service = EarningsManager(db, SettingsManager(db), local_tz=UTC)
        basso = next(brand for brand in service.list_brands() if brand["name"] == "Basso")
        history = db.conn.execute(
            """
            INSERT INTO processing_history (brand, product_code, status, processed_at)
            VALUES ('Basso', 'APP', 'success', CURRENT_TIMESTAMP)
            """
        )
        db.conn.commit()
        now = datetime.now(UTC)
        service.create_entry("OUTSIDE", "bicycle", brand_id=basso["id"], now=now)
        service.create_entry(
            "APP", "bicycle", brand_id=basso["id"], source="regular_upload",
            processing_history_id=int(history.lastrowid), now=now,
        )

        main = _Main(db)
        screen = AnalyticsScreen(main)
        screen.show()
        app.processEvents()

        assert screen.products_metric.value_label.text() == "2"
        assert screen.earnings_metric.value_label.text() == "€2.00"
        assert screen.manual_products_metric.value_label.text() == "1"
        assert screen.app_products_metric.value_label.text() == "1"
        assert screen.brand_chart.labels == ["Basso"]
        assert set(screen.source_chart.labels) == {"Outside app", "Regular app"}
        assert screen.success_donut.percentage == 100

        screen._view_earnings()
        assert main.opened == ["earnings"]

        export_path = tmp_path / "combined-analytics.xlsx"
        monkeypatch.setattr(
            "GUI_Qt.screens.AnalyticsScreen.QFileDialog.getSaveFileName",
            lambda *_args, **_kwargs: (str(export_path), "Excel Files (*.xlsx)"),
        )
        screen._export_analytics()
        workbook = load_workbook(export_path, data_only=True)
        assert workbook.sheetnames == [
            "Summary", "Brands", "Sources", "Product types", "App issues"
        ]
        assert workbook["Summary"]["B3"].value == 2
        assert workbook["Summary"]["B4"].value == 2

        screen.close()
        screen.deleteLater()
        main.deleteLater()
        app.processEvents()
    finally:
        db.close()
