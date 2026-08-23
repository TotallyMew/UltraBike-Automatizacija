from __future__ import annotations

import csv
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from tools.orbea_automation import (
    CancellationToken,
    CatalogueEntry,
    CatalogueIndex,
    OrbeaAutomationService,
    OrbeaRunConfig,
    PimboFilterSpec,
    RunCheckpoint,
    create_run_directory,
    extract_template_codes,
    find_latest_compatible_run,
    select_representative_variant,
)
from tools.orbea_automation.report import (
    sort_existing_match_workbook,
    write_image_manifest,
    write_report,
)
from tools.orbea_automation.pimbo import PimboBrowserClient


CATALOGUE_HEADERS = (
    "Model",
    "Year",
    "Category",
    "Subcategory",
    "Regional Listings",
    "Product Link",
    "Other Links",
    "Identifiers",
    "Unique Model ID",
)


def build_catalogue(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Unique Models"
    worksheet.append(CATALOGUE_HEADERS)
    worksheet.append(
        (
            "ORCA M30i",
            2026,
            "Road",
            "Performance",
            8,
            "https://cms.orbea.com/en-be/orca-m30i",
            "",
            "11486 • U107TTCC",
            "orca-m30i-2026",
        )
    )
    worksheet.append(
        (
            "ORCA M20i",
            2026,
            "Road",
            "Performance",
            5,
            "https://cms.orbea.com/en-be/orca-m20i",
            "",
            "11487 • U107TTCC",
            "orca-m20i-2026",
        )
    )
    worksheet.append(
        (
            "RISE LT H10",
            2026,
            "MTB",
            "Electric",
            4,
            "https://cms.orbea.com/en-be/rise-lt-h10",
            "",
            "T123TTCC",
            "rise-lt-h10-2026",
        )
    )
    workbook.save(path)
    workbook.close()


def result_row(
    row_key: str,
    status: str,
    *,
    sku: str = "",
    title: str = "",
    catalogue_url: str = "",
    catalogue_model: str = "",
) -> dict:
    return {
        "row_key": row_key,
        "page": 1,
        "row": 1,
        "title": title,
        "visible_code": sku,
        "brand": "Orbea",
        "variant_count": 2,
        "list_stock": 3,
        "list_status": "Draft",
        "status": status,
        "match_method": "variant SKU prefix" if status == "code_match" else "title only",
        "note": "",
        "product_url": f"https://pim.bo.ultrabike.lt/dashboard/products/{row_key}",
        "product_id": row_key,
        "sku": sku,
        "variant_stock": 2,
        "variant_count_found": 2,
        "candidate_reason": "catalogue model title",
        "catalogue_prefix": sku[:4] if sku else "",
        "catalogue_code": f"{sku[:4]}TTCC" if sku else "",
        "catalogue_model": catalogue_model,
        "catalogue_year": 2026,
        "catalogue_category": "Road",
        "catalogue_subcategory": "Performance",
        "catalogue_url": catalogue_url,
        "catalogue_unique_model_id": row_key,
    }


class CatalogueTests(unittest.TestCase):
    def test_compound_identifier_extracts_ttcc_only(self) -> None:
        self.assertEqual(extract_template_codes("11486 • U107TTCC"), ("U107TTCC",))
        self.assertEqual(
            extract_template_codes("U107TTCC • U107TTCC / T123TTCC"),
            ("U107TTCC", "T123TTCC"),
        )

    def test_workbook_matching_uses_prefix_then_title(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "catalogue.xlsx"
            build_catalogue(path)
            index = CatalogueIndex.from_workbook(path)
            narrowed = index.match("U10707SV", "Orbea ORCA M30i Black")
            self.assertEqual(narrowed.status, "code_match")
            self.assertEqual(narrowed.entry.template_code, "U107TTCC")
            self.assertEqual(narrowed.entry.model, "ORCA M30i")
            ambiguous = index.match("U10707SV", "Unknown Orbea bicycle")
            self.assertEqual(ambiguous.status, "ambiguous")
            title_only = index.match("ZZZZ9999", "Orbea RISE LT H10")
            self.assertEqual(title_only.status, "title_only")

    def test_representative_variant_prefers_first_in_stock_sku(self) -> None:
        chosen = select_representative_variant(
            (
                {"sku": "U10701AA", "stock": 0},
                {"sku": "U10707SV", "stock": "3"},
                {"sku": "U10708SV", "stock": 6},
            )
        )
        self.assertEqual(chosen["sku"], "U10707SV")
        fallback = select_representative_variant(
            ({"sku": "U10701AA", "stock": 0}, {"sku": "U10702AA", "stock": None})
        )
        self.assertEqual(fallback["sku"], "U10701AA")


class CheckpointTests(unittest.TestCase):
    def test_atomic_checkpoint_resume_and_filter_compatibility(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            build_catalogue(catalogue)
            config = OrbeaRunConfig(catalogue, root / "runs")
            run_dir = create_run_directory(
                config.output_root, datetime(2026, 8, 5, 12, 30, 15)
            )
            checkpoint = RunCheckpoint.create(run_dir, config)
            checkpoint.upsert_result(result_row("p1", "unmatched", title="Unknown"))
            checkpoint.mark_cancelled()
            checkpoint.data["compatibility"].pop("download_product_photos")
            checkpoint.save()

            self.assertEqual(find_latest_compatible_run(config), run_dir)
            resumed = RunCheckpoint.load(run_dir, config)
            self.assertTrue(resumed.resumed)
            self.assertEqual(resumed.processed_row_keys(), {"p1"})
            self.assertFalse(resumed.data["cancelled"])

            changed = OrbeaRunConfig(
                catalogue,
                config.output_root,
                filters=PimboFilterSpec(statuses=("Published",), stock="Any"),
            )
            self.assertIsNone(find_latest_compatible_run(changed))

            counts = resumed.counts()
            self.assertEqual(counts["review"], 1)
            self.assertEqual(counts["matched"], 0)
            self.assertEqual(counts["images"], 0)
            self.assertEqual(counts["unavailable"], 0)
            self.assertEqual(counts["errors"], 0)

    def test_new_run_folder_never_overwrites_same_second(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            moment = datetime(2026, 8, 5, 12, 30, 15)
            first = create_run_directory(Path(directory), moment)
            second = create_run_directory(Path(directory), moment)
            self.assertEqual(first.name, "20260805-123015")
            self.assertEqual(second.name, "20260805-123015-2")


class PimboFilterTests(unittest.TestCase):
    def test_apply_filters_sets_every_field_resets_brand_and_returns_to_page_one(self) -> None:
        class SpyClient(PimboBrowserClient):
            def __init__(self):
                super().__init__(object())
                self.calls = []

            def ensure_products_page(self, timeout=20.0):
                self.calls.append(("ensure_products_page",))

            def _set_search(self):
                self.calls.append(("fixed_search", "orbea"))

            def _open_filter_dialog(self):
                self.calls.append(("open_dialog",))
                return object()

            def _close_filter_dialog(self):
                self.calls.append(("close_dialog",))

            def _set_button(self, label, active):
                self.calls.append(("button", label, active))

            def _set_select_value(self, label, value):
                self.calls.append(("select", label, value))

            def go_to_page(self, page_number):
                self.calls.append(("page", page_number))

            def _verify_filters(self, spec):
                self.calls.append(("verify", spec))

        spec = PimboFilterSpec(
            statuses=("Draft", "Published"),
            family_id="family-opaque-id",
            category_id="category-opaque-id",
            source_id="manual",
            stock="Out of stock",
            completeness_locale="en",
            completeness_buckets=("<40%", "100%"),
            sort="Most complete",
        )
        client = SpyClient()
        client.apply_filters(spec)

        self.assertIn(("fixed_search", "orbea"), client.calls)
        self.assertIn(("select", "Brand", ""), client.calls)
        self.assertIn(("select", "Family", "family-opaque-id"), client.calls)
        self.assertIn(("select", "Category", "category-opaque-id"), client.calls)
        self.assertIn(("select", "Source", "manual"), client.calls)
        self.assertIn(("select", "Locale", "en"), client.calls)
        button_calls = {
            (call[1], call[2])
            for call in client.calls
            if len(call) == 3 and call[0] == "button"
        }
        self.assertTrue(
            {
                ("Draft", True),
                ("In Review", False),
                ("Published", True),
                ("Disabled", False),
                ("Out of stock", True),
                ("<40%", True),
                ("40–80%", False),
                ("≥80%", False),
                ("100%", True),
                ("Most complete", True),
            }.issubset(button_calls)
        )
        self.assertIn(("page", 1), client.calls)
        self.assertEqual(client.calls[-1], ("verify", spec))


class ReportTests(unittest.TestCase):
    def test_fallback_row_key_keeps_duplicate_looking_rows_distinct(self) -> None:
        snapshot = {"title": "Orbea Custom", "visible_code": "CUSTOM", "row_href": ""}
        first = PimboBrowserClient._row_key(snapshot, 1, 1)
        second = PimboBrowserClient._row_key(snapshot, 1, 2)
        linked = PimboBrowserClient._row_key(
            {**snapshot, "row_href": "https://pim.bo/products/stable"}, 3, 9
        )
        linked_again = PimboBrowserClient._row_key(
            {**snapshot, "row_href": "https://pim.bo/products/stable"}, 4, 1
        )
        self.assertNotEqual(first, second)
        self.assertEqual(linked, linked_again)

    def test_old_negative_probe_is_reset_once_but_downloads_are_preserved(self) -> None:
        record = {
            "availability_probe_version": 1,
            "geometry_status": "not_available",
            "size_guide_status": "downloaded",
        }
        refresh_required = OrbeaAutomationService._upgrade_probe_record(record, 2)
        self.assertTrue(refresh_required)
        self.assertEqual(record["geometry_status"], "pending")
        self.assertEqual(record["size_guide_status"], "downloaded")
        self.assertEqual(record["availability_probe_version"], 1)
        self.assertTrue(record["probe_refresh_required"])

    def test_retry_failed_processes_a_legacy_transient_probe_once(self) -> None:
        class DummyReporter:
            def emit(self, *_args, **_kwargs):
                pass

        class DummyDriver:
            def __init__(self):
                self.quit_calls = 0

            def quit(self):
                self.quit_calls += 1

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            build_catalogue(catalogue)
            config = OrbeaRunConfig(catalogue, root / "runs")
            checkpoint = RunCheckpoint.create(
                create_run_directory(config.output_root), config
            )
            url = "https://cms.orbea.com/en-be/rise-lt-h10"
            checkpoint.upsert_result(
                result_row(
                    "p1",
                    "code_match",
                    sku="T12309AA",
                    title="Orbea RISE LT H10",
                    catalogue_url=url,
                    catalogue_model="RISE LT H10",
                )
            )
            checkpoint.upsert_image(
                url,
                {
                    "availability_probe_version": 1,
                    "geometry_status": "transient_error",
                    "size_guide_status": "transient_error",
                    "attempts": 8,
                    "retryable": True,
                },
            )
            driver = DummyDriver()
            service = OrbeaAutomationService(
                object(), image_driver_factory=lambda *_args: driver
            )
            capture_result = {
                "availability_probe_version": 2,
                "geometry_status": "not_available",
                "size_guide_status": "not_available",
                "geometry_dimensions": None,
                "size_guide_dimensions": None,
                "geometry_error": "",
                "size_guide_error": "",
                "errors": [],
                "retryable": False,
            }
            with patch(
                "tools.orbea_table_image_downloader.capture_orbea_tables",
                return_value=capture_result,
            ) as capture:
                service._download_images(
                    config,
                    checkpoint,
                    DummyReporter(),
                    CancellationToken(),
                    None,
                    retry_failed=True,
                )

            capture.assert_called_once()
            self.assertEqual(driver.quit_calls, 1)
            record = checkpoint.images[url]
            self.assertEqual(record["geometry_status"], "not_available")
            self.assertEqual(record["size_guide_status"], "not_available")
            self.assertEqual(record["attempts"], 9)

    def test_integrated_run_upgrades_old_single_geometry_capture(self) -> None:
        from tools.orbea_table_image_downloader import (
            AVAILABILITY_PROBE_VERSION,
            GEOMETRY_CAPTURE_VERSION,
        )

        class DummyReporter:
            def emit(self, *_args, **_kwargs):
                pass

        class DummyDriver:
            def quit(self):
                pass

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            build_catalogue(catalogue)
            config = OrbeaRunConfig(catalogue, root / "runs")
            checkpoint = RunCheckpoint.create(
                create_run_directory(config.output_root), config
            )
            url = "https://www.orbea.com/en-be/onna-20"
            checkpoint.upsert_result(
                result_row(
                    "p1",
                    "code_match",
                    sku="ONNA20",
                    title="Orbea Onna 20",
                    catalogue_url=url,
                    catalogue_model="ONNA 20",
                )
            )
            checkpoint.upsert_image(
                url,
                {
                    "availability_probe_version": AVAILABILITY_PROBE_VERSION,
                    "geometry_capture_version": GEOMETRY_CAPTURE_VERSION - 1,
                    "geometry_status": "downloaded",
                    "size_guide_status": "not_available",
                    "attempts": 1,
                    "retryable": False,
                },
            )
            service = OrbeaAutomationService(
                object(), image_driver_factory=lambda *_args: DummyDriver()
            )
            capture_result = {
                "availability_probe_version": AVAILABILITY_PROBE_VERSION,
                "geometry_status": "downloaded",
                "size_guide_status": "not_available",
                "geometry_dimensions": (640, 480),
                "size_guide_dimensions": None,
                "geometry_position": "",
                "geometry_position_supported": False,
                "geometry_size_selector_supported": True,
                "geometry_variants": [
                    {
                        "size": "XS",
                        "wheel_size": '27.5"',
                        "filename": "geometry-xs.png",
                        "status": "downloaded",
                    }
                ],
                "geometry_error": "",
                "size_guide_error": "",
                "errors": [],
                "retryable": False,
            }
            with patch(
                "tools.orbea_table_image_downloader.capture_orbea_tables",
                return_value=capture_result,
            ) as capture:
                service._download_images(
                    config,
                    checkpoint,
                    DummyReporter(),
                    CancellationToken(),
                    None,
                    retry_failed=False,
                )

            capture.assert_called_once()
            record = checkpoint.images[url]
            self.assertEqual(
                record["geometry_capture_version"], GEOMETRY_CAPTURE_VERSION
            )
            self.assertEqual(record["geometry_variants"][0]["size"], "XS")

    def test_product_photo_option_uses_unique_matched_catalogue_links(self) -> None:
        class DummyReporter:
            def __init__(self):
                self.updates = []

            def emit(self, *args):
                self.updates.append(args)

        class FakePhotoService:
            def __init__(self):
                self.calls = []

            def run_many(self, urls, output_dir, **kwargs):
                self.calls.append((tuple(urls), Path(output_dir)))
                kwargs["progress"](
                    SimpleNamespace(
                        current=1000,
                        total=1000,
                        message="Product photos complete",
                    )
                )
                return SimpleNamespace(
                    products=1,
                    variants=3,
                    views=4,
                    files=(Path(output_dir) / "one.png", Path(output_dir) / "two.png"),
                    failures=(),
                    unavailable=("missing-view",),
                    cancelled=False,
                )

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            build_catalogue(catalogue)
            config = OrbeaRunConfig(
                catalogue,
                root / "runs",
                download_product_photos=True,
            )
            checkpoint = RunCheckpoint.create(
                create_run_directory(config.output_root), config
            )
            url = "https://cms.orbea.com/en-be/orca-m30i"
            for row_key, sku in (("p1", "U10707SV"), ("p2", "U10708SV")):
                checkpoint.upsert_result(
                    result_row(
                        row_key,
                        "code_match",
                        sku=sku,
                        title="Orbea ORCA M30i",
                        catalogue_url=url,
                        catalogue_model="ORCA M30i",
                    )
                )

            fake = FakePhotoService()
            reporter = DummyReporter()
            service = OrbeaAutomationService(
                object(), photo_service_factory=lambda: fake
            )
            service._download_product_photos(
                checkpoint, reporter, CancellationToken(), None
            )

            self.assertEqual(fake.calls, [((url,), checkpoint.run_dir / "product-photos")])
            self.assertTrue(checkpoint.data["product_photos_completed"])
            self.assertEqual(checkpoint.data["product_photos"]["products"], 1)
            self.assertEqual(checkpoint.data["product_photos"]["files"], 2)
            self.assertEqual(checkpoint.counts()["product_photos"], 2)
            self.assertEqual(reporter.updates[-1][0], "product_photos")

    def test_formatted_workbook_manifest_and_duplicate_url_job(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            build_catalogue(catalogue)
            config = OrbeaRunConfig(catalogue, root / "runs")
            checkpoint = RunCheckpoint.create(
                create_run_directory(config.output_root), config
            )
            url_a = "https://cms.orbea.com/en-be/orca-m30i"
            url_b = "https://cms.orbea.com/en-be/orca-m30i/?locale=ignored"
            checkpoint.upsert_result(
                result_row(
                    "p1",
                    "code_match",
                    sku="U10707SV",
                    title="Orbea ORCA M30i",
                    catalogue_url=url_a,
                    catalogue_model="ORCA M30i",
                )
            )
            checkpoint.upsert_result(
                result_row(
                    "p2",
                    "code_match",
                    sku="U10708SV",
                    title="Orbea ORCA M30i",
                    catalogue_url=url_b,
                    catalogue_model="ORCA M30i",
                )
            )
            checkpoint.upsert_result(
                result_row("p3", "title_only", sku="BAD01", title="Orbea ORCA")
            )
            checkpoint.upsert_result(
                result_row(
                    "p4",
                    "code_match",
                    sku="T12309AA",
                    title="Orbea RISE Custom Frame",
                    catalogue_url="invalid-url",
                    catalogue_model="RISE Custom",
                )
            )
            canonical = "https://cms.orbea.com/en-be/orca-m30i"
            checkpoint.upsert_image(
                canonical,
                {
                    "canonical_url": canonical,
                    "geometry_status": "downloaded",
                    "geometry_image": "images/orca/geometry.png",
                    "geometry_variants": [
                        {
                            "size": "XS",
                            "wheel_size": '27.5"',
                            "filename": "geometry-xs.png",
                            "status": "downloaded",
                        },
                        {
                            "size": "M",
                            "wheel_size": '29"',
                            "filename": "geometry-m.png",
                            "status": "downloaded",
                        },
                    ],
                    "folder": "images/orca",
                    "size_guide_status": "not_available",
                    "size_guide_image": "images/orca/size-guide-cm.png",
                    "retryable": False,
                    "attempts": 1,
                    "errors": [],
                },
            )

            service = OrbeaAutomationService(object())
            jobs = service._image_jobs(checkpoint)
            self.assertEqual(len(jobs), 1)
            self.assertEqual(jobs[0]["variant_skus"], ["U10707SV", "U10708SV"])

            workbook_path = write_report(checkpoint)
            manifest_path = write_image_manifest(checkpoint)
            workbook = load_workbook(workbook_path, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames, ["Summary", "Matches", "Review", "Raw Scan"]
                )
                self.assertEqual(workbook.active.title, "Matches")
                matches = workbook["Matches"]
                self.assertEqual(matches.freeze_panes, "A2")
                self.assertTrue(matches.auto_filter.ref)
                self.assertFalse(matches.sheet_view.showGridLines)
                header = {cell.value: cell.column for cell in matches[1]}
                self.assertEqual(
                    list(header),
                    [
                        "Variant SKU",
                        "Pimbo Product",
                        "Catalogue Code",
                        "Year",
                        "Orbea URL",
                    ],
                )
                self.assertEqual(matches.max_row, 4)
                self.assertEqual(
                    [matches.cell(row, header["Pimbo Product"]).value for row in range(2, 5)],
                    ["Orbea ORCA M30i", "Orbea ORCA M30i", "Orbea RISE Custom Frame"],
                )
                self.assertTrue(matches.cell(2, header["Orbea URL"]).hyperlink)
                self.assertEqual(workbook["Review"].max_row, 2)
            finally:
                workbook.close()

            with manifest_path.open("r", encoding="utf-8-sig", newline="") as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 3)
            self.assertEqual(rows[0]["Geometry Status"], "downloaded")
            self.assertEqual(rows[0]["Geometry Sizes"], "XS; M")
            self.assertEqual(rows[0]["Geometry Wheel Sizes"], '27.5"; 29"')
            self.assertEqual(
                rows[0]["Geometry PNGs"],
                f"{Path('images/orca/geometry-xs.png')}; "
                f"{Path('images/orca/geometry-m.png')}",
            )
            self.assertEqual(rows[2]["Geometry Status"], "not_available")

    def test_existing_match_excel_is_cleaned_and_sorted(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "orbea_pimbo_variant_matches.xlsx"
            destination = root / "orbea_pimbo_variant_matches_sorted.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Matches"
            worksheet.append(
                (
                    "Pimbo Product",
                    "Unwanted",
                    "Year",
                    "Orbea URL",
                    "Variant SKU",
                    "Catalogue Code",
                    "Catalogue Model",
                )
            )
            worksheet.append(
                ("Pimbo A", "remove", 2025, "https://orbea.test/rise20", "SKU20", "RISE20", "RISE 20")
            )
            worksheet.append(
                ("Pimbo Z", "remove", 2024, "https://orbea.test/orca10-old", "SKU10B", "ORCA10", "ORCA 10")
            )
            worksheet.append(
                ("Pimbo Y", "remove", 2026, "https://orbea.test/orca10", "SKU10A", "ORCA10", "ORCA 10")
            )
            worksheet.append(
                ("Pimbo X", "remove", 2026, "https://orbea.test/orca2", "SKU2", "ORCA2", "ORCA 2")
            )
            workbook.save(source)
            workbook.close()

            result = sort_existing_match_workbook(source, destination)
            self.assertEqual(result, destination)
            sorted_workbook = load_workbook(result)
            try:
                self.assertEqual(sorted_workbook.sheetnames, ["Matches"])
                sheet = sorted_workbook["Matches"]
                self.assertEqual(
                    [cell.value for cell in sheet[1]],
                    [
                        "Variant SKU",
                        "Pimbo Product",
                        "Catalogue Code",
                        "Year",
                        "Orbea URL",
                    ],
                )
                self.assertEqual(
                    [sheet.cell(row, 1).value for row in range(2, 6)],
                    ["SKU2", "SKU10A", "SKU10B", "SKU20"],
                )
                self.assertTrue(sheet["E2"].hyperlink)
            finally:
                sorted_workbook.close()

    def test_cancel_before_browser_work_still_writes_partial_excel(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            catalogue = root / "catalogue.xlsx"
            catalogue.write_bytes(b"catalogue fingerprint only")
            config = OrbeaRunConfig(catalogue, root / "runs")
            cancellation = CancellationToken()
            cancellation.cancel()
            result = OrbeaAutomationService(object()).run(
                config, cancellation=cancellation
            )
            self.assertTrue(result.cancelled)
            self.assertFalse(result.completed)
            self.assertTrue(result.workbook_path.is_file())
            self.assertTrue(result.manifest_path.is_file())
            self.assertTrue(result.checkpoint_path.is_file())


if __name__ == "__main__":
    unittest.main()
