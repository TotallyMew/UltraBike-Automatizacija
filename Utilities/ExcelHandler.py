"""
Excel Handler Utility
Shared Excel import/export functionality for batch screens
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Optional, Tuple
from datetime import datetime


class ExcelHandler:
    """Utility class for Excel import/export operations in batch screens"""

    @staticmethod
    def create_workbook_with_headers(headers: List[str], title: str = "Batch Data") -> openpyxl.Workbook:
        """
        Create a new Excel workbook with formatted headers

        Args:
            headers: List of column header names
            title: Sheet title

        Returns:
            openpyxl.Workbook object
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

        # Write headers with formatting
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        return wb

    @staticmethod
    def export_table_data(headers: List[str], rows_data: List[Dict[str, any]],
                         filename: str, title: str = "Batch Data") -> bool:
        """
        Export table data to Excel file

        Args:
            headers: List of column header names
            rows_data: List of dictionaries containing row data
            filename: Output filename
            title: Sheet title

        Returns:
            True if successful, False otherwise
        """
        try:
            wb = ExcelHandler.create_workbook_with_headers(headers, title)
            ws = wb.active

            # Write data rows
            for row_idx, row_data in enumerate(rows_data, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    # Use header key (lowercase) to get value from row_data
                    key = header.lower().replace(' ', '_')
                    value = row_data.get(key, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return False

    @staticmethod
    def export_failed_items(headers: List[str], failed_items: List[Dict[str, any]],
                           filename: str) -> bool:
        """
        Export failed items with error details to Excel

        Args:
            headers: List of column header names (should include 'Status' and 'Error')
            failed_items: List of dictionaries containing failed row data with errors
            filename: Output filename

        Returns:
            True if successful, False otherwise
        """
        try:
            wb = ExcelHandler.create_workbook_with_headers(headers, "Failed Items")
            ws = wb.active

            # Write failed items with error highlighting
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row_idx, item in enumerate(failed_items, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    key = header.lower().replace(' ', '_')
                    value = item.get(key, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Highlight error column
                    if header.lower() == 'error' and value:
                        cell.fill = error_fill

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            return True
        except Exception as e:
            print(f"Error exporting failed items: {e}")
            return False

    @staticmethod
    def read_excel_with_preview(filename: str, max_preview_rows: int = 5) -> Tuple[bool, List[str], List[List[any]], List[List[any]]]:
        """
        Read Excel file and return headers, preview data, and full data

        Args:
            filename: Excel file path
            max_preview_rows: Maximum number of rows to include in preview

        Returns:
            Tuple of (success, headers, preview_data, full_data)
            - success: bool indicating if read was successful
            - headers: List of column headers
            - preview_data: First N rows for preview
            - full_data: All data rows
        """
        try:
            wb = openpyxl.load_workbook(filename, data_only=True)
            ws = wb.active

            # Read headers (first row)
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else "")

            # Read all data rows
            full_data = []
            preview_data = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                # Convert row tuple to list, handling None values
                row_list = [str(cell) if cell is not None else "" for cell in row]
                full_data.append(row_list)

                # Add to preview if within limit
                if row_idx < max_preview_rows:
                    preview_data.append(row_list)

            return True, headers, preview_data, full_data

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return False, [], [], []

    @staticmethod
    def map_columns(excel_headers: List[str], expected_headers: List[str]) -> Dict[int, int]:
        """
        Create a mapping between Excel columns and expected columns

        Args:
            excel_headers: Headers from the Excel file
            expected_headers: Expected headers in the correct order

        Returns:
            Dictionary mapping expected column index to Excel column index
            Returns empty dict if critical columns are missing
        """
        mapping = {}

        # Normalize headers for comparison (lowercase, no spaces)
        def normalize(header: str) -> str:
            return header.lower().strip().replace(' ', '_')

        normalized_excel = {normalize(h): idx for idx, h in enumerate(excel_headers)}

        for expected_idx, expected_header in enumerate(expected_headers):
            normalized_expected = normalize(expected_header)

            if normalized_expected in normalized_excel:
                mapping[expected_idx] = normalized_excel[normalized_expected]

        return mapping

    @staticmethod
    def remap_row_data(row_data: List[any], column_mapping: Dict[int, int],
                      num_expected_columns: int) -> List[any]:
        """
        Remap row data according to column mapping

        Args:
            row_data: Original row data from Excel
            column_mapping: Mapping from expected index to Excel index
            num_expected_columns: Number of expected columns

        Returns:
            List of values in the expected order
        """
        result = [""] * num_expected_columns

        for expected_idx, excel_idx in column_mapping.items():
            if excel_idx < len(row_data):
                result[expected_idx] = row_data[excel_idx]

        return result

    @staticmethod
    def validate_row(row_data: List[any], required_columns: List[int]) -> Tuple[bool, str]:
        """
        Validate that required columns have values

        Args:
            row_data: Row data to validate
            required_columns: List of column indices that must have values

        Returns:
            Tuple of (is_valid, error_message)
        """
        for col_idx in required_columns:
            if col_idx < len(row_data):
                value = str(row_data[col_idx]).strip()
                if not value or value.lower() in ('none', 'null', ''):
                    return False, f"Missing required value in column {col_idx + 1}"
            else:
                return False, f"Missing column {col_idx + 1}"

        return True, ""
