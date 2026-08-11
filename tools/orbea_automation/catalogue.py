from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import load_workbook


IDENTIFIER_RE = re.compile(r"\b([A-Z][A-Z0-9]*TTCC)\b", re.IGNORECASE)
SKU_RE = re.compile(r"\b([A-Z][A-Z0-9]{4,})\b", re.IGNORECASE)
YEAR_SUFFIX_RE = re.compile(r"\s+20\d{2}$")


def normalize_code(value: Any) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(value or "").upper())


def extract_template_codes(value: Any) -> tuple[str, ...]:
    """Extract unique ``...TTCC`` identifiers from a compound catalogue cell."""

    return tuple(
        dict.fromkeys(match.group(1).upper() for match in IDENTIFIER_RE.finditer(str(value or "")))
    )


def first_code_token(value: Any) -> str:
    match = SKU_RE.search(str(value or "").upper())
    return normalize_code(match.group(1)) if match else ""


def normalize_name(value: Any) -> str:
    text = str(value or "").upper()
    text = re.sub(r"\bORBEA\b", " ", text)
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def model_key(value: Any) -> str:
    return YEAR_SUFFIX_RE.sub("", normalize_name(value))


def parse_number(value: Any) -> float | int | None:
    if value is None:
        return None
    text = str(value).strip().replace("\u00a0", "").replace(",", ".")
    if not text or text in {"—", "-", "N/A"}:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    return int(number) if number.is_integer() else number


@dataclass(frozen=True)
class CatalogueEntry:
    prefix: str
    template_code: str
    model: str
    year: int | None = None
    category: str = ""
    subcategory: str = ""
    product_link: str = ""
    identifiers: str = ""
    unique_model_id: str = ""
    regional_listings: int = 0

    @property
    def model_key(self) -> str:
        return model_key(self.model)


@dataclass(frozen=True)
class MatchResult:
    status: str
    method: str
    entry: CatalogueEntry | None
    note: str = ""

    def catalogue_fields(self) -> dict[str, Any]:
        if self.entry is None:
            return {
                "catalogue_prefix": "",
                "catalogue_code": "",
                "catalogue_model": "",
                "catalogue_year": None,
                "catalogue_category": "",
                "catalogue_subcategory": "",
                "catalogue_url": "",
                "catalogue_unique_model_id": "",
            }
        entry = self.entry
        return {
            "catalogue_prefix": entry.prefix,
            "catalogue_code": entry.template_code,
            "catalogue_model": entry.model,
            "catalogue_year": entry.year,
            "catalogue_category": entry.category,
            "catalogue_subcategory": entry.subcategory,
            "catalogue_url": entry.product_link,
            "catalogue_unique_model_id": entry.unique_model_id,
        }


class CatalogueIndex:
    def __init__(self, entries: Iterable[CatalogueEntry]):
        entries = tuple(entries)
        if not entries:
            raise ValueError("No catalogue identifiers ending in TTCC were found")

        grouped: dict[str, dict[str, list[CatalogueEntry]]] = {}
        for entry in entries:
            grouped.setdefault(entry.prefix, {}).setdefault(entry.model_key, []).append(entry)

        self.by_prefix: dict[str, list[CatalogueEntry]] = {}
        for prefix, model_groups in grouped.items():
            canonical: list[CatalogueEntry] = []
            for candidates in model_groups.values():
                canonical.append(
                    max(
                        candidates,
                        key=lambda item: (
                            item.regional_listings,
                            item.year is not None,
                            item.year or 0,
                            bool(item.product_link),
                        ),
                    )
                )
            self.by_prefix[prefix] = canonical

        self.prefixes = sorted(self.by_prefix, key=len, reverse=True)
        self.by_model: dict[str, list[CatalogueEntry]] = {}
        for candidates in self.by_prefix.values():
            for entry in candidates:
                self.by_model.setdefault(entry.model_key, []).append(entry)
        self.model_keys = sorted(self.by_model, key=lambda value: (-len(value), value))

    @classmethod
    def from_workbook(cls, path: Path) -> "CatalogueIndex":
        path = Path(path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            if "Unique Models" not in workbook.sheetnames:
                raise ValueError("The catalogue has no 'Unique Models' worksheet")
            worksheet = workbook["Unique Models"]
            rows = worksheet.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration as error:
                raise ValueError("The catalogue's Unique Models sheet is empty") from error

            headers = {
                str(name).strip(): index
                for index, name in enumerate(header_row)
                if name is not None and str(name).strip()
            }
            required = {
                "Model",
                "Year",
                "Category",
                "Subcategory",
                "Regional Listings",
                "Product Link",
                "Other Links",
                "Identifiers",
                "Unique Model ID",
            }
            missing = sorted(required - set(headers))
            if missing:
                raise ValueError(f"Catalogue columns are missing: {', '.join(missing)}")

            def get(row: Sequence[Any], name: str) -> Any:
                index = headers[name]
                return row[index] if index < len(row) else None

            entries: list[CatalogueEntry] = []
            for row in rows:
                identifiers = str(get(row, "Identifiers") or "")
                product_link = str(get(row, "Product Link") or "").strip()
                if not product_link.lower().startswith(("http://", "https://")):
                    other_links = re.findall(
                        r"https?://[^\s\u2022]+", str(get(row, "Other Links") or "")
                    )
                    product_link = next(
                        (url for url in other_links if "/catalog" not in url.lower()),
                        other_links[0] if other_links else "",
                    )
                year_value = parse_number(get(row, "Year"))
                for template_code in extract_template_codes(identifiers):
                    entries.append(
                        CatalogueEntry(
                            prefix=template_code[:-4],
                            template_code=template_code,
                            model=str(get(row, "Model") or "").strip(),
                            year=int(year_value) if year_value is not None else None,
                            category=str(get(row, "Category") or "").strip(),
                            subcategory=str(get(row, "Subcategory") or "").strip(),
                            product_link=product_link,
                            identifiers=identifiers.strip(),
                            unique_model_id=str(get(row, "Unique Model ID") or "").strip(),
                            regional_listings=int(
                                parse_number(get(row, "Regional Listings")) or 0
                            ),
                        )
                    )
        finally:
            workbook.close()
        return cls(entries)

    @staticmethod
    def _dedupe_entries(entries: Iterable[CatalogueEntry]) -> list[CatalogueEntry]:
        unique: dict[tuple[str, str], CatalogueEntry] = {}
        for entry in entries:
            key = (entry.prefix, entry.model_key)
            incumbent = unique.get(key)
            if incumbent is None or entry.regional_listings > incumbent.regional_listings:
                unique[key] = entry
        return list(unique.values())

    def title_candidates(self, title: str) -> list[CatalogueEntry]:
        normalized = normalize_name(title)
        matches: list[CatalogueEntry] = []
        best_length = 0
        for key in self.model_keys:
            if normalized == key or normalized.startswith(f"{key} "):
                if len(key) > best_length:
                    matches = list(self.by_model[key])
                    best_length = len(key)
                elif len(key) == best_length:
                    matches.extend(self.by_model[key])
        return self._dedupe_entries(matches)

    def code_candidates(self, sku: str) -> list[CatalogueEntry]:
        code = normalize_code(sku)
        matching_prefixes = [prefix for prefix in self.prefixes if code.startswith(prefix)]
        if not matching_prefixes:
            return []
        longest = len(matching_prefixes[0])
        candidates: list[CatalogueEntry] = []
        for prefix in matching_prefixes:
            if len(prefix) != longest:
                break
            candidates.extend(self.by_prefix[prefix])
        return self._dedupe_entries(candidates)

    def is_likely_bicycle(self, visible_code: str, title: str) -> tuple[bool, str]:
        if self.code_candidates(first_code_token(visible_code)):
            return True, "catalogue code prefix"
        if self.title_candidates(title):
            return True, "catalogue model title"
        return False, "not in bicycle catalogue"

    def match(self, sku: str, title: str) -> MatchResult:
        code_matches = self.code_candidates(sku)
        if len(code_matches) == 1:
            return MatchResult("code_match", "variant SKU prefix", code_matches[0])
        if len(code_matches) > 1:
            title_matches = {
                (entry.prefix, entry.model_key): entry
                for entry in self.title_candidates(title)
            }
            narrowed = [
                entry
                for entry in code_matches
                if (entry.prefix, entry.model_key) in title_matches
            ]
            if len(narrowed) == 1:
                return MatchResult(
                    "code_match", "variant SKU prefix + title", narrowed[0]
                )
            return MatchResult(
                "ambiguous",
                "variant SKU prefix",
                None,
                "More than one catalogue model uses this code prefix",
            )

        title_matches = self.title_candidates(title)
        if len(title_matches) == 1:
            return MatchResult(
                "title_only",
                "title only",
                title_matches[0],
                "Model title matches, but the variant SKU prefix does not",
            )
        if len(title_matches) > 1:
            return MatchResult(
                "ambiguous",
                "title only",
                None,
                "More than one catalogue model matches this title",
            )
        return MatchResult(
            "unmatched",
            "none",
            None,
            "No catalogue code prefix or model title matched",
        )


def select_representative_variant(
    variants: Iterable[Mapping[str, Any]],
) -> Mapping[str, Any] | None:
    """Prefer the first in-stock SKU, then the first valid SKU."""

    valid = [item for item in variants if normalize_code(item.get("sku", ""))]
    if not valid:
        return None
    return next(
        (
            item
            for item in valid
            if (parse_number(item.get("stock")) or 0) > 0
        ),
        valid[0],
    )

