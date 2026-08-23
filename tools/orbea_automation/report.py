from __future__ import annotations

import csv
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .checkpoint import RunCheckpoint, utc_now
from .utils import canonicalize_url


NAVY = "17365D"
BLUE = "2F75B5"
PALE_BLUE = "D9EAF7"
PALE_GREEN = "E2F0D9"
PALE_AMBER = "FFF2CC"
PALE_RED = "FCE4D6"
PALE_GREY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "1F2937"
LIGHT_BORDER = Side(style="thin", color="D9E2F3")


MATCH_HEADERS = (
    "Variant SKU",
    "Pimbo Product",
    "Catalogue Code",
    "Year",
    "Orbea URL",
)

REVIEW_HEADERS = (
    "Status",
    "Variant SKU",
    "Pimbo Product",
    "Visible Code",
    "Stock",
    "Match Method",
    "Note",
    "Pimbo URL",
    "Catalogue Candidate",
    "Orbea URL",
    "Page",
)

RAW_HEADERS = (
    "Page",
    "Row",
    "Pimbo Product",
    "Visible Code",
    "Brand",
    "Pimbo Status",
    "Listed Stock",
    "Listed Variants",
    "Variant SKU",
    "Variant Stock",
    "Variants Found",
    "Product ID",
    "Pimbo URL",
    "Candidate Reason",
    "Result Status",
    "Match Method",
    "Note",
    "Catalogue Code",
    "Catalogue Model",
    "Orbea URL",
)


def _image_record(checkpoint: RunCheckpoint, result: dict[str, Any]) -> dict[str, Any]:
    record = checkpoint.images.get(
        canonicalize_url(result.get("catalogue_url", "")), {}
    )
    if record:
        return record
    return {
        "geometry_status": result.get("geometry_status", "pending"),
        "size_guide_status": result.get("size_guide_status", "pending"),
        "errors": [result.get("image_note", "")] if result.get("image_note") else [],
    }


def _record_error(record: dict[str, Any]) -> str:
    values: list[str] = []
    for key in ("errors", "geometry_error", "size_guide_error", "error"):
        value = record.get(key)
        if isinstance(value, (list, tuple)):
            values.extend(str(item) for item in value if item)
        elif value:
            values.append(str(value))
    return " | ".join(dict.fromkeys(values))


def _natural_sort_key(value: Any) -> tuple[tuple[int, Any], ...]:
    """Return a case-insensitive key that keeps embedded numbers intuitive."""

    parts = re.split(r"(\d+)", str(value or "").strip().casefold())
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part)
        for part in parts
        if part
    )


def _year_sort_key(value: Any) -> tuple[int, float]:
    try:
        return (0, -float(value))
    except (TypeError, ValueError):
        return (1, 0)


def _match_row_sort_key(
    row: Sequence[Any], catalogue_model: Any = ""
) -> tuple[Any, ...]:
    return (
        _natural_sort_key(catalogue_model or row[1]),
        _year_sort_key(row[3]),
        _natural_sort_key(row[2]),
        _natural_sort_key(row[0]),
    )


def _matches_rows(checkpoint: RunCheckpoint) -> Iterable[tuple[Any, ...]]:
    rows: list[tuple[tuple[Any, ...], Any]] = []
    for result in checkpoint.results:
        if result.get("status") != "code_match":
            continue
        rows.append(
            (
                (
                    result.get("sku", ""),
                    result.get("title", ""),
                    result.get("catalogue_code", ""),
                    result.get("catalogue_year"),
                    result.get("catalogue_url", ""),
                ),
                result.get("catalogue_model", ""),
            )
        )
    rows.sort(key=lambda item: _match_row_sort_key(item[0], item[1]))
    return [row for row, _model in rows]


def _review_rows(checkpoint: RunCheckpoint) -> Iterable[tuple[Any, ...]]:
    for result in checkpoint.results:
        if result.get("status") in {"code_match", "excluded"}:
            continue
        yield (
            result.get("status", ""),
            result.get("sku", ""),
            result.get("title", ""),
            result.get("visible_code", ""),
            result.get("variant_stock"),
            result.get("match_method", ""),
            result.get("note", ""),
            result.get("product_url", ""),
            result.get("catalogue_model", ""),
            result.get("catalogue_url", ""),
            result.get("page"),
        )


def _raw_rows(checkpoint: RunCheckpoint) -> Iterable[tuple[Any, ...]]:
    for result in checkpoint.results:
        yield (
            result.get("page"),
            result.get("row"),
            result.get("title", ""),
            result.get("visible_code", ""),
            result.get("brand", ""),
            result.get("list_status", ""),
            result.get("list_stock"),
            result.get("variant_count"),
            result.get("sku", ""),
            result.get("variant_stock"),
            result.get("variant_count_found"),
            result.get("product_id", ""),
            result.get("product_url", ""),
            result.get("candidate_reason", ""),
            result.get("status", ""),
            result.get("match_method", ""),
            result.get("note", ""),
            result.get("catalogue_code", ""),
            result.get("catalogue_model", ""),
            result.get("catalogue_url", ""),
        )


def _style_header(worksheet, row: int = 1) -> None:
    for cell in worksheet[row]:
        if cell.value is None:
            continue
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=BLUE))
    worksheet.row_dimensions[row].height = 30


def _widths(worksheet, widths: Sequence[float]) -> None:
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _make_data_sheet(
    workbook: Workbook,
    name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    widths: Sequence[float],
) -> Any:
    worksheet = workbook.create_sheet(name)
    worksheet.sheet_view.showGridLines = False
    worksheet.append(tuple(headers))
    for row in rows:
        worksheet.append(tuple(row))
    _style_header(worksheet)
    _widths(worksheet, widths)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(worksheet.max_row, 1)}"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=TEXT)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
    return worksheet


def _add_web_links(worksheet, header_names: Sequence[str]) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for name in header_names:
        column = headers.get(name)
        if not column:
            continue
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, column)
            if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                cell.hyperlink = cell.value
                cell.style = "Hyperlink"


def _format_matches_sheet(worksheet) -> None:
    """Keep long product names legible without making short rows oversized."""

    for row_index in range(2, worksheet.max_row + 1):
        product_cell = worksheet.cell(row_index, 2)
        product_cell.alignment = Alignment(vertical="top", wrap_text=True)
        line_count = max(1, (len(str(product_cell.value or "")) + 57) // 58)
        worksheet.row_dimensions[row_index].height = min(45, max(18, line_count * 15))
        worksheet.cell(row_index, 4).number_format = "0"


def _add_file_links(worksheet, header_names: Sequence[str], run_dir: Path) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    for name in header_names:
        column = headers.get(name)
        if not column:
            continue
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row, column)
            if not cell.value:
                continue
            path = Path(str(cell.value))
            if not path.is_absolute():
                path = run_dir / path
            cell.hyperlink = path.resolve().as_uri()
            cell.style = "Hyperlink"


def _add_status_formatting(worksheet, header: str) -> None:
    headers = {cell.value: cell.column for cell in worksheet[1]}
    column = headers.get(header)
    if not column or worksheet.max_row < 2:
        return
    letter = get_column_letter(column)
    target = f"{letter}2:{letter}{worksheet.max_row}"
    for status, color in {
        "code_match": PALE_GREEN,
        "downloaded": PALE_GREEN,
        "not_available": PALE_GREY,
        "title_only": PALE_AMBER,
        "ambiguous": PALE_AMBER,
        "unmatched": PALE_RED,
        "no_variant": PALE_RED,
        "error": PALE_RED,
        "transient_error": PALE_RED,
    }.items():
        worksheet.conditional_formatting.add(
            target,
            FormulaRule(
                formula=[f'LOWER(${letter}2)="{status}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )


def _summary_sheet(workbook: Workbook, checkpoint: RunCheckpoint) -> Any:
    worksheet = workbook.active
    worksheet.title = "Summary"
    worksheet.sheet_view.showGridLines = False
    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = "Orbea Pimbo Match Report"
    worksheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    worksheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34

    compatibility = checkpoint.data.get("compatibility", {})
    filters = compatibility.get("filters", {})
    result_counts = Counter(str(item.get("status") or "unknown") for item in checkpoint.results)
    image_counts = Counter()
    for record in checkpoint.images.values():
        image_counts[f"Geometry: {record.get('geometry_status', 'pending')}"] += 1
        image_counts[f"Size guide: {record.get('size_guide_status', 'pending')}"] += 1

    summary_rows: list[tuple[str, Any]] = [
        ("Generated (UTC)", utc_now()),
        ("Run", checkpoint.data.get("run_id", checkpoint.run_dir.name)),
        ("Run status", "Completed" if checkpoint.data.get("completed") else "Partial"),
        ("Search", compatibility.get("search", "orbea")),
        ("Status filters", ", ".join(filters.get("statuses", [])) or "Any"),
        ("Stock", filters.get("stock", "Any")),
        ("Family ID", filters.get("family_id", "") or "All families"),
        ("Category ID", filters.get("category_id", "") or "All categories"),
        ("Source ID", filters.get("source_id", "") or "All sources"),
        ("Completeness locale", filters.get("completeness_locale", "Overall")),
        (
            "Completeness buckets",
            ", ".join(filters.get("completeness_buckets", [])) or "Any",
        ),
        ("Sort", filters.get("sort", "Recent")),
        ("Products scanned", len(checkpoint.results)),
        ("Confirmed matches", result_counts["code_match"]),
        (
            "Needs review",
            sum(
                result_counts[name]
                for name in ("title_only", "ambiguous", "unmatched", "no_variant", "error")
            ),
        ),
        ("Excluded", result_counts["excluded"]),
        ("Unique matched Orbea URLs", len(checkpoint.images)),
    ]
    summary_rows.extend(sorted(image_counts.items()))

    worksheet["A3"] = "Run and filters"
    worksheet["A3"].font = Font(name="Aptos", size=12, bold=True, color=NAVY)
    for row_index, (label, value) in enumerate(summary_rows, start=4):
        worksheet.cell(row_index, 1, label)
        worksheet.cell(row_index, 2, value)
        worksheet.cell(row_index, 1).font = Font(name="Aptos", bold=True, color=TEXT)
        worksheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=PALE_BLUE)
        worksheet.cell(row_index, 1).border = Border(bottom=LIGHT_BORDER)
        worksheet.cell(row_index, 2).border = Border(bottom=LIGHT_BORDER)
        worksheet.cell(row_index, 2).alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.column_dimensions["A"].width = 27
    worksheet.column_dimensions["B"].width = 54
    worksheet.column_dimensions["C"].width = 3
    worksheet.column_dimensions["D"].width = 3
    worksheet.freeze_panes = "A4"
    return worksheet


def write_report(checkpoint: RunCheckpoint) -> Path:
    """Build the four-sheet formatted report and atomically replace the prior copy."""

    workbook = Workbook()
    _summary_sheet(workbook, checkpoint)
    matches = _make_data_sheet(
        workbook,
        "Matches",
        MATCH_HEADERS,
        _matches_rows(checkpoint),
        (18, 58, 20, 10, 52),
    )
    review = _make_data_sheet(
        workbook,
        "Review",
        REVIEW_HEADERS,
        _review_rows(checkpoint),
        (16, 16, 34, 16, 10, 22, 46, 42, 28, 42, 8),
    )
    raw = _make_data_sheet(
        workbook,
        "Raw Scan",
        RAW_HEADERS,
        _raw_rows(checkpoint),
        (8, 8, 36, 16, 12, 15, 12, 14, 16, 12, 14, 20, 42, 24, 16, 22, 44, 18, 28, 42),
    )

    _add_web_links(matches, ("Orbea URL",))
    _format_matches_sheet(matches)
    _add_web_links(review, ("Pimbo URL", "Orbea URL"))
    _add_web_links(raw, ("Pimbo URL", "Orbea URL"))
    _add_status_formatting(review, "Status")
    _add_status_formatting(raw, "Result Status")
    workbook.active = workbook.index(matches)

    destination = checkpoint.workbook_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.stem}.tmp{destination.suffix}")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, destination)
    return destination


def _next_sorted_path(source: Path) -> Path:
    base = source.with_name(f"{source.stem}_sorted{source.suffix}")
    if not base.exists():
        return base
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return source.with_name(f"{source.stem}_sorted_{timestamp}{source.suffix}")


def _header_lookup(worksheet) -> dict[str, int]:
    return {
        re.sub(r"\s+", " ", str(cell.value or "").strip()).casefold(): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }


def sort_existing_match_workbook(
    source_path: Path | str,
    destination_path: Path | str | None = None,
) -> Path:
    """Create a clean, sorted five-column copy of an existing match workbook."""

    source = Path(source_path)
    if not source.is_file() or source.suffix.lower() != ".xlsx":
        raise ValueError("Choose an existing .xlsx file.")
    destination = Path(destination_path) if destination_path else _next_sorted_path(source)

    from openpyxl import load_workbook

    original = load_workbook(source, data_only=False, read_only=False)
    try:
        worksheet = original["Matches"] if "Matches" in original.sheetnames else original.active
        lookup = _header_lookup(worksheet)
        missing = [header for header in MATCH_HEADERS if header.casefold() not in lookup]
        if missing:
            raise ValueError(
                "The Excel file is missing required columns: " + ", ".join(missing)
            )

        indexes = [lookup[header.casefold()] for header in MATCH_HEADERS]
        model_column = lookup.get("catalogue model") or lookup.get("model")
        sortable_rows = []
        for row_index in range(2, worksheet.max_row + 1):
            row = tuple(worksheet.cell(row_index, column).value for column in indexes)
            if not any(value not in (None, "") for value in row):
                continue
            model = (
                worksheet.cell(row_index, model_column).value if model_column else ""
            )
            sortable_rows.append((row, model))
        sortable_rows.sort(key=lambda item: _match_row_sort_key(item[0], item[1]))
        rows = [row for row, _model in sortable_rows]
    finally:
        original.close()

    output = Workbook()
    output.remove(output.active)
    matches = _make_data_sheet(
        output,
        "Matches",
        MATCH_HEADERS,
        rows,
        (18, 58, 20, 10, 52),
    )
    _add_web_links(matches, ("Orbea URL",))
    _format_matches_sheet(matches)

    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.stem}.tmp{destination.suffix}")
    output.save(temporary)
    output.close()
    os.replace(temporary, destination)
    return destination


def write_image_manifest(checkpoint: RunCheckpoint) -> Path:
    headers = (
        "Variant SKU",
        "Pimbo Product",
        "Catalogue Code",
        "Catalogue Model",
        "Pimbo URL",
        "Orbea URL",
        "Geometry Status",
        "Geometry PNG",
        "Geometry Sizes",
        "Geometry Wheel Sizes",
        "Geometry PNGs",
        "Size Guide Status",
        "Size Guide CM PNG",
        "Retryable",
        "Attempts",
        "Image Error",
    )
    destination = checkpoint.manifest_path
    temporary = destination.with_name(f".{destination.name}.tmp")
    destination.parent.mkdir(parents=True, exist_ok=True)
    with temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for result in checkpoint.results:
            if result.get("status") != "code_match":
                continue
            image = _image_record(checkpoint, result)
            geometry_variants = list(image.get("geometry_variants", []) or [])
            geometry_folder = Path(str(image.get("folder", "")))
            writer.writerow(
                {
                    "Variant SKU": result.get("sku", ""),
                    "Pimbo Product": result.get("title", ""),
                    "Catalogue Code": result.get("catalogue_code", ""),
                    "Catalogue Model": result.get("catalogue_model", ""),
                    "Pimbo URL": result.get("product_url", ""),
                    "Orbea URL": result.get("catalogue_url", ""),
                    "Geometry Status": image.get("geometry_status", "pending"),
                    "Geometry PNG": image.get("geometry_image", ""),
                    "Geometry Sizes": "; ".join(
                        str(variant.get("size", "")).strip()
                        for variant in geometry_variants
                        if str(variant.get("size", "")).strip()
                    ),
                    "Geometry Wheel Sizes": "; ".join(
                        str(variant.get("wheel_size", "")).strip()
                        for variant in geometry_variants
                        if str(variant.get("wheel_size", "")).strip()
                    ),
                    "Geometry PNGs": "; ".join(
                        str(geometry_folder / str(variant.get("filename", "")))
                        for variant in geometry_variants
                        if str(variant.get("filename", "")).strip()
                    ),
                    "Size Guide Status": image.get("size_guide_status", "pending"),
                    "Size Guide CM PNG": image.get(
                        "size_guide_image", image.get("size_guide_cm_image", "")
                    ),
                    "Retryable": bool(image.get("retryable")),
                    "Attempts": image.get("attempts", 0),
                    "Image Error": _record_error(image),
                }
            )
    os.replace(temporary, destination)
    return destination
